# -*- coding: utf-8 -*-
import sqlite3, sys, os
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "mcnex_ai.db")
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "엠씨넥스_기업분석보고서.xlsx")
conn = sqlite3.connect(DB)

wb = Workbook()

# === STYLE DEFINITIONS ===
NAVY = "1B2A4A"
DARK_BLUE = "2C3E6B"
MID_BLUE = "3A5BA0"
LIGHT_BLUE = "D6E4F0"
LIGHTER_BLUE = "EBF1F8"
WHITE = "FFFFFF"
ACCENT_GOLD = "D4A843"
ACCENT_RED = "C0392B"
ACCENT_GREEN = "27AE60"
LIGHT_GRAY = "F2F2F2"
MED_GRAY = "D9D9D9"

title_font = Font(name="맑은 고딕", size=22, bold=True, color=WHITE)
subtitle_font = Font(name="맑은 고딕", size=11, color="B0C4DE")
section_font = Font(name="맑은 고딕", size=14, bold=True, color=NAVY)
header_font = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
data_font = Font(name="맑은 고딕", size=10)
data_font_bold = Font(name="맑은 고딕", size=10, bold=True)
blue_font = Font(name="맑은 고딕", size=10, color="0000FF")
green_font = Font(name="맑은 고딕", size=10, color=ACCENT_GREEN)
red_font = Font(name="맑은 고딕", size=10, color=ACCENT_RED)
pct_font_green = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT_GREEN)
pct_font_red = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT_RED)
small_font = Font(name="맑은 고딕", size=9, color="666666")

title_fill = PatternFill("solid", fgColor=NAVY)
header_fill = PatternFill("solid", fgColor=DARK_BLUE)
sub_header_fill = PatternFill("solid", fgColor=MID_BLUE)
light_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
lighter_fill = PatternFill("solid", fgColor=LIGHTER_BLUE)
gray_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
white_fill = PatternFill("solid", fgColor=WHITE)
gold_fill = PatternFill("solid", fgColor="FFF3CD")
red_fill = PatternFill("solid", fgColor="F8D7DA")
green_fill = PatternFill("solid", fgColor="D4EDDA")

thin_border = Border(
    left=Side(style='thin', color=MED_GRAY),
    right=Side(style='thin', color=MED_GRAY),
    top=Side(style='thin', color=MED_GRAY),
    bottom=Side(style='thin', color=MED_GRAY)
)
bottom_border = Border(bottom=Side(style='medium', color=NAVY))

center_al = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_al = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_al = Alignment(horizontal='right', vertical='center')
wrap_al = Alignment(vertical='top', wrap_text=True)

NUM_FMT = '#,##0'
NUM_FMT_BILL = '#,##0'
PCT_FMT = '0.0%'
PCT_FMT2 = '0.00%'

def style_range(ws, row, col_start, col_end, font=None, fill=None, alignment=None, border=None, number_format=None):
    for c in range(col_start, col_end+1):
        cell = ws.cell(row=row, column=c)
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        if border: cell.border = border
        if number_format: cell.number_format = number_format

def write_header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start+i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_al
        cell.border = thin_border

def write_data_row(ws, row, data, col_start=1, fonts=None, fills=None, alignments=None, number_formats=None):
    for i, d in enumerate(data):
        cell = ws.cell(row=row, column=col_start+i, value=d)
        cell.font = (fonts[i] if fonts and i < len(fonts) else data_font)
        cell.fill = (fills[i] if fills and i < len(fills) else white_fill)
        cell.alignment = (alignments[i] if alignments and i < len(alignments) else right_al)
        cell.border = thin_border
        if number_formats and i < len(number_formats) and number_formats[i]:
            cell.number_format = number_formats[i]

def set_col_widths(ws, widths):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

def add_section_title(ws, row, title, col_end=11):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = section_font
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = bottom_border
    return row + 1

# ============================================================
# SHEET 1: 표지 (Cover)
# ============================================================
ws1 = wb.active
ws1.title = "표지"
ws1.sheet_properties.tabColor = NAVY
set_col_widths(ws1, [3, 20, 20, 20, 20, 20, 3])

for r in range(1, 35):
    for c in range(1, 8):
        ws1.cell(row=r, column=c).fill = title_fill

ws1.merge_cells('B6:F6')
ws1.cell(row=6, column=2, value="엠씨넥스(MCNEX)").font = Font(name="맑은 고딕", size=32, bold=True, color=WHITE)
ws1.cell(row=6, column=2).alignment = Alignment(horizontal='center', vertical='center')

ws1.merge_cells('B8:F8')
ws1.cell(row=8, column=2, value="심층 기업분석 보고서").font = Font(name="맑은 고딕", size=20, color=ACCENT_GOLD)
ws1.cell(row=8, column=2).alignment = Alignment(horizontal='center')

ws1.merge_cells('B11:F11')
ws1.cell(row=11, column=2, value="종목코드: 097520 (유가증권시장)  |  업종: 기타 영상기기 제조업").font = subtitle_font
ws1.cell(row=11, column=2).alignment = Alignment(horizontal='center')

ws1.merge_cells('B12:F12')
ws1.cell(row=12, column=2, value="CCM(카메라모듈) 기술 기반 영상전문기업").font = subtitle_font
ws1.cell(row=12, column=2).alignment = Alignment(horizontal='center')

info_data = [
    (15, "대표이사", "민동욱"),
    (16, "설립일", "2004년 12월 22일"),
    (17, "본사", "인천 연수구 송도과학로16번길 13-39 엠씨넥스타워"),
    (18, "시장구분", "유가증권시장 (2021.07 코스닥→이전상장)"),
    (19, "신용등급", "A- (나이스평가정보, 2024.05)"),
    (20, "주요고객", "삼성전자(모바일), 현대모비스(전장)"),
    (21, "생산기지", "한국(본사R&D+생산), 베트남(VINA 1~3공장)"),
    (22, "분석기준일", "2026년 2월 6일"),
]
for r, label, val in info_data:
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws1.cell(row=r, column=2, value=label).font = Font(name="맑은 고딕", size=11, color="8899AA")
    ws1.cell(row=r, column=2).alignment = Alignment(horizontal='right', vertical='center')
    ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    ws1.cell(row=r, column=4, value=val).font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
    ws1.cell(row=r, column=4).alignment = Alignment(horizontal='left', vertical='center')

ws1.merge_cells('B25:F25')
ws1.cell(row=25, column=2, value="데이터 출처: OpenDART 공시 574건, 사업보고서 13년치 전수분석").font = Font(name="맑은 고딕", size=9, color="6688AA")
ws1.cell(row=25, column=2).alignment = Alignment(horizontal='center')

ws1.merge_cells('B27:F30')
cell = ws1.cell(row=27, column=2)
cell.value = ("핵심 요약:\n"
    "• PER 5.5배, PBR 0.9배의 저평가 가치주\n"
    "• 2024년 영업이익 +144% YoY, 순이익 636억(역대 최대)\n"
    "• 전장(ADAS) 매출 비중 27%→성장 가속, 배당수익률 4.0%")
cell.font = Font(name="맑은 고딕", size=10, color=ACCENT_GOLD)
cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

print("  [1/9] 표지 완료")

# ============================================================
# SHEET 2: 핵심 실적 (10년 재무)
# ============================================================
ws2 = wb.create_sheet("핵심실적")
ws2.sheet_properties.tabColor = "2C3E6B"
set_col_widths(ws2, [14, 14, 14, 14, 12, 14, 14, 14, 12, 12, 14])

perf = conn.execute("SELECT * FROM v_annual_performance").fetchall()

row = 1
ws2.merge_cells('A1:K1')
ws2.cell(row=1, column=1, value="10년 연결 재무실적 (단위: 억원)").font = section_font
ws2.cell(row=1, column=1).border = bottom_border
row = 3

headers = ["연도", "매출액", "영업이익", "순이익", "EPS(원)", "총자산", "총부채", "총자본", "부채비율", "영업이익률", "ROE"]
write_header_row(ws2, row, headers)
row += 1

prev_equity = None
for p in perf:
    yr = p[0]
    rev = int(p[1])//100000000 if p[1] else 0
    op = int(p[2])//100000000 if p[2] else 0
    ni = int(p[3])//100000000 if p[3] else 0
    eps = int(p[4]) if p[4] else 0
    ta = int(p[5])//100000000 if p[5] else 0
    tl = int(p[6])//100000000 if p[6] else 0
    te = int(p[7])//100000000 if p[7] else 0

    data = [yr, rev, op, ni, eps, ta, tl, te, None, None, None]
    nf = [None, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, '0.0%', '0.0%', '0.0%']

    fonts_row = [data_font_bold] + [data_font]*10
    fills_row = [light_fill] + [white_fill]*10

    if op < 0:
        fonts_row[2] = red_font
    if ni < 0:
        fonts_row[3] = red_font

    als = [center_al] + [right_al]*10
    write_data_row(ws2, row, data, fonts=fonts_row, fills=fills_row, alignments=als, number_formats=nf)

    # Formulas
    col_te = get_column_letter(8)
    col_tl = get_column_letter(7)
    col_rev = get_column_letter(2)
    col_op = get_column_letter(3)
    col_ni = get_column_letter(4)

    ws2.cell(row=row, column=9).value = f"={col_tl}{row}/{col_te}{row}"
    ws2.cell(row=row, column=9).number_format = '0.0%'
    ws2.cell(row=row, column=9).font = data_font
    ws2.cell(row=row, column=9).alignment = right_al
    ws2.cell(row=row, column=9).border = thin_border

    ws2.cell(row=row, column=10).value = f"={col_op}{row}/{col_rev}{row}"
    ws2.cell(row=row, column=10).number_format = '0.0%'
    ws2.cell(row=row, column=10).font = data_font
    ws2.cell(row=row, column=10).alignment = right_al
    ws2.cell(row=row, column=10).border = thin_border

    if prev_equity is not None and prev_equity > 0:
        avg_eq_formula = f"=({col_te}{row}+{prev_equity})/2"
        ws2.cell(row=row, column=11).value = f"={col_ni}{row}/(({col_te}{row}+{col_te}{row-1})/2)"
    else:
        ws2.cell(row=row, column=11).value = f"={col_ni}{row}/{col_te}{row}"
    ws2.cell(row=row, column=11).number_format = '0.0%'
    ws2.cell(row=row, column=11).font = data_font
    ws2.cell(row=row, column=11).alignment = right_al
    ws2.cell(row=row, column=11).border = thin_border

    prev_equity = te
    row += 1

# YoY growth section
row += 2
row = add_section_title(ws2, row, "전년대비 성장률 (YoY)")
headers2 = ["연도", "매출 YoY", "영업이익 YoY", "순이익 YoY"]
write_header_row(ws2, row, headers2)
row += 1

data_start = 4  # row 4 is first data row
for i in range(1, len(perf)):
    r = data_start + i
    yr = perf[i][0]
    ws2.cell(row=row, column=1, value=yr).font = data_font_bold
    ws2.cell(row=row, column=1).fill = light_fill
    ws2.cell(row=row, column=1).alignment = center_al
    ws2.cell(row=row, column=1).border = thin_border

    for col_idx, src_col in [(2, 'B'), (3, 'C'), (4, 'D')]:
        prev_r = data_start + i - 1
        curr_r = data_start + i
        formula = f'=IF({src_col}{prev_r}=0,"-",({src_col}{curr_r}-{src_col}{prev_r})/{src_col}{prev_r})'
        cell = ws2.cell(row=row, column=col_idx, value=formula)
        cell.number_format = '0.0%'
        cell.font = data_font
        cell.alignment = right_al
        cell.border = thin_border
    row += 1

print("  [2/9] 핵심실적 완료")

# ============================================================
# SHEET 3: 2025년 분기실적
# ============================================================
ws3 = wb.create_sheet("2025실적")
ws3.sheet_properties.tabColor = "27AE60"
set_col_widths(ws3, [14, 14, 14, 14, 14, 16, 16])

row = 1
ws3.merge_cells('A1:G1')
ws3.cell(row=1, column=1, value="2025년 분기별 잠정실적 (단위: 백만원)").font = section_font
ws3.cell(row=1, column=1).border = bottom_border

row = 3
headers = ["분기", "매출액", "영업이익", "순이익", "영업이익률", "매출 YoY", "영업이익 YoY"]
write_header_row(ws3, row, headers)
row += 1

q_data = [
    ("2024.1Q", 287835, 15205, 15930),
    ("2024.2Q", 225759, 7650, 18423),
    ("2024.3Q", 245071, 3077, 2505),
    ("2024.4Q", 298393, 18451, 26747),
    ("2024 합계", 1057058, 44385, 63605),
    ("2025.1Q", 347917, 20918, 20068),
    ("2025.2Q", 326424, 16171, 13050),
    ("2025.3Q", 303537, 6708, 9537),
    ("2025 누계", 977877, 43797, 42656),
]

for i, (qtr, rev, op, ni) in enumerate(q_data):
    is_total = "합계" in qtr or "누계" in qtr
    f_main = data_font_bold if is_total else data_font
    f_fill = gold_fill if is_total else (lighter_fill if "2025" in qtr else white_fill)

    data_row = [qtr, rev, op, ni, None, None, None]
    nf = [None, NUM_FMT, NUM_FMT, NUM_FMT, '0.0%', '0.0%', '0.0%']
    fonts_r = [f_main]*7
    fills_r = [f_fill]*7
    als = [center_al] + [right_al]*6
    write_data_row(ws3, row, data_row, fonts=fonts_r, fills=fills_r, alignments=als, number_formats=nf)

    # OPM formula
    ws3.cell(row=row, column=5).value = f"=C{row}/B{row}"
    ws3.cell(row=row, column=5).number_format = '0.0%'
    ws3.cell(row=row, column=5).border = thin_border

    # YoY: 2025 vs 2024
    if i >= 5 and i <= 7:
        prev_row = row - 5
        ws3.cell(row=row, column=6).value = f"=(B{row}-B{prev_row})/B{prev_row}"
        ws3.cell(row=row, column=6).number_format = '0.0%'
        ws3.cell(row=row, column=6).border = thin_border
        ws3.cell(row=row, column=7).value = f"=(C{row}-C{prev_row})/C{prev_row}"
        ws3.cell(row=row, column=7).number_format = '0.0%'
        ws3.cell(row=row, column=7).border = thin_border

    row += 1

row += 2
row = add_section_title(ws3, row, "핵심 포인트", col_end=7)
points = [
    "2025년 3분기 누계 매출 9,779억 → 이미 2023년 연간(9,325억) 초과",
    "2025년 연간 매출 1.3조원대 역대 최고 매출 갱신 전망",
    "3Q 영업이익 67억(OPM 2.2%)으로 급감 → 원가율 상승/제품믹스 변화 모니터링 필요",
    "1Q~2Q는 고부가 플래그십 모델 효과로 높은 마진 시현",
]
for pt in points:
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws3.cell(row=row, column=1, value=f"• {pt}").font = data_font
    ws3.cell(row=row, column=1).alignment = left_al
    row += 1

print("  [3/9] 2025실적 완료")

# ============================================================
# SHEET 4: 사업구조
# ============================================================
ws4 = wb.create_sheet("사업구조")
ws4.sheet_properties.tabColor = "D4A843"
set_col_widths(ws4, [18, 14, 14, 14, 14, 14, 14, 14])

row = 1
ws4.merge_cells('A1:H1')
ws4.cell(row=1, column=1, value="사업부문별 매출 구조 (연결, 단위: 억원)").font = section_font
ws4.cell(row=1, column=1).border = bottom_border

row = 3
headers = ["품목", "2020 수출", "2020 내수", "2021 수출", "2021 내수", "2022 수출", "2022 내수", "2023 수출"]
# Simplified structure
headers2 = ["구분", "2020", "2021", "2022", "2023", "비중(23)", "추세"]
write_header_row(ws4, row, headers2)
row += 1

seg_data = [
    ("휴대폰용", 11460, 8373, 9246, 6583, None),
    ("자동차용", 1534, 1602, 1701, 2527, None),
    ("기타(상품/개발/식음)", 119, 118, 127, 215, None),
    ("합계", 13113, 10093, 11086, 9325, None),
]

for nm, y20, y21, y22, y23, _ in seg_data:
    is_total = nm == "합계"
    f = data_font_bold if is_total else data_font
    fl = gold_fill if is_total else white_fill

    data_row = [nm, y20, y21, y22, y23, None, None]
    nf = [None, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, '0.0%', None]
    write_data_row(ws4, row, data_row, fonts=[f]*7, fills=[fl]*7,
                   alignments=[left_al]+[right_al]*6, number_formats=nf)

    total_row_abs = row + (3 - (row - 4)) if not is_total else row
    if not is_total:
        total_ref = 4 + 3
        ws4.cell(row=row, column=6).value = f"=E{row}/E{total_ref}"
        ws4.cell(row=row, column=6).number_format = '0.0%'
        ws4.cell(row=row, column=6).border = thin_border

    if y22 and y23 and not is_total:
        trend = "↑" if y23 > y22 else ("↓" if y23 < y22 else "→")
        ws4.cell(row=row, column=7, value=trend).font = green_font if trend == "↑" else red_font
        ws4.cell(row=row, column=7).alignment = center_al
        ws4.cell(row=row, column=7).border = thin_border

    row += 1

# Subsidiaries
row += 2
row = add_section_title(ws4, row, "종속회사 현황 (2024)", col_end=8)
headers = ["자회사명", "지분율", "역할", "장부가(억)", "순이익(억)", "총자산(억)"]
write_header_row(ws4, row, headers)
row += 1

subs = [
    ("엠씨넥스VINA(베트남)", "100%", "핵심 생산기지 (1~3공장)", 1216, 309, 3355),
    ("엠씨넥스상해(중국)", "100%", "수출입 영업/CS", 1, -1, 14),
    ("엠씨넥스에프앤비", "100%", "구내식당", 10, -0.4, 10),
]

for nm, pct, role, bv, ni, ta in subs:
    data_row = [nm, pct, role, bv, ni, ta]
    nf = [None, None, None, NUM_FMT, '#,##0.0', NUM_FMT]
    f_ni = green_font if ni > 0 else red_font
    write_data_row(ws4, row, data_row,
                   fonts=[data_font_bold, data_font, data_font, data_font, f_ni, data_font],
                   alignments=[left_al, center_al, left_al, right_al, right_al, right_al],
                   number_formats=nf)
    row += 1

# Production capacity
row += 2
row = add_section_title(ws4, row, "생산능력 및 가동률", col_end=8)
headers = ["구분", "생산CAPA(만개)", "2021 생산", "2021 가동률", "2022 생산", "2022 가동률", "2023 생산", "2023 가동률"]
write_header_row(ws4, row, headers)
row += 1

prod = [
    ("휴대폰용", "20,400만", "13,752만", "67.4%", "13,209만", "64.8%", "11,867만", "58.2%"),
    ("자동차용", "600만", "523만", "87.1%", "504만", "84.0%", "527만", "87.8%"),
]
for item in prod:
    write_data_row(ws4, row, list(item),
                   alignments=[left_al]+[center_al]*7,
                   fonts=[data_font_bold]+[data_font]*7)
    row += 1

print("  [4/9] 사업구조 완료")

# ============================================================
# SHEET 5: 배당/주주환원
# ============================================================
ws5 = wb.create_sheet("주주환원")
ws5.sheet_properties.tabColor = "E74C3C"
set_col_widths(ws5, [12, 14, 12, 14, 14, 14])

row = 1
ws5.merge_cells('A1:F1')
ws5.cell(row=1, column=1, value="배당 및 주주환원 정책").font = section_font
ws5.cell(row=1, column=1).border = bottom_border

row = 3
headers = ["연도", "주당배당금(원)", "EPS(원)", "배당성향", "배당수익률", "비고"]
write_header_row(ws5, row, headers)
row += 1

div_data = [
    ("2015", 330, 1833, "17.9%", "1.17%", ""),
    ("2016", 0, -2824, "-", "-", "적자→무배당"),
    ("2017", 260, -100, "-321.8%", "1.27%", "적자에도 배당"),
    ("2018", 300, 2161, "17.1%", "2.40%", ""),
    ("2019", 500, 5151, "10.4%", "1.70%", "최대실적"),
    ("2020", 500, 2180, "23.9%", "1.30%", ""),
    ("2021", 500, 2229, "22.6%", "0.90%", "유가증권 이전"),
    ("2022", 500, 1293, "38.5%", "1.80%", "배당성향 확대"),
    ("2023", 600, 1585, "37.9%", "2.00%", "배당금 인상"),
    ("2024", 800, 3661, "21.2%", "4.00%", "대폭 인상"),
    ("2025E", 1000, "-", "-", "3.74%", "결산배당 결정"),
]

for d in div_data:
    yr, dps, eps_v, payout, yld, note = d
    nf = [None, NUM_FMT, NUM_FMT, None, None, None]
    f_dps = green_font if isinstance(dps, (int,float)) and dps > 0 else red_font
    als = [center_al, right_al, right_al, center_al, center_al, left_al]

    write_data_row(ws5, row, [yr, dps if dps else "-", eps_v, payout, yld, note],
                   fonts=[data_font_bold, f_dps, data_font, data_font, data_font, small_font],
                   alignments=als, number_formats=nf)
    row += 1

# Shareholder return events
row += 2
row = add_section_title(ws5, row, "최근 주주환원 이벤트", col_end=6)
events = [
    ("2024.07", "자기주식 취득 결정"),
    ("2024.11", "자기주식 추가 취득 결정"),
    ("2024.12", "결산배당 주당 800원 결정"),
    ("2025.01", "자기주식 취득 결정"),
    ("2025.09", "자기주식 취득 결정"),
    ("2025.11", "주식 583,482주 소각 결정 (약 170억원)"),
    ("2025.12", "결산배당 주당 1,000원 결정"),
]
for dt, evt in events:
    ws5.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws5.cell(row=row, column=1, value=dt).font = data_font_bold
    ws5.cell(row=row, column=1).alignment = center_al
    ws5.cell(row=row, column=1).fill = lighter_fill
    ws5.cell(row=row, column=1).border = thin_border
    ws5.cell(row=row, column=2, value=evt).font = data_font
    ws5.cell(row=row, column=2).alignment = left_al
    ws5.cell(row=row, column=2).border = thin_border
    row += 1

# Share structure
row += 2
row = add_section_title(ws5, row, "주식 구조 (2024.12.31)", col_end=6)
share_info = [
    ("발행가능주식수", "50,000,000주"),
    ("발행주식수", "17,977,732주"),
    ("자기주식", "1,110,000주"),
    ("유통주식수", "16,867,732주"),
    ("최대주주(민동욱)", "4,661,000주 (25.93%)"),
    ("특수관계인 합계", "5,196,495주 (28.91%)"),
    ("소액주주 비율", "66.33% (58,267명)"),
]
for label, val in share_info:
    ws5.cell(row=row, column=1, value=label).font = data_font
    ws5.cell(row=row, column=1).alignment = left_al
    ws5.cell(row=row, column=1).fill = lighter_fill
    ws5.cell(row=row, column=1).border = thin_border
    ws5.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws5.cell(row=row, column=2, value=val).font = data_font_bold
    ws5.cell(row=row, column=2).alignment = left_al
    ws5.cell(row=row, column=2).border = thin_border
    row += 1

print("  [5/9] 주주환원 완료")

# ============================================================
# SHEET 6: R&D / 특허
# ============================================================
ws6 = wb.create_sheet("R&D_특허")
ws6.sheet_properties.tabColor = "8E44AD"
set_col_widths(ws6, [14, 50, 14, 14])

row = 1
ws6.merge_cells('A1:D1')
ws6.cell(row=1, column=1, value="연구개발 및 특허 현황").font = section_font
ws6.cell(row=1, column=1).border = bottom_border

row = 3
row = add_section_title(ws6, row, "연구개발비 추이", col_end=4)
headers = ["연도", "R&D비용(억원)", "매출대비", "비고"]
write_header_row(ws6, row, headers)
row += 1

rd_data = [
    ("2022", 397, "3.58%", "자산처리 22억 포함"),
    ("2023", 356, "3.82%", "자산처리 10억 포함"),
    ("2024", 327, "3.10%", "자산처리 3억 포함"),
]
for d in rd_data:
    write_data_row(ws6, row, list(d), fonts=[data_font_bold, data_font, data_font, small_font],
                   alignments=[center_al, right_al, center_al, left_al],
                   number_formats=[None, NUM_FMT, None, None])
    row += 1

row += 1
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws6.cell(row=row, column=1, value="연구인력 284명 (전체의 57.7%) | 임원15 + 수석107 + 선임106 + 연구원56").font = Font(name="맑은 고딕", size=10, bold=True, color=MID_BLUE)
row += 2

row = add_section_title(ws6, row, "특허 포트폴리오 (24건)", col_end=4)
headers = ["취득일", "특허명/내용", "분류", "활용"]
write_header_row(ws6, row, headers)
row += 1

patents = conn.execute("SELECT rcept_dt, patent_name, patent_detail, patent_plan FROM patents ORDER BY rcept_dt").fetchall()
for p in patents:
    dt = p[0]
    nm = (p[1] or "")[:60]
    detail = (p[2] or "")[:80]
    plan = (p[3] or "")[:40]

    category = "액츄에이터" if "액" in detail or "Actuator" in detail or "actuator" in detail else \
               "전장카메라" if "차량" in detail or "주차" in detail or "AVM" in detail else \
               "광학/제조" if "광축" in detail or "카메라모듈" in detail or "프리즘" in detail else \
               "보안/인식" if "홍채" in detail or "보안" in detail or "터치" in detail else "기타"

    display = nm if nm and "특허" not in nm[:3] else detail[:60]
    write_data_row(ws6, row, [dt, display, category, plan[:30] if plan else ""],
                   fonts=[data_font, data_font, data_font_bold, small_font],
                   alignments=[center_al, left_al, center_al, left_al])
    row += 1

print("  [6/9] R&D_특허 완료")

# ============================================================
# SHEET 7: 투자지표
# ============================================================
ws7 = wb.create_sheet("투자지표")
ws7.sheet_properties.tabColor = "E67E22"
set_col_widths(ws7, [20, 16, 40])

row = 1
ws7.merge_cells('A1:C1')
ws7.cell(row=1, column=1, value="핵심 투자지표 (2024 기준, 주가 20,000원 가정)").font = section_font
ws7.cell(row=1, column=1).border = bottom_border

row = 3
metrics = [
    ("PER", "5.5배", "EPS 3,661원 기준. 동종업 평균 10~15배 대비 크게 저평가"),
    ("PBR", "0.91배", "시총 약 3,400억 / 자본 3,722억. 장부가 이하"),
    ("ROE", "18.3%", "순이익 636억 / 평균자본 3,479억. 높은 수준"),
    ("영업이익률", "4.2%", "조립업 특성상 구조적 저마진이나 2024년 양호"),
    ("부채비율", "54.2%", "2016년 363%에서 극적 개선. 재무 건전"),
    ("순차입금비율", "2.7%", "사실상 무차입 경영에 근접"),
    ("배당수익률", "4.0%", "주당 800원 (2024). 2025년 1,000원으로 인상"),
    ("시가총액", "약 3,400억원", "유통시총 약 3,000억원 (자기주식 제외)"),
]

headers = ["지표", "값", "해석"]
write_header_row(ws7, row, headers)
row += 1

for label, val, desc in metrics:
    write_data_row(ws7, row, [label, val, desc],
                   fonts=[data_font_bold, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), data_font],
                   fills=[lighter_fill, gold_fill, white_fill],
                   alignments=[left_al, center_al, left_al])
    row += 1

row += 2
row = add_section_title(ws7, row, "저평가 요인 분석", col_end=3)
factors = [
    ("삼성전자 1사 의존", "매출의 70%+ 삼성전자 의존 → 고객 편중 디스카운트"),
    ("조립 사업 저마진", "원재료(센서/렌즈)가 원가의 70%+ → 구조적 마진 한계"),
    ("스마트폰 성숙기", "글로벌 스마트폰 물량 성장 제한적 → 성장 프리미엄 부재"),
    ("중소형주 디스카운트", "기관/외국인 관심 제한적, 유동성 프리미엄 부재"),
]
for factor, desc in factors:
    ws7.cell(row=row, column=1, value=factor).font = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT_RED)
    ws7.cell(row=row, column=1).alignment = left_al
    ws7.cell(row=row, column=1).fill = red_fill
    ws7.cell(row=row, column=1).border = thin_border
    ws7.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws7.cell(row=row, column=2, value=desc).font = data_font
    ws7.cell(row=row, column=2).alignment = left_al
    ws7.cell(row=row, column=2).border = thin_border
    row += 1

print("  [7/9] 투자지표 완료")

# ============================================================
# SHEET 8: 시나리오 분석
# ============================================================
ws8 = wb.create_sheet("시나리오")
ws8.sheet_properties.tabColor = "2ECC71"
set_col_widths(ws8, [16, 18, 18, 18, 18])

row = 1
ws8.merge_cells('A1:E1')
ws8.cell(row=1, column=1, value="향후 시나리오 분석").font = section_font
ws8.cell(row=1, column=1).border = bottom_border

scenarios = [
    ("강세 (Bull)", "27AE60", [
        ("전제", "갤럭시 카메라 대폭 업그레이드 + 전장 35%+ + 주주환원 강화"),
        ("매출", "1.35조원"),
        ("영업이익률", "5.0%+"),
        ("순이익", "700억+"),
        ("EPS", "4,000원"),
        ("목표 PER", "10배"),
        ("목표주가", "40,000원"),
        ("상승여력", "~100%"),
    ]),
    ("기본 (Base)", "F39C12", [
        ("전제", "모바일 유지 + 전장 연 10% 성장 + 배당 유지"),
        ("매출", "1.3조원"),
        ("영업이익률", "4.0%"),
        ("순이익", "500억"),
        ("EPS", "2,800원"),
        ("목표 PER", "8배"),
        ("목표주가", "22,400원"),
        ("상승여력", "~12%"),
    ]),
    ("약세 (Bear)", "E74C3C", [
        ("전제", "삼성 부진 + 전장 정체 + 마진 하락"),
        ("매출", "1.1조원"),
        ("영업이익률", "2.0%"),
        ("순이익", "250억"),
        ("EPS", "1,400원"),
        ("목표 PER", "7배"),
        ("목표주가", "9,800원"),
        ("하락위험", "~50%"),
    ]),
]

row = 3
headers = ["항목", "강세 (Bull)", "기본 (Base)", "약세 (Bear)"]
write_header_row(ws8, row, headers)
row += 1

items = ["전제조건", "매출 전망", "영업이익률", "순이익 전망", "EPS 전망", "적용 PER", "목표주가", "현주가 대비"]
bull = ["갤럭시 업그레이드+전장35%", "1.35조원", "5.0%+", "700억+", "4,000원", "10배", "40,000원", "상승 ~100%"]
base = ["모바일유지+전장10%성장", "1.3조원", "4.0%", "500억", "2,800원", "8배", "22,400원", "상승 ~12%"]
bear = ["삼성부진+전장정체", "1.1조원", "2.0%", "250억", "1,400원", "7배", "9,800원", "하락 ~50%"]

for i in range(len(items)):
    bull_color = green_fill
    base_color = gold_fill
    bear_color = red_fill

    write_data_row(ws8, row, [items[i], bull[i], base[i], bear[i]],
                   fonts=[data_font_bold, data_font, data_font, data_font],
                   fills=[lighter_fill, bull_color, base_color, bear_color],
                   alignments=[left_al, center_al, center_al, center_al])
    row += 1

# SWOT
row += 2
row = add_section_title(ws8, row, "SWOT 분석", col_end=5)

ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
ws8.cell(row=row, column=1, value="강점 (Strengths)").font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
ws8.cell(row=row, column=1).fill = PatternFill("solid", fgColor="27AE60")
ws8.cell(row=row, column=1).alignment = center_al
ws8.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
ws8.cell(row=row, column=3, value="약점 (Weaknesses)").font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
ws8.cell(row=row, column=3).fill = PatternFill("solid", fgColor="E74C3C")
ws8.cell(row=row, column=3).alignment = center_al
row += 1

strengths = [
    "삼성전자/현대차 핵심 공급사 지위",
    "베트남 3개 공장 원가경쟁력",
    "연구인력 57.7%, 특허 24건",
    "부채비율 54% 재무건전",
    "액츄에이터 내재화 수익력",
]
weaknesses = [
    "삼성전자 매출 편중 70%+",
    "조립 중심 구조적 저마진 3~5%",
    "자체 브랜드/최종제품 부재",
    "원재료 가격 통제력 없음",
    "스마트폰 시장 성숙기",
]

for i in range(max(len(strengths), len(weaknesses))):
    s = strengths[i] if i < len(strengths) else ""
    w = weaknesses[i] if i < len(weaknesses) else ""
    ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws8.cell(row=row, column=1, value=f"• {s}" if s else "").font = data_font
    ws8.cell(row=row, column=1).alignment = left_al
    ws8.cell(row=row, column=1).fill = green_fill
    ws8.cell(row=row, column=1).border = thin_border
    ws8.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws8.cell(row=row, column=3, value=f"• {w}" if w else "").font = data_font
    ws8.cell(row=row, column=3).alignment = left_al
    ws8.cell(row=row, column=3).fill = red_fill
    ws8.cell(row=row, column=3).border = thin_border
    row += 1

row += 1
ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
ws8.cell(row=row, column=1, value="기회 (Opportunities)").font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
ws8.cell(row=row, column=1).fill = PatternFill("solid", fgColor="2980B9")
ws8.cell(row=row, column=1).alignment = center_al
ws8.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
ws8.cell(row=row, column=3, value="위협 (Threats)").font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
ws8.cell(row=row, column=3).fill = PatternFill("solid", fgColor="7F8C8D")
ws8.cell(row=row, column=3).alignment = center_al
row += 1

opportunities = [
    "ADAS 의무화 → 차량당 카메라 증가",
    "자율주행 L3/L4 본격화",
    "프리즘줌 등 고부가 모듈 확대",
    "삼성 XR기기 등 신규 카테고리",
]
threats = [
    "중국 카메라모듈 업체 추격",
    "삼성 자체 모듈 내재화 가능성",
    "환율/베트남 인건비 상승",
    "글로벌 경기침체 시 수요 감소",
]

for i in range(max(len(opportunities), len(threats))):
    o = opportunities[i] if i < len(opportunities) else ""
    t = threats[i] if i < len(threats) else ""
    ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws8.cell(row=row, column=1, value=f"• {o}" if o else "").font = data_font
    ws8.cell(row=row, column=1).alignment = left_al
    ws8.cell(row=row, column=1).fill = PatternFill("solid", fgColor="D6EAF8")
    ws8.cell(row=row, column=1).border = thin_border
    ws8.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws8.cell(row=row, column=3, value=f"• {t}" if t else "").font = data_font
    ws8.cell(row=row, column=3).alignment = left_al
    ws8.cell(row=row, column=3).fill = PatternFill("solid", fgColor="E5E7E9")
    ws8.cell(row=row, column=3).border = thin_border
    row += 1

print("  [8/9] 시나리오 완료")

# ============================================================
# SHEET 9: 모니터링 체크리스트
# ============================================================
ws9 = wb.create_sheet("모니터링")
ws9.sheet_properties.tabColor = "1ABC9C"
set_col_widths(ws9, [6, 30, 40, 20])

row = 1
ws9.merge_cells('A1:D1')
ws9.cell(row=1, column=1, value="향후 이익 판단을 위한 핵심 모니터링 지표").font = section_font
ws9.cell(row=1, column=1).border = bottom_border

row = 3
headers = ["순위", "모니터링 항목", "세부 내용", "확인 시기"]
write_header_row(ws9, row, headers)
row += 1

monitors = [
    ("1", "삼성 갤럭시 카메라 스펙/물량",
     "갤럭시 S/Z 시리즈 카메라 화소, 개수, OIS/줌 사양이 ASP와 매출을 직접 결정. IDC/Counterpoint 스마트폰 출하량도 참고.",
     "1~2월, 7~8월 (언팩)"),
    ("2", "전장(자동차) 매출 비중",
     "전장 비중이 30~40% 넘으면 밸류에이션 리레이팅. 현대차 ADAS 채택률, 차량당 카메라수 확인.",
     "매 분기"),
    ("3", "분기별 영업이익률",
     "2~3%=저수익 | 4~5%=양호 | 6%+=호황. 원재료(센서/렌즈) 가격, 환율, 제품믹스가 핵심 변수.",
     "2/5/8/11월 (잠정실적)"),
    ("4", "베트남 VINA 법인 실적",
     "연결 순이익의 ~50% 창출. 가동률/수율/인건비 상승 모니터링. 수출입은행 채무보증 규모도 확인.",
     "연간 사업보고서"),
    ("5", "자기주식/배당 정책",
     "2024~25년 적극적 주주환원(배당 증가+자사주 매입+소각). 이 기조 지속 시 주가 하방 지지력 강화.",
     "3월 주총, 수시"),
]

for rank, title, detail, timing in monitors:
    write_data_row(ws9, row, [rank, title, detail, timing],
                   fonts=[Font(name="맑은 고딕", size=14, bold=True, color=NAVY), data_font_bold, data_font, data_font],
                   fills=[gold_fill, lighter_fill, white_fill, lighter_fill],
                   alignments=[center_al, left_al, left_al, center_al])
    ws9.row_dimensions[row].height = 45
    row += 1

row += 2
row = add_section_title(ws9, row, "모니터링 캘린더", col_end=4)
calendar = [
    ("1~2월", "삼성 갤럭시 S시리즈 언팩 (카메라 사양 확인)", "★★★"),
    ("2월", "전년도 잠정실적 공시 (4Q + 연간)", "★★★"),
    ("3월", "사업보고서 제출 / 정기주총 / 배당·자사주 정책", "★★★"),
    ("5월", "1분기 잠정실적 공시", "★★☆"),
    ("7~8월", "삼성 갤럭시 Z/폴드 언팩 + 2분기 잠정실적", "★★★"),
    ("11월", "3분기 잠정실적 공시", "★★☆"),
    ("12월", "결산배당 결정 공시", "★★☆"),
    ("수시", "자기주식 취득/처분/소각 공시", "★☆☆"),
    ("수시", "현대차 ADAS 로드맵 (CES 등)", "★★☆"),
]

headers = ["시기", "이벤트", "중요도"]
write_header_row(ws9, row, headers)
row += 1

for timing, event, importance in calendar:
    write_data_row(ws9, row, [timing, event, importance],
                   fonts=[data_font_bold, data_font, data_font_bold],
                   fills=[lighter_fill, white_fill, gold_fill],
                   alignments=[center_al, left_al, center_al])
    row += 1

row += 2
row = add_section_title(ws9, row, "주요 이벤트 히스토리 (최근)", col_end=4)
headers = ["일자", "유형", "내용"]
write_header_row(ws9, row, headers)
row += 1

events = conn.execute("""SELECT rcept_dt, event_type, SUBSTR(event_summary, 1, 80)
    FROM key_events WHERE rcept_dt >= '20230101' ORDER BY rcept_dt DESC LIMIT 20""").fetchall()
for dt, etype, summary in events:
    summary_clean = summary.replace('\n', ' ').strip()[:70]
    write_data_row(ws9, row, [dt, etype, summary_clean],
                   fonts=[data_font, data_font_bold, data_font],
                   alignments=[center_al, center_al, left_al])
    row += 1

print("  [9/9] 모니터링 완료")

# ============================================================
# SAVE
# ============================================================
wb.save(OUT)
conn.close()
print(f"\n보고서 생성 완료: {OUT}")
print(f"시트 구성: {wb.sheetnames}")
