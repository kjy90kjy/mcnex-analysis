# -*- coding: utf-8 -*-
"""엠씨넥스 모바일용 보고서 — 좁은 화면 최적화 (3열, 큰 글씨, 세로 스크롤)"""
import sqlite3, sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "mcnex_ai.db")
conn = sqlite3.connect(DB)

# === CONSTANTS ===
PRICE = 28100; SHARES = 17977732; SHARES_TREASURY = 1110000; SHARES_WA = 17373105
억 = 100_000_000; MCAP = PRICE * SHARES
REV24=1057058423929; OP24=44384671816; NI24=63604930146; EPS24=3661
EQ24=372224146153; EQ23=323612702499; LIAB24=201879404879
CASH24=35881602429; ST_DEBT=44402388000; LT_DEBT=1308000000
DA24=44255618814+462932726+1740746569; OPCF24=82470298068; CAPEX24=36541908724
DIV_PAID=10570639200; TREAS_BUY=15656989850; DPS24=800
Q24=[{"r":287835,"o":15205,"n":15930},{"r":225759,"o":7650,"n":18423},
     {"r":245071,"o":3077,"n":2505},{"r":298393,"o":18451,"n":26747}]
Q25=[{"r":347917,"o":20918,"n":20068},{"r":326424,"o":16171,"n":13050},{"r":303537,"o":6708,"n":9537}]
CUM25_NI=42656e6; CUM25_OP=43797e6; CUM25_REV=977877e6
BPS=EQ24/SHARES; NET_DEBT=ST_DEBT+LT_DEBT-CASH24; EBITDA24=OP24+DA24
FCF24=OPCF24-CAPEX24; EV=MCAP+NET_DEBT
TRAIL_NI=Q24[3]["n"]*1e6+Q25[0]["n"]*1e6+Q25[1]["n"]*1e6+Q25[2]["n"]*1e6
TRAIL_EPS=TRAIL_NI/SHARES_WA
E25_ANN_EPS=(CUM25_NI*4/3)/SHARES_WA
E25_CONS_EPS=(CUM25_NI+(Q25[1]["n"]*1e6+Q25[2]["n"]*1e6)/2)/SHARES_WA
ROE24=NI24/((EQ24+EQ23)/2)

def fmt(v): return f"{v/억:,.0f}억"
def fw(v): return f"{v:,.0f}원"
def pct(v): return f"{v*100:.1f}%"

# === MOBILE STYLES ===
# 모바일: 3열 기본, 큰 폰트, 넓은 행 높이, wrap text
NAVY="1B2A4A"; DARK="2C3E6B"; W="FFFFFF"; GOLD_C="D4A843"; RED_C="C0392B"; GREEN_C="27AE60"

# 모바일은 더 큰 폰트
t1=Font(name="맑은 고딕",size=16,bold=True,color=W)       # title
t2=Font(name="맑은 고딕",size=14,bold=True,color=NAVY)     # section
hf_=Font(name="맑은 고딕",size=11,bold=True,color=W)       # header
df_=Font(name="맑은 고딕",size=11)                          # data
db_=Font(name="맑은 고딕",size=11,bold=True)               # data bold
bl_=Font(name="맑은 고딕",size=11,bold=True,color="0000FF") # blue
gn_=Font(name="맑은 고딕",size=11,bold=True,color=GREEN_C)
rd_=Font(name="맑은 고딕",size=11,bold=True,color=RED_C)
sm_=Font(name="맑은 고딕",size=10,color="666666")
big_=Font(name="맑은 고딕",size=13,bold=True,color=NAVY)
huge_=Font(name="맑은 고딕",size=18,bold=True,color=NAVY)

tf_=PatternFill("solid",fgColor=NAVY)
hfl=PatternFill("solid",fgColor=DARK)
lf_=PatternFill("solid",fgColor="D6E4F0")
llf_=PatternFill("solid",fgColor="EBF1F8")
wf_=PatternFill("solid",fgColor=W)
gld_=PatternFill("solid",fgColor="FFF3CD")
gnf_=PatternFill("solid",fgColor="D4EDDA")
rdf_=PatternFill("solid",fgColor="F8D7DA")
blf_=PatternFill("solid",fgColor="D6EAF8")
grf_=PatternFill("solid",fgColor="F2F2F2")

ca_=Alignment(horizontal='center',vertical='center',wrap_text=True)
la_=Alignment(horizontal='left',vertical='center',wrap_text=True)
ra_=Alignment(horizontal='right',vertical='center',wrap_text=True)
tb_=Border(left=Side('thin',color="D0D0D0"),right=Side('thin',color="D0D0D0"),
           top=Side('thin',color="D0D0D0"),bottom=Side('thin',color="D0D0D0"))
bb_=Border(bottom=Side('medium',color=NAVY))

# 모바일 컬럼: 3열 (12+16+16=44)
COL3 = [14, 18, 18]
COL2 = [18, 32]

def sw(ws, w):
    for i, v in enumerate(w, 1): ws.column_dimensions[get_column_letter(i)].width = v

def mhdr(ws, r, vals):
    """모바일 헤더"""
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = hf_; c.fill = hfl; c.alignment = ca_; c.border = tb_

def mrow(ws, r, vals, fonts=None, fills=None, als=None, h=None):
    """모바일 데이터 행"""
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = fonts[i-1] if fonts else df_
        c.fill = fills[i-1] if fills else wf_
        c.alignment = als[i-1] if als else ca_
        c.border = tb_
    if h: ws.row_dimensions[r].height = h

def msec(ws, r, title, cols=3):
    """모바일 섹션 타이틀"""
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
    c = ws.cell(row=r, column=1, value=title)
    c.font = t2; c.border = bb_
    ws.row_dimensions[r].height = 28
    return r + 1

def minfo(ws, r, label, value, cols=3):
    """모바일 라벨-값 쌍"""
    ws.cell(row=r, column=1, value=label).font = db_
    ws.cell(row=r, column=1).fill = llf_; ws.cell(row=r, column=1).alignment = la_; ws.cell(row=r, column=1).border = tb_
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=cols)
    ws.cell(row=r, column=2, value=value).font = bl_
    ws.cell(row=r, column=2).fill = wf_; ws.cell(row=r, column=2).alignment = ra_; ws.cell(row=r, column=2).border = tb_
    ws.row_dimensions[r].height = 24
    return r + 1

wb = Workbook()

# ============================================================
# 단일 시트: 모든 내용을 세로로 길게 배치
# ============================================================
ws = wb.active
ws.title = "엠씨넥스 분석"
ws.sheet_properties.tabColor = NAVY
sw(ws, COL3)

row = 1

# ---- TITLE ----
ws.merge_cells('A1:C2')
c = ws.cell(row=1, column=1, value="엠씨넥스(097520)\n종합 기업분석 보고서")
c.font = t1; c.fill = tf_; c.alignment = ca_
ws.row_dimensions[1].height = 24; ws.row_dimensions[2].height = 24
row = 3
ws.merge_cells('A3:C3')
ws.cell(row=3, column=1, value=f"현재가 {PRICE:,}원 | 시총 {fmt(MCAP)} | 2026.02.06").font = Font(name="맑은 고딕", size=11, bold=True, color=DARK)
ws.cell(row=3, column=1).alignment = ca_; ws.row_dimensions[3].height = 22

# ==== 1. 기본정보 ====
row = 5
row = msec(ws, row, "기본 정보")
for lbl, val in [
    ("현재주가", fw(PRICE)), ("시가총액", fmt(MCAP)),
    ("유통시총", fmt(PRICE*(SHARES-SHARES_TREASURY))),
    ("발행주식수", f"{SHARES:,}주"), ("자기주식", f"{SHARES_TREASURY:,}주"),
    ("업종", "카메라모듈(CCM) 전문기업"),
    ("주요고객", "삼성전자, 현대모비스"),
    ("대표이사", "민동욱"), ("설립", "2004.12.22")]:
    row = minfo(ws, row, lbl, val)

# ==== 2. 핵심 밸류에이션 ====
row += 1
row = msec(ws, row, "핵심 밸류에이션")
mhdr(ws, row, ["지표", "값", "판정"]); row += 1

vals = [
    ("PER (2024)", f"{PRICE/EPS24:.1f}배", "저평가"),
    ("PER (T4Q)", f"{PRICE/TRAIL_EPS:.1f}배", "저평가"),
    ("PER (25E 보수적)", f"{PRICE/E25_CONS_EPS:.1f}배", "적정"),
    ("PBR", f"{PRICE/BPS:.2f}배", "적정"),
    ("EV/EBITDA", f"{EV/EBITDA24:.1f}배", "저평가"),
    ("PSR", f"{MCAP/REV24:.2f}배", "매우저평가"),
    ("배당수익률", pct(DPS24/PRICE), "양호"),
    ("FCF수익률", pct(FCF24/SHARES/PRICE), "매우양호"),
    ("총주주환원", pct((DIV_PAID+TREAS_BUY)/MCAP), "양호"),
    ("ROE", pct(ROE24), "양호"),
    ("부채비율", pct(LIAB24/EQ24), "건전"),
]
for lbl, val, judge in vals:
    if "매우" in judge: jf, jfl = gn_, gnf_
    elif "저" in judge or "양호" in judge or "건전" in judge: jf, jfl = Font(name="맑은 고딕",size=11,bold=True,color="2E86C1"), blf_
    elif "적정" in judge: jf, jfl = db_, gld_
    else: jf, jfl = rd_, rdf_
    mrow(ws, row, [lbl, val, judge], fonts=[db_, bl_, jf], fills=[llf_, gld_, jfl], als=[la_, ca_, ca_], h=24)
    row += 1

# ==== 3. 2024 확정실적 ====
row += 1
row = msec(ws, row, "2024년 확정 실적")
for lbl, val in [
    ("매출액", fmt(REV24)), ("영업이익", f"{fmt(OP24)} (OPM {OP24/REV24*100:.1f}%)"),
    ("순이익", f"{fmt(NI24)} (역대 최대)"), ("EPS", fw(EPS24)), ("BPS", fw(int(BPS))),
    ("DPS", f"{fw(DPS24)} (전년 600원)"),
    ("총자산", fmt(574103551032)), ("총부채", fmt(LIAB24)),
    ("총자본", fmt(EQ24)), ("영업CF", fmt(OPCF24)),
    ("FCF", fmt(FCF24)), ("EBITDA", fmt(EBITDA24))]:
    row = minfo(ws, row, lbl, val)

# ==== 4. 2025 분기실적 ====
row += 1
row = msec(ws, row, "2025년 분기별 잠정실적 (백만원)")
mhdr(ws, row, ["분기", "매출", "영업이익"]); row += 1
for lbl, rv, op in [("24Q1",287835,15205),("24Q2",225759,7650),("24Q3",245071,3077),("24Q4",298393,18451),
                     ("25Q1",347917,20918),("25Q2",326424,16171),("25Q3",303537,6708)]:
    is25 = "25" in lbl; fl = gnf_ if is25 else wf_
    mrow(ws, row, [lbl, f"{rv:,}", f"{op:,}"],
         fonts=[db_, df_, df_], fills=[blf_ if is25 else llf_, fl, fl], als=[ca_, ra_, ra_]); row += 1
# 합산
mrow(ws, row, ["25누계", f"{int(CUM25_REV/1e6):,}", f"{int(CUM25_OP/1e6):,}"],
     fonts=[db_]*3, fills=[gld_]*3, als=[ca_, ra_, ra_]); row += 1
mrow(ws, row, ["T4Q합산", f"{int((Q24[3]['r']*1e6+Q25[0]['r']*1e6+Q25[1]['r']*1e6+Q25[2]['r']*1e6)/1e6):,}",
     f"{int((Q24[3]['o']*1e6+Q25[0]['o']*1e6+Q25[1]['o']*1e6+Q25[2]['o']*1e6)/1e6):,}"],
     fonts=[db_]*3, fills=[gld_]*3, als=[ca_, ra_, ra_]); row += 1

# Key points
row += 1
for pt in ["9M 매출 YoY +29%, 영업이익 +69%",
           "Q3 OPM 2.2% 급락 → 모니터링 필요",
           "상반기가 연간이익 85% 비중"]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"• {pt}").font = df_; ws.cell(row=row, column=1).alignment = la_
    ws.row_dimensions[row].height = 22; row += 1

# ==== 5. PER 다각도 ====
row += 1
row = msec(ws, row, "PER 다각도 분석")
mhdr(ws, row, ["기준", "EPS", "PER"]); row += 1
for lbl, eps_v in [("2024 확정", EPS24), ("Trailing 4Q", int(TRAIL_EPS)),
                    ("25E 연환산", int(E25_ANN_EPS)), ("25E 보수적", int(E25_CONS_EPS))]:
    per = PRICE / eps_v
    pf = gn_ if per < 7.5 else (bl_ if per < 9 else db_)
    pfl = gnf_ if per < 7.5 else (blf_ if per < 9 else gld_)
    mrow(ws, row, [lbl, fw(eps_v), f"{per:.1f}배"], fonts=[db_, bl_, pf], fills=[llf_, gld_, pfl]); row += 1

# Target PER
row += 1
row = msec(ws, row, "목표PER별 적정주가")
mhdr(ws, row, ["PER배수", "2024 EPS", "T4Q EPS"]); row += 1
for m in [7, 8, 10, 12]:
    mrow(ws, row, [f"{m}배", fw(EPS24*m), fw(int(TRAIL_EPS*m))],
         fonts=[db_, df_, df_], fills=[llf_, gnf_ if EPS24*m>PRICE else wf_, gnf_ if TRAIL_EPS*m>PRICE else wf_],
         als=[ca_, ra_, ra_]); row += 1

# ==== 6. EV/EBITDA & FCF ====
row += 1
row = msec(ws, row, "EV/EBITDA & FCF")
for lbl, val in [
    ("EV (시총+순차입금)", fmt(EV)),
    ("EBITDA (2024)", fmt(EBITDA24)),
    ("EV/EBITDA", f"{EV/EBITDA24:.1f}배"),
    ("", ""),
    ("영업CF", fmt(OPCF24)),
    ("CAPEX", fmt(CAPEX24)),
    ("FCF", fmt(FCF24)),
    ("FCF/주", fw(int(FCF24/SHARES))),
    ("FCF 수익률", pct(FCF24/SHARES/PRICE)),
    ("", ""),
    ("배당지급", f"{fmt(DIV_PAID)} ({fw(DPS24)}/주)"),
    ("자사주매입", fmt(TREAS_BUY)),
    ("총주주환원", fmt(DIV_PAID+TREAS_BUY)),
    ("주주환원율", pct((DIV_PAID+TREAS_BUY)/MCAP))]:
    if not lbl: row += 1; continue
    row = minfo(ws, row, lbl, val)

# ==== 7. RIM 적정주가 ====
row += 1
row = msec(ws, row, "잔여이익모델(RIM) 적정주가")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="BPS × (1+(ROE-ke)/(ke-g)) | ke=10%").font = sm_
ws.cell(row=row, column=1).alignment = la_; row += 1
mhdr(ws, row, ["시나리오", "적정주가", "현재가대비"]); row += 1
for lbl, roe, ke, g in [("보수적(ROE=ke)", 0.10, 0.10, 0.02), ("기본(ROE12%)", 0.12, 0.10, 0.02),
                          ("적극(ROE15%)", 0.15, 0.10, 0.02), ("낙관(ROE18%)", 0.18, 0.10, 0.03)]:
    fv = BPS * (1 + (roe - ke) / (ke - g)); up = (fv - PRICE) / PRICE
    uf = gn_ if up > 0 else rd_; ufl = gnf_ if up > 0 else rdf_
    mrow(ws, row, [lbl, fw(int(fv)), f"{up*100:+.1f}%"], fonts=[db_, bl_, uf], fills=[llf_, gld_, ufl]); row += 1

# ==== 8. 시나리오 ====
row += 1
row = msec(ws, row, "시나리오별 목표주가")
bull_t = int(TRAIL_EPS*1.15)*12; base_t = int(TRAIL_EPS*10)
bear_t = int(E25_CONS_EPS*0.85)*7
exp_v = int(bull_t*0.2 + base_t*0.5 + bear_t*0.3)

mhdr(ws, row, ["시나리오", "목표주가", "현재가대비"]); row += 1
mrow(ws, row, ["강세(Bull)", fw(bull_t), f"{(bull_t-PRICE)/PRICE*100:+.1f}%"],
     fonts=[db_, Font(name="맑은 고딕",size=12,bold=True,color=GREEN_C), gn_], fills=[gnf_]*3, h=28); row += 1
mrow(ws, row, ["기본(Base)", fw(base_t), f"{(base_t-PRICE)/PRICE*100:+.1f}%"],
     fonts=[db_, Font(name="맑은 고딕",size=12,bold=True,color=NAVY), bl_], fills=[gld_]*3, h=28); row += 1
mrow(ws, row, ["약세(Bear)", fw(bear_t), f"{(bear_t-PRICE)/PRICE*100:+.1f}%"],
     fonts=[db_, Font(name="맑은 고딕",size=12,bold=True,color=RED_C), rd_], fills=[rdf_]*3, h=28); row += 1
row += 1
mrow(ws, row, ["기대값", fw(exp_v), f"{(exp_v-PRICE)/PRICE*100:+.1f}%"],
     fonts=[Font(name="맑은 고딕",size=12,bold=True,color=NAVY)]*3, fills=[blf_]*3, h=28); row += 1

# 전제조건
row += 1
for sc, cond in [("Bull", "갤럭시 업사이클 + 전장35% + PER리레이팅"),
                  ("Base", "모바일유지 + 전장10%성장 + 배당확대"),
                  ("Bear", "삼성부진 + 전장정체 + 마진악화")]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"{sc}: {cond}").font = df_
    ws.cell(row=row, column=1).alignment = la_; ws.row_dimensions[row].height = 22; row += 1

# ==== 9. 배당 이력 ====
row += 1
row = msec(ws, row, "배당 이력")
mhdr(ws, row, ["연도", "DPS(원)", "배당률"]); row += 1
for yr, dps, yl in [("2019",500,"1.70%"),("2020",500,"1.30%"),("2021",500,"0.90%"),
                     ("2022",500,"1.80%"),("2023",600,"2.00%"),("2024",800,f"{800/PRICE*100:.1f}%"),("2025E",1000,f"{1000/PRICE*100:.1f}%")]:
    mrow(ws, row, [yr, f"{dps:,}" if isinstance(dps, int) else dps, yl],
         fonts=[db_, bl_, df_], fills=[llf_, gld_, wf_]); row += 1

# ==== 10. 사업구조 ====
row += 1
row = msec(ws, row, "매출 구조 (2023년)")
for lbl, val in [("휴대폰용 CCM", "6,583억 (70.6%)"), ("자동차용 CCM", "2,527억 (27.1%)"),
                  ("기타", "215억 (2.3%)"), ("합계", "9,325억")]:
    row = minfo(ws, row, lbl, val)

row += 1
row = msec(ws, row, "종속회사")
for nm, desc in [("VINA(베트남)", "핵심 생산기지 100%, 순이익 309억"),
                  ("상해(중국)", "수출입/CS 100%"),
                  ("에프앤비", "구내식당 100%")]:
    row = minfo(ws, row, nm, desc)

# ==== 11. R&D ====
row += 1
row = msec(ws, row, "R&D / 특허")
for lbl, val in [("연구인력", "284명 (전체 57.7%)"),
                  ("R&D비용(24)", "327억 (매출 3.1%)"),
                  ("특허건수", "24건"),
                  ("핵심분야", "액츄에이터, 전장카메라, 광학")]:
    row = minfo(ws, row, lbl, val)

# ==== 12. SWOT ====
row += 1
row = msec(ws, row, "SWOT 분석")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="강점 (S)").font = Font(name="맑은 고딕",size=12,bold=True,color=W)
ws.cell(row=row, column=1).fill = PatternFill("solid",fgColor=GREEN_C); ws.cell(row=row, column=1).alignment = ca_; row += 1
for s in ["삼성/현대차 핵심 공급사","베트남 3공장 원가경쟁력","연구인력 57.7%, 특허24건","부채비율 54% 재무건전","액츄에이터 내재화 수익"]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"• {s}").font = df_; ws.cell(row=row, column=1).fill = gnf_
    ws.cell(row=row, column=1).alignment = la_; ws.cell(row=row, column=1).border = tb_; row += 1

ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="약점 (W)").font = Font(name="맑은 고딕",size=12,bold=True,color=W)
ws.cell(row=row, column=1).fill = PatternFill("solid",fgColor=RED_C); ws.cell(row=row, column=1).alignment = ca_; row += 1
for w_ in ["삼성전자 매출 편중 70%+","조립중심 저마진 3~5%","자체 브랜드 부재","원재료 가격 통제력 없음","스마트폰 성숙기"]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"• {w_}").font = df_; ws.cell(row=row, column=1).fill = rdf_
    ws.cell(row=row, column=1).alignment = la_; ws.cell(row=row, column=1).border = tb_; row += 1

ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="기회 (O)").font = Font(name="맑은 고딕",size=12,bold=True,color=W)
ws.cell(row=row, column=1).fill = PatternFill("solid",fgColor="2980B9"); ws.cell(row=row, column=1).alignment = ca_; row += 1
for o_ in ["ADAS 의무화→카메라수 증가","자율주행 L3/L4 본격화","프리즘줌 고부가 모듈","삼성 XR 신규 카테고리"]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"• {o_}").font = df_; ws.cell(row=row, column=1).fill = blf_
    ws.cell(row=row, column=1).alignment = la_; ws.cell(row=row, column=1).border = tb_; row += 1

ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="위협 (T)").font = Font(name="맑은 고딕",size=12,bold=True,color=W)
ws.cell(row=row, column=1).fill = PatternFill("solid",fgColor="7F8C8D"); ws.cell(row=row, column=1).alignment = ca_; row += 1
for t_ in ["중국 업체 추격","삼성 모듈 내재화 가능성","환율/인건비 상승","글로벌 경기침체"]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"• {t_}").font = df_; ws.cell(row=row, column=1).fill = PatternFill("solid",fgColor="E5E7E9")
    ws.cell(row=row, column=1).alignment = la_; ws.cell(row=row, column=1).border = tb_; row += 1

# ==== 13. 10년 실적 요약 ====
row += 1
row = msec(ws, row, "10년 실적 요약 (억원)")
mhdr(ws, row, ["연도", "매출/영업이익", "순이익/EPS"]); row += 1
hist = [(2015,5029,263,162,1833),(2016,4125,-239,-247,-2824),(2017,6685,197,-9,-100),
        (2018,6970,411,298,2161),(2019,12677,1131,845,5151),(2020,13113,592,384,2180),
        (2021,10093,239,395,2229),(2022,11086,107,230,1293),(2023,9325,182,279,1585),(2024,10571,444,636,3661)]
for yr, rev, op, ni, eps in hist:
    nif = gn_ if ni > 0 else rd_
    mrow(ws, row, [str(yr), f"{rev:,} / {op:,}", f"{ni:,} / {eps:,}"],
         fonts=[db_, df_, nif], fills=[llf_, wf_, gnf_ if ni > 200 else (rdf_ if ni < 0 else wf_)],
         als=[ca_, ra_, ra_]); row += 1

# ==== 14. 모니터링 ====
row += 1
row = msec(ws, row, "핵심 모니터링 항목")
monitors = [
    ("1", "갤럭시 카메라 스펙", "S/Z 시리즈 카메라 사양이 매출 직접 결정. 1~2월, 7~8월 언팩"),
    ("2", "전장 매출비중", "30~40% 넘으면 리레이팅. 현대차 ADAS 채택률"),
    ("3", "분기 영업이익률", "2~3%=저수익|4~5%=양호|6%+=호황"),
    ("4", "VINA 법인실적", "연결순이익 ~50% 창출. 가동률/인건비"),
    ("5", "주주환원 정책", "배당증가+자사주+소각 지속 여부"),
]
for rank, title, detail in monitors:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"{rank}. {title}").font = db_
    ws.cell(row=row, column=1).fill = gld_; ws.cell(row=row, column=1).alignment = la_; ws.cell(row=row, column=1).border = tb_
    ws.row_dimensions[row].height = 22; row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=detail).font = df_
    ws.cell(row=row, column=1).alignment = la_; ws.cell(row=row, column=1).border = tb_
    ws.row_dimensions[row].height = 36; row += 1

# Calendar
row += 1
row = msec(ws, row, "모니터링 캘린더")
mhdr(ws, row, ["시기", "이벤트", "중요도"]); row += 1
for t_, e_, imp in [("1~2월","갤럭시S 언팩","★★★"),("2월","4Q+연간 실적","★★★"),
                     ("3월","사업보고서/주총","★★★"),("5월","1Q 실적","★★☆"),
                     ("7~8월","갤럭시Z+2Q실적","★★★"),("11월","3Q 실적","★★☆"),
                     ("12월","배당 결정","★★☆")]:
    mrow(ws, row, [t_, e_, imp], fonts=[db_, df_, db_], fills=[llf_, wf_, gld_]); row += 1

# ==== Footer ====
row += 2
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="* 데이터: OpenDART 574건 공시 전수분석 | 분석일: 2026.02.06").font = sm_
ws.cell(row=row, column=1).alignment = ca_
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="* 본 보고서는 투자 참고용이며 투자판단의 책임은 투자자에게 있습니다").font = sm_
ws.cell(row=row, column=1).alignment = ca_

# === SAVE ===
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "엠씨넥스_모바일용.xlsx")
wb.save(OUT)
conn.close()
print(f"모바일용 보고서 생성: {OUT}")
print(f"총 {row}행, 3열 구성")
