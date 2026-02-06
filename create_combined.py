# -*- coding: utf-8 -*-
"""엠씨넥스 종합 기업분석 + 밸류에이션 보고서 (현재가 28,100원 기준)"""
import sqlite3, sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "mcnex_ai.db")
BASE = os.path.dirname(os.path.abspath(__file__))
conn = sqlite3.connect(DB)

# === KEY CONSTANTS ===
PRICE = 28100
SHARES = 17977732
SHARES_TREASURY = 1110000
SHARES_WA = 17373105
억 = 100_000_000
MCAP = PRICE * SHARES  # ~5,052억

# 2024 Annual
REV24 = 1057058423929; OP24 = 44384671816; NI24 = 63604930146; EPS24 = 3661
EQ24 = 372224146153; EQ23 = 323612702499; ASSETS24 = 574103551032; LIAB24 = 201879404879
CASH24 = 35881602429; ST_DEBT = 44402388000; LT_DEBT = 1308000000
DA24 = 44255618814 + 462932726 + 1740746569
OPCF24 = 82470298068; CAPEX24 = 36541908724
DIV_PAID = 10570639200; TREAS_BUY = 15656989850; DPS24 = 800

# Quarters
Q24 = [{"rev":287835057669,"op":15204634567,"ni":15929655236},  # Q1
       {"rev":225759412794,"op":7650453061,"ni":18422813873},   # Q2
       {"rev":245070688152,"op":3077406264,"ni":2504912286},    # Q3
       {"rev":REV24-287835057669-225759412794-245070688152,      # Q4
        "op":OP24-15204634567-7650453061-3077406264,
        "ni":NI24-15929655236-18422813873-2504912286}]
Q25 = [{"rev":347917e6,"op":20918e6,"ni":20068e6},
       {"rev":326424e6,"op":16171e6,"ni":13050e6},
       {"rev":303537e6,"op":6708e6,"ni":9537e6}]
CUM25 = {"rev":977877e6,"op":43797e6,"ni":42656e6}

# Derived
BPS = EQ24 / SHARES
NET_DEBT = ST_DEBT + LT_DEBT - CASH24
EBITDA24 = OP24 + DA24; FCF24 = OPCF24 - CAPEX24; EV = MCAP + NET_DEBT
TRAIL_NI = Q24[3]["ni"]+Q25[0]["ni"]+Q25[1]["ni"]+Q25[2]["ni"]
TRAIL_OP = Q24[3]["op"]+Q25[0]["op"]+Q25[1]["op"]+Q25[2]["op"]
TRAIL_REV = Q24[3]["rev"]+Q25[0]["rev"]+Q25[1]["rev"]+Q25[2]["rev"]
TRAIL_EPS = TRAIL_NI / SHARES_WA
E25_ANN_EPS = (CUM25["ni"]*4/3) / SHARES_WA
E25_Q4LIKE_EPS = (CUM25["ni"]+Q24[3]["ni"]) / SHARES_WA
E25_CONS_EPS = (CUM25["ni"]+(Q25[1]["ni"]+Q25[2]["ni"])/2) / SHARES_WA
AVG_EQ = (EQ24+EQ23)/2; ROE24 = NI24/AVG_EQ

# === STYLE ===
NAVY="1B2A4A"; DARK="2C3E6B"; MID="3A5BA0"; LB="D6E4F0"; LLB="EBF1F8"; W="FFFFFF"
GOLD_C="D4A843"; RED_C="C0392B"; GREEN_C="27AE60"; GRAY_C="F2F2F2"

title_font=Font(name="맑은 고딕",size=22,bold=True,color=W)
sub_font=Font(name="맑은 고딕",size=11,color="B0C4DE")
sec_font=Font(name="맑은 고딕",size=14,bold=True,color=NAVY)
hdr_font=Font(name="맑은 고딕",size=10,bold=True,color=W)
df=Font(name="맑은 고딕",size=10)
db=Font(name="맑은 고딕",size=10,bold=True)
d_blue=Font(name="맑은 고딕",size=10,bold=True,color="0000FF")
d_grn=Font(name="맑은 고딕",size=10,bold=True,color=GREEN_C)
d_red=Font(name="맑은 고딕",size=10,bold=True,color=RED_C)
d_navy=Font(name="맑은 고딕",size=12,bold=True,color=NAVY)
sm=Font(name="맑은 고딕",size=9,color="666666")

tf=PatternFill("solid",fgColor=NAVY)
hf=PatternFill("solid",fgColor=DARK)
mf=PatternFill("solid",fgColor=MID)
lf=PatternFill("solid",fgColor=LB)
llf=PatternFill("solid",fgColor=LLB)
gf=PatternFill("solid",fgColor=GRAY_C)
wf=PatternFill("solid",fgColor=W)
gld=PatternFill("solid",fgColor="FFF3CD")
rdf=PatternFill("solid",fgColor="F8D7DA")
gnf=PatternFill("solid",fgColor="D4EDDA")
blf=PatternFill("solid",fgColor="D6EAF8")

ca=Alignment(horizontal='center',vertical='center',wrap_text=True)
la=Alignment(horizontal='left',vertical='center',wrap_text=True)
ra=Alignment(horizontal='right',vertical='center')
tb=Border(left=Side('thin',color="D9D9D9"),right=Side('thin',color="D9D9D9"),
          top=Side('thin',color="D9D9D9"),bottom=Side('thin',color="D9D9D9"))
bb=Border(bottom=Side('medium',color=NAVY))
NF='#,##0'; PF='0.0%'

def sw(ws,w):
    for i,v in enumerate(w,1): ws.column_dimensions[get_column_letter(i)].width=v
def wh(ws,r,h,fills=None):
    for i,v in enumerate(h,1):
        c=ws.cell(row=r,column=i,value=v); c.font=hdr_font; c.fill=fills[i-1] if fills else hf; c.alignment=ca; c.border=tb
def wr(ws,r,d,fonts=None,fills=None,als=None,nfs=None):
    for i,v in enumerate(d,1):
        c=ws.cell(row=r,column=i,value=v)
        c.font=fonts[i-1] if fonts else df; c.fill=fills[i-1] if fills else wf
        c.alignment=als[i-1] if als else ca; c.border=tb
        if nfs and i<=len(nfs) and nfs[i-1]: c.number_format=nfs[i-1]
def st(ws,r,t,ce=11):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ce)
    c=ws.cell(row=r,column=1,value=t); c.font=sec_font; c.border=bb; return r+1
def fmt(v): return f"{v/억:,.0f}억"
def fw(v): return f"{v:,.0f}원"
def pct(v): return f"{v*100:.1f}%"

wb = Workbook()

# ============================================================
# SHEET 1: 표지
# ============================================================
ws = wb.active; ws.title="표지"; ws.sheet_properties.tabColor=NAVY
sw(ws,[3,20,20,20,20,20,3])
for r in range(1,35):
    for c in range(1,8): ws.cell(row=r,column=c).fill=tf
ws.merge_cells('B6:F6')
ws.cell(row=6,column=2,value="엠씨넥스(MCNEX)").font=Font(name="맑은 고딕",size=32,bold=True,color=W)
ws.cell(row=6,column=2).alignment=Alignment(horizontal='center',vertical='center')
ws.merge_cells('B8:F8')
ws.cell(row=8,column=2,value="심층 기업분석 + 밸류에이션 종합보고서").font=Font(name="맑은 고딕",size=18,color=GOLD_C)
ws.cell(row=8,column=2).alignment=ca
ws.merge_cells('B11:F11')
ws.cell(row=11,column=2,value="097520 (유가증권)  |  기타 영상기기 제조업  |  CCM(카메라모듈) 전문기업").font=sub_font
ws.cell(row=11,column=2).alignment=ca
info=[
    (14,"현재주가",f"{PRICE:,}원 (2026.02.06 기준)"),
    (15,"시가총액",f"{fmt(MCAP)} (유통시총 {fmt(PRICE*(SHARES-SHARES_TREASURY))})"),
    (16,"대표이사 / 설립일","민동욱 / 2004.12.22"),
    (17,"본사","인천 연수구 송도과학로16번길 13-39 엠씨넥스타워"),
    (18,"주요고객","삼성전자(모바일 CCM), 현대모비스(전장 카메라)"),
    (19,"생산기지","한국(R&D+생산), 베트남(VINA 1~3공장)"),
    (20,"분석기준일","2026년 2월 6일 | 데이터: OpenDART 574건 전수분석"),
]
for r,lbl,val in info:
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
    ws.cell(row=r,column=2,value=lbl).font=Font(name="맑은 고딕",size=11,color="8899AA")
    ws.cell(row=r,column=2).alignment=Alignment(horizontal='right',vertical='center')
    ws.merge_cells(start_row=r,start_column=4,end_row=r,end_column=6)
    ws.cell(row=r,column=4,value=val).font=Font(name="맑은 고딕",size=11,bold=True,color=W)
    ws.cell(row=r,column=4).alignment=la
ws.merge_cells('B23:F26')
c=ws.cell(row=23,column=2)
c.value=(f"핵심 밸류에이션 (현재가 {PRICE:,}원 기준):\n"
    f"• PER {PRICE/EPS24:.1f}배(2024) / {PRICE/TRAIL_EPS:.1f}배(T4Q) — 업종 10~15배 대비 저평가\n"
    f"• PBR {PRICE/BPS:.2f}배 | EV/EBITDA {EV/EBITDA24:.1f}배 | FCF수익률 {FCF24/SHARES/PRICE*100:.1f}%\n"
    f"• 2025 9M: 매출 YoY+29%, 영업이익 YoY+69%. 연환산 매출 1.3조 역대 최고")
c.font=Font(name="맑은 고딕",size=10,color=GOLD_C)
c.alignment=Alignment(horizontal='left',vertical='top',wrap_text=True)
ws.merge_cells('B28:F28')
ws.cell(row=28,column=2,value="시나리오별 목표: Bull ~55,000원(+96%) | Base ~40,000원(+42%) | Bear ~18,500원(-34%)").font=Font(name="맑은 고딕",size=10,bold=True,color="AED6F1")
ws.cell(row=28,column=2).alignment=ca
print("  [1/12] 표지")

# ============================================================
# SHEET 2: 핵심실적 (from original)
# ============================================================
ws2=wb.create_sheet("핵심실적"); ws2.sheet_properties.tabColor=DARK
sw(ws2,[14,14,14,14,12,14,14,14,12,12,14])
perf=conn.execute("SELECT * FROM v_annual_performance").fetchall()
row=1; ws2.merge_cells('A1:K1')
ws2.cell(row=1,column=1,value="10년 연결 재무실적 (단위: 억원)").font=sec_font
ws2.cell(row=1,column=1).border=bb; row=3
wh(ws2,row,["연도","매출액","영업이익","순이익","EPS(원)","총자산","총부채","총자본","부채비율","영업이익률","ROE"])
row+=1; prev_eq=None
for p in perf:
    yr=p[0]; rev=int(p[1])//억 if p[1] else 0; op=int(p[2])//억 if p[2] else 0
    ni=int(p[3])//억 if p[3] else 0; eps=int(p[4]) if p[4] else 0
    ta=int(p[5])//억 if p[5] else 0; tl=int(p[6])//억 if p[6] else 0; te=int(p[7])//억 if p[7] else 0
    fts=[db]+[df]*10; fls=[lf]+[wf]*10
    if op<0: fts[2]=d_red
    if ni<0: fts[3]=d_red
    wr(ws2,row,[yr,rev,op,ni,eps,ta,tl,te,None,None,None],fonts=fts,fills=fls,
       als=[ca]+[ra]*10,nfs=[None]+[NF]*7+[PF,PF,PF])
    cl=get_column_letter
    ws2.cell(row=row,column=9,value=f"={cl(7)}{row}/{cl(8)}{row}").number_format=PF
    ws2.cell(row=row,column=10,value=f"={cl(3)}{row}/{cl(2)}{row}").number_format=PF
    if prev_eq is not None:
        ws2.cell(row=row,column=11,value=f"={cl(4)}{row}/(({cl(8)}{row}+{cl(8)}{row-1})/2)").number_format=PF
    else:
        ws2.cell(row=row,column=11,value=f"={cl(4)}{row}/{cl(8)}{row}").number_format=PF
    for c in [9,10,11]: ws2.cell(row=row,column=c).font=df; ws2.cell(row=row,column=c).alignment=ra; ws2.cell(row=row,column=c).border=tb
    prev_eq=te; row+=1
row+=2; st(ws2,row,"전년대비 성장률 (YoY)"); row+=1
wh(ws2,row,["연도","매출 YoY","영업이익 YoY","순이익 YoY"]); row+=1
ds=4
for i in range(1,len(perf)):
    r=ds+i; ws2.cell(row=row,column=1,value=perf[i][0]).font=db
    ws2.cell(row=row,column=1).fill=lf; ws2.cell(row=row,column=1).alignment=ca; ws2.cell(row=row,column=1).border=tb
    for ci,sc in [(2,'B'),(3,'C'),(4,'D')]:
        c=ws2.cell(row=row,column=ci,value=f'=IF({sc}{ds+i-1}=0,"-",({sc}{ds+i}-{sc}{ds+i-1})/{sc}{ds+i-1})')
        c.number_format=PF; c.font=df; c.alignment=ra; c.border=tb
    row+=1
print("  [2/12] 핵심실적")

# ============================================================
# SHEET 3: 2025실적 (from original)
# ============================================================
ws3=wb.create_sheet("2025실적"); ws3.sheet_properties.tabColor=GREEN_C
sw(ws3,[14,14,14,14,14,16,16])
row=1; ws3.merge_cells('A1:G1')
ws3.cell(row=1,column=1,value="2025년 분기별 잠정실적 (단위: 백만원)").font=sec_font
ws3.cell(row=1,column=1).border=bb; row=3
wh(ws3,row,["분기","매출액","영업이익","순이익","영업이익률","매출 YoY","영업이익 YoY"]); row+=1
qd=[("24Q1",287835,15205,15930),("24Q2",225759,7650,18423),("24Q3",245071,3077,2505),
    ("24Q4",298393,18451,26747),("24합계",1057058,44385,63605),
    ("25Q1",347917,20918,20068),("25Q2",326424,16171,13050),("25Q3",303537,6708,9537),("25누계",977877,43797,42656)]
for i,(q,rv,op,ni) in enumerate(qd):
    tot="합계" in q or "누계" in q; is25="25" in q
    fl=gld if tot else (gnf if is25 else wf); fn=db if tot else df
    wr(ws3,row,[q,rv,op,ni,None,None,None],fonts=[fn]*7,fills=[fl]*7,als=[ca]+[ra]*6,nfs=[None,NF,NF,NF,PF,PF,PF])
    ws3.cell(row=row,column=5,value=f"=C{row}/B{row}").number_format=PF; ws3.cell(row=row,column=5).border=tb
    if 5<=i<=7:
        pr=row-5
        ws3.cell(row=row,column=6,value=f"=(B{row}-B{pr})/B{pr}").number_format=PF; ws3.cell(row=row,column=6).border=tb
        ws3.cell(row=row,column=7,value=f"=(C{row}-C{pr})/C{pr}").number_format=PF; ws3.cell(row=row,column=7).border=tb
    row+=1
row+=2; st(ws3,row,"핵심 포인트",7); row+=1
for pt in ["9M 누적 매출 9,779억 → 2023년 연간(9,325억) 이미 초과. 역대최고 매출 확실시",
           "3Q OPM 2.2% 급락 → 계절적 제품믹스 변화 or 일회성 비용 확인 필요",
           "1Q~2Q 갤럭시S 효과로 고마진. 상반기 이익이 연간의 85% 비중"]:
    ws3.merge_cells(start_row=row,start_column=1,end_row=row,end_column=7)
    ws3.cell(row=row,column=1,value=f"• {pt}").font=df; ws3.cell(row=row,column=1).alignment=la; row+=1
print("  [3/12] 2025실적")

# ============================================================
# SHEET 4: 사업구조 (from original)
# ============================================================
ws4=wb.create_sheet("사업구조"); ws4.sheet_properties.tabColor=GOLD_C
sw(ws4,[18,14,14,14,14,14,14,14])
row=1; ws4.merge_cells('A1:H1')
ws4.cell(row=1,column=1,value="사업부문별 매출 구조 (연결, 단위: 억원)").font=sec_font
ws4.cell(row=1,column=1).border=bb; row=3
wh(ws4,row,["구분","2020","2021","2022","2023","비중(23)","추세"]); row+=1
for nm,y20,y21,y22,y23 in [("휴대폰용",11460,8373,9246,6583),("자동차용",1534,1602,1701,2527),
                             ("기타",119,118,127,215),("합계",13113,10093,11086,9325)]:
    tot=nm=="합계"; fl=gld if tot else wf
    wr(ws4,row,[nm,y20,y21,y22,y23,None,None],fonts=[db if tot else df]*7,fills=[fl]*7,
       als=[la]+[ra]*6,nfs=[None]+[NF]*4+[PF,None])
    if not tot:
        ws4.cell(row=row,column=6,value=f"=E{row}/E{row+(3-(row-4))}").number_format=PF; ws4.cell(row=row,column=6).border=tb
        t="↑" if y23>y22 else "↓"; ws4.cell(row=row,column=7,value=t).font=d_grn if t=="↑" else d_red
        ws4.cell(row=row,column=7).alignment=ca; ws4.cell(row=row,column=7).border=tb
    row+=1
row+=2; st(ws4,row,"종속회사 현황 (2024)",8); row+=1
wh(ws4,row,["자회사명","지분율","역할","장부가(억)","순이익(억)","총자산(억)"]); row+=1
for nm,p,role,bv,ni,ta in [("엠씨넥스VINA(베트남)","100%","핵심 생산기지",1216,309,3355),
                             ("엠씨넥스상해(중국)","100%","수출입/CS",1,-1,14),
                             ("엠씨넥스에프앤비","100%","구내식당",10,-0.4,10)]:
    wr(ws4,row,[nm,p,role,bv,ni,ta],fonts=[db,df,df,df,d_grn if ni>0 else d_red,df],als=[la,ca,la,ra,ra,ra])
    row+=1
print("  [4/12] 사업구조")

# ============================================================
# SHEET 5: 주주환원 (from original)
# ============================================================
ws5=wb.create_sheet("주주환원"); ws5.sheet_properties.tabColor="E74C3C"
sw(ws5,[12,14,12,14,14,14])
row=1; ws5.merge_cells('A1:F1')
ws5.cell(row=1,column=1,value="배당 및 주주환원 정책").font=sec_font; ws5.cell(row=1,column=1).border=bb; row=3
wh(ws5,row,["연도","주당배당(원)","EPS(원)","배당성향","배당수익률","비고"]); row+=1
for yr,dps,eps,po,yl,note in [
    ("2015",330,1833,"17.9%","1.17%",""),("2016",0,-2824,"-","-","적자"),
    ("2017",260,-100,"-","1.27%",""),("2018",300,2161,"17.1%","2.40%",""),
    ("2019",500,5151,"10.4%","1.70%","최대실적"),("2020",500,2180,"23.9%","1.30%",""),
    ("2021",500,2229,"22.6%","0.90%",""),("2022",500,1293,"38.5%","1.80%",""),
    ("2023",600,1585,"37.9%","2.00%",""),("2024",800,3661,"21.2%",f"{800/PRICE*100:.2f}%","대폭 인상"),
    ("2025E",1000,"-","-",f"{1000/PRICE*100:.2f}%","")]:
    wr(ws5,row,[yr,dps if dps else "-",eps,po,yl,note],
       fonts=[db,d_grn if isinstance(dps,int) and dps>0 else d_red,df,df,df,sm],als=[ca,ra,ra,ca,ca,la])
    row+=1
row+=2; st(ws5,row,"주주환원 이벤트",6); row+=1
for dt,evt in [("2024.07","자기주식 취득 결정"),("2024.11","자기주식 추가 취득"),
               ("2024.12","결산배당 주당 800원"),("2025.01","자기주식 취득"),
               ("2025.09","자기주식 취득"),("2025.11","주식 583,482주 소각 (~170억)"),("2025.12","결산배당 주당 1,000원")]:
    ws5.merge_cells(start_row=row,start_column=2,end_row=row,end_column=6)
    ws5.cell(row=row,column=1,value=dt).font=db; ws5.cell(row=row,column=1).alignment=ca
    ws5.cell(row=row,column=1).fill=llf; ws5.cell(row=row,column=1).border=tb
    ws5.cell(row=row,column=2,value=evt).font=df; ws5.cell(row=row,column=2).alignment=la; ws5.cell(row=row,column=2).border=tb
    row+=1
row+=2; st(ws5,row,f"주식 구조 (2024.12.31) — 현재가 {PRICE:,}원 기준",6); row+=1
for lbl,val in [("발행주식수(보통주)",f"{SHARES:,}주"),("자기주식",f"{SHARES_TREASURY:,}주"),
                ("유통주식수",f"{SHARES-SHARES_TREASURY:,}주"),("시가총액",fmt(MCAP)),
                ("유통시총",fmt(PRICE*(SHARES-SHARES_TREASURY))),
                ("최대주주(민동욱)","4,661,000주 (25.93%)"),("소액주주","66.33% (58,267명)")]:
    ws5.cell(row=row,column=1,value=lbl).font=df; ws5.cell(row=row,column=1).alignment=la
    ws5.cell(row=row,column=1).fill=llf; ws5.cell(row=row,column=1).border=tb
    ws5.merge_cells(start_row=row,start_column=2,end_row=row,end_column=3)
    ws5.cell(row=row,column=2,value=val).font=db; ws5.cell(row=row,column=2).alignment=la; ws5.cell(row=row,column=2).border=tb
    row+=1
print("  [5/12] 주주환원")

# ============================================================
# SHEET 6: R&D/특허 (from original)
# ============================================================
ws6=wb.create_sheet("R&D_특허"); ws6.sheet_properties.tabColor="8E44AD"
sw(ws6,[14,50,14,14])
row=1; ws6.merge_cells('A1:D1')
ws6.cell(row=1,column=1,value="연구개발 및 특허 현황").font=sec_font; ws6.cell(row=1,column=1).border=bb
row=3; st(ws6,row,"연구개발비 추이",4); row+=1
wh(ws6,row,["연도","R&D비용(억)","매출대비","비고"]); row+=1
for d in [("2022",397,"3.58%",""),("2023",356,"3.82%",""),("2024",327,"3.10%","")]:
    wr(ws6,row,list(d),fonts=[db,df,df,sm],als=[ca,ra,ca,la]); row+=1
row+=1
ws6.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
ws6.cell(row=row,column=1,value="연구인력 284명 (57.7%) | 임원15+수석107+선임106+연구원56").font=Font(name="맑은 고딕",size=10,bold=True,color=MID)
row+=2; st(ws6,row,"특허 포트폴리오 (24건)",4); row+=1
wh(ws6,row,["취득일","특허명/내용","분류","활용"]); row+=1
for p in conn.execute("SELECT rcept_dt,patent_name,patent_detail,patent_plan FROM patents ORDER BY rcept_dt").fetchall():
    dt=p[0]; nm=(p[1]or"")[:60]; det=(p[2]or"")[:80]; plan=(p[3]or"")[:40]
    cat="액츄에이터" if "액" in det or "ctuator" in det else "전장" if "차량" in det or "주차" in det or "AVM" in det else "광학/제조" if "광축" in det or "프리즘" in det else "보안" if "홍채" in det else "기타"
    disp=nm if nm and "특허" not in nm[:3] else det[:60]
    wr(ws6,row,[dt,disp,cat,plan[:30] if plan else ""],fonts=[df,df,db,sm],als=[ca,la,ca,la]); row+=1
print("  [6/12] R&D_특허")

# ============================================================
# SHEET 7: 밸류에이션 종합 (NEW - replaces old 투자지표)
# ============================================================
ws7=wb.create_sheet("밸류에이션"); ws7.sheet_properties.tabColor="E67E22"
sw(ws7,[22,18,18,18,18,22])
row=1; ws7.merge_cells('A1:F2')
c=ws7.cell(row=1,column=1,value=f"엠씨넥스 밸류에이션 종합 (현재가 {PRICE:,}원)")
c.font=Font(name="맑은 고딕",size=18,bold=True,color=W); c.fill=tf; c.alignment=ca
row=4
# Key info
for d in [["현재 주가",fw(PRICE),"발행주식수",f"{SHARES:,}주","시가총액",fmt(MCAP)],
          ["2024 EPS",fw(EPS24),"2024 BPS",fw(int(BPS)),"2024 DPS",fw(DPS24)],
          ["Trailing EPS",fw(int(TRAIL_EPS)),"ROE (2024)",pct(ROE24),"부채비율",pct(LIAB24/EQ24)]]:
    for i in range(0,6,2):
        ws7.cell(row=row,column=i+1,value=d[i]).font=db; ws7.cell(row=row,column=i+1).fill=blf
        ws7.cell(row=row,column=i+1).alignment=la; ws7.cell(row=row,column=i+1).border=tb
        ws7.cell(row=row,column=i+2,value=d[i+1]).font=d_blue; ws7.cell(row=row,column=i+2).fill=wf
        ws7.cell(row=row,column=i+2).alignment=ca; ws7.cell(row=row,column=i+2).border=tb
    row+=1
row+=1; st(ws7,row,"멀티플 종합",6); row+=1
wh(ws7,row,["지표","산출 방식","값","판정","업종평균(참고)","비고"]); row+=1

vals_list = [
    ["PER (2024 확정)",f"주가/EPS({fw(EPS24)})",f"{PRICE/EPS24:.1f}배","저평가","10~15배","확정 사업보고서"],
    ["PER (Trailing 4Q)",f"주가/T4Q EPS({fw(int(TRAIL_EPS))})",f"{PRICE/TRAIL_EPS:.1f}배","저평가","10~15배","24Q4+25Q1~Q3"],
    ["PER (2025E 연환산)",f"주가/연환산EPS({fw(int(E25_ANN_EPS))})",f"{PRICE/E25_ANN_EPS:.1f}배","적정~저평가","10~15배","9M×4/3"],
    ["PER (2025E 보수적)",f"주가/EPS({fw(int(E25_CONS_EPS))})",f"{PRICE/E25_CONS_EPS:.1f}배","적정","10~15배","Q4=Q2Q3평균"],
    ["PBR",f"주가/BPS({fw(int(BPS))})",f"{PRICE/BPS:.2f}배","적정","1.0~2.0배",""],
    ["EV/EBITDA",f"EV({fmt(EV)})/EBITDA({fmt(EBITDA24)})",f"{EV/EBITDA24:.1f}배","저평가","7~12배",""],
    ["PSR (2024)",f"시총/매출({fmt(REV24)})",f"{MCAP/REV24:.2f}배","매우저평가","0.5~1.5배",""],
    ["PSR (Trailing)",f"시총/T4Q매출({fmt(TRAIL_REV)})",f"{MCAP/TRAIL_REV:.2f}배","매우저평가","",""],
    ["PCR",f"주가/OpCF({fw(int(OPCF24/SHARES))})",f"{PRICE/(OPCF24/SHARES):.1f}배","저평가","8~15배",""],
    ["배당수익률",f"DPS({fw(DPS24)})/주가",pct(DPS24/PRICE),"양호","1~3%",""],
    ["FCF 수익률",f"FCF({fw(int(FCF24/SHARES))})/주가",pct(FCF24/SHARES/PRICE),"매우양호","3~6%",""],
    ["총주주환원율",f"(배당+자사주)/시총",pct((DIV_PAID+TREAS_BUY)/MCAP),"양호","2~5%",""],
    ["ROE",f"순이익/평균자본",pct(ROE24),"양호","8~15%",""],
]
for v in vals_list:
    j=v[3]
    if "매우" in j: jf,jfl=d_grn,gnf
    elif "저평가" in j or "양호" in j: jf,jfl=Font(name="맑은 고딕",size=10,bold=True,color="2E86C1"),blf
    elif "적정" in j: jf,jfl=db,gld
    else: jf,jfl=d_red,rdf
    wr(ws7,row,v,fonts=[db,df,d_blue,jf,df,sm],fills=[llf,wf,gld,jfl,gf,wf],als=[la,la,ca,ca,ca,la]); row+=1
print("  [7/12] 밸류에이션 종합")

# ============================================================
# SHEET 8: PER 상세 (from valuation)
# ============================================================
ws8=wb.create_sheet("PER분석"); ws8.sheet_properties.tabColor="2980B9"
sw(ws8,[24,16,16,16,16,20])
row=1; st(ws8,row,f"PER 다각도 분석 (현재가 {PRICE:,}원)",6); row+=1
st(ws8,row,"A. EPS 산출 방식별 PER",6); row+=1
wh(ws8,row,["산출 방식","순이익(억)","EPS(원)","PER(배)","의미","비고"]); row+=1
for lbl,ni_v,eps_v,desc,note in [
    ("2024 확정",int(NI24/억),EPS24,"확정치 기준","사업보고서"),
    ("Trailing 4Q",int(TRAIL_NI/억),int(TRAIL_EPS),"직전4분기","24Q4+25Q1~Q3"),
    ("2025E 연환산",int(CUM25["ni"]*4/3/억),int(E25_ANN_EPS),"단순 연율화","9M×4/3"),
    ("2025E (24Q4 대입)",int((CUM25["ni"]+Q24[3]["ni"])/억),int(E25_Q4LIKE_EPS),"전년Q4 대입","Q4 통상 강세"),
    ("2025E 보수적",int((CUM25["ni"]+(Q25[1]["ni"]+Q25[2]["ni"])/2)/억),int(E25_CONS_EPS),"하반기 둔화","Q4=Q2Q3평균")]:
    per=PRICE/eps_v
    pf2=d_grn if per<7.5 else (d_blue if per<9 else db)
    pfl=gnf if per<7.5 else (blf if per<9 else gld)
    wr(ws8,row,[lbl,ni_v,f"{eps_v:,}",f"{per:.2f}배",desc,note],
       fonts=[db,df,d_blue,pf2,df,sm],fills=[llf,wf,gld,pfl,wf,gf],als=[la,ra,ra,ca,la,la]); row+=1

# B. Quarterly
row+=1; st(ws8,row,"B. 분기별 실적 추이 (백만원)",6); row+=1
wh(ws8,row,["분기","매출액","영업이익","OPM","순이익","YoY 순이익"]); row+=1
qs=[("24Q1",Q24[0]),("24Q2",Q24[1]),("24Q3",Q24[2]),("24Q4",Q24[3]),
    ("25Q1",Q25[0]),("25Q2",Q25[1]),("25Q3",Q25[2])]
pq={"25Q1":Q24[0],"25Q2":Q24[1],"25Q3":Q24[2]}
for n,q in qs:
    rm=int(q["rev"]/1e6); om=int(q["op"]/1e6); nm_=int(q["ni"]/1e6)
    opm=q["op"]/q["rev"] if q["rev"] else 0
    yoy=""
    if n in pq:
        p=pq[n]
        if p["ni"]>0: yoy=f"{(q['ni']-p['ni'])/p['ni']*100:+.1f}%"
    is25=n.startswith("25"); fl2=gnf if is25 else wf
    wr(ws8,row,[n,f"{rm:,}",f"{om:,}",f"{opm*100:.1f}%",f"{nm_:,}",yoy],
       fonts=[db,df,df,d_grn if opm>0.05 else(d_red if opm<0.03 else df),df,db],
       fills=[blf if is25 else llf]+[fl2]*5,als=[ca,ra,ra,ca,ra,ca]); row+=1
# Trailing sum
wr(ws8,row,["Trailing 4Q",f"{int(TRAIL_REV/1e6):,}",f"{int(TRAIL_OP/1e6):,}",
            f"{TRAIL_OP/TRAIL_REV*100:.1f}%",f"{int(TRAIL_NI/1e6):,}",""],
   fonts=[db]*6,fills=[gld]*6,als=[ca,ra,ra,ca,ra,ca]); row+=1

# C. Target PER
row+=1; st(ws8,row,"C. 목표PER별 적정주가",6); row+=1
wh(ws8,row,["기준","EPS(원)","PER 7배","PER 8배","PER 10배","PER 12배"]); row+=1
for lbl,eps in [("2024 확정",EPS24),("Trailing 4Q",int(TRAIL_EPS)),
                ("2025E (24Q4대입)",int(E25_Q4LIKE_EPS)),("2025E 보수적",int(E25_CONS_EPS))]:
    wr(ws8,row,[lbl,f"{eps:,}",fw(eps*7),fw(eps*8),fw(eps*10),fw(eps*12)],
       fonts=[db,d_blue,df,df,d_grn,d_grn],fills=[llf,gld,wf,gnf if eps*8>PRICE else wf,gnf,gnf],
       als=[la,ra,ra,ra,ra,ra]); row+=1
row+=1
ws8.merge_cells(start_row=row,start_column=1,end_row=row,end_column=6)
ws8.cell(row=row,column=1,value=f"→ 현재 {PRICE:,}원 = Trailing EPS 기준 PER {PRICE/TRAIL_EPS:.1f}배. 목표PER 10배 시 적정가 ~{fw(int(TRAIL_EPS*10))}").font=db
print("  [8/12] PER분석")

# ============================================================
# SHEET 9: PBR/ROE/RIM
# ============================================================
ws9=wb.create_sheet("PBR_ROE_RIM"); ws9.sheet_properties.tabColor="8E44AD"
sw(ws9,[20,16,16,16,16,22])
row=1; st(ws9,row,"PBR / ROE / 잔여이익모델(RIM)",6); row+=1

# Historical
st(ws9,row,"A. 연도별 BPS/ROE 추이",6); row+=1
wh(ws9,row,["연도","자본(억)","BPS(원)","ROE","EPS(원)","순이익(억)"]); row+=1
HIST={2019:(1267712135419,113067104268,84507307910,5151,224567081980),
      2020:(1311342738149,59167161589,38443583445,2180,252934743419),
      2021:(1009273639991,23878951946,39505208461,2229,298966196208),
      2022:(1108637950325,10702196295,22973446783,1293,309066577971),
      2023:(932490110758,18224255538,27923690827,1585,323612702499),
      2024:(REV24,OP24,NI24,EPS24,EQ24)}
peq=None
for yr in [2019,2020,2021,2022,2023,2024]:
    rv,op,ni,eps,eq=HIST[yr]; bps=int(eq/SHARES)
    roe=ni/((eq+peq)/2) if peq else ni/eq; peq=eq
    rf=d_grn if roe>0.12 else(d_red if roe<0.05 else df)
    wr(ws9,row,[str(yr),int(eq/억),f"{bps:,}",f"{roe*100:.1f}%",f"{eps:,}",int(ni/억)],
       fonts=[db,df,d_blue,rf,df,df],fills=[llf,wf,gld,wf,wf,wf],als=[ca,ra,ra,ca,ra,ra]); row+=1

# PBR
row+=1; st(ws9,row,"B. 현재 PBR",6); row+=1
for lbl,val in [(f"BPS ({fw(int(BPS))})",f"PBR = {PRICE/BPS:.2f}배"),
                ("PBR 1.0배 주가",fw(int(BPS))),("PBR 1.5배 주가",fw(int(BPS*1.5))),("PBR 2.0배 주가",fw(int(BPS*2)))]:
    ws9.cell(row=row,column=1,value=lbl).font=db; ws9.cell(row=row,column=1).fill=llf
    ws9.cell(row=row,column=1).alignment=la; ws9.cell(row=row,column=1).border=tb
    ws9.cell(row=row,column=2,value=val).font=d_blue; ws9.cell(row=row,column=2).fill=gld
    ws9.cell(row=row,column=2).alignment=ca; ws9.cell(row=row,column=2).border=tb; row+=1

# RIM
row+=1; st(ws9,row,"C. 잔여이익모델(RIM) 적정주가",6); row+=1
ws9.cell(row=row,column=1,value="산식: BPS × (1 + (ROE-ke)/(ke-g)) | ke=10% (무위험3.5%+β1.0×ERP6.5%)").font=sm; row+=1
wh(ws9,row,["시나리오","지속ROE","ke","성장률(g)","적정주가","현재가 대비"]); row+=1
for lbl,roe,ke,g in [("보수적(ROE=ke)",0.10,0.10,0.02),("기본(과거평균)",0.12,0.10,0.02),
                      ("적극적(최근)",0.15,0.10,0.02),("낙관적(ROE유지)",0.18,0.10,0.03)]:
    fv=BPS*(1+(roe-ke)/(ke-g)); up=(fv-PRICE)/PRICE
    uf=d_grn if up>0 else d_red; ufl=gnf if up>0 else rdf
    wr(ws9,row,[lbl,f"{roe*100:.0f}%",f"{ke*100:.0f}%",f"{g*100:.0f}%",fw(int(fv)),f"{up*100:+.1f}%"],
       fonts=[db,df,df,df,d_blue,uf],fills=[llf,wf,wf,wf,gld,ufl],als=[la,ca,ca,ca,ra,ca]); row+=1
row+=1
ws9.merge_cells(start_row=row,start_column=1,end_row=row,end_column=6)
ws9.cell(row=row,column=1,value=f"→ 기본(ROE12%) 적정가 ~{fw(int(BPS*(1+0.02/0.08)))}. 적극적(15%) ~{fw(int(BPS*(1+0.05/0.08)))}. 현재가는 ROE 12~15% 반영.").font=db
print("  [9/12] PBR/ROE/RIM")

# ============================================================
# SHEET 10: EV/EBITDA & FCF
# ============================================================
ws10=wb.create_sheet("EV_EBITDA_FCF"); ws10.sheet_properties.tabColor="E67E22"
sw(ws10,[24,18,18,18,24])
row=1; st(ws10,row,"EV/EBITDA & FCF 밸류에이션",5); row+=1

# EV
st(ws10,row,"A. Enterprise Value",5); row+=1
wh(ws10,row,["항목","금액(억)","비고","",""]); row+=1
for lbl,amt,note in [("시가총액",int(MCAP/억),f"주가{PRICE:,}×{SHARES:,}주"),
                      ("(+) 총차입금",int((ST_DEBT+LT_DEBT)/억),f"단기{int(ST_DEBT/억)}+장기{int(LT_DEBT/억)}"),
                      ("(-) 현금",int(CASH24/억),""),("(=) 순차입금",int(NET_DEBT/억),""),
                      ("(=) EV",int(EV/억),"시총+순차입금")]:
    tot="(=)" in lbl
    wr(ws10,row,[lbl,f"{amt:,}",note,"",""],fonts=[db if tot else df,d_blue if tot else df,sm,df,df],
       fills=[gld if tot else llf,gld if tot else wf,wf,wf,wf],als=[la,ra,la,la,la]); row+=1

# EBITDA
row+=1; st(ws10,row,"B. EBITDA",5); row+=1
wh(ws10,row,["항목","2024(억)","비고","",""]); row+=1
for lbl,amt,note in [("영업이익",int(OP24/억),""),("(+) 감가상각비",int(44255618814/억),"유형자산"),
                      ("(+) 투자부동산상각",int(462932726/억),""),("(+) 무형자산상각",int(1740746569/억),""),
                      ("(=) EBITDA",int(EBITDA24/억),f"마진 {EBITDA24/REV24*100:.1f}%")]:
    tot="(=)" in lbl
    wr(ws10,row,[lbl,f"{amt:,}",note,"",""],fonts=[db if tot else df,d_blue if tot else df,sm,df,df],
       fills=[gld if tot else llf,gld if tot else wf,wf,wf,wf],als=[la,ra,la,la,la]); row+=1

# EV/EBITDA scenarios
row+=1; st(ws10,row,"C. EV/EBITDA 적정주가",5); row+=1
wh(ws10,row,["기준","EBITDA(억)","현재배수","적정EV(8배)","적정주가(8배)"]); row+=1
trail_ebitda=int(TRAIL_OP/억)+int(DA24/억)
e25_ebitda=int(CUM25["op"]*4/3/억)+int(DA24/억)
for lbl,eb in [("2024 확정",int(EBITDA24/억)),("Trailing 4Q",trail_ebitda),("2025E 연환산",e25_ebitda)]:
    ev_eb=(EV/억)/eb; fev8=eb*8; fp=int((fev8-int(NET_DEBT/억))*억/SHARES)
    wr(ws10,row,[lbl,f"{eb:,}",f"{ev_eb:.1f}배",f"{fev8:,}억",fw(fp)],
       fonts=[db,df,d_blue,df,d_grn],fills=[llf,wf,gld,wf,gnf],als=[la,ra,ca,ra,ra]); row+=1

# FCF
row+=1; st(ws10,row,"D. FCF & 주주환원",5); row+=1
wh(ws10,row,["항목","2024(억)","비고","",""]); row+=1
for lbl,amt,note in [("영업활동CF",int(OPCF24/억),""),("(-) CAPEX",int(CAPEX24/억),"유형자산취득"),
                      ("(=) FCF",int(FCF24/억),""),("FCF/주","",fw(int(FCF24/SHARES))),
                      ("FCF 수익률","",pct(FCF24/SHARES/PRICE)),("","",""),
                      ("배당금 지급",int(DIV_PAID/억),f"{fw(DPS24)}/주"),
                      ("자기주식 취득",int(TREAS_BUY/억),""),
                      ("총 주주환원",int((DIV_PAID+TREAS_BUY)/억),""),
                      ("총주주환원율","",pct((DIV_PAID+TREAS_BUY)/MCAP))]:
    if not lbl: row+=1; continue
    tot="(=)" in lbl or "수익률" in lbl or "환원율" in lbl or "총" in lbl
    v2=f"{amt:,}" if isinstance(amt,int) and amt>0 else note
    v3=note if isinstance(amt,int) and amt>0 else ""
    wr(ws10,row,[lbl,v2,v3,"",""],fonts=[db if tot else df,d_blue if tot else df,sm,df,df],
       fills=[gld if tot else llf,gld if tot else wf,wf,wf,wf],als=[la,ra,la,la,la]); row+=1
row+=1
ws10.merge_cells(start_row=row,start_column=1,end_row=row,end_column=5)
ws10.cell(row=row,column=1,value=f"→ FCF수익률 {FCF24/SHARES/PRICE*100:.1f}%, 총주주환원 {pct((DIV_PAID+TREAS_BUY)/MCAP)}. 시총 대비 연간 460억 잉여현금 창출.").font=db
print("  [10/12] EV/EBITDA/FCF")

# ============================================================
# SHEET 11: 시나리오 + SWOT (updated prices)
# ============================================================
ws11=wb.create_sheet("시나리오_SWOT"); ws11.sheet_properties.tabColor="2ECC71"
sw(ws11,[22,18,18,18,22])
row=1; st(ws11,row,f"시나리오별 목표주가 (현재가 {PRICE:,}원)",5); row+=1

# Scenario table
st(ws11,row,"A. 방법론별 적정가 레인지",5); row+=1
wh(ws11,row,["방법론","보수적","기본","적극적","산출 근거"]); row+=1
for lbl,c_,b_,a_,note in [
    ("PER 방식",fw(int(E25_CONS_EPS*7)),fw(int(TRAIL_EPS*10)),fw(int(TRAIL_EPS*12)),"EPS×목표PER"),
    ("PBR 방식",fw(int(BPS)),fw(int(BPS*1.5)),fw(int(BPS*2)),"BPS×목표PBR"),
    ("EV/EBITDA",fw(int((int(EBITDA24/억)*6-int(NET_DEBT/억))*억/SHARES)),
     fw(int((int(EBITDA24/억)*8-int(NET_DEBT/억))*억/SHARES)),
     fw(int((int(EBITDA24/억)*10-int(NET_DEBT/억))*억/SHARES)),"EBITDA×배수"),
    ("RIM",fw(int(BPS)),fw(int(BPS*1.25)),fw(int(BPS*1.625)),"ROE기반"),
    ("FCF기반",fw(int(FCF24*8/SHARES)),fw(int(FCF24*10/SHARES)),fw(int(FCF24*14/SHARES)),"FCF×배수")]:
    wr(ws11,row,[lbl,c_,b_,a_,note],fonts=[db,df,d_blue,d_grn,sm],fills=[llf,rdf,gld,gnf,wf],als=[la,ra,ra,ra,la]); row+=1

# Bull/Base/Bear
row+=1; st(ws11,row,"B. 종합 시나리오",5); row+=1
wh(ws11,row,["시나리오","목표주가","현재가 대비","전제조건","확률(주관)"],
   fills=[PatternFill("solid",fgColor=RED_C)]*5); row+=1
bull_eps=int(TRAIL_EPS*1.15); bull_t=bull_eps*12; bull_up=(bull_t-PRICE)/PRICE
base_t=int(TRAIL_EPS*10); base_up=(base_t-PRICE)/PRICE
bear_eps=int(E25_CONS_EPS*0.85); bear_t=bear_eps*7; bear_up=(bear_t-PRICE)/PRICE
exp=int(bull_t*0.2+base_t*0.5+bear_t*0.3); exp_up=(exp-PRICE)/PRICE

wr(ws11,row,["강세(Bull)",fw(bull_t),f"{bull_up*100:+.1f}%","갤럭시 업사이클+전장35%+PER리레이팅","20%"],
   fonts=[db,Font(name="맑은 고딕",size=12,bold=True,color=GREEN_C),d_grn,df,df],fills=[gnf]*5,als=[ca,ra,ca,la,ca]); row+=1
wr(ws11,row,["기본(Base)",fw(base_t),f"{base_up*100:+.1f}%","모바일유지+전장10%성장+배당확대","50%"],
   fonts=[db,d_navy,d_blue,df,df],fills=[gld]*5,als=[ca,ra,ca,la,ca]); row+=1
wr(ws11,row,["약세(Bear)",fw(bear_t),f"{bear_up*100:+.1f}%","삼성부진+전장정체+마진악화","30%"],
   fonts=[db,Font(name="맑은 고딕",size=12,bold=True,color=RED_C),d_red,df,df],fills=[rdf]*5,als=[ca,ra,ca,la,ca]); row+=1
row+=1
wr(ws11,row,["확률가중 기대값",fw(exp),f"{exp_up*100:+.1f}%","Bull×20%+Base×50%+Bear×30%",""],
   fonts=[Font(name="맑은 고딕",size=12,bold=True,color=NAVY)]*5,fills=[blf]*5,als=[ca,ra,ca,la,ca]); row+=1

# SWOT
row+=2; st(ws11,row,"C. SWOT 분석",5); row+=1
ws11.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)
ws11.cell(row=row,column=1,value="강점 (S)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws11.cell(row=row,column=1).fill=PatternFill("solid",fgColor=GREEN_C); ws11.cell(row=row,column=1).alignment=ca
ws11.merge_cells(start_row=row,start_column=3,end_row=row,end_column=4)
ws11.cell(row=row,column=3,value="약점 (W)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws11.cell(row=row,column=3).fill=PatternFill("solid",fgColor=RED_C); ws11.cell(row=row,column=3).alignment=ca; row+=1
SW_S=["삼성/현대차 핵심 공급사","베트남 3공장 원가경쟁력","연구인력 57.7%, 특허24건","부채비율 54% 재무건전","액츄에이터 내재화"]
SW_W=["삼성전자 매출편중 70%+","조립중심 저마진 3~5%","자체 브랜드 부재","원재료 가격 통제력 없음","스마트폰 성숙기"]
for i in range(5):
    ws11.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)
    ws11.cell(row=row,column=1,value=f"• {SW_S[i]}").font=df; ws11.cell(row=row,column=1).fill=gnf
    ws11.cell(row=row,column=1).alignment=la; ws11.cell(row=row,column=1).border=tb
    ws11.merge_cells(start_row=row,start_column=3,end_row=row,end_column=4)
    ws11.cell(row=row,column=3,value=f"• {SW_W[i]}").font=df; ws11.cell(row=row,column=3).fill=rdf
    ws11.cell(row=row,column=3).alignment=la; ws11.cell(row=row,column=3).border=tb; row+=1
row+=1
ws11.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)
ws11.cell(row=row,column=1,value="기회 (O)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws11.cell(row=row,column=1).fill=PatternFill("solid",fgColor="2980B9"); ws11.cell(row=row,column=1).alignment=ca
ws11.merge_cells(start_row=row,start_column=3,end_row=row,end_column=4)
ws11.cell(row=row,column=3,value="위협 (T)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws11.cell(row=row,column=3).fill=PatternFill("solid",fgColor="7F8C8D"); ws11.cell(row=row,column=3).alignment=ca; row+=1
O=["ADAS 의무화→카메라수 증가","자율주행 L3/L4 본격화","프리즘줌 고부가 모듈","삼성 XR 신규 카테고리"]
T=["중국 업체 추격","삼성 모듈 내재화 가능성","환율/인건비 상승","글로벌 경기침체"]
for i in range(4):
    ws11.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)
    ws11.cell(row=row,column=1,value=f"• {O[i]}").font=df; ws11.cell(row=row,column=1).fill=blf
    ws11.cell(row=row,column=1).alignment=la; ws11.cell(row=row,column=1).border=tb
    ws11.merge_cells(start_row=row,start_column=3,end_row=row,end_column=4)
    ws11.cell(row=row,column=3,value=f"• {T[i]}").font=df; ws11.cell(row=row,column=3).fill=PatternFill("solid",fgColor="E5E7E9")
    ws11.cell(row=row,column=3).alignment=la; ws11.cell(row=row,column=3).border=tb; row+=1
print("  [11/12] 시나리오/SWOT")

# ============================================================
# SHEET 12: 모니터링 (from original)
# ============================================================
ws12=wb.create_sheet("모니터링"); ws12.sheet_properties.tabColor="1ABC9C"
sw(ws12,[6,30,40,20])
row=1; ws12.merge_cells('A1:D1')
ws12.cell(row=1,column=1,value="향후 핵심 모니터링 지표").font=sec_font; ws12.cell(row=1,column=1).border=bb; row=3
wh(ws12,row,["#","항목","세부 내용","확인 시기"]); row+=1
for rank,title,detail,timing in [
    ("1","삼성 갤럭시 카메라 스펙/물량","갤럭시 S/Z 카메라 화소·개수·OIS/줌 사양이 ASP와 매출 직접 결정","1~2월, 7~8월"),
    ("2","전장(자동차) 매출 비중","30~40% 넘으면 밸류에이션 리레이팅. 현대차 ADAS 채택률 확인","매 분기"),
    ("3","분기별 영업이익률","2~3%=저수익|4~5%=양호|6%+=호황. 원재료·환율·믹스가 핵심","2/5/8/11월"),
    ("4","베트남 VINA 실적","연결순이익 ~50% 창출. 가동률/수율/인건비 모니터링","연간"),
    ("5","자기주식/배당 정책","적극적 주주환원(배당증가+자사주+소각) 지속 여부","3월 주총, 수시")]:
    wr(ws12,row,[rank,title,detail,timing],
       fonts=[Font(name="맑은 고딕",size=14,bold=True,color=NAVY),db,df,df],
       fills=[gld,llf,wf,llf],als=[ca,la,la,ca])
    ws12.row_dimensions[row].height=45; row+=1
row+=2; st(ws12,row,"모니터링 캘린더",4); row+=1
wh(ws12,row,["시기","이벤트","중요도"]); row+=1
for t,e,imp in [("1~2월","갤럭시S 언팩 (카메라사양)","★★★"),("2월","4Q+연간 잠정실적","★★★"),
                ("3월","사업보고서/주총/배당정책","★★★"),("5월","1Q 잠정실적","★★☆"),
                ("7~8월","갤럭시Z 언팩+2Q실적","★★★"),("11월","3Q 잠정실적","★★☆"),
                ("12월","결산배당 결정","★★☆"),("수시","자기주식 취득/소각","★☆☆"),("수시","ADAS 로드맵","★★☆")]:
    wr(ws12,row,[t,e,imp],fonts=[db,df,db],fills=[llf,wf,gld],als=[ca,la,ca]); row+=1
row+=2; st(ws12,row,"주요 이벤트 히스토리",4); row+=1
wh(ws12,row,["일자","유형","내용"]); row+=1
for dt,et,s in conn.execute("SELECT rcept_dt,event_type,SUBSTR(event_summary,1,80) FROM key_events WHERE rcept_dt>='20230101' ORDER BY rcept_dt DESC LIMIT 20").fetchall():
    wr(ws12,row,[dt,et,s.replace('\n',' ').strip()[:70]],fonts=[df,db,df],als=[ca,ca,la]); row+=1
print("  [12/12] 모니터링")

# === SAVE ===
OUT1=os.path.join(BASE,"엠씨넥스_종합보고서.xlsx")
wb.save(OUT1); conn.close()
print(f"\n종합보고서 생성 완료: {OUT1}")
print(f"시트 ({len(wb.sheetnames)}개): {wb.sheetnames}")
