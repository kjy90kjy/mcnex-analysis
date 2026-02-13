# -*- coding: utf-8 -*-
"""대한항공 투자 구루 분석 보고서 (Buffett/Munger 한국형 Four Filters)"""
import sqlite3, sys, os, statistics
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ai.db")
BASE = os.path.dirname(os.path.abspath(__file__))
conn = sqlite3.connect(DB)

# === KEY CONSTANTS ===
PRICE = 23700
SHARES = 368220612  # 보통주 발행주식수
SHARES_TREASURY = 49  # 자사주 (거의 없음)
SHARES_FLOAT = SHARES - SHARES_TREASURY
억 = 100_000_000
MCAP = PRICE * SHARES  # ~87,300억

# 2024 Annual (연결)
REV24 = 17_870_718_495_804; OP24 = 2_110_200_077_994; NI24 = 1_381_858_075_058; EPS24 = 3566
EQ24 = 10_963_191_867_177; EQ23 = 9_815_208_241_633
ASSETS24 = 47_012_065_940_089; LIAB24 = 36_048_874_072_912
CASH24 = 2_215_624_563_052
ST_DEBT = 2_517_021_740_000  # 단기차입금
LT_DEBT = 1_819_300_917_987 + 1_463_107_639_009 + 158_439_252_418  # 장기차입금+사채+자산유동화
DA24 = 1_737_223_223_952 + 58_806_701_481  # 감가상각비+무형자산상각비
OPCF24 = 4_558_915_095_322; CAPEX24 = 2_894_099_745_681
DIV_PAID = 277_054_000_000 * 억 // 억  # 배당금총액(백만원→원환산 근사)
DIV_PAID = 278_229_104_875  # 2024 CF상 배당금지급액
TREAS_BUY = 0  # 자사주 매입 없음
DPS24 = 750  # 보통주 주당배당

# Derived
BPS = EQ24 / SHARES
NET_DEBT = ST_DEBT + LT_DEBT - CASH24
EBITDA24 = OP24 + DA24; FCF24 = OPCF24 - CAPEX24; EV = MCAP + NET_DEBT
AVG_EQ = (EQ24 + EQ23) / 2; ROE24 = NI24 / AVG_EQ

# === STYLE ===
NAVY="1B2A4A"; DARK="2C3E6B"; MID="3A5BA0"; LB="D6E4F0"; LLB="EBF1F8"; W="FFFFFF"
GOLD_C="D4A843"; RED_C="C0392B"; GREEN_C="27AE60"; GRAY_C="F2F2F2"

title_font=Font(name="맑은 고딕",size=22,bold=True,color=W)
sub_font=Font(name="맑은 고딕",size=11,color="B0C4DE")
sec_font=Font(name="맑은 고딕",size=14,bold=True,color=NAVY)
hdr_font=Font(name="맑은 고딕",size=10,bold=True,color=W)
df=Font(name="맑은 고딕",size=10)
db=Font(name="맑은 고딕",size=10,bold=True)
d_blue=Font(name="맑은 고딕",size=10,bold=True,color="0000FF")
d_grn=Font(name="맑은 고딕",size=10,bold=True,color=GREEN_C)
d_red=Font(name="맑은 고딕",size=10,bold=True,color=RED_C)
d_navy=Font(name="맑은 고딕",size=12,bold=True,color=NAVY)
sm=Font(name="맑은 고딕",size=9,color="666666")

tf=PatternFill("solid",fgColor=NAVY)
hf=PatternFill("solid",fgColor=DARK)
mf=PatternFill("solid",fgColor=MID)
lf=PatternFill("solid",fgColor=LB)
llf=PatternFill("solid",fgColor=LLB)
gf=PatternFill("solid",fgColor=GRAY_C)
wf=PatternFill("solid",fgColor=W)
gld=PatternFill("solid",fgColor="FFF3CD")
rdf=PatternFill("solid",fgColor="F8D7DA")
gnf=PatternFill("solid",fgColor="D4EDDA")
blf=PatternFill("solid",fgColor="D6EAF8")

ca=Alignment(horizontal='center',vertical='center',wrap_text=True)
la=Alignment(horizontal='left',vertical='center',wrap_text=True)
ra=Alignment(horizontal='right',vertical='center')
tb=Border(left=Side('thin',color="D9D9D9"),right=Side('thin',color="D9D9D9"),
          top=Side('thin',color="D9D9D9"),bottom=Side('thin',color="D9D9D9"))
bb=Border(bottom=Side('medium',color=NAVY))
NF='#,##0'; PF='0.0%'

def sw(ws,w):
    for i,v in enumerate(w,1): ws.column_dimensions[get_column_letter(i)].width=v
def setup_print(ws):
    ws.page_setup.paperSize = Worksheet.PAPERSIZE_LETTER
    ws.page_setup.orientation = Worksheet.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.75, bottom=0.75,
        header=0.3, footer=0.3
    )
    ws.print_options.horizontalCentered = True
def wh(ws,r,h,fills=None):
    for i,v in enumerate(h,1):
        c=ws.cell(row=r,column=i,value=v); c.font=hdr_font; c.fill=fills[i-1] if fills else hf; c.alignment=ca; c.border=tb
def wr(ws,r,d,fonts=None,fills=None,als=None,nfs=None):
    for i,v in enumerate(d,1):
        c=ws.cell(row=r,column=i,value=v)
        c.font=fonts[i-1] if fonts else df; c.fill=fills[i-1] if fills else wf
        c.alignment=als[i-1] if als else ca; c.border=tb
        if nfs and i<=len(nfs) and nfs[i-1]: c.number_format=nfs[i-1]
def st(ws,r,t,ce=8):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ce)
    c=ws.cell(row=r,column=1,value=t); c.font=sec_font; c.border=bb; return r+1
def fmt(v): return f"{v/억:,.0f}억"
def fw(v): return f"{v:,.0f}원"
def pct(v): return f"{v*100:.1f}%"

# === LOAD DATA ===
# v_annual_performance (Note: revenue is NULL for 대한항공)
perf = conn.execute("SELECT * FROM v_annual_performance").fetchall()

# Revenue from financial_summary (first entry per year = 연결)
rev_rows = conn.execute("""SELECT bsns_year, MIN(thstrm_amount) FROM financial_summary
    WHERE account_nm='매출액' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    GROUP BY bsns_year ORDER BY bsns_year""").fetchall()
rev_map = {}
for r in rev_rows:
    try: rev_map[r[0]] = int(str(r[1]).replace(',',''))
    except: pass

# Operating income from financial_summary
op_rows = conn.execute("""SELECT bsns_year, MIN(thstrm_amount) FROM financial_summary
    WHERE account_nm='영업이익' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    GROUP BY bsns_year ORDER BY bsns_year""").fetchall()
op_map = {}
for r in op_rows:
    try: op_map[r[0]] = int(str(r[1]).replace(',',''))
    except: pass

# Net income from financial_summary
ni_rows = conn.execute("""SELECT bsns_year, MIN(thstrm_amount) FROM financial_summary
    WHERE account_nm='당기순이익(손실)' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    GROUP BY bsns_year ORDER BY bsns_year""").fetchall()
ni_map = {}
for r in ni_rows:
    try: ni_map[r[0]] = int(str(r[1]).replace(',',''))
    except: pass

# 매출총이익
gp_rows = conn.execute("""SELECT bsns_year, thstrm_amount FROM financial_statements
    WHERE account_nm='매출총이익' AND sj_div='CIS' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    ORDER BY bsns_year""").fetchall()
gp_map = {r[0]: int(r[1]) for r in gp_rows if r[1]}

# 영업CF
opcf_rows = conn.execute("""SELECT bsns_year, thstrm_amount FROM financial_statements
    WHERE account_nm='영업활동으로 인한 현금흐름' AND sj_div='CF' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    ORDER BY bsns_year""").fetchall()
opcf_map = {r[0]: int(r[1]) for r in opcf_rows if r[1]}

# CAPEX (유형자산 취득 - sign varies)
capex_rows = conn.execute("""SELECT bsns_year, thstrm_amount FROM financial_statements
    WHERE account_nm LIKE '%유형자산%취득%' AND sj_div='CF' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    ORDER BY bsns_year""").fetchall()
capex_map = {r[0]: abs(int(r[1])) for r in capex_rows if r[1]}

# 감가상각비 (유형+무형)
da_rows = conn.execute("""SELECT bsns_year, account_nm, thstrm_amount FROM financial_statements
    WHERE account_nm IN ('감가상각비','투자부동산감가상각비','무형자산상각비')
    AND sj_div='CF' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    ORDER BY bsns_year""").fetchall()
da_map = {}
for r in da_rows:
    if r[2]:
        da_map[r[0]] = da_map.get(r[0], 0) + int(r[2])

# 배당 (주당 현금배당금)
div_rows = conn.execute("""SELECT bsns_year, thstrm FROM dividends
    WHERE se='주당 현금배당금(원)' AND stock_knd='보통주' ORDER BY bsns_year""").fetchall()
div_map = {}
for r in div_rows:
    try:
        v = str(r[1]).replace(',','').replace('-','0')
        div_map[r[0]] = int(v) if v and v != '' else 0
    except: div_map[r[0]] = 0

# 배당성향
dpayout_rows = conn.execute("""SELECT bsns_year, thstrm FROM dividends
    WHERE se='(연결)현금배당성향(%)' ORDER BY bsns_year""").fetchall()
dpayout_map = {}
for r in dpayout_rows:
    try:
        v = str(r[1]).replace(',','').replace('-','')
        if v: dpayout_map[r[0]] = float(v)
    except: pass

# 자사주 총계
ts_rows = conn.execute("""SELECT bsns_year, bsis_qy, change_qy_acqs, change_qy_dsps, change_qy_incnr, trmend_qy
    FROM treasury_stock WHERE acqs_mth1='총계' AND stock_knd='보통주' ORDER BY bsns_year""").fetchall()

# 유상증자/CB
cap_all = conn.execute("""SELECT DISTINCT isu_dcrs_de, isu_dcrs_stle, isu_dcrs_stock_knd, isu_dcrs_qy
    FROM capital_changes WHERE isu_dcrs_stle LIKE '%유상증자%' OR isu_dcrs_stle LIKE '%전환%' OR isu_dcrs_stle LIKE '%무상%'
    ORDER BY isu_dcrs_de""").fetchall()

# 임원 (2024) - executives table has major shareholders mixed in
exec_rows = conn.execute("""SELECT nm, ofcps, rgist_exctv_at, fte_at FROM executives
    WHERE bsns_year='2024' LIMIT 20""").fetchall()

# CEO 보수
pay_rows = conn.execute("""SELECT bsns_year, nm, ofcps, mendng_totamt FROM individual_pay
    WHERE nm='조원태' ORDER BY bsns_year""").fetchall()

# All individual pay 2024
pay_all_2024 = conn.execute("""SELECT nm, ofcps, mendng_totamt FROM individual_pay
    WHERE bsns_year='2024' ORDER BY CAST(REPLACE(REPLACE(mendng_totamt,',',''),' ','') AS INTEGER) DESC""").fetchall()

# 특허
patent_rows = conn.execute("SELECT rcept_dt, patent_name, patent_detail, patent_plan FROM patents ORDER BY rcept_dt").fetchall()

# 주요 이벤트
event_rows = conn.execute("""SELECT rcept_dt, event_type, SUBSTR(event_summary,1,80) FROM key_events
    ORDER BY rcept_dt DESC LIMIT 20""").fetchall()

# Build annual data dict
annual = {}
for p in perf:
    yr = p[0]
    # Use financial_summary for revenue/op/ni since v_annual_performance has NULL revenue
    rev = rev_map.get(yr, 0)
    op = int(p[2]) if p[2] else op_map.get(yr, 0)
    ni = int(p[3]) if p[3] else ni_map.get(yr, 0)
    eps = int(p[4]) if p[4] else 0
    ta = int(p[5]) if p[5] else 0
    tl = int(p[6]) if p[6] else 0
    te = int(p[7]) if p[7] else 0
    annual[yr] = {"rev":rev,"op":op,"ni":ni,"eps":eps,"ta":ta,"tl":tl,"te":te,
                  "gp":gp_map.get(yr,0),"opcf":opcf_map.get(yr,0),
                  "capex":capex_map.get(yr,0),"da":da_map.get(yr,0)}

# Last 5 years for trend analysis
YEARS5 = [y for y in sorted(annual.keys()) if y >= '2020']

wb = Workbook()

# ============================================================
# SHEET 1: Four Filters 종합 대시보드
# ============================================================
ws1 = wb.active; ws1.title="Four Filters"; ws1.sheet_properties.tabColor=NAVY
sw(ws1,[4,22,14,14,14,14,14,14,4])
setup_print(ws1)

# Title banner
for r in range(1,6):
    for c in range(1,10): ws1.cell(row=r,column=c).fill=tf
ws1.merge_cells('B2:H2')
ws1.cell(row=2,column=2,value="대한항공(Korean Air) 투자 구루 분석").font=Font(name="맑은 고딕",size=24,bold=True,color=W)
ws1.cell(row=2,column=2).alignment=ca
ws1.merge_cells('B4:H4')
ws1.cell(row=4,column=2,value=f"Four Filters 종합 대시보드  |  현재가 {PRICE:,}원  |  2026.02.13 기준").font=sub_font
ws1.cell(row=4,column=2).alignment=ca

row = 7

# === Four Filters Score ===
st(ws1,row,"A. Four Filters 스코어카드 (100점 만점)",8); row+=1

# Filter 1: 사업 이해도
f1_items = [
    ("사업모델 명확도","여객+화물 항공운송, 명확한 사업모델","높음",5),
    ("매출구조 단순도","여객·화물·기타 3개 사업부문, 비교적 단순","보통",4),
    ("10년 일관성","코로나 급락 후 회복, 변동 큼","보통",3),
    ("기술/규제 예측성","유가·환율·규제 불확실성","보통",3),
    ("해외 사업 이해도","글로벌 항공사, 노선·실적 투명","높음",5),
]
f1_score = sum(x[3] for x in f1_items)

# Filter 2: 경제적 해자
f2_items = [
    ("전환비용","마일리지·제휴 항공사 네트워크, 노선 친숙도","강함",4),
    ("원가우위","규모의 경제 있으나 유가에 취약","보통",3),
    ("무형자산","브랜드, 노선권, 공항 슬롯","강함",4),
    ("규모의 경제","대형항공사이나 LCC 경쟁 심화","보통",3),
    ("이익 지속성","경기 민감, 유가·환율 변동 큼","보통",3),
]
f2_score = sum(x[3] for x in f2_items)

# Filter 3: 경영진
f3_items = [
    ("오너 지분율","한진칼 경유 간접 지배, 직접 지분 제한적","보통",3),
    ("보수 합리성","조원태 51억, 매출 17.8조 대비 적정","보통",3),
    ("자사주 정책","자사주 거의 없음 (49주), 소극적","미흡",2),
    ("배당 이력","최근 3년 750원 꾸준, 그 전 무배당 기간 존재","보통",3),
    ("소통/IR","분기별 IR, 애널리스트 미팅 활발","양호",4),
]
f3_score = sum(x[3] for x in f3_items)

# Filter 4: 안전마진
per24 = PRICE/EPS24
pbr = PRICE/BPS
f4_items = [
    ("PER 수준",f"{per24:.1f}배 (기준 8~15배, 항공업 저PER)","저평가",4),
    ("PBR 수준",f"{pbr:.2f}배 (기준 0.8~1.5배)","저평가",4),
    ("FCF 수익률",f"{FCF24/MCAP*100:.1f}% (기준 >5%)","우수",4),
    ("EV/EBITDA",f"{EV/EBITDA24:.1f}배 (기준 6~10배)","보통",3),
    ("Owner Earnings",f"높으나 CAPEX 변동 큼 (항공기 투자)","보통",3),
]
f4_score = sum(x[3] for x in f4_items)

total_score = f1_score + f2_score + f3_score + f4_score
grade = "A" if total_score>=80 else ("B" if total_score>=60 else ("C" if total_score>=40 else "D"))
grade_color = GREEN_C if grade in ["A","B"] else (GOLD_C if grade=="C" else RED_C)

# Score summary
wh(ws1,row,["","필터","소항목1","소항목2","소항목3","소항목4","소항목5","소계"]); row+=1
for fname,items,score in [("F1: 사업 이해도",f1_items,f1_score),
                           ("F2: 경제적 해자",f2_items,f2_score),
                           ("F3: 경영진 품질",f3_items,f3_score),
                           ("F4: 안전마진",f4_items,f4_score)]:
    scores = [x[3] for x in items]
    sc_fill = gnf if score>=20 else (gld if score>=15 else rdf)
    wr(ws1,row,["",fname]+scores+[f"{score}/25"],
       fonts=[df,db]+[df]*5+[Font(name="맑은 고딕",size=11,bold=True,color=GREEN_C if score>=20 else NAVY)],
       fills=[wf,llf]+[wf]*5+[sc_fill],als=[ca,la]+[ca]*6)
    row+=1

# Total
row+=1
ws1.merge_cells(start_row=row,start_column=2,end_row=row,end_column=5)
ws1.cell(row=row,column=2,value="종합점수").font=d_navy; ws1.cell(row=row,column=2).alignment=la; ws1.cell(row=row,column=2).border=tb
ws1.merge_cells(start_row=row,start_column=6,end_row=row,end_column=7)
ws1.cell(row=row,column=6,value=f"{total_score}/100 (등급: {grade})").font=Font(name="맑은 고딕",size=14,bold=True,color=grade_color)
ws1.cell(row=row,column=6).fill=gnf if grade in ["A","B"] else gld
ws1.cell(row=row,column=6).alignment=ca; ws1.cell(row=row,column=6).border=tb
row+=2

# === 투자 테제 ===
st(ws1,row,"B. 투자 테제",8); row+=1
ws1.merge_cells(start_row=row,start_column=2,end_row=row+1,end_column=8)
c=ws1.cell(row=row,column=2)
c.value=("대한항공은 국내 1위 FSC(Full Service Carrier)로 여객·화물 양 축의 글로벌 네트워크를 보유한 "
         "항공운송 대표기업. 아시아나항공 합병으로 규모의 경제 극대화가 기대되나, "
         f"부채비율 329%의 고레버리지 구조가 핵심 리스크. "
         f"현재 PER {per24:.1f}배, PBR {pbr:.2f}배로 밸류에이션은 매력적이나, "
         "재무구조 개선 확인 시 매수 전환이 적절.")
c.font=Font(name="맑은 고딕",size=10,bold=True,color=NAVY); c.alignment=Alignment(horizontal='left',vertical='top',wrap_text=True)
row+=3

# === 핵심 강점/리스크 ===
st(ws1,row,"C. 핵심 강점 vs 리스크",8); row+=1

# Strengths
ws1.merge_cells(start_row=row,start_column=2,end_row=row,end_column=4)
ws1.cell(row=row,column=2,value="핵심 강점 (3)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws1.cell(row=row,column=2).fill=PatternFill("solid",fgColor=GREEN_C); ws1.cell(row=row,column=2).alignment=ca
ws1.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
ws1.cell(row=row,column=5,value="핵심 리스크 (3)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws1.cell(row=row,column=5).fill=PatternFill("solid",fgColor=RED_C); ws1.cell(row=row,column=5).alignment=ca
row+=1

strengths = ["글로벌 노선 네트워크 + 인천허브 (아시아-미주 최적 위치)",
             "화물 사업 세계 3위 → 경기 민감하나 고수익 사업",
             "아시아나 합병 → 규모의 경제·노선 최적화 시너지"]
risks = ["부채비율 329% — 항공업 특성이나 재무 리스크 높음",
         "유가·환율 변동 → 영업이익 급등락 불가피",
         "아시아나 합병 후 통합 리스크 (인력·노선·시스템)"]

for i in range(3):
    ws1.merge_cells(start_row=row,start_column=2,end_row=row,end_column=4)
    ws1.cell(row=row,column=2,value=f"  {i+1}. {strengths[i]}").font=df
    ws1.cell(row=row,column=2).fill=gnf; ws1.cell(row=row,column=2).alignment=la; ws1.cell(row=row,column=2).border=tb
    ws1.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
    ws1.cell(row=row,column=5,value=f"  {i+1}. {risks[i]}").font=df
    ws1.cell(row=row,column=5).fill=rdf; ws1.cell(row=row,column=5).alignment=la; ws1.cell(row=row,column=5).border=tb
    row+=1

# === 소항목 상세 ===
row+=1; st(ws1,row,"D. 필터별 소항목 상세",8); row+=1
wh(ws1,row,["","항목","평가 근거","판정","점수","","",""]); row+=1
for fname,items in [("Filter 1: 사업 이해도",f1_items),("Filter 2: 경제적 해자",f2_items),
                     ("Filter 3: 경영진 품질",f3_items),("Filter 4: 안전마진",f4_items)]:
    ws1.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
    ws1.cell(row=row,column=1,value=fname).font=Font(name="맑은 고딕",size=11,bold=True,color=MID)
    ws1.cell(row=row,column=1).fill=llf; ws1.cell(row=row,column=1).border=tb; row+=1
    for item_nm,evidence,judgment,score in items:
        sf = d_grn if score>=4 else (df if score>=3 else d_red)
        sfill = gnf if score>=4 else (wf if score>=3 else rdf)
        wr(ws1,row,["",item_nm,evidence,judgment,f"{score}/5","","",""],
           fonts=[df,db,df,sf,sf,df,df,df],fills=[wf,llf,wf,sfill,sfill,wf,wf,wf],
           als=[ca,la,la,ca,ca,ca,ca,ca]); row+=1

print("  [1/7] Four Filters 대시보드")

# ============================================================
# SHEET 2: 경제적 해자 분석
# ============================================================
ws2 = wb.create_sheet("경제적_해자"); ws2.sheet_properties.tabColor=DARK
sw(ws2,[20,16,16,16,16,16,16,16])
setup_print(ws2)

row=1; st(ws2,row,"경제적 해자(Economic Moat) 분석",8); row+=1

# 해자 유형 판정표
st(ws2,row,"A. 해자 유형 판정",8); row+=1
wh(ws2,row,["해자 유형","보유 여부","강도","근거","지속가능성","","",""]); row+=1
moat_types = [
    ("비용우위","△","보통","규모의 경제 있으나 유가·인건비에 노출","3~5년"),
    ("전환비용","O","강함","마일리지 적립, 스카이팀 제휴, 노선 친숙도","10년+"),
    ("네트워크 효과","O","보통","인천허브 + 스카이팀 글로벌 네트워크","10년+"),
    ("무형자산(브랜드)","O","강함","대한항공 브랜드, 노선권, 공항 슬롯","10년+"),
    ("규모의 경제","O","보통","국내 1위 + 아시아나 합병 → 확대","5~10년"),
]
for tp,has,strength,basis,duration in moat_types:
    hf2 = d_grn if has=="O" else (d_red if has=="X" else db)
    hfl = gnf if has=="O" else (rdf if has=="X" else gld)
    wr(ws2,row,[tp,has,strength,basis,duration,"","",""],
       fonts=[db,hf2,hf2,df,df,df,df,df],fills=[llf,hfl,hfl,wf,wf,wf,wf,wf],
       als=[la,ca,ca,la,ca,ca,ca,ca]); row+=1

# 수익성 추이
row+=1; st(ws2,row,"B. 수익성 추이 (5년)",8); row+=1
wh(ws2,row,["연도","매출액(억)","매출총이익(억)","GPM","영업이익(억)","OPM","ROIC","비고"]); row+=1

for yr in YEARS5:
    d = annual[yr]
    rev_b = d["rev"]//억; gp_b = d["gp"]//억; op_b = d["op"]//억
    gpm = d["gp"]/d["rev"] if d["rev"] else 0
    opm = d["op"]/d["rev"] if d["rev"] else 0
    nopat = d["op"] * 0.78
    roic = nopat / d["te"] if d["te"] else 0

    gpm_f = d_grn if gpm>0.20 else (d_red if gpm<0.10 else df)
    opm_f = d_grn if opm>0.10 else (d_red if opm<0.05 else df)
    note = ""
    if yr=='2024': note = "역대 최고 매출"
    elif yr=='2020': note = "코로나 타격"
    elif yr=='2022': note = "최대 영업이익"

    wr(ws2,row,[yr,rev_b,gp_b,f"{gpm*100:.1f}%",op_b,f"{opm*100:.1f}%",f"{roic*100:.1f}%",note],
       fonts=[db,df,df,gpm_f,df,opm_f,df,sm],fills=[lf]+[wf]*7,
       als=[ca,ra,ra,ca,ra,ca,ca,la],nfs=[None,NF,NF,None,NF,None,None,None]); row+=1

# R&D / 항공 특성
row+=1; st(ws2,row,"C. 항공업 진입장벽 및 경쟁우위",8); row+=1
wh(ws2,row,["항목","내용","비고","","","","",""]); row+=1
for item,content,note in [
    ("노선권/슬롯","인천-미주·유럽 핵심 슬롯 보유, 신규 진입 사실상 불가","규제 해자"),
    ("마일리지 락인","스카이패스 회원 3,700만+, 전환비용 높음","고객 충성도"),
    ("화물 경쟁력","화물 매출 세계 3위, 전용기 20+대 보유","차별화"),
    ("스카이팀 제휴","20개 항공사 글로벌 네트워크, 코드쉐어","네트워크 해자"),
    ("아시아나 합병","2024년 합병 완료, 노선·슬롯·정비 시너지","규모 확대"),
    ("특허","항공업 특성상 특허 없음 (기술 해자 아닌 규제·규모 해자)","N/A"),
]:
    wr(ws2,row,[item,content,note,"","","","",""],
       fonts=[db,df,sm,df,df,df,df,df],fills=[llf,wf,wf,wf,wf,wf,wf,wf],
       als=[la,la,la,ca,ca,ca,ca,ca]); row+=1

# 해자 종합 평가
row+=1; st(ws2,row,"D. 해자 종합 평가",8); row+=1
for item,val in [("해자 폭","Narrow ~ Medium (전환비용 + 브랜드 + 규모)"),
                  ("해자 추세","확대 중 (아시아나 합병 → 국내 독과점 구조)"),
                  ("핵심 해자","노선권/슬롯 + 마일리지 전환비용 + 스카이팀 네트워크"),
                  ("위협 요인","LCC 확대, 유가 급등, 글로벌 경기침체"),
                  ("결론","규제+규모 기반 Narrow Moat 보유. 합병 시너지 가시화 시 Medium으로 확대 가능")]:
    ws2.cell(row=row,column=1,value=item).font=db; ws2.cell(row=row,column=1).fill=llf
    ws2.cell(row=row,column=1).alignment=la; ws2.cell(row=row,column=1).border=tb
    ws2.merge_cells(start_row=row,start_column=2,end_row=row,end_column=8)
    ws2.cell(row=row,column=2,value=val).font=df if "위협" not in item else d_red
    ws2.cell(row=row,column=2).alignment=la; ws2.cell(row=row,column=2).border=tb; row+=1

print("  [2/7] 경제적 해자")

# ============================================================
# SHEET 3: 경영진 평가
# ============================================================
ws3 = wb.create_sheet("경영진_평가"); ws3.sheet_properties.tabColor=GOLD_C
sw(ws3,[20,16,16,16,16,16,16,16])
setup_print(ws3)

row=1; st(ws3,row,"경영진 평가 (Management Quality)",8); row+=1

# 주요 경영진 명단
st(ws3,row,"A. 주요 경영진 (2024년 사업보고서 기준)",8); row+=1
wh(ws3,row,["성명","직위","구분","비고","","","",""]); row+=1
mgmt_list = [
    ("조원태","회장(대표이사)","오너경영","한진칼 대주주, 경영권 승계"),
    ("우기홍","부회장(대표이사)","전문경영인","실질적 경영 총괄"),
    ("유종석","임원","등기임원",""),
    ("김세진","임원","등기임원",""),
    ("이기광","고문","비등기","전 경영진"),
]
for nm,pos,cat,note in mgmt_list:
    is_ceo = "대표" in pos or "회장" in pos
    wr(ws3,row,[nm,pos,cat,note,"","","",""],
       fonts=[d_navy if is_ceo else db,df,df,sm,df,df,df,df],
       fills=[gld if is_ceo else llf]+[wf]*7,als=[la,la,ca,la,ca,ca,ca,ca]); row+=1

# CEO 보수 추이
row+=1; st(ws3,row,"B. 대표이사(조원태) 보수 추이",8); row+=1
wh(ws3,row,["연도","성명","직위","총보수","비고","","",""]); row+=1
for pr in pay_rows:
    yr,nm,ofcps,amt = pr[0],pr[1],pr[2],pr[3]
    try:
        amt_v = int(str(amt).replace(',',''))
        amt_str = f"{amt_v:,}원"
        amt_m = f"(월 ~{amt_v//12:,}원)"
    except:
        amt_str = str(amt); amt_m = ""
    wr(ws3,row,[yr,nm,ofcps,amt_str,amt_m,"","",""],
       fonts=[db,db,df,d_blue,sm,df,df,df],fills=[lf,wf,wf,gld,wf,wf,wf,wf],
       als=[ca,la,la,ra,la,ca,ca,ca]); row+=1

# 2024 개별보수 전체
row+=1; st(ws3,row,"B-2. 2024년 개별보수 공시 대상자",8); row+=1
wh(ws3,row,["성명","직위","총보수(원)","비고","","","",""]); row+=1
for p in pay_all_2024:
    nm,ofcps,amt = p[0],p[1],p[2]
    try:
        amt_v = int(str(amt).replace(',',''))
        amt_str = f"{amt_v:,}원"
    except:
        amt_str = str(amt)
    is_ceo = "회장" in (ofcps or "") or "대표" in (ofcps or "")
    wr(ws3,row,[nm,ofcps,amt_str,"","","","",""],
       fonts=[d_navy if is_ceo else db,df,d_blue,df,df,df,df,df],
       fills=[gld if is_ceo else llf,wf,gld,wf,wf,wf,wf,wf],
       als=[la,la,ra,la,ca,ca,ca,ca]); row+=1

# 오너경영 체크리스트
row+=1; st(ws3,row,"C. 오너경영 체크리스트",8); row+=1
wh(ws3,row,["","평가 항목","현황","판정","근거","","",""]); row+=1
owner_checks = [
    ("1","오너 지분/지배구조","한진칼 경유 간접 지배, 직접 지분 제한적","보통","간접 지배 구조 — 이해관계 일치도 제한적"),
    ("2","경영 안정성","조원태 체제 안정화 (2019년 경영권 분쟁 해소)","보통","2019년 분쟁 후 경영권 안정"),
    ("3","보수 합리성","조원태 51억 (2024)","보통","매출 17.8조·순이익 1.38조 대비 0.37%"),
    ("4","자사주 매입/소각","자사주 거의 없음 (49주), 매입/소각 이력 없음","미흡","주주환원 수단으로 미활용"),
    ("5","배당 정책","750원 (2022~2024), 배당수익률 3.2%","보통","꾸준하나 인상 없음, 과거 무배당 이력"),
    ("6","자본배분","대규모 항공기 투자 (40대 신규 구매 결정 등)","보통","성장투자 우선, 주주환원 후순위"),
    ("7","소통/IR","분기별 컨퍼런스콜, 기관 대상 IR 활발","양호","정기적·체계적 IR 활동"),
]
for num,item,status,judgment,basis in owner_checks:
    jf = d_grn if judgment=="우수" else (d_blue if judgment=="양호" else (db if judgment=="보통" else d_red))
    jfl = gnf if judgment=="우수" else (blf if judgment=="양호" else (gld if judgment=="보통" else rdf))
    wr(ws3,row,[num,item,status,judgment,basis,"","",""],
       fonts=[df,db,df,jf,sm,df,df,df],fills=[wf,llf,wf,jfl,wf,wf,wf,wf],
       als=[ca,la,la,ca,la,ca,ca,ca]); row+=1

# 주식 구조
row+=1; st(ws3,row,"D. 주식 구조",8); row+=1
for lbl,val in [("발행주식수(보통주)",f"{SHARES:,}주"),
                ("자기주식",f"{SHARES_TREASURY:,}주 (0.00%)"),
                ("유통주식수",f"{SHARES_FLOAT:,}주"),
                ("최대주주(한진칼 외 특수관계인)",f"129,885,869주 (~29.3%)"),
                ("국민연금","36,811,447주 (~10.0%)"),
                ("외국인","~30%")]:
    ws3.cell(row=row,column=1,value=lbl).font=db; ws3.cell(row=row,column=1).fill=llf
    ws3.cell(row=row,column=1).alignment=la; ws3.cell(row=row,column=1).border=tb
    ws3.merge_cells(start_row=row,start_column=2,end_row=row,end_column=4)
    ws3.cell(row=row,column=2,value=val).font=d_blue; ws3.cell(row=row,column=2).alignment=la
    ws3.cell(row=row,column=2).border=tb; row+=1

# 경영진 종합 평가
row+=1; st(ws3,row,"E. 경영진 종합 평가",8); row+=1
for item,val in [("오너십 점수","3/5 (간접 지배 구조, 이해관계 일치도 제한적)"),
                  ("자본배분 점수","2/5 (자사주 매입/소각 없음, 배당 동결)"),
                  ("투명성 점수","4/5 (IR 활발, 실적 공시 충실)"),
                  ("승계 리스크","낮음 (조원태 체제 안정, 조현민 부사장 차세대)"),
                  ("결론","안정적 오너경영이나 주주환원은 보수적. IR은 업계 상위 수준.")]:
    ws3.cell(row=row,column=1,value=item).font=db; ws3.cell(row=row,column=1).fill=llf
    ws3.cell(row=row,column=1).alignment=la; ws3.cell(row=row,column=1).border=tb
    ws3.merge_cells(start_row=row,start_column=2,end_row=row,end_column=8)
    ws3.cell(row=row,column=2,value=val).font=df; ws3.cell(row=row,column=2).alignment=la
    ws3.cell(row=row,column=2).border=tb; row+=1

print("  [3/7] 경영진 평가")

# ============================================================
# SHEET 4: 자본배분 이력
# ============================================================
ws4 = wb.create_sheet("자본배분"); ws4.sheet_properties.tabColor="E74C3C"
sw(ws4,[14,14,14,14,14,14,14,14])
setup_print(ws4)

row=1; st(ws4,row,"자본배분 이력 (Capital Allocation History)",8); row+=1

# 배당 추이
st(ws4,row,"A. 배당 추이",8); row+=1
wh(ws4,row,["연도","주당배당(원)","EPS(원)","배당성향(%)","변화","비고","",""]); row+=1
div_years = ['2015','2016','2017','2018','2019','2020','2021','2022','2023','2024']
prev_dps = None
for yr in div_years:
    dps = div_map.get(yr,0)
    eps_yr = annual[yr]["eps"] if yr in annual else 0
    # For years without EPS in financial_statements, calculate from NI
    if eps_yr == 0 and yr in annual:
        ni_yr = annual[yr]["ni"]
        eps_yr = int(ni_yr / SHARES) if ni_yr else 0
    po = dpayout_map.get(yr, 0)
    chg = ""
    if prev_dps is not None and prev_dps > 0 and dps > 0:
        chg = "↑" if dps > prev_dps else ("↓" if dps < prev_dps else "→")
    elif dps > 0 and (prev_dps is None or prev_dps == 0):
        chg = "재개"
    note = ""
    if yr=='2024': note = "3년 연속 750원"
    elif yr=='2023': note = "동결"
    elif yr=='2022': note = "대폭 재개 (코로나 이후)"
    elif yr=='2020': note = "코로나 무배당"
    elif yr=='2019': note = "적자 무배당"
    elif yr=='2017': note = "배당 재개"
    elif yr=='2015': note = "적자 무배당"

    dps_f = d_grn if dps > 0 else d_red
    chg_f = d_grn if chg in ["↑","재개"] else (d_red if chg=="↓" else df)
    wr(ws4,row,[yr,dps if dps else "-",f"{eps_yr:,}" if eps_yr else "-",f"{po:.1f}%" if po else "-",chg,note,"",""],
       fonts=[db,dps_f,df,df,chg_f,sm,df,df],fills=[lf]+[wf]*7,
       als=[ca,ra,ra,ca,ca,la,ca,ca]); row+=1
    prev_dps = dps

# 자사주 이력
row+=1; st(ws4,row,"B. 자기주식 현황 (보통주 총계)",8); row+=1
ws4.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
ws4.cell(row=row,column=1,value="대한항공은 자사주 매입·소각을 주주환원 수단으로 활용하지 않음. 보유 자사주는 합병 등 기타 사유로 취득한 단주(49주)에 불과.").font=df
ws4.cell(row=row,column=1).fill=gld; ws4.cell(row=row,column=1).alignment=la; ws4.cell(row=row,column=1).border=tb
row+=2

# 유상증자/CB 이력
st(ws4,row,"C. 유상증자/전환사채 이력",8); row+=1
wh(ws4,row,["일자","변동사유","주식종류","수량(주)","영향","비고","",""]); row+=1
cap_events = [
    ("2015.03","유상증자(주주배정)","보통주","14,164,306","희석","재무구조 개선"),
    ("2017.03","유상증자(주주배정)","보통주","22,004,890","희석","재무구조 개선"),
    ("2020.07","유상증자(주주배정)","보통주","79,365,079","대폭 희석","코로나 긴급 자금"),
    ("2021.03","유상증자(주주배정)","보통주","173,611,112","대폭 희석","아시아나 인수 자금"),
    ("2022.06","전환권행사","보통주","20,399,836","희석","전환사채 행사"),
]
for dt,reason,kind,qty,impact,note in cap_events:
    imp_f = d_red if "대폭" in impact else db
    wr(ws4,row,[dt,reason,kind,qty,impact,note,"",""],
       fonts=[db,d_red,df,df,imp_f,sm,df,df],fills=[lf,rdf,wf,wf,rdf if "대폭" in impact else gld,wf,wf,wf],
       als=[ca,la,ca,ra,ca,la,ca,ca]); row+=1

row+=1
ws4.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
ws4.cell(row=row,column=1,value="주의: 2015~2022년 대규모 유상증자로 주당 가치 지속 희석. 주주 관점에서 부정적 자본배분 이력.").font=d_red
ws4.cell(row=row,column=1).fill=rdf; ws4.cell(row=row,column=1).alignment=la; ws4.cell(row=row,column=1).border=tb
row+=2

# 항공기 투자 (CAPEX)
st(ws4,row,"D. 항공기 투자(CAPEX) 추이",8); row+=1
wh(ws4,row,["연도","영업CF(억)","CAPEX(억)","FCF(억)","CAPEX/영업CF","비고","",""]); row+=1
for yr in YEARS5:
    d = annual[yr]
    opcf_b = d["opcf"]//억; capex_b = d["capex"]//억; fcf_b = opcf_b - capex_b
    cap_ratio = d["capex"]/d["opcf"] if d["opcf"]>0 else 0
    note = ""
    if yr=='2024': note = "신규 항공기 40대 발주"
    elif yr=='2023': note = "투자 확대"
    elif yr=='2020': note = "코로나 투자 축소"
    wr(ws4,row,[yr,opcf_b,capex_b,fcf_b,f"{cap_ratio*100:.0f}%",note,"",""],
       fonts=[db,df,df,d_grn if fcf_b>0 else d_red,df,sm,df,df],
       fills=[lf,wf,wf,gnf if fcf_b>0 else rdf,wf,wf,wf,wf],
       als=[ca,ra,ra,ra,ca,la,ca,ca],nfs=[None,NF,NF,NF,None,None,None,None]); row+=1

# 자본배분 점수카드
row+=1; st(ws4,row,"E. 자본배분 점수카드",8); row+=1
for item,score,comment in [
    ("배당 일관성","3/5","10년 중 4년 무배당 (적자/코로나). 최근 3년 안정적이나 인상 없음"),
    ("자사주 정책","1/5","자사주 매입·소각 이력 없음. 주주환원 수단 미활용"),
    ("유상증자/CB","1/5","5회 유상증자로 주당 가치 대폭 희석 — 주주에게 불리"),
    ("투자(CAPEX)","3/5",f"CAPEX/영업CF {CAPEX24/OPCF24*100:.0f}% — 항공업 특성이나 높은 수준"),
    ("총주주환원율","2/5",f"{pct(DIV_PAID/MCAP)} (배당만) — 자사주 미실시로 업종 평균 이하"),
]:
    sf = d_grn if "5/5" in score else (d_blue if "4/5" in score else (df if "3/5" in score else d_red))
    sfill = gnf if "5/5" in score or "4/5" in score else (gld if "3/5" in score else rdf)
    wr(ws4,row,[item,score,comment,"","","","",""],
       fonts=[db,sf,df,df,df,df,df,df],fills=[llf,sfill,wf,wf,wf,wf,wf,wf],
       als=[la,ca,la,ca,ca,ca,ca,ca]); row+=1

print("  [4/7] 자본배분")

# ============================================================
# SHEET 5: 수익성 품질 분석
# ============================================================
ws5 = wb.create_sheet("수익성_품질"); ws5.sheet_properties.tabColor="8E44AD"
sw(ws5,[14,14,14,14,14,14,14,14])
setup_print(ws5)

row=1; st(ws5,row,"수익성 품질 분석 (Earnings Quality)",8); row+=1

# DuPont 분해
st(ws5,row,"A. DuPont 3단계 분해 (5년)",8); row+=1
wh(ws5,row,["연도","순이익률","자산회전율","레버리지","ROE","ROE (산식확인)","비고",""]); row+=1

roe_list = []
for yr in YEARS5:
    d = annual[yr]
    npm = d["ni"]/d["rev"] if d["rev"] else 0
    ato = d["rev"]/d["ta"] if d["ta"] else 0
    lev = d["ta"]/d["te"] if d["te"] else 0
    roe_calc = npm * ato * lev
    roe_list.append(roe_calc)

    npm_f = d_grn if npm>0.05 else (d_red if npm<0 else df)
    roe_f = d_grn if roe_calc>0.12 else (d_red if roe_calc<0 else df)
    note = ""
    if yr=='2020': note = "코로나 적자 (중단영업이익 제외)"
    elif yr=='2022': note = "최대 실적"
    elif yr=='2024': note = "아시아나 합병 원년"

    wr(ws5,row,[yr,f"{npm*100:.1f}%",f"{ato:.2f}회",f"{lev:.2f}배",f"{roe_calc*100:.1f}%",
                f"= {npm*100:.1f}x{ato:.2f}x{lev:.2f}",note,""],
       fonts=[db,npm_f,df,df,roe_f,sm,sm,df],fills=[lf]+[wf]*7,
       als=[ca,ca,ca,ca,ca,la,la,ca]); row+=1

# ROE summary
row+=1
# Filter out deeply negative years for meaningful average
roe_positive = [r for r in roe_list if r > -0.5]
avg_roe = statistics.mean(roe_positive) if roe_positive else 0
ws5.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
ws5.cell(row=row,column=1,value=f"5년 평균 ROE: {avg_roe*100:.1f}% | 판정: {'양호 (>10%)' if avg_roe>0.10 else '보통'}  |  주의: 고레버리지(4배+)가 ROE 드라이버").font=db
ws5.cell(row=row,column=1).fill=gnf if avg_roe>0.10 else gld; ws5.cell(row=row,column=1).alignment=la
ws5.cell(row=row,column=1).border=tb; row+=2

# FCF/NI 비율
st(ws5,row,"B. FCF 대 순이익 비율 (현금이익 품질)",8); row+=1
wh(ws5,row,["연도","순이익(억)","영업CF(억)","CAPEX(억)","FCF(억)","FCF/NI","판정",""]); row+=1
for yr in YEARS5:
    d = annual[yr]
    ni_b = d["ni"]//억; opcf_b = d["opcf"]//억; capex_b = d["capex"]//억
    fcf_b = opcf_b - capex_b
    fcf_ni = (d["opcf"]-d["capex"])/d["ni"] if d["ni"]>0 else 0

    j = "양호" if fcf_ni>0.8 else ("주의" if fcf_ni>0.5 else "경고")
    if d["ni"] <= 0: j = "적자"
    jf = d_grn if j=="양호" else (db if j=="주의" else d_red)
    jfl = gnf if j=="양호" else (gld if j=="주의" else rdf)

    wr(ws5,row,[yr,ni_b,opcf_b,capex_b,fcf_b,f"{fcf_ni*100:.0f}%" if d["ni"]>0 else "N/A",j,""],
       fonts=[db,df,df,df,d_blue,jf,jf,df],fills=[lf,wf,wf,wf,gld,jfl,jfl,wf],
       als=[ca,ra,ra,ra,ra,ca,ca,ca],nfs=[None,NF,NF,NF,NF,None,None,None]); row+=1

# 발생액 비율
row+=1; st(ws5,row,"C. 발생액 비율 (Accrual Ratio)",8); row+=1
ws5.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
ws5.cell(row=row,column=1,value="발생액 비율 = (순이익 - 영업CF) / 총자산  |  +-5% 이내 양호, +-10% 초과 경고").font=sm
ws5.cell(row=row,column=1).alignment=la; row+=1

wh(ws5,row,["연도","순이익(억)","영업CF(억)","차이(억)","총자산(억)","발생액비율","판정",""]); row+=1
for yr in YEARS5:
    d = annual[yr]
    ni_b = d["ni"]//억; opcf_b = d["opcf"]//억; diff = (d["ni"]-d["opcf"])//억
    ta_b = d["ta"]//억
    accrual = (d["ni"]-d["opcf"])/d["ta"] if d["ta"] else 0

    j = "양호" if abs(accrual)<0.05 else ("주의" if abs(accrual)<0.10 else "경고")
    jf = d_grn if j=="양호" else (db if j=="주의" else d_red)
    jfl = gnf if j=="양호" else (gld if j=="주의" else rdf)

    wr(ws5,row,[yr,ni_b,opcf_b,diff,ta_b,f"{accrual*100:.1f}%",j,""],
       fonts=[db,df,df,d_red if diff>0 else d_grn,df,jf,jf,df],
       fills=[lf,wf,wf,wf,wf,jfl,jfl,wf],als=[ca,ra,ra,ra,ra,ca,ca,ca],
       nfs=[None,NF,NF,NF,NF,None,None,None]); row+=1

# 이익 변동성
row+=1; st(ws5,row,"D. 이익 변동성 (Earnings Volatility)",8); row+=1

op_list = [annual[yr]["op"] for yr in YEARS5]
if len(op_list) >= 2:
    op_mean = statistics.mean(op_list)
    op_stdev = statistics.stdev(op_list)
    cv = op_stdev / op_mean if op_mean > 0 else 99
else:
    cv = 0; op_mean = 0; op_stdev = 0

cv_j = "안정적" if cv<0.3 else ("변동성 있음" if cv<0.6 else "높은 변동성")
cv_f = d_grn if cv<0.3 else (db if cv<0.6 else d_red)

wh(ws5,row,["지표","값","판정","기준","","","",""]); row+=1
wr(ws5,row,["영업이익 평균 (5년)",f"{int(op_mean/억):,}억","","","","","",""],
   fonts=[db,d_blue,df,df,df,df,df,df],fills=[llf,gld,wf,wf,wf,wf,wf,wf],als=[la,ra,ca,la,ca,ca,ca,ca]); row+=1
wr(ws5,row,["영업이익 표준편차",f"{int(op_stdev/억):,}억","","","","","",""],
   fonts=[db,df,df,df,df,df,df,df],fills=[llf,wf,wf,wf,wf,wf,wf,wf],als=[la,ra,ca,la,ca,ca,ca,ca]); row+=1
wr(ws5,row,["변동계수 (CV)",f"{cv:.2f}",cv_j,"<0.3 안정 / 0.3~0.6 보통 / >0.6 높음","","","",""],
   fonts=[db,cv_f,cv_f,sm,df,df,df,df],
   fills=[llf,gnf if cv<0.3 else (gld if cv<0.6 else rdf),gnf if cv<0.3 else (gld if cv<0.6 else rdf),wf,wf,wf,wf,wf],
   als=[la,ca,ca,la,ca,ca,ca,ca]); row+=1

# 종합 판정
row+=1; st(ws5,row,"E. 수익성 품질 종합",8); row+=1
for item,val in [
    ("DuPont 분석",f"평균 ROE {avg_roe*100:.1f}% — 고레버리지(4배+)가 ROE 드라이버. 순이익률 변동이 핵심"),
    ("FCF 품질","영업CF > 순이익 (항공업 감가상각 반영). CAPEX 부담 높아 FCF 변동 큼"),
    ("발생액","대부분 양호 — 영업CF가 순이익을 크게 상회 (감가상각 효과)"),
    ("이익 변동성",f"CV {cv:.2f} — 경기민감 항공업 특성. 유가·환율·수요에 따라 급등락"),
    ("결론","이익 품질은 양호(CF 우수). 다만 경기민감성·고레버리지로 이익 변동성이 핵심 리스크"),
]:
    ws5.cell(row=row,column=1,value=item).font=db; ws5.cell(row=row,column=1).fill=llf
    ws5.cell(row=row,column=1).alignment=la; ws5.cell(row=row,column=1).border=tb
    ws5.merge_cells(start_row=row,start_column=2,end_row=row,end_column=8)
    ws5.cell(row=row,column=2,value=val).font=df
    ws5.cell(row=row,column=2).alignment=la; ws5.cell(row=row,column=2).border=tb; row+=1

print("  [5/7] 수익성 품질")

# ============================================================
# SHEET 6: 내재가치 & 안전마진
# ============================================================
ws6 = wb.create_sheet("내재가치_안전마진"); ws6.sheet_properties.tabColor="2980B9"
sw(ws6,[22,16,16,16,16,16,16,16])
setup_print(ws6)

row=1; st(ws6,row,"내재가치 & 안전마진 분석",8); row+=1

# Owner Earnings
OE24 = NI24 + DA24 - CAPEX24
OE_per_share = OE24 / SHARES

st(ws6,row,"A. Owner Earnings (버핏 방식)",8); row+=1
wh(ws6,row,["항목","2024(억)","주당(원)","비고","","","",""]); row+=1
for lbl,amt,per_share,note in [
    ("순이익",int(NI24/억),f"{int(NI24/SHARES):,}",""),
    ("(+) 감가상각비",int(DA24/억),f"{int(DA24/SHARES):,}","유형+무형자산상각"),
    ("(-) CAPEX",int(CAPEX24/억),f"{int(CAPEX24/SHARES):,}","유형자산 취득"),
    ("(=) Owner Earnings",int(OE24/억),f"{int(OE_per_share):,}","버핏이 보는 진정한 이익"),
]:
    tot = "(=)" in lbl
    wr(ws6,row,[lbl,f"{amt:,}",per_share,note,"","","",""],
       fonts=[db if tot else df,d_blue if tot else df,sm,df,df,df,df,df],
       fills=[gld if tot else llf,gld if tot else wf,wf,wf,wf,wf,wf,wf],
       als=[la,ra,la,la,ca,ca,ca,ca]); row+=1

row+=1
oe_per = MCAP / OE24 if OE24 > 0 else 0
ws6.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
ws6.cell(row=row,column=1,value=f"현재 시총/Owner Earnings = {oe_per:.1f}배  |  Owner Earnings 수익률 = {OE24/MCAP*100:.1f}%  |  주의: CAPEX 변동 큼(항공기 투자 주기)").font=db
ws6.cell(row=row,column=1).fill=blf; ws6.cell(row=row,column=1).alignment=la; ws6.cell(row=row,column=1).border=tb
row+=2

# 자산가치
st(ws6,row,"B. 자산가치 (BPS 기반)",8); row+=1
wh(ws6,row,["지표","값","판정","","","","",""]); row+=1
for lbl,val,j in [
    ("BPS",f"{int(BPS):,}원",""),
    ("현재 PBR",f"{PRICE/BPS:.2f}배","적정" if 0.8<=PRICE/BPS<=1.5 else ("저평가" if PRICE/BPS<0.8 else "고평가")),
    ("PBR 0.8배 (하단)",fw(int(BPS*0.8)),""),
    ("PBR 1.0배",fw(int(BPS*1.0)),""),
    ("PBR 1.5배 (상단)",fw(int(BPS*1.5)),""),
]:
    wr(ws6,row,[lbl,val,j,"","","","",""],
       fonts=[db,d_blue,d_grn if "저평가" in j else (d_red if "고평가" in j else df),df,df,df,df,df],
       fills=[llf,gld,gnf if "저평가" in j else (rdf if "고평가" in j else wf),wf,wf,wf,wf,wf],
       als=[la,ra,ca,ca,ca,ca,ca,ca]); row+=1

# PER/PBR 밴드
row+=1; st(ws6,row,"C. PER/PBR 밴드 (과거 vs 현재)",8); row+=1
wh(ws6,row,["연도","EPS(원)","PER(현재가)","BPS(원)","PBR(현재가)","ROE","비고",""]); row+=1
# Historical equity for BPS calc
HIST = {
    2019: (-1691, 2_780_792_692_635),
    2020: (-625, None),  # equity is None in v_annual_performance for 2020
    2021: (1743, 6_865_689_155_499),
    2022: (4787, 9_292_460_206_825),
    2023: (2866, EQ23),
    2024: (EPS24, EQ24),
}
# 2020 equity not in view - use from financial statements (total_assets - total_liabilities)
eq_2020 = 25_190_061_074_278 - 21_878_336_959_960  # 3,311,724,114,318
HIST[2020] = (-625, eq_2020)

for yr_i in [2019,2020,2021,2022,2023,2024]:
    eps_v, eq_v = HIST[yr_i]
    # For years with negative earnings, use NI-based EPS
    if eps_v == 0 and str(yr_i) in annual:
        eps_v = int(annual[str(yr_i)]["ni"] / SHARES)
    bps_yr = int(eq_v / SHARES) if eq_v else 0
    per_yr = PRICE / eps_v if eps_v > 0 else 0
    pbr_yr = PRICE / bps_yr if bps_yr > 0 else 0
    roe_yr = annual[str(yr_i)]["ni"] / eq_v if eq_v else 0

    note = ""
    if yr_i == 2020: note = "코로나 적자"
    elif yr_i == 2024: note = "아시아나 합병"

    wr(ws6,row,[str(yr_i),f"{eps_v:,}" if eps_v != 0 else "적자",
                f"{per_yr:.1f}배" if per_yr>0 else "적자",f"{bps_yr:,}",
                f"{pbr_yr:.2f}배",f"{roe_yr*100:.1f}%",note,""],
       fonts=[db,d_blue if eps_v>0 else d_red,df,df,df,d_grn if roe_yr>0.10 else df,sm,df],
       fills=[lf]+[wf]*7,als=[ca,ra,ca,ra,ca,ca,la,ca]); row+=1

# 시나리오별 적정가
row+=1; st(ws6,row,"D. 시나리오별 적정주가",8); row+=1
wh(ws6,row,["시나리오","방법론","적정가","현재가 대비","전제","","",""]); row+=1

scenarios = [
    ("보수적","PER 5배 x EPS",int(EPS24*5),f"EPS {EPS24:,}원 x 5배 (항공업 하단)"),
    ("보수적","BPS x 0.7배",int(BPS*0.7),"PBR 하단 (부채 고려)"),
    ("기본","PER 7배 x EPS",int(EPS24*7),f"EPS {EPS24:,}원 x 7배 (항공업 적정)"),
    ("기본","EV/EBITDA 5배",int((EBITDA24*5 - NET_DEBT)/SHARES),"EV/EBITDA 기반"),
    ("낙관적","PER 10배 x EPS",int(EPS24*10),f"EPS {EPS24:,}원 x 10배 (합병 시너지 반영)"),
    ("낙관적","BPS x 1.2배",int(BPS*1.2),"부채 감축 + 자산 재평가"),
]
for scen,method,target,basis in scenarios:
    upside = (target-PRICE)/PRICE
    uf = d_grn if upside>0 else d_red
    ufl = gnf if upside>0 else rdf
    scen_fl = gnf if "낙관" in scen else (gld if "기본" in scen else rdf)
    wr(ws6,row,[scen,method,fw(target),f"{upside*100:+.1f}%",basis,"","",""],
       fonts=[db,df,d_blue,uf,sm,df,df,df],fills=[scen_fl,wf,gld,ufl,wf,wf,wf,wf],
       als=[ca,la,ra,ca,la,ca,ca,ca]); row+=1

# 안전마진 테스트
row+=1; st(ws6,row,"E. 안전마진 테스트",8); row+=1

conserv_values = [s[2] for s in scenarios if "보수" in s[0]]
conserv_avg = int(statistics.mean(conserv_values)) if conserv_values else PRICE
base_values = [s[2] for s in scenarios if "기본" in s[0]]
base_avg = int(statistics.mean(base_values)) if base_values else PRICE
opt_values = [s[2] for s in scenarios if "낙관" in s[0]]
opt_avg = int(statistics.mean(opt_values)) if opt_values else PRICE

for lbl,iv,verdict in [
    ("보수적 내재가치",conserv_avg,"하방 리스크 확인" if PRICE>conserv_avg else "안전마진 확보"),
    ("기본 내재가치",base_avg,"적정" if abs(PRICE-base_avg)/base_avg<0.15 else ("상승여력" if PRICE<base_avg else "과대평가")),
    ("낙관적 내재가치",opt_avg,"업사이드" if PRICE<opt_avg else "이미 반영"),
]:
    margin = (iv - PRICE) / iv if iv else 0
    mf2 = d_grn if margin>0.15 else (df if margin>0 else d_red)
    mfl = gnf if margin>0.15 else (gld if margin>0 else rdf)
    wr(ws6,row,[lbl,fw(iv),f"안전마진 {margin*100:+.1f}%",verdict,f"현재가 {PRICE:,}원","","",""],
       fonts=[db,d_blue,mf2,mf2,sm,df,df,df],fills=[llf,gld,mfl,mfl,wf,wf,wf,wf],
       als=[la,ra,ca,ca,la,ca,ca,ca]); row+=1

row+=1
ws6.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
verdict_overall = (f"기본 시나리오({fw(base_avg)}) 대비 {(base_avg-PRICE)/base_avg*100:+.1f}% 안전마진. "
                   "부채비율 300% 이하 + 합병 시너지 가시화 시 낙관 시나리오 달성 가능. "
                   "재무구조 개선이 확인되기 전까지는 HOLD 유지.")
ws6.cell(row=row,column=1,value=f"-> {verdict_overall}").font=db
ws6.cell(row=row,column=1).fill=blf; ws6.cell(row=row,column=1).alignment=la; ws6.cell(row=row,column=1).border=tb

print("  [6/7] 내재가치/안전마진")

# ============================================================
# SHEET 7: 투자판정 & 모니터링
# ============================================================
ws7 = wb.create_sheet("투자판정"); ws7.sheet_properties.tabColor="1ABC9C"
sw(ws7,[22,16,16,16,16,16,16,16])
setup_print(ws7)

row=1
# Title banner
for r in range(1,5):
    for c in range(1,9): ws7.cell(row=r,column=c).fill=tf
ws7.merge_cells('A2:H2')
ws7.cell(row=2,column=1,value="최종 투자 판정").font=Font(name="맑은 고딕",size=22,bold=True,color=W)
ws7.cell(row=2,column=1).alignment=ca
row=6

# 최종 판정
st(ws7,row,"A. 최종 판정",8); row+=1

# Determine verdict (70 = B grade, HOLD)
if total_score >= 75:
    verdict = "BUY (매수)"
    verdict_color = GREEN_C
    verdict_fill = gnf
elif total_score >= 55:
    verdict = "HOLD (관망 - 매수 근접)"
    verdict_color = GOLD_C
    verdict_fill = gld
else:
    verdict = "REJECT (거부)"
    verdict_color = RED_C
    verdict_fill = rdf

ws7.merge_cells(start_row=row,start_column=1,end_row=row+1,end_column=3)
c = ws7.cell(row=row,column=1,value=verdict)
c.font=Font(name="맑은 고딕",size=20,bold=True,color=verdict_color)
c.fill=verdict_fill; c.alignment=ca; c.border=tb

ws7.merge_cells(start_row=row,start_column=4,end_row=row+1,end_column=8)
c = ws7.cell(row=row,column=4)
c.value=(f"Four Filters 종합 {total_score}/100 (등급 {grade})\n"
         f"PER {per24:.1f}배 | PBR {pbr:.2f}배 | FCF Yield {FCF24/MCAP*100:.1f}% | "
         f"배당수익률 {DPS24/PRICE*100:.1f}% | 부채비율 {LIAB24/EQ24*100:.0f}%")
c.font=Font(name="맑은 고딕",size=10,bold=True,color=NAVY)
c.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True); c.border=tb
row+=3

# 매수 전환 조건
st(ws7,row,"B. 매수 전환 조건 (재무구조 개선 확인 시)",8); row+=1
ws7.merge_cells(start_row=row,start_column=1,end_row=row+1,end_column=8)
c = ws7.cell(row=row,column=1)
c.value=("현재 판정: HOLD (매수 근접). 아래 조건 충족 시 매수 전환 권고:\n"
         "1) 부채비율 300% 이하 확인  2) 아시아나 합병 시너지 가시화 (원가 절감, 노선 최적화)  "
         "3) 2년 연속 FCF 양(+) 유지  4) 배당 인상 또는 자사주 매입 시작")
c.font=Font(name="맑은 고딕",size=10,bold=True,color=NAVY)
c.alignment=Alignment(horizontal='left',vertical='top',wrap_text=True); c.border=tb; c.fill=blf
row+=3

# 판정 근거
st(ws7,row,"C. 판정 근거",8); row+=1
reasons = [
    ("매수 근거 1","글로벌 네트워크 + 인천허브: 아시아-미주 최적 위치, 스카이팀 20개 항공사 제휴"),
    ("매수 근거 2",f"밸류에이션 매력: PER {per24:.1f}배, PBR {pbr:.2f}배, FCF Yield {FCF24/MCAP*100:.1f}%"),
    ("매수 근거 3","아시아나 합병 시너지: 노선 최적화·비용 절감·슬롯 통합으로 국내 독과점"),
    ("매수 근거 4","화물 사업 경쟁력: 세계 3위, e커머스 확대와 함께 구조적 성장"),
    ("주의 사항 1","부채비율 329%: 항공업 특성이나, 금리 상승기 이자부담 가중"),
    ("주의 사항 2","유상증자 이력 5회: 과거 주주 희석 패턴. 재발 시 주당가치 훼손"),
    ("주의 사항 3","유가·환율 민감: WTI $120+ 또는 원화 약세 시 수익성 급락"),
]
for title,detail in reasons:
    is_risk = "주의" in title
    tf2 = d_red if is_risk else d_grn
    tfl = rdf if is_risk else gnf
    ws7.cell(row=row,column=1,value=title).font=tf2; ws7.cell(row=row,column=1).fill=tfl
    ws7.cell(row=row,column=1).alignment=la; ws7.cell(row=row,column=1).border=tb
    ws7.merge_cells(start_row=row,start_column=2,end_row=row,end_column=8)
    ws7.cell(row=row,column=2,value=detail).font=df
    ws7.cell(row=row,column=2).alignment=la; ws7.cell(row=row,column=2).border=tb; row+=1

# SWOT (구루 관점)
row+=1; st(ws7,row,"D. 구루 관점 SWOT",8); row+=1

ws7.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
ws7.cell(row=row,column=1,value="강점 (Strengths)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws7.cell(row=row,column=1).fill=PatternFill("solid",fgColor=GREEN_C); ws7.cell(row=row,column=1).alignment=ca
ws7.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
ws7.cell(row=row,column=5,value="약점 (Weaknesses)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws7.cell(row=row,column=5).fill=PatternFill("solid",fgColor=RED_C); ws7.cell(row=row,column=5).alignment=ca
row+=1

SW_S = ["국내 1위 FSC + 아시아나 합병 → 독과점",
        "인천허브 + 스카이팀 글로벌 네트워크",
        "화물 세계 3위 → 고수익 사업부",
        "강력한 브랜드 + 마일리지 락인"]
SW_W = ["부채비율 329% — 고레버리지 구조",
        "유가·환율 변동에 극히 민감",
        "유상증자 5회 — 주주 희석 이력",
        "자사주 매입·소각 전무 — 보수적 주주환원"]
for i in range(4):
    ws7.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
    ws7.cell(row=row,column=1,value=f"  {SW_S[i]}").font=df; ws7.cell(row=row,column=1).fill=gnf
    ws7.cell(row=row,column=1).alignment=la; ws7.cell(row=row,column=1).border=tb
    ws7.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
    ws7.cell(row=row,column=5,value=f"  {SW_W[i]}").font=df; ws7.cell(row=row,column=5).fill=rdf
    ws7.cell(row=row,column=5).alignment=la; ws7.cell(row=row,column=5).border=tb; row+=1

row+=1
ws7.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
ws7.cell(row=row,column=1,value="기회 (Opportunities)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws7.cell(row=row,column=1).fill=PatternFill("solid",fgColor="2980B9"); ws7.cell(row=row,column=1).alignment=ca
ws7.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
ws7.cell(row=row,column=5,value="위협 (Threats)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws7.cell(row=row,column=5).fill=PatternFill("solid",fgColor="7F8C8D"); ws7.cell(row=row,column=5).alignment=ca
row+=1

O = ["아시아나 합병 시너지 본격화 (비용 절감 1조+)",
     "중국 노선 회복 + 동남아 수요 증가",
     "e커머스 확대 → 항공화물 구조적 성장",
     "부채 감축 → 밸류에이션 리레이팅"]
T = ["유가 $120+ 장기 지속 → 수익성 급락",
     "글로벌 경기침체 → 여행 수요 감소",
     "LCC 시장 확대 → 단거리 노선 경쟁 심화",
     "환율 급등 → 외화부채 부담 가중"]
for i in range(4):
    ws7.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
    ws7.cell(row=row,column=1,value=f"  {O[i]}").font=df; ws7.cell(row=row,column=1).fill=blf
    ws7.cell(row=row,column=1).alignment=la; ws7.cell(row=row,column=1).border=tb
    ws7.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
    ws7.cell(row=row,column=5,value=f"  {T[i]}").font=df
    ws7.cell(row=row,column=5).fill=PatternFill("solid",fgColor="E5E7E9")
    ws7.cell(row=row,column=5).alignment=la; ws7.cell(row=row,column=5).border=tb; row+=1

# 모니터링 지표
row+=1; st(ws7,row,"E. 핵심 모니터링 지표 (5개)",8); row+=1
wh(ws7,row,["#","지표","세부 내용","확인 시기","구루 관점","","",""]); row+=1
monitors = [
    ("1","부채비율/순차입금","300% 이하 진입 여부, 순차입금 감소 추세","분기 실적","재무 안전성 핵심"),
    ("2","아시아나 합병 시너지","노선 통합·비용 절감·인력 효율화 진척도","반기","해자 확대 여부"),
    ("3","유가·환율 추이","WTI 유가 $80~100 적정, $120+ 경고 / 원달러 환율","수시","수익성 직결"),
    ("4","여객·화물 수요","국제선 여객 RPK, 화물 FTK 성장률","월간 통계","매출 성장 확인"),
    ("5","배당·주주환원 변화","배당 인상 또는 자사주 매입 시작 여부","3월 주총","경영진 의지 확인"),
]
for num,title,detail,timing,guru in monitors:
    wr(ws7,row,[num,title,detail,timing,guru,"","",""],
       fonts=[Font(name="맑은 고딕",size=12,bold=True,color=NAVY),db,df,df,sm,df,df,df],
       fills=[gld,llf,wf,llf,wf,wf,wf,wf],als=[ca,la,la,ca,la,ca,ca,ca])
    ws7.row_dimensions[row].height=40; row+=1

# 매도 트리거
row+=1; st(ws7,row,"F. 매도 트리거 조건",8); row+=1
sell_triggers = [
    ("부채비율 400% 초과","재무 리스크 급증 → 즉시 매도"),
    ("유가 $120+ 장기 지속 (6개월+)","구조적 수익성 훼손 → 비중 축소"),
    ("영업이익률 5% 이하 2년 연속","합병 시너지 실패 → 해자 약화"),
    ("아시아나 합병 시너지 미달 (통합 지연)","기대 대비 실망 → 밸류에이션 하향"),
    ("경영권 분쟁 재발 또는 유상증자 재개","주주 가치 훼손 → 즉시 매도"),
]
for trigger,action in sell_triggers:
    ws7.merge_cells(start_row=row,start_column=1,end_row=row,end_column=5)
    ws7.cell(row=row,column=1,value=f"  {trigger}").font=df
    ws7.cell(row=row,column=1).fill=rdf; ws7.cell(row=row,column=1).alignment=la; ws7.cell(row=row,column=1).border=tb
    ws7.merge_cells(start_row=row,start_column=6,end_row=row,end_column=8)
    ws7.cell(row=row,column=6,value=action).font=d_red
    ws7.cell(row=row,column=6).alignment=la; ws7.cell(row=row,column=6).border=tb; row+=1

# 리스크 시나리오
row+=1; st(ws7,row,"G. 리스크 시나리오 (최악의 경우)",8); row+=1
ws7.merge_cells(start_row=row,start_column=1,end_row=row+3,end_column=8)
c = ws7.cell(row=row,column=1)
c.value=("최악 시나리오: 유가 $150+ 장기화 + 글로벌 경기침체 + 아시아나 통합 실패\n"
         "-> 매출 12조(-33%), 영업적자 전환, 순적자 1조+\n"
         "-> 부채비율 500%+ 돌파, 유상증자 불가피\n"
         "-> PBR 0.3배 적용 시 주가 ~9,000원 (현재가 대비 -62%). 이 시나리오 확률: 5~10%")
c.font=Font(name="맑은 고딕",size=10,color=RED_C)
c.fill=PatternFill("solid",fgColor="FDE8E8")
c.alignment=Alignment(horizontal='left',vertical='top',wrap_text=True); c.border=tb

# 주요 이벤트 타임라인
row+=5; st(ws7,row,"H. 주요 이벤트 타임라인 (최근)",8); row+=1
wh(ws7,row,["일자","유형","내용","","","","",""]); row+=1
for ev in event_rows[:15]:
    dt = ev[0]; tp = ev[1]; content = (ev[2] or "")[:70]
    tp_f = d_grn if "배당" in tp else (d_blue if "IR" in tp else (d_red if "매출변동" in tp else db))
    wr(ws7,row,[dt,tp,content,"","","","",""],
       fonts=[db,tp_f,df,df,df,df,df,df],
       fills=[llf,gld if "배당" in tp else wf,wf,wf,wf,wf,wf,wf],
       als=[ca,ca,la,ca,ca,ca,ca,ca]); row+=1

print("  [7/7] 투자판정/모니터링")

# === SAVE ===
OUT = os.path.join(BASE, "대한항공_투자구루분석.xlsx")
wb.save(OUT)
conn.close()
print(f"\n투자 구루 분석 보고서 생성 완료: {OUT}")
print(f"시트 ({len(wb.sheetnames)}개): {wb.sheetnames}")
