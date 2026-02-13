# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "대한항공_밸류에이션.xlsx")
wb = Workbook()

# === CONSTANTS ===
PRICE = 23700
SHARES_OUTSTANDING = 368220612  # 보통주 유통주식수
SHARES_PREFERRED = 1110791  # 우선주 유통
SHARES_WA = 368220612  # 가중평균 (EPS 역산 기준)
MARKET_CAP = PRICE * SHARES_OUTSTANDING  # 약 87,309억원

# 2024 Annual (연결)
REV_2024 = 17870718495804  # 매출(영업수익) 17.9조
OP_2024 = 2110200077994   # 영업이익 2.1조
NI_2024 = 1317261688939   # 당기순이익(지배) 1.3조
EPS_2024 = 3566           # EPS(보통주)
EQUITY_TOTAL_2024 = 10963191867177  # 자본총계 10.9조
EQUITY_CTRL_2024 = 10472800000000   # 지배지분 약 10.47조 (BPS 역산: 28,400원 × 368,220,612주)
EQUITY_2023 = 9700000000000  # 2023 지배지분 추정 (ROE 계산용)
ASSETS_2024 = 47012065940089  # 자산총계 47조
LIAB_2024 = 36048874072912   # 부채총계 36조
CAPITAL_2024 = 1846657275000  # 자본금

# 현금 & 차입금 (2024, 원)
CASH_2024 = 2215624563052      # 현금
ST_FINANCIAL_2024 = 4475200000000000 // 1000  # 단기금융상품 (4,475,200,000,000)
# 다시 올바르게 입력
CASH_2024 = 2215624563052
ST_FINANCIAL = 4475200000000   # 단기금융상품
ST_DEBT_2024 = 2517021740000   # 단기차입금
LT_DEBT_2024 = 1819300917987   # 장기차입금
CURRENT_LT_DEBT = 2541418556700  # 유동성장기부채
BONDS_2024 = 1463107639009     # 사채
LEASE_NONCURRENT = 8744563527885  # 리스부채(비유동)
LEASE_CURRENT = 2182080664669    # 유동성리스부채

# CF 항목 (2024)
DA_2024 = 1737200000000    # 감가상각비
IA_2024 = 58800000000      # 무형자산상각비
OPCF_2024 = 4558900000000  # 영업활동CF 4.6조
CAPEX_2024 = 2894000000000 # CAPEX 2.9조
INTEREST_2024 = 514100000000  # 이자비용
DIV_PAID_2024 = 278200000000  # 배당금지급

# Historical (연결, 대한항공)
HIST = {
    2015: {"rev": 11900000000000, "op": 1300000000000, "ni": 300000000000, "eps": 815, "equity": 3500000000000},
    2016: {"rev": 11700000000000, "op": 1560000000000, "ni": 127000000000, "eps": 345, "equity": 3300000000000},
    2017: {"rev": 12100000000000, "op": 1100000000000, "ni": 489000000000, "eps": 1328, "equity": 3700000000000},
    2018: {"rev": 12700000000000, "op": 530000000000, "ni": -200000000000, "eps": -543, "equity": 3400000000000},
    2019: {"rev": 12400000000000, "op": 360000000000, "ni": -660000000000, "eps": -1792, "equity": 2700000000000},
    2020: {"rev": 7400000000000, "op": -740000000000, "ni": -2060000000000, "eps": -5593, "equity": 600000000000},
    2021: {"rev": 10100000000000, "op": 880000000000, "ni": 150000000000, "eps": 407, "equity": 5200000000000},
    2022: {"rev": 14700000000000, "op": 2230000000000, "ni": 1090000000000, "eps": 2959, "equity": 7700000000000},
    2023: {"rev": 16300000000000, "op": 1760000000000, "ni": 1210000000000, "eps": 3286, "equity": 9700000000000},
    2024: {"rev": REV_2024, "op": OP_2024, "ni": NI_2024, "eps": EPS_2024, "equity": EQUITY_CTRL_2024},
}

억 = 100_000_000

# === DERIVED CALCULATIONS ===
BPS = 28400  # 지배지분 / 보통주유통주식수
DPS_2024 = int(DIV_PAID_2024 / SHARES_OUTSTANDING)  # 약 755원

# 순차입금 (항공업: 리스부채 포함)
TOTAL_DEBT = ST_DEBT_2024 + LT_DEBT_2024 + CURRENT_LT_DEBT + BONDS_2024 + LEASE_NONCURRENT + LEASE_CURRENT
TOTAL_CASH = CASH_2024 + ST_FINANCIAL
NET_DEBT = TOTAL_DEBT - TOTAL_CASH  # 약 12.6조
EBITDA_2024 = OP_2024 + DA_2024 + IA_2024  # 영업이익 + 감가상각비 + 무형자산상각비
FCF_2024 = OPCF_2024 - CAPEX_2024  # 영업CF - CAPEX
EV = MARKET_CAP + NET_DEBT  # 시총 + 순차입금

# ROE
AVG_EQ = (EQUITY_CTRL_2024 + EQUITY_2023) / 2
ROE_2024 = NI_2024 / AVG_EQ

# Ke (자기자본비용) for RIM
Ke = 0.10  # 무위험 3.5% + beta(1.2) × ERP(5.4%) ≈ 10%
GROWTH = 0.02  # 장기성장률


# === STYLES ===
NAVY = "1B2A4A"
DARK = "2C3E6B"
MID = "3A5BA0"
WHITE = "FFFFFF"
GOLD_BG = "FFF3CD"
GREEN_BG = "D4EDDA"
RED_BG = "F8D7DA"
BLUE_BG = "D6E4F0"
LBLUE = "EBF1F8"

title_font = Font(name="맑은 고딕", size=18, bold=True, color=WHITE)
section_font = Font(name="맑은 고딕", size=13, bold=True, color=NAVY)
h_font = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
d_font = Font(name="맑은 고딕", size=10)
d_bold = Font(name="맑은 고딕", size=10, bold=True)
d_blue = Font(name="맑은 고딕", size=10, bold=True, color="0000FF")
d_green = Font(name="맑은 고딕", size=10, bold=True, color="27AE60")
d_red = Font(name="맑은 고딕", size=10, bold=True, color="C0392B")
note_font = Font(name="맑은 고딕", size=9, color="666666")
big_font = Font(name="맑은 고딕", size=14, bold=True, color=NAVY)
huge_font = Font(name="맑은 고딕", size=20, bold=True, color=NAVY)

title_fill = PatternFill("solid", fgColor=NAVY)
header_fill = PatternFill("solid", fgColor=DARK)
mid_fill = PatternFill("solid", fgColor=MID)
blue_fill = PatternFill("solid", fgColor=BLUE_BG)
lblue_fill = PatternFill("solid", fgColor=LBLUE)
gold_fill = PatternFill("solid", fgColor=GOLD_BG)
green_fill = PatternFill("solid", fgColor=GREEN_BG)
red_fill = PatternFill("solid", fgColor=RED_BG)
white_fill = PatternFill("solid", fgColor=WHITE)
gray_fill = PatternFill("solid", fgColor="F2F2F2")

center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")
thin_border = Border(
    left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"))
bottom_border = Border(bottom=Side(style="medium", color=NAVY))


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def setup_print(ws):
    """Letter 용지 가로(Landscape) 인쇄 최적화 설정"""
    ws.page_setup.paperSize = Worksheet.PAPERSIZE_LETTER
    ws.page_setup.orientation = Worksheet.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.5, bottom=0.5,
        header=0.3, footer=0.3
    )
    ws.print_options.horizontalCentered = True

def write_header(ws, row, vals, fills=None):
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = h_font
        c.fill = fills[i-1] if fills else header_fill
        c.alignment = center
        c.border = thin_border

def write_row(ws, row, vals, fonts=None, fills=None, aligns=None):
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = fonts[i-1] if fonts else d_font
        c.fill = fills[i-1] if fills else white_fill
        c.alignment = aligns[i-1] if aligns else center
        c.border = thin_border

def section_title(ws, row, title, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=title)
    c.font = section_font
    c.border = bottom_border
    return row + 1

def fmt(v):
    """숫자를 억원 문자열로"""
    return f"{v/억:,.0f}억"

def fmt_won(v):
    return f"{v:,.0f}원"

def pct(v):
    return f"{v*100:.1f}%"


# ============================================================
# SHEET 1: 밸류에이션 종합 대시보드
# ============================================================
ws1 = wb.active
ws1.title = "종합"
ws1.sheet_properties.tabColor = "1B2A4A"
set_widths(ws1, [22, 18, 18, 18, 18, 22])
setup_print(ws1)

# Title bar
row = 1
ws1.merge_cells('A1:F2')
c = ws1.cell(row=1, column=1, value="대한항공(003490) 밸류에이션 분석")
c.font = title_font
c.fill = title_fill
c.alignment = Alignment(horizontal="center", vertical="center")

row = 3
ws1.merge_cells('A3:F3')
ws1.cell(row=3, column=1, value=f"기준일: 2026.02.13 | 현재가: {PRICE:,}원 | 시가총액: {fmt(MARKET_CAP)}").font = Font(name="맑은 고딕", size=11, bold=True, color=DARK)
ws1.cell(row=3, column=1).alignment = center

# Key metrics box
row = 5
row = section_title(ws1, row, "기본 정보", 6)
info_data = [
    ["현재 주가", fmt_won(PRICE), "보통주 유통주식수", f"{SHARES_OUTSTANDING:,}주", "시가총액(보통주)", fmt(MARKET_CAP)],
    ["우선주 유통", f"{SHARES_PREFERRED:,}주", "자본금", fmt(CAPITAL_2024), "부채비율", f"{LIAB_2024/EQUITY_TOTAL_2024*100:.0f}%"],
    ["2024 EPS", fmt_won(EPS_2024), "2024 BPS", fmt_won(BPS), "2024 DPS", fmt_won(DPS_2024)],
    ["2024 매출", fmt(REV_2024), "2024 영업이익", fmt(OP_2024), "2024 순이익(지배)", fmt(NI_2024)],
]
for data in info_data:
    for i in range(0, 6, 2):
        ws1.cell(row=row, column=i+1, value=data[i]).font = d_bold
        ws1.cell(row=row, column=i+1).fill = blue_fill
        ws1.cell(row=row, column=i+1).alignment = left
        ws1.cell(row=row, column=i+1).border = thin_border
        ws1.cell(row=row, column=i+2, value=data[i+1]).font = d_blue
        ws1.cell(row=row, column=i+2).fill = white_fill
        ws1.cell(row=row, column=i+2).alignment = center
        ws1.cell(row=row, column=i+2).border = thin_border
    row += 1

# Valuation Summary
row += 1
row = section_title(ws1, row, "밸류에이션 멀티플 종합", 6)
write_header(ws1, row, ["지표", "산출 방식", "값", "판정", "업종 평균(참고)", "비고"])
row += 1

PER_2024 = PRICE / EPS_2024
PBR_2024 = PRICE / BPS
EV_EBITDA_2024 = EV / EBITDA_2024
FCF_PER_SHARE = FCF_2024 / SHARES_OUTSTANDING
OPCF_PER_SHARE = OPCF_2024 / SHARES_OUTSTANDING

valuation_rows = [
    ["PER (2024)", f"주가 / EPS({fmt_won(EPS_2024)})", f"{PER_2024:.1f}배", "저평가", "8~12배", "확정 실적 기준"],
    ["PBR (2024)", f"주가 / BPS({fmt_won(BPS)})", f"{PBR_2024:.2f}배", "저평가", "1.0~1.5배", "지배지분 기준"],
    ["EV/EBITDA (2024)", f"EV({fmt(EV)}) / EBITDA({fmt(EBITDA_2024)})", f"{EV_EBITDA_2024:.1f}배", "저평가", "6~8배", "리스부채 포함 EV"],
    ["PSR (2024)", f"시총 / 매출({fmt(REV_2024)})", f"{MARKET_CAP/REV_2024:.2f}배", "매우 저평가", "0.5~1.0배", "매출 17.9조 기업"],
    ["PCR", f"주가 / OpCF/주({fmt_won(int(OPCF_PER_SHARE))})", f"{PRICE/OPCF_PER_SHARE:.1f}배", "저평가", "5~10배", "영업CF 기준"],
    ["ROE (2024)", f"순이익 / 평균자본", pct(ROE_2024), "양호", "8~15%", "지배지분 기준"],
    ["배당수익률", f"DPS({fmt_won(DPS_2024)}) / 주가", pct(DPS_2024/PRICE), "보통", "2~4%", "2024년 배당"],
    ["FCF 수익률", f"FCF/주({fmt_won(int(FCF_PER_SHARE))}) / 주가", pct(FCF_PER_SHARE/PRICE), "양호", "3~6%", "CAPEX 2.9조 차감"],
    ["순차입금/EBITDA", f"순차입금({fmt(NET_DEBT)}) / EBITDA", f"{NET_DEBT/EBITDA_2024:.1f}배", "주의", "2배 이하", "리스부채 포함"],
    ["이자보상배율", f"영업이익 / 이자비용({fmt(INTEREST_2024)})", f"{OP_2024/INTEREST_2024:.1f}배", "양호", "3배 이상", "개선 추세"],
]

for vals in valuation_rows:
    judge = vals[3]
    if "매우 저" in judge or "매우 양" in judge:
        judge_font = d_green
        judge_fill = green_fill
    elif "저평가" in judge or "양호" in judge:
        judge_font = Font(name="맑은 고딕", size=10, bold=True, color="2E86C1")
        judge_fill = PatternFill("solid", fgColor="D6EAF8")
    elif "적정" in judge or "보통" in judge:
        judge_font = d_bold
        judge_fill = gold_fill
    elif "주의" in judge:
        judge_font = d_red
        judge_fill = red_fill
    else:
        judge_font = d_red
        judge_fill = red_fill

    write_row(ws1, row, vals,
              fonts=[d_bold, d_font, d_blue, judge_font, d_font, note_font],
              fills=[lblue_fill, white_fill, gold_fill, judge_fill, gray_fill, white_fill],
              aligns=[left, left, center, center, center, left])
    row += 1

row += 1
ws1.cell(row=row, column=1, value="* 판정 기준: 항공업(FSC) 글로벌 피어 및 국내 대형주 평균 대비 상대 평가").font = note_font
row += 1
ws1.cell(row=row, column=1, value="* 항공업 특성: 리스부채를 EV에 포함해야 정확한 밸류에이션 가능 (IFRS 16)").font = note_font
row += 1
ws1.cell(row=row, column=1, value=f"* 순차입금 산출: 총차입금(리스 포함) {fmt(TOTAL_DEBT)} - 현금성자산 {fmt(TOTAL_CASH)} = {fmt(NET_DEBT)}").font = note_font

print("  [1/5] 종합 대시보드 완료")


# ============================================================
# SHEET 2: PER 다각도 분석
# ============================================================
ws2 = wb.create_sheet("PER분석")
ws2.sheet_properties.tabColor = "2980B9"
set_widths(ws2, [24, 16, 16, 16, 16, 20])
setup_print(ws2)

row = 1
row = section_title(ws2, row, f"PER 다각도 분석 (현재가 {PRICE:,}원 기준)", 6)

# A. EPS 산출 방식별 PER
row = section_title(ws2, row, "A. EPS 산출 방식별 PER", 6)
write_header(ws2, row, ["산출 방식", "순이익(억)", "EPS(원)", "PER(배)", "의미", "비고"])
row += 1

per_methods = [
    ["2024 확정 (사업보고서)", int(NI_2024/억), EPS_2024, PRICE/EPS_2024,
     "가장 신뢰 높은 확정치", "reprt_code 11011"],
    ["2024 영업이익 기반", int(OP_2024/억), int(OP_2024/SHARES_OUTSTANDING), PRICE/(OP_2024/SHARES_OUTSTANDING),
     "영업이익 기준 PER (참고)", "순이익 변동성 큰 항공업 보완"],
    ["Forward PER (10% 성장)", int(NI_2024*1.1/억), int(EPS_2024*1.1), PRICE/(EPS_2024*1.1),
     "보수적 이익성장 가정", "아시아나 합병 시너지"],
    ["Forward PER (20% 성장)", int(NI_2024*1.2/억), int(EPS_2024*1.2), PRICE/(EPS_2024*1.2),
     "적극적 이익성장 가정", "화물+여객 동반 성장"],
]

for vals in per_methods:
    per_val = vals[3]
    if per_val < 6:
        per_font = d_green
        per_fill = green_fill
    elif per_val < 8:
        per_font = d_blue
        per_fill = PatternFill("solid", fgColor="D6EAF8")
    else:
        per_font = d_bold
        per_fill = gold_fill

    write_row(ws2, row,
              [vals[0], vals[1], f"{vals[2]:,}", f"{per_val:.2f}배", vals[4], vals[5]],
              fonts=[d_bold, d_font, d_blue, per_font, d_font, note_font],
              fills=[lblue_fill, white_fill, gold_fill, per_fill, white_fill, gray_fill],
              aligns=[left, right, right, center, left, left])
    row += 1

# B. Historical EPS & PER 추이
row += 1
row = section_title(ws2, row, "B. 역사적 EPS 추이 (2015-2024)", 6)
write_header(ws2, row, ["연도", "매출(억)", "영업이익(억)", "순이익(억)", "EPS(원)", "비고"])
row += 1

notes_by_year = {
    2015: "정상영업",
    2016: "유가 하락 수혜",
    2017: "양호한 실적",
    2018: "유가 급등, 적자전환",
    2019: "코로나 직전, 적자 지속",
    2020: "코로나 팬데믹 충격",
    2021: "화물 호황, 흑자전환",
    2022: "역대 최대 실적",
    2023: "여객 정상화, 고실적 지속",
    2024: "아시아나 합병, 매출 최대",
}

for yr in range(2015, 2025):
    h = HIST[yr]
    eps_font = d_green if h["eps"] > 0 else d_red
    ni_font = d_green if h["ni"] > 0 else d_red

    write_row(ws2, row,
              [str(yr), int(h["rev"]/억), int(h["op"]/억), int(h["ni"]/억),
               f"{h['eps']:,}", notes_by_year.get(yr, "")],
              fonts=[d_bold, d_font, d_font, ni_font, eps_font, note_font],
              fills=[lblue_fill, white_fill, white_fill, white_fill,
                     green_fill if h["eps"] > 0 else red_fill, gray_fill],
              aligns=[center, right, right, right, right, left])
    row += 1

# C. 업종 비교 PER
row += 1
row = section_title(ws2, row, "C. 항공업 피어 PER 비교 (참고)", 6)
write_header(ws2, row, ["회사", "시가총액", "PER(배)", "PBR(배)", "EV/EBITDA", "비고"])
row += 1

peers = [
    ["대한항공", f"{int(MARKET_CAP/억):,}억", f"{PER_2024:.1f}", f"{PBR_2024:.2f}", f"{EV_EBITDA_2024:.1f}", "분석 대상"],
    ["ANA Holdings", "약 20조원", "10~12", "1.5~2.0", "6~8", "일본 1위"],
    ["Singapore Airlines", "약 18조원", "8~10", "1.5~2.0", "5~7", "프리미엄 캐리어"],
    ["Delta Air Lines", "약 50조원", "7~9", "3~5", "5~7", "미국 Big 3"],
    ["Cathay Pacific", "약 7조원", "6~8", "0.8~1.2", "4~6", "홍콩 FSC"],
    ["항공업 평균", "-", "8~12", "1.0~2.0", "5~8", "글로벌 FSC 기준"],
]

for vals in peers:
    is_target = vals[0] == "대한항공"
    is_avg = vals[0] == "항공업 평균"
    write_row(ws2, row, vals,
              fonts=[d_bold, d_font, d_blue if is_target else d_font, d_font, d_font, note_font],
              fills=[gold_fill if is_target else (blue_fill if is_avg else lblue_fill)] + \
                    [gold_fill if is_target else (blue_fill if is_avg else white_fill)] * 5,
              aligns=[left, right, center, center, center, left])
    row += 1

row += 1
ws2.cell(row=row, column=1, value="→ 현재 PER 6.6배는 글로벌 항공업 평균(8~12배) 대비 상당한 할인. 아시아나 합병 효과 반영 시 업사이드 여력 존재.").font = d_bold

print("  [2/5] PER 분석 완료")


# ============================================================
# SHEET 3: PBR/ROE/RIM 분석
# ============================================================
ws3 = wb.create_sheet("PBR_ROE")
ws3.sheet_properties.tabColor = "8E44AD"
set_widths(ws3, [20, 16, 16, 16, 16, 22])
setup_print(ws3)

row = 1
row = section_title(ws3, row, "PBR / ROE / 잔여이익모델(RIM) 분석", 6)

# A. Historical BPS & ROE
row = section_title(ws3, row, "A. 연도별 자본/BPS 추이", 6)
write_header(ws3, row, ["연도", "지배지분(억)", "BPS(원)", "ROE", "EPS(원)", "순이익(억)"])
row += 1

prev_eq = None
for yr in [2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024]:
    h = HIST[yr]
    eq = h["equity"]
    bps = int(eq / SHARES_OUTSTANDING)
    ni = h["ni"]
    eps = h["eps"]
    if prev_eq and prev_eq > 0 and eq > 0:
        avg_eq = (eq + prev_eq) / 2
        if avg_eq > 0:
            roe = ni / avg_eq
        else:
            roe = 0
    else:
        roe = ni / eq if eq > 0 else 0
    prev_eq = eq

    roe_font = d_green if roe > 0.12 else (d_red if roe < 0 else d_font)
    write_row(ws3, row,
              [str(yr), int(eq/억), f"{bps:,}", f"{roe*100:.1f}%", f"{eps:,}", int(ni/억)],
              fonts=[d_bold, d_font, d_blue, roe_font, d_font if eps >= 0 else d_red, d_font if ni >= 0 else d_red],
              fills=[lblue_fill, white_fill, gold_fill, white_fill, white_fill, white_fill],
              aligns=[center, right, right, center, right, right])
    row += 1

# Current PBR
row += 1
row = section_title(ws3, row, "B. 현재 PBR 분석", 6)
write_header(ws3, row, ["항목", "값", "", "", "", ""])
row += 1

pbr_info = [
    ("2024말 지배지분", fmt(EQUITY_CTRL_2024)),
    ("BPS (지배지분/보통주유통)", fmt_won(BPS)),
    ("현재 PBR", f"{PBR_2024:.2f}배"),
    ("PBR 0.5배 주가", fmt_won(int(BPS * 0.5))),
    ("PBR 1.0배 주가", fmt_won(int(BPS * 1.0))),
    ("PBR 1.5배 주가", fmt_won(int(BPS * 1.5))),
    ("PBR 2.0배 주가", fmt_won(int(BPS * 2.0))),
]
for label, val in pbr_info:
    ws3.cell(row=row, column=1, value=label).font = d_bold
    ws3.cell(row=row, column=1).fill = lblue_fill
    ws3.cell(row=row, column=1).alignment = left
    ws3.cell(row=row, column=1).border = thin_border
    ws3.cell(row=row, column=2, value=val).font = d_blue
    ws3.cell(row=row, column=2).fill = gold_fill
    ws3.cell(row=row, column=2).alignment = center
    ws3.cell(row=row, column=2).border = thin_border
    row += 1

# C. RIM (Residual Income Model)
row += 1
row = section_title(ws3, row, "C. 잔여이익모델(RIM) 적정주가", 6)
ws3.cell(row=row, column=1, value="산식: 적정가 = BPS + BPS × (ROE - ke) / (ke - g)").font = note_font
row += 1
ws3.cell(row=row, column=1, value="ke(자기자본비용) = 무위험이자율 3.5% + β(1.2) × ERP(5.4%) = 10.0%").font = note_font
row += 1
ws3.cell(row=row, column=1, value="항공업 β는 1.0~1.5 범위. 고레버리지 산업 특성 반영.").font = note_font
row += 1

write_header(ws3, row, ["시나리오", "지속ROE", "ke", "성장률(g)", "적정주가", "현재가 대비"])
row += 1

rim_scenarios = [
    ("보수적 (ROE=ke)", 0.10, 0.10, 0.02, "ROE가 자본비용과 동일"),
    ("기본 (ROE 13%)", 0.13, 0.10, 0.02, "2024년 ROE 수준"),
    ("적극적 (합병 시너지)", 0.15, 0.10, 0.02, "아시아나 합병 효과 반영"),
    ("낙관적 (ROE 확대)", 0.18, 0.10, 0.03, "글로벌 1위 항공사 수준"),
]

for label, roe, ke, g, note in rim_scenarios:
    if roe == ke:
        fair = BPS  # ROE = ke이면 초과이익 없음
    else:
        fair = BPS * (1 + (roe - ke) / (ke - g))
    upside = (fair - PRICE) / PRICE
    upside_str = f"{upside*100:+.1f}%"
    up_font = d_green if upside > 0 else d_red
    up_fill = green_fill if upside > 0 else red_fill

    write_row(ws3, row,
              [label, f"{roe*100:.0f}%", f"{ke*100:.0f}%", f"{g*100:.0f}%", fmt_won(int(fair)), upside_str],
              fonts=[d_bold, d_font, d_font, d_font, d_blue, up_font],
              fills=[lblue_fill, white_fill, white_fill, white_fill, gold_fill, up_fill],
              aligns=[left, center, center, center, right, center])
    row += 1

# RIM with EPS-based calculation
row += 1
rim_eps = BPS + (EPS_2024 - BPS * Ke) / (Ke - GROWTH)
ws3.cell(row=row, column=1, value=f"→ RIM (EPS기반): BPS({fmt_won(BPS)}) + 초과이익({fmt_won(EPS_2024)}-{fmt_won(int(BPS*Ke))})/(10%-2%) = {fmt_won(int(rim_eps))}").font = d_bold
row += 1
ws3.cell(row=row, column=1, value=f"→ 기본 시나리오(ROE 13%) 적정가 약 {fmt_won(int(BPS*(1+(0.13-0.10)/(0.10-0.02))))}. 현재가 {fmt_won(PRICE)}은 보수적~기본 사이 반영.").font = d_bold

print("  [3/5] PBR/ROE 완료")


# ============================================================
# SHEET 4: EV/EBITDA & FCF
# ============================================================
ws4 = wb.create_sheet("EV_EBITDA_FCF")
ws4.sheet_properties.tabColor = "E67E22"
set_widths(ws4, [24, 18, 18, 18, 24])
setup_print(ws4)

row = 1
row = section_title(ws4, row, "EV/EBITDA & FCF 밸류에이션", 5)

# A. EV 산출 (항공업 특화: 리스부채 포함)
row = section_title(ws4, row, "A. Enterprise Value 산출 (리스부채 포함)", 5)
write_header(ws4, row, ["항목", "금액(억)", "비고", "", ""])
row += 1

ev_items = [
    ("시가총액 (보통주)", int(MARKET_CAP/억), f"주가 {PRICE:,}원 × {SHARES_OUTSTANDING:,}주"),
    ("(+) 단기차입금", int(ST_DEBT_2024/억), ""),
    ("(+) 장기차입금", int(LT_DEBT_2024/억), ""),
    ("(+) 유동성장기부채", int(CURRENT_LT_DEBT/억), ""),
    ("(+) 사채", int(BONDS_2024/억), ""),
    ("(+) 리스부채(비유동)", int(LEASE_NONCURRENT/억), "항공기 리스 (IFRS 16)"),
    ("(+) 유동성리스부채", int(LEASE_CURRENT/억), ""),
    ("(=) 총차입금", int(TOTAL_DEBT/억), "리스부채 포함"),
    ("(-) 현금성자산", int(CASH_2024/억), "현금및현금성자산"),
    ("(-) 단기금융상품", int(ST_FINANCIAL/억), ""),
    ("(=) 순차입금", int(NET_DEBT/억), "차입금 - 현금성"),
    ("(=) EV", int(EV/억), "시가총액 + 순차입금"),
]
for label, amt, note in ev_items:
    is_total = label.startswith("(=)")
    write_row(ws4, row, [label, f"{amt:,}", note, "", ""],
              fonts=[d_bold if is_total else d_font, d_blue if is_total else d_font, note_font, d_font, d_font],
              fills=[gold_fill if is_total else lblue_fill, gold_fill if is_total else white_fill, white_fill, white_fill, white_fill],
              aligns=[left, right, left, left, left])
    row += 1

# B. EBITDA
row += 1
row = section_title(ws4, row, "B. EBITDA 산출", 5)
write_header(ws4, row, ["항목", "2024(억)", "비고", "", ""])
row += 1

ebitda_items = [
    ("영업이익", int(OP_2024/억), "CIS 영업이익"),
    ("(+) 감가상각비", int(DA_2024/억), "유형자산 (항공기 등)"),
    ("(+) 무형자산상각비", int(IA_2024/억), "노선권, 소프트웨어 등"),
    ("(=) EBITDA", int(EBITDA_2024/억), ""),
    ("EBITDA 마진", "", f"{EBITDA_2024/REV_2024*100:.1f}%"),
]
for label, amt, note in ebitda_items:
    is_total = label.startswith("(=)") or label == "EBITDA 마진"
    val = f"{amt:,}" if isinstance(amt, int) else note
    write_row(ws4, row, [label, val if isinstance(amt, int) else "", note if isinstance(amt, int) else val, "", ""],
              fonts=[d_bold if is_total else d_font, d_blue if is_total else d_font, note_font, d_font, d_font],
              fills=[gold_fill if is_total else lblue_fill, gold_fill if is_total else white_fill, white_fill, white_fill, white_fill],
              aligns=[left, right, left, left, left])
    row += 1

# C. EV/EBITDA 밸류에이션
row += 1
row = section_title(ws4, row, "C. EV/EBITDA 밸류에이션", 5)
write_header(ws4, row, ["목표 배수", "적정 EV(억)", "적정 시총(억)", "적정 주가", "현재가 대비"])
row += 1

for mult in [4.0, 5.0, 5.5, 6.0, 7.0, 8.0]:
    fair_ev = int(EBITDA_2024/억) * mult
    fair_mcap = fair_ev - int(NET_DEBT/억)
    fair_price = int(fair_mcap * 억 / SHARES_OUTSTANDING) if fair_mcap > 0 else 0
    upside = (fair_price - PRICE) / PRICE if fair_price > 0 else -1

    is_current = abs(mult - EV_EBITDA_2024) < 0.5
    up_font = d_green if upside > 0 else d_red
    row_fill = gold_fill if is_current else white_fill

    write_row(ws4, row,
              [f"EV/EBITDA {mult:.1f}배" + (" (현재)" if is_current else ""),
               f"{int(fair_ev):,}", f"{int(fair_mcap):,}", fmt_won(fair_price), f"{upside*100:+.1f}%"],
              fonts=[d_bold if is_current else d_font, d_font, d_font, d_blue, up_font],
              fills=[gold_fill if is_current else lblue_fill, row_fill, row_fill, row_fill,
                     green_fill if upside > 0 else red_fill],
              aligns=[left, right, right, right, center])
    row += 1

# D. FCF 분석
row += 1
row = section_title(ws4, row, "D. FCF(잉여현금흐름) 분석", 5)
write_header(ws4, row, ["항목", "2024(억)", "비고", "", ""])
row += 1

fcf_items = [
    ("영업활동현금흐름", int(OPCF_2024/억), "4.6조원"),
    ("(-) 설비투자(CAPEX)", int(CAPEX_2024/억), "유형자산 취득 (항공기 등)"),
    ("(=) FCF", int(FCF_2024/억), "잉여현금흐름"),
    ("FCF/주", "", fmt_won(int(FCF_PER_SHARE))),
    ("FCF 수익률", "", f"{FCF_PER_SHARE/PRICE*100:.1f}%"),
    ("", "", ""),
    ("이자비용", int(INTEREST_2024/억), "CF 상 이자비용"),
    ("배당금 지급", int(DIV_PAID_2024/억), f"DPS {DPS_2024:,}원"),
    ("FCF - 이자 - 배당", int((FCF_2024-INTEREST_2024-DIV_PAID_2024)/억), "순잉여현금"),
]

for label, amt, note in fcf_items:
    if not label:
        row += 1
        continue
    is_total = "(=)" in label or "수익률" in label or "순잉여" in label
    val_str = f"{amt:,}" if isinstance(amt, int) and amt != 0 else note
    write_row(ws4, row, [label, val_str if isinstance(amt, int) and amt != 0 else "",
                          note if isinstance(amt, int) and amt != 0 else val_str, "", ""],
              fonts=[d_bold if is_total else d_font, d_blue if is_total else d_font, note_font, d_font, d_font],
              fills=[gold_fill if is_total else lblue_fill, gold_fill if is_total else white_fill, white_fill, white_fill, white_fill],
              aligns=[left, right, left, left, left])
    row += 1

row += 1
ws4.cell(row=row, column=1, value=f"→ FCF {fmt(FCF_2024)}으로 양호. 다만 항공업 특성상 대규모 항공기 투자(CAPEX {fmt(CAPEX_2024)}) 지속 필요.").font = d_bold
row += 1
ws4.cell(row=row, column=1, value=f"→ 이자비용 {fmt(INTEREST_2024)} 차감 후에도 연간 {fmt(FCF_2024-INTEREST_2024-DIV_PAID_2024)} 순잉여현금 창출.").font = d_bold

print("  [4/5] EV/EBITDA/FCF 완료")


# ============================================================
# SHEET 5: 시나리오별 목표주가 종합
# ============================================================
ws5 = wb.create_sheet("목표주가")
ws5.sheet_properties.tabColor = "2ECC71"
set_widths(ws5, [22, 18, 18, 18, 22])
setup_print(ws5)

row = 1
row = section_title(ws5, row, f"시나리오별 목표주가 종합 (현재가 {PRICE:,}원)", 5)

# A. 방법론별 적정가 범위
row = section_title(ws5, row, "A. 밸류에이션 방법론별 적정가 레인지", 5)
write_header(ws5, row, ["방법론", "보수적", "기본", "적극적", "산출 근거"])
row += 1

# PER 방식
per_cons = EPS_2024 * 5   # PER 5배 (약세)
per_base = EPS_2024 * 8   # PER 8배 (기본)
per_aggr = EPS_2024 * 10  # PER 10배 (적극)

# PBR 방식
pbr_cons = int(BPS * 0.5)
pbr_base = int(BPS * 1.0)
pbr_aggr = int(BPS * 1.5)

# EV/EBITDA 방식
def ev_ebitda_price(mult):
    fair_ev = int(EBITDA_2024/억) * mult
    fair_mcap = fair_ev - int(NET_DEBT/억)
    return int(fair_mcap * 억 / SHARES_OUTSTANDING) if fair_mcap > 0 else 0

ev_cons = ev_ebitda_price(4.5)
ev_base = ev_ebitda_price(6.0)
ev_aggr = ev_ebitda_price(8.0)

# RIM 방식
rim_cons = BPS  # ROE = ke
rim_base = int(BPS * (1 + (0.13 - 0.10) / (0.10 - 0.02)))
rim_aggr = int(BPS * (1 + (0.15 - 0.10) / (0.10 - 0.02)))

# FCF 방식
fcf_cons = int(FCF_2024 * 6 / SHARES_OUTSTANDING)
fcf_base = int(FCF_2024 * 10 / SHARES_OUTSTANDING)
fcf_aggr = int(FCF_2024 * 14 / SHARES_OUTSTANDING)

methods = [
    ("PER 방식", fmt_won(per_cons), fmt_won(per_base), fmt_won(per_aggr),
     "EPS×목표PER (5/8/10배)"),
    ("PBR 방식", fmt_won(pbr_cons), fmt_won(pbr_base), fmt_won(pbr_aggr),
     "BPS×목표PBR (0.5/1.0/1.5배)"),
    ("EV/EBITDA 방식", fmt_won(ev_cons), fmt_won(ev_base), fmt_won(ev_aggr),
     "EBITDA×목표배수-순차입금 (4.5/6/8배)"),
    ("RIM 방식", fmt_won(rim_cons), fmt_won(rim_base), fmt_won(rim_aggr),
     "BPS×(1+(ROE-ke)/(ke-g))"),
    ("FCF 기반", fmt_won(fcf_cons), fmt_won(fcf_base), fmt_won(fcf_aggr),
     "FCF×목표배수 (6/10/14배)"),
]

for label, cons, base, aggr, note in methods:
    write_row(ws5, row, [label, cons, base, aggr, note],
              fonts=[d_bold, d_font, d_blue, d_green, note_font],
              fills=[lblue_fill, red_fill, gold_fill, green_fill, white_fill],
              aligns=[left, right, right, right, left])
    row += 1

# B. 종합 판단
row += 1
row = section_title(ws5, row, "B. 종합 시나리오", 5)
write_header(ws5, row, ["시나리오", "목표주가", "현재가 대비", "전제조건", "확률(주관)"],
             fills=[PatternFill("solid", fgColor="C0392B"),
                    PatternFill("solid", fgColor="C0392B"),
                    PatternFill("solid", fgColor="C0392B"),
                    PatternFill("solid", fgColor="C0392B"),
                    PatternFill("solid", fgColor="C0392B")])
row += 1

# Bull case: PER 10배 → 35,000원
bull_target = 35000
bull_upside = (bull_target - PRICE) / PRICE
write_row(ws5, row, ["강세 (Bull)", fmt_won(bull_target), f"{bull_upside*100:+.1f}%",
                      "아시아나 시너지 본격화 + 화물호황 + PER 리레이팅 10배", "20%"],
          fonts=[d_bold, Font(name="맑은 고딕", size=12, bold=True, color="27AE60"), d_green, d_font, d_font],
          fills=[green_fill]*5, aligns=[center, right, center, left, center])
row += 1

# Base case: PER 8배 → 28,000원
base_target = 28000
base_upside = (base_target - PRICE) / PRICE
write_row(ws5, row, ["기본 (Base)", fmt_won(base_target), f"{base_upside*100:+.1f}%",
                      "여객 정상화 유지 + 합병 시너지 점진적 반영 + PER 8배", "50%"],
          fonts=[d_bold, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), d_blue, d_font, d_font],
          fills=[gold_fill]*5, aligns=[center, right, center, left, center])
row += 1

# Bear case: PER 4.5배 → 16,000원
bear_target = 16000
bear_upside = (bear_target - PRICE) / PRICE
write_row(ws5, row, ["약세 (Bear)", fmt_won(bear_target), f"{bear_upside*100:+.1f}%",
                      "경기 침체 + 유가 급등 + 합병 비용 부담 + PER 4.5배", "30%"],
          fonts=[d_bold, Font(name="맑은 고딕", size=12, bold=True, color="C0392B"), d_red, d_font, d_font],
          fills=[red_fill]*5, aligns=[center, right, center, left, center])
row += 1

# Expected value
exp_val = int(bull_target * 0.20 + base_target * 0.50 + bear_target * 0.30)
exp_upside = (exp_val - PRICE) / PRICE
row += 1
write_row(ws5, row, ["확률가중 기대값", fmt_won(exp_val), f"{exp_upside*100:+.1f}%",
                      "Bull×20% + Base×50% + Bear×30%", ""],
          fonts=[Font(name="맑은 고딕", size=12, bold=True, color=NAVY)] * 5,
          fills=[PatternFill("solid", fgColor="AED6F1")] * 5,
          aligns=[center, right, center, left, center])

# C. Key Metrics Summary Box
row += 2
row = section_title(ws5, row, "C. 핵심 체크포인트", 5)

checkpoints = [
    ("현재 PER 6.6배", "글로벌 항공업 평균 PER 8~12배 대비 30~45% 할인. 한국 디스카운트 + 합병 불확실성 반영"),
    ("PBR 0.83배", "자본총계 대비 저평가. 2020년 자본잠식 위기에서 빠르게 자본 확충 (2020년 6천억→2024년 10.5조)"),
    ("EV/EBITDA 5.5배", "리스부채 포함 시 EV가 크지만, EBITDA 3.9조로 업종 평균 수준. 항공기 투자 대비 수익성 양호"),
    ("ROE 13.3%", "자본비용(10%) 대비 초과수익 창출 중. 항공업치고 높은 수준"),
    ("아시아나 합병 효과", "2024년 아시아나 합병 완료. 노선 최적화, 중복비용 제거 등 시너지 본격화 기대"),
    ("화물사업 경쟁력", "글로벌 Top 3 화물 수송. B747-8F 등 대형 화물기 보유. 이커머스 성장 수혜"),
    ("리스크: 고레버리지", f"순차입금/EBITDA {NET_DEBT/EBITDA_2024:.1f}배. 항공업 특성이나 리스부채 12.6조 부담"),
    ("리스크: 유가 민감도", "제트유 가격 변동이 영업이익에 직접 영향. 유가 급등 시 마진 압박"),
    ("리스크: 환율", "달러 매출 비중 높으나 유류비·리스료도 달러. 원화 약세 시 이중 효과"),
    ("리스크: 경기 순환", "항공 수요는 경기에 민감. 글로벌 경기 침체 시 여객·화물 모두 타격"),
]

for label, desc in checkpoints:
    ws5.cell(row=row, column=1, value=label).font = d_bold
    ws5.cell(row=row, column=1).fill = lblue_fill
    ws5.cell(row=row, column=1).alignment = left
    ws5.cell(row=row, column=1).border = thin_border
    ws5.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws5.cell(row=row, column=2, value=desc).font = d_font
    ws5.cell(row=row, column=2).alignment = left
    ws5.cell(row=row, column=2).border = thin_border
    row += 1

print("  [5/5] 목표주가 완료")


# === SAVE ===
wb.save(OUT)
print(f"\n밸류에이션 보고서 생성 완료: {OUT}")
print(f"시트: {wb.sheetnames}")
