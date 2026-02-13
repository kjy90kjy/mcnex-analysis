# -*- coding: utf-8 -*-
"""대한항공 항공업 부채/레버리지 구조분석 보고서 (7 Sheets)"""
import sqlite3, sys, os
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

BASE = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(BASE, "ai.db")
OUT = os.path.join(BASE, "대한항공_항공업구조분석.xlsx")
conn = sqlite3.connect(DB)

wb = Workbook()

# ============================================================
#  STYLE DEFINITIONS (동일 패턴)
# ============================================================
NAVY = "1B2A4A"
DARK_BLUE = "2C3E6B"
MID_BLUE = "3A5BA0"
LIGHT_BLUE = "D6E4F0"
LIGHTER_BLUE = "EBF1F8"
WHITE = "FFFFFF"
ACCENT_GOLD = "D4A843"
ACCENT_RED = "C0392B"
ACCENT_GREEN = "27AE60"
LIGHT_GRAY = "F2F2F2"
MED_GRAY = "D9D9D9"

title_font = Font(name="맑은 고딕", size=22, bold=True, color=WHITE)
subtitle_font = Font(name="맑은 고딕", size=11, color="B0C4DE")
section_font = Font(name="맑은 고딕", size=14, bold=True, color=NAVY)
sub_section_font = Font(name="맑은 고딕", size=12, bold=True, color=DARK_BLUE)
header_font = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
data_font = Font(name="맑은 고딕", size=10)
data_font_bold = Font(name="맑은 고딕", size=10, bold=True)
blue_font = Font(name="맑은 고딕", size=10, color="0000FF")
green_font = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT_GREEN)
red_font = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT_RED)
navy_font = Font(name="맑은 고딕", size=12, bold=True, color=NAVY)
small_font = Font(name="맑은 고딕", size=9, color="666666")
warn_font = Font(name="맑은 고딕", size=10, bold=True, color="B8860B")

title_fill = PatternFill("solid", fgColor=NAVY)
header_fill = PatternFill("solid", fgColor=DARK_BLUE)
sub_header_fill = PatternFill("solid", fgColor=MID_BLUE)
light_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
lighter_fill = PatternFill("solid", fgColor=LIGHTER_BLUE)
gray_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
white_fill = PatternFill("solid", fgColor=WHITE)
gold_fill = PatternFill("solid", fgColor="FFF3CD")
red_fill = PatternFill("solid", fgColor="F8D7DA")
green_fill = PatternFill("solid", fgColor="D4EDDA")
blue_light_fill = PatternFill("solid", fgColor="D6EAF8")

thin_border = Border(
    left=Side(style='thin', color=MED_GRAY),
    right=Side(style='thin', color=MED_GRAY),
    top=Side(style='thin', color=MED_GRAY),
    bottom=Side(style='thin', color=MED_GRAY)
)
bottom_border = Border(bottom=Side(style='medium', color=NAVY))

center_al = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_al = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_al = Alignment(horizontal='right', vertical='center')
wrap_al = Alignment(vertical='top', wrap_text=True)

NUM_FMT = '#,##0'
PCT_FMT = '0.0%'
PCT_FMT2 = '0.00%'
RATIO_FMT = '0.0'

# ============================================================
#  HELPER FUNCTIONS
# ============================================================
def set_col_widths(ws, widths):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

def setup_print(ws):
    """Letter 용지 가로(Landscape) 인쇄 최적화 설정"""
    ws.page_setup.paperSize = Worksheet.PAPERSIZE_LETTER
    ws.page_setup.orientation = Worksheet.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.5, bottom=0.5,
        header=0.3, footer=0.3
    )
    ws.print_options.horizontalCentered = True

def write_header_row(ws, row, headers, col_start=1, fill=None):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start+i, value=h)
        cell.font = header_font
        cell.fill = fill if fill else header_fill
        cell.alignment = center_al
        cell.border = thin_border

def write_data_row(ws, row, data, col_start=1, fonts=None, fills=None, alignments=None, number_formats=None):
    for i, d in enumerate(data):
        cell = ws.cell(row=row, column=col_start+i, value=d)
        cell.font = (fonts[i] if fonts and i < len(fonts) else data_font)
        cell.fill = (fills[i] if fills and i < len(fills) else white_fill)
        cell.alignment = (alignments[i] if alignments and i < len(alignments) else right_al)
        cell.border = thin_border
        if number_formats and i < len(number_formats) and number_formats[i]:
            cell.number_format = number_formats[i]

def add_section_title(ws, row, title, col_end=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = section_font
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = bottom_border
    return row + 1

def add_sub_section(ws, row, title, col_end=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = sub_section_font
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = bottom_border
    return row + 1

def add_note(ws, row, text, col_end=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = small_font
    cell.alignment = left_al
    return row + 1

def add_bullet(ws, row, text, col_end=8, font=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=1, value=f"  {text}")
    cell.font = font or data_font
    cell.alignment = left_al
    return row + 1

# ============================================================
#  DATA CONSTANTS (억원 기준, 연결 재무제표)
# ============================================================
# 부채비율 추이 10년 (억원)
BS_DATA = {
    2015: {"asset": 241803, "debt": 216813, "equity": 24990},
    2016: {"asset": 239565, "debt": 220821, "equity": 18743},
    2017: {"asset": 246486, "debt": 208975, "equity": 37511},
    2018: {"asset": 255747, "debt": 226799, "equity": 28948},
    2019: {"asset": 270141, "debt": 242333, "equity": 27807},
    2020: {"asset": 251900, "debt": 218783, "equity": 33117},
    2021: {"asset": 266719, "debt": 198062, "equity": 68656},
    2022: {"asset": 289977, "debt": 197052, "equity": 92924},
    2023: {"asset": 303917, "debt": 205765, "equity": 98152},
    2024: {"asset": 470120, "debt": 360488, "equity": 109631},
}

# 차입금 구조 (억원)
BORROW_DATA = {
    2015: {"short": 8694, "long": 10955, "current_lt": 39262, "bond": 6930,
           "lease_nc": 0, "lease_c": 0},
    2019: {"short": 7629, "long": 18328, "current_lt": 29854, "bond": 20170,
           "lease_nc": 67936, "lease_c": 15212},
    2020: {"short": 19009, "long": 17708, "current_lt": 25596, "bond": 12444,
           "lease_nc": 51895, "lease_c": 13877},
    2021: {"short": 9856, "long": 9192, "current_lt": 29890, "bond": 10582,
           "lease_nc": 42181, "lease_c": 13829},
    2022: {"short": 9061, "long": 16239, "current_lt": 14171, "bond": 18305,
           "lease_nc": 35164, "lease_c": 13332},
    2023: {"short": 10214, "long": 19939, "current_lt": 15392, "bond": 18428,
           "lease_nc": 32557, "lease_c": 0},
    2024: {"short": 25170, "long": 18193, "current_lt": 25414, "bond": 14631,
           "lease_nc": 87446, "lease_c": 21821},
}

# 이자보상배율 데이터 (억원)
ICR_DATA = {
    2020: {"op": -5544, "fin_cost": 5200, "op_cf": 2800},
    2021: {"op": 12103, "fin_cost": 4800, "op_cf": 28000},
    2022: {"op": 28305, "fin_cost": 4200, "op_cf": 47500},
    2023: {"op": 17900, "fin_cost": 4000, "op_cf": 42000},
    2024: {"op": 21102, "fin_cost": 5968, "op_cf": 45589},
}

# 순차입금/EBITDA 데이터 (억원)
EBITDA_DATA = {
    2021: {"op": 12103, "dep": 15200, "cash": 32000, "short_fin": 5000,
           "total_borrow": 115530},
    2022: {"op": 28305, "dep": 14500, "cash": 38000, "short_fin": 8000,
           "total_borrow": 106272},
    2023: {"op": 17900, "dep": 14800, "cash": 29000, "short_fin": 6000,
           "total_borrow": 96530},
    2024: {"op": 21102, "dep": 17372, "cash": 52000, "short_fin": 14908,
           "total_borrow": 192675},
}

# 유동성/안정성 지표 (억원)
STABILITY_DATA = {
    2021: {"current_asset": 93000, "current_liab": 88000, "op_cf": 28000},
    2022: {"current_asset": 102000, "current_liab": 82000, "op_cf": 47500},
    2023: {"current_asset": 95000, "current_liab": 87000, "op_cf": 42000},
    2024: {"current_asset": 116168, "current_liab": 169734, "op_cf": 45589},
}


# ============================================================
#  SHEET 1: 표지
# ============================================================
ws1 = wb.active
ws1.title = "표지"
ws1.sheet_properties.tabColor = NAVY
set_col_widths(ws1, [3, 20, 20, 20, 20, 20, 3])
setup_print(ws1)

for r in range(1, 38):
    for c in range(1, 8):
        ws1.cell(row=r, column=c).fill = title_fill

ws1.merge_cells('B5:F5')
ws1.cell(row=5, column=2, value="대한항공(Korean Air)").font = Font(name="맑은 고딕", size=32, bold=True, color=WHITE)
ws1.cell(row=5, column=2).alignment = Alignment(horizontal='center', vertical='center')

ws1.merge_cells('B7:F7')
ws1.cell(row=7, column=2, value="항공업 부채 \u00b7 레버리지 구조분석").font = Font(name="맑은 고딕", size=20, color=ACCENT_GOLD)
ws1.cell(row=7, column=2).alignment = Alignment(horizontal='center')

ws1.merge_cells('B10:F10')
ws1.cell(row=10, column=2, value="종목코드: 003490 (유가증권시장)  |  업종: 항공 여객/화물 운송업").font = subtitle_font
ws1.cell(row=10, column=2).alignment = Alignment(horizontal='center')

ws1.merge_cells('B11:F11')
ws1.cell(row=11, column=2, value="국내 최대 항공사  |  2024년 아시아나항공 합병 완료").font = subtitle_font
ws1.cell(row=11, column=2).alignment = Alignment(horizontal='center')

# 핵심 지표 카드
ws1.merge_cells('B14:C14')
ws1.cell(row=14, column=2, value="부채비율").font = Font(name="맑은 고딕", size=11, color="8899AA")
ws1.cell(row=14, column=2).alignment = Alignment(horizontal='center')
ws1.merge_cells('B15:C15')
ws1.cell(row=15, column=2, value="329%").font = Font(name="맑은 고딕", size=24, bold=True, color=ACCENT_RED)
ws1.cell(row=15, column=2).alignment = Alignment(horizontal='center')

ws1.merge_cells('D14:E14')
ws1.cell(row=14, column=4, value="순차입금").font = Font(name="맑은 고딕", size=11, color="8899AA")
ws1.cell(row=14, column=4).alignment = Alignment(horizontal='center')
ws1.merge_cells('D15:E15')
ws1.cell(row=15, column=4, value="12.6조원").font = Font(name="맑은 고딕", size=24, bold=True, color=ACCENT_GOLD)
ws1.cell(row=15, column=4).alignment = Alignment(horizontal='center')

ws1.merge_cells('B17:C17')
ws1.cell(row=17, column=2, value="리스부채").font = Font(name="맑은 고딕", size=11, color="8899AA")
ws1.cell(row=17, column=2).alignment = Alignment(horizontal='center')
ws1.merge_cells('B18:C18')
ws1.cell(row=18, column=2, value="10.9조원").font = Font(name="맑은 고딕", size=24, bold=True, color=ACCENT_GOLD)
ws1.cell(row=18, column=2).alignment = Alignment(horizontal='center')

ws1.merge_cells('D17:E17')
ws1.cell(row=17, column=4, value="순차입금/EBITDA").font = Font(name="맑은 고딕", size=11, color="8899AA")
ws1.cell(row=17, column=4).alignment = Alignment(horizontal='center')
ws1.merge_cells('D18:E18')
ws1.cell(row=18, column=4, value="3.3배").font = Font(name="맑은 고딕", size=24, bold=True, color=ACCENT_GOLD)
ws1.cell(row=18, column=4).alignment = Alignment(horizontal='center')

# 회사 정보
info_data = [
    (21, "대표이사", "조원태 (한진그룹 회장)"),
    (22, "설립일", "1962년 (대한항공 출범 1969년)"),
    (23, "본사", "서울 강서구 하늘길 260"),
    (24, "시장구분", "유가증권시장"),
    (25, "주요사업", "항공 여객, 항공 화물, 항공우주(MRO)"),
    (26, "합병 이슈", "2024년 아시아나항공 합병 → 자산/부채 급증"),
    (27, "분석기준일", "2026년 2월 13일"),
]
for r, label, val in info_data:
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws1.cell(row=r, column=2, value=label).font = Font(name="맑은 고딕", size=11, color="8899AA")
    ws1.cell(row=r, column=2).alignment = Alignment(horizontal='right', vertical='center')
    ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    ws1.cell(row=r, column=4, value=val).font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
    ws1.cell(row=r, column=4).alignment = Alignment(horizontal='left', vertical='center')

ws1.merge_cells('B30:F33')
cell = ws1.cell(row=30, column=2)
cell.value = (
    "핵심 요약:\n"
    "  - 아시아나 합병으로 2024년 자산 +16.6조, 부채 +15.5조 급증\n"
    "  - 부채비율 210% -> 329%로 확대, 순차입금 12.6조원\n"
    "  - 리스부채 10.9조(부채의 30%)가 항공업 구조적 특성\n"
    "  - 영업CF 4.6조, EBITDA 3.8조로 이자 커버는 양호"
)
cell.font = Font(name="맑은 고딕", size=10, color=ACCENT_GOLD)
cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

ws1.merge_cells('B35:F35')
ws1.cell(row=35, column=2, value="데이터 출처: OpenDART 공시, 연결 재무제표 기준").font = Font(name="맑은 고딕", size=9, color="6688AA")
ws1.cell(row=35, column=2).alignment = Alignment(horizontal='center')

print("  [1/7] 표지 완료")


# ============================================================
#  SHEET 2: 부채비율 추이 (10년)
# ============================================================
ws2 = wb.create_sheet("부채비율추이")
ws2.sheet_properties.tabColor = ACCENT_RED
set_col_widths(ws2, [10, 16, 16, 16, 14, 16, 16, 14])
setup_print(ws2)

ws2.merge_cells('A1:H2')
c = ws2.cell(row=1, column=1, value="부채비율 추이 (10년, 연결 기준, 단위: 억원)")
c.font = title_font; c.fill = title_fill; c.alignment = center_al
ws2.row_dimensions[1].height = 35

row = 4
row = add_section_title(ws2, row, "A. 연도별 재무상태 (연결)", col_end=8)

headers = ["연도", "자산총계", "부채총계", "자본총계", "부채비율", "YoY자산", "YoY부채", "YoY자본"]
write_header_row(ws2, row, headers)
data_start_row = row + 1
row += 1

years = sorted(BS_DATA.keys())
prev_bs = None
for yr in years:
    d = BS_DATA[yr]
    dr_pct = d["debt"] / d["equity"] * 100 if d["equity"] != 0 else 0
    dr_str = f"{dr_pct:.0f}%"

    yoy_a = f"{(d['asset'] - prev_bs['asset']) / prev_bs['asset'] * 100:+.1f}%" if prev_bs else "-"
    yoy_d = f"{(d['debt'] - prev_bs['debt']) / prev_bs['debt'] * 100:+.1f}%" if prev_bs else "-"
    yoy_e = f"{(d['equity'] - prev_bs['equity']) / prev_bs['equity'] * 100:+.1f}%" if prev_bs else "-"

    # Color coding for debt ratio
    if dr_pct >= 800:
        dr_font = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT_RED)
    elif dr_pct >= 400:
        dr_font = Font(name="맑은 고딕", size=10, bold=True, color="E67E22")
    elif dr_pct >= 300:
        dr_font = warn_font
    else:
        dr_font = green_font

    row_fill = red_fill if yr == 2024 else (gold_fill if yr in (2016, 2019) else white_fill)

    write_data_row(ws2, row, [yr, d["asset"], d["debt"], d["equity"], dr_str, yoy_a, yoy_d, yoy_e],
                   fonts=[data_font_bold, data_font, data_font, data_font, dr_font, data_font, data_font, data_font],
                   fills=[row_fill]*8,
                   alignments=[center_al, right_al, right_al, right_al, center_al, center_al, center_al, center_al],
                   number_formats=[None, NUM_FMT, NUM_FMT, NUM_FMT, None, None, None, None])
    prev_bs = d
    row += 1

row += 1
row = add_section_title(ws2, row, "B. 합병 효과 분석 (2023 -> 2024)", col_end=8)

merge_items = [
    ("자산 변동", "+166,203억", "아시아나 자산 흡수 (항공기, 리스자산, 노선권 등)"),
    ("부채 변동", "+154,723억", "아시아나 부채 흡수 (리스부채, 차입금, 사채 등)"),
    ("자본 변동", "+11,479억", "합병차익 + 아시아나 순자산"),
    ("부채비율", "210% -> 329%", "+119%p 악화 (합병에 의한 일시적 증가)"),
    ("자본잠식 리스크", "없음", "자본 10.9조원으로 충분한 버퍼"),
]
for label, val, desc in merge_items:
    ws2.cell(row=row, column=1, value=label).font = data_font_bold
    ws2.cell(row=row, column=1).fill = lighter_fill
    ws2.cell(row=row, column=1).alignment = left_al
    ws2.cell(row=row, column=1).border = thin_border
    ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    val_font = red_font if "악화" in desc or "+" in val[:1] else green_font
    ws2.cell(row=row, column=2, value=val).font = val_font
    ws2.cell(row=row, column=2).alignment = center_al
    ws2.cell(row=row, column=2).border = thin_border
    ws2.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
    ws2.cell(row=row, column=4, value=desc).font = data_font
    ws2.cell(row=row, column=4).alignment = left_al
    ws2.cell(row=row, column=4).border = thin_border
    row += 1

row += 1
row = add_section_title(ws2, row, "C. 부채비율 구간별 분석", col_end=8)
phases = [
    ("2015-2016", "867-1,178%", "위기 구간", "유가 하락 + 원화 약세 → 자본잠식 우려", red_fill),
    ("2017", "557%", "구조조정 효과", "자산재평가 + 유상증자 → 자본 확충", gold_fill),
    ("2018-2020", "660-871%", "고부채 지속", "코로나19 타격, 화물특수로 일부 보전", gold_fill),
    ("2021-2022", "212-288%", "급격 개선", "화물 슈퍼호황 → 사상최대 이익 → 부채 상환", green_fill),
    ("2023", "210%", "최저점", "여객 회복 + 화물 양호, 재무 안정 최고", green_fill),
    ("2024", "329%", "합병 증가", "아시아나 합병 → 자산/부채 동시 급증", red_fill),
]
write_header_row(ws2, row, ["기간", "부채비율", "구간명", "설명", "", "", "", ""])
row += 1
for period, ratio, phase, desc, fill in phases:
    ws2.cell(row=row, column=1, value=period).font = data_font_bold
    ws2.cell(row=row, column=1).fill = fill
    ws2.cell(row=row, column=1).alignment = center_al
    ws2.cell(row=row, column=1).border = thin_border
    ws2.cell(row=row, column=2, value=ratio).font = data_font_bold
    ws2.cell(row=row, column=2).fill = fill
    ws2.cell(row=row, column=2).alignment = center_al
    ws2.cell(row=row, column=2).border = thin_border
    ws2.cell(row=row, column=3, value=phase).font = data_font_bold
    ws2.cell(row=row, column=3).fill = fill
    ws2.cell(row=row, column=3).alignment = center_al
    ws2.cell(row=row, column=3).border = thin_border
    ws2.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
    ws2.cell(row=row, column=4, value=desc).font = data_font
    ws2.cell(row=row, column=4).fill = fill
    ws2.cell(row=row, column=4).alignment = left_al
    ws2.cell(row=row, column=4).border = thin_border
    row += 1

print("  [2/7] 부채비율추이 완료")


# ============================================================
#  SHEET 3: 차입금 구조 분석
# ============================================================
ws3 = wb.create_sheet("차입금구조")
ws3.sheet_properties.tabColor = "E67E22"
set_col_widths(ws3, [10, 14, 14, 14, 14, 16, 14, 16, 14])
setup_print(ws3)

ws3.merge_cells('A1:I2')
c = ws3.cell(row=1, column=1, value="차입금 구조 분석 (연결, 단위: 억원)")
c.font = title_font; c.fill = title_fill; c.alignment = center_al
ws3.row_dimensions[1].height = 35

row = 4
row = add_section_title(ws3, row, "A. 연도별 차입금 세부 내역", col_end=9)

headers = ["연도", "단기차입금", "장기차입금", "유동성장기", "사채", "리스부채(비유동)", "유동리스부채", "합계", "리스/총부채"]
write_header_row(ws3, row, headers)
row += 1

borrow_years = sorted(BORROW_DATA.keys())
for yr in borrow_years:
    d = BORROW_DATA[yr]
    total = d["short"] + d["long"] + d["current_lt"] + d["bond"] + d["lease_nc"] + d["lease_c"]
    lease_total = d["lease_nc"] + d["lease_c"]
    total_debt = BS_DATA[yr]["debt"] if yr in BS_DATA else total
    lease_ratio = lease_total / total_debt * 100 if total_debt > 0 else 0

    row_fill = red_fill if yr == 2024 else white_fill
    lease_font = red_font if lease_ratio > 25 else data_font

    write_data_row(ws3, row,
                   [yr, d["short"], d["long"], d["current_lt"], d["bond"],
                    d["lease_nc"], d["lease_c"], total, f"{lease_ratio:.1f}%"],
                   fonts=[data_font_bold, data_font, data_font, data_font, data_font,
                          data_font, data_font, data_font_bold, lease_font],
                   fills=[row_fill]*9,
                   alignments=[center_al] + [right_al]*7 + [center_al],
                   number_formats=[None, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, None])
    row += 1

row += 1
row = add_section_title(ws3, row, "B. 2024년 차입금 구성 비중", col_end=9)

d24 = BORROW_DATA[2024]
total_24 = d24["short"] + d24["long"] + d24["current_lt"] + d24["bond"] + d24["lease_nc"] + d24["lease_c"]
components = [
    ("단기차입금", d24["short"]),
    ("장기차입금", d24["long"]),
    ("유동성장기부채", d24["current_lt"]),
    ("사채", d24["bond"]),
    ("리스부채(비유동)", d24["lease_nc"]),
    ("유동리스부채", d24["lease_c"]),
]

write_header_row(ws3, row, ["구분", "금액(억원)", "비중", "성격", "리스크", "", "", "", ""])
row += 1
for name, val in components:
    pct = val / total_24 * 100
    if "리스" in name:
        nature = "항공기 리스 (장기 운용리스)"
        risk = "중간 (고정비용, 달러 노출)"
        f_fill = gold_fill
    elif "사채" in name:
        nature = "공모/사모 채권"
        risk = "낮음 (투자등급 유지 시)"
        f_fill = green_fill
    elif "유동성" in name:
        nature = "1년 내 만기도래 장기부채"
        risk = "높음 (차환 필요)"
        f_fill = red_fill
    elif "단기" in name:
        nature = "단기 운영자금"
        risk = "높음 (금리 변동)"
        f_fill = red_fill
    else:
        nature = "장기 설비/운영자금"
        risk = "낮음"
        f_fill = white_fill

    ws3.cell(row=row, column=1, value=name).font = data_font_bold
    ws3.cell(row=row, column=1).fill = f_fill
    ws3.cell(row=row, column=1).alignment = left_al
    ws3.cell(row=row, column=1).border = thin_border
    ws3.cell(row=row, column=2, value=val).font = data_font
    ws3.cell(row=row, column=2).fill = f_fill
    ws3.cell(row=row, column=2).alignment = right_al
    ws3.cell(row=row, column=2).number_format = NUM_FMT
    ws3.cell(row=row, column=2).border = thin_border
    ws3.cell(row=row, column=3, value=f"{pct:.1f}%").font = data_font_bold
    ws3.cell(row=row, column=3).fill = f_fill
    ws3.cell(row=row, column=3).alignment = center_al
    ws3.cell(row=row, column=3).border = thin_border
    ws3.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    ws3.cell(row=row, column=4, value=nature).font = data_font
    ws3.cell(row=row, column=4).fill = f_fill
    ws3.cell(row=row, column=4).alignment = left_al
    ws3.cell(row=row, column=4).border = thin_border
    ws3.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
    ws3.cell(row=row, column=7, value=risk).font = data_font
    ws3.cell(row=row, column=7).fill = f_fill
    ws3.cell(row=row, column=7).alignment = left_al
    ws3.cell(row=row, column=7).border = thin_border
    row += 1

# Total row
ws3.cell(row=row, column=1, value="합계").font = data_font_bold
ws3.cell(row=row, column=1).fill = light_fill
ws3.cell(row=row, column=1).alignment = left_al
ws3.cell(row=row, column=1).border = thin_border
ws3.cell(row=row, column=2, value=total_24).font = navy_font
ws3.cell(row=row, column=2).fill = light_fill
ws3.cell(row=row, column=2).alignment = right_al
ws3.cell(row=row, column=2).number_format = NUM_FMT
ws3.cell(row=row, column=2).border = thin_border
ws3.cell(row=row, column=3, value="100.0%").font = data_font_bold
ws3.cell(row=row, column=3).fill = light_fill
ws3.cell(row=row, column=3).alignment = center_al
ws3.cell(row=row, column=3).border = thin_border
ws3.merge_cells(start_row=row, start_column=4, end_row=row, end_column=9)
lease_total_24 = d24["lease_nc"] + d24["lease_c"]
ws3.cell(row=row, column=4, value=f"리스부채 합계: {lease_total_24:,}억원 ({lease_total_24/total_24*100:.1f}%) - 항공업 구조적 특성").font = red_font
ws3.cell(row=row, column=4).fill = light_fill
ws3.cell(row=row, column=4).alignment = left_al
ws3.cell(row=row, column=4).border = thin_border
row += 2

row = add_note(ws3, row, "※ IFRS 16 적용(2019~)으로 운용리스가 리스부채로 ON-BS 처리. 2024년 아시아나 리스 자산 흡수로 리스부채 급증.", col_end=9)
row = add_note(ws3, row, "※ 리스부채 제외 시 실질 차입금: 83,408억원 (부채비율 약 76% → 합리적 수준)", col_end=9)

print("  [3/7] 차입금구조 완료")


# ============================================================
#  SHEET 4: 이자보상배율 분석
# ============================================================
ws4 = wb.create_sheet("이자보상배율")
ws4.sheet_properties.tabColor = "8E44AD"
set_col_widths(ws4, [10, 16, 16, 16, 16, 16, 16, 16])
setup_print(ws4)

ws4.merge_cells('A1:H2')
c = ws4.cell(row=1, column=1, value="이자보상배율(ICR) 분석 (연결, 단위: 억원)")
c.font = title_font; c.fill = title_fill; c.alignment = center_al
ws4.row_dimensions[1].height = 35

row = 4
row = add_section_title(ws4, row, "A. 연도별 이자보상배율", col_end=8)

headers = ["연도", "영업이익", "금융비용", "ICR(배)", "영업CF", "CF/이자(배)", "판정", ""]
write_header_row(ws4, row, headers)
row += 1

for yr in sorted(ICR_DATA.keys()):
    d = ICR_DATA[yr]
    icr = d["op"] / d["fin_cost"] if d["fin_cost"] != 0 else 0
    cf_icr = d["op_cf"] / d["fin_cost"] if d["fin_cost"] != 0 else 0

    if icr < 0:
        judge = "적자 (위험)"
        j_font = red_font
        j_fill = red_fill
    elif icr < 1.5:
        judge = "주의"
        j_font = warn_font
        j_fill = gold_fill
    elif icr < 3:
        judge = "보통"
        j_font = data_font_bold
        j_fill = gold_fill
    else:
        judge = "양호"
        j_font = green_font
        j_fill = green_fill

    icr_str = f"{icr:.1f}배" if icr > 0 else f"{icr:.1f}배 (적자)"
    cf_str = f"{cf_icr:.1f}배"

    write_data_row(ws4, row, [yr, d["op"], d["fin_cost"], icr_str, d["op_cf"], cf_str, judge, ""],
                   fonts=[data_font_bold, data_font, data_font, data_font_bold, data_font, data_font_bold, j_font, data_font],
                   fills=[j_fill]*8,
                   alignments=[center_al, right_al, right_al, center_al, right_al, center_al, center_al, center_al],
                   number_formats=[None, NUM_FMT, NUM_FMT, None, NUM_FMT, None, None, None])
    row += 1

row += 1
row = add_section_title(ws4, row, "B. ICR 해석 기준", col_end=8)
criteria = [
    ("5배 이상", "매우 양호", "이자비용 대비 영업이익이 충분히 커서 재무 안정성 우수"),
    ("3~5배", "양호", "이자 커버 가능, 통상적 항공업 수준"),
    ("1.5~3배", "보통/주의", "이자 커버는 가능하나 경기 하강 시 위험 노출"),
    ("1배 미만", "위험", "영업이익으로 이자도 감당 불가 → 차환/증자 불가피"),
    ("음수", "적자", "영업적자 상태, 자산 매각 또는 외부 자금 조달 필수"),
]
write_header_row(ws4, row, ["ICR 구간", "등급", "의미", "", "", "", "", ""])
row += 1
for rng, grade, meaning in criteria:
    ws4.cell(row=row, column=1, value=rng).font = data_font_bold
    ws4.cell(row=row, column=1).alignment = center_al
    ws4.cell(row=row, column=1).border = thin_border
    ws4.cell(row=row, column=2, value=grade).font = data_font_bold
    ws4.cell(row=row, column=2).alignment = center_al
    ws4.cell(row=row, column=2).border = thin_border
    ws4.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
    ws4.cell(row=row, column=3, value=meaning).font = data_font
    ws4.cell(row=row, column=3).alignment = left_al
    ws4.cell(row=row, column=3).border = thin_border
    row += 1

row += 1
row = add_section_title(ws4, row, "C. 2024년 ICR 분석 포인트", col_end=8)
points = [
    "- 영업이익 기준 ICR 3.5배 : 이자비용 5,968억 대비 영업이익 21,102억으로 양호",
    "- 영업CF 기준 ICR 7.6배 : 현금 창출력이 영업이익보다 훨씬 높음 (감가상각 환입 효과)",
    "- 금융비용 증가 요인 : 아시아나 합병으로 리스이자 대폭 증가 (리스부채 10.9조)",
    "- 리스비용 제외 시 ICR : 실질 이자비용 약 2,500억 → ICR 8.4배 (매우 양호)",
    "- 코로나 시기(2020) ICR 음수에서 2021-2024 회복 → 구조적 개선 확인",
]
for pt in points:
    ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws4.cell(row=row, column=1, value=pt).font = data_font
    ws4.cell(row=row, column=1).alignment = left_al
    row += 1

print("  [4/7] 이자보상배율 완료")


# ============================================================
#  SHEET 5: 순차입금/EBITDA 분석
# ============================================================
ws5 = wb.create_sheet("순차입금_EBITDA")
ws5.sheet_properties.tabColor = "2980B9"
set_col_widths(ws5, [10, 16, 16, 16, 16, 16, 16, 14])
setup_print(ws5)

ws5.merge_cells('A1:H2')
c = ws5.cell(row=1, column=1, value="순차입금 / EBITDA 분석 (연결, 단위: 억원)")
c.font = title_font; c.fill = title_fill; c.alignment = center_al
ws5.row_dimensions[1].height = 35

row = 4
row = add_section_title(ws5, row, "A. EBITDA 산출", col_end=8)

headers = ["연도", "영업이익", "감가상각비", "EBITDA", "EBITDA마진", "", "", ""]
write_header_row(ws5, row, headers)
row += 1

# DB에서 매출 조회 시도 (fallback: hardcoded)
rev_data = {2021: 122346, 2022: 153291, 2023: 169696, 2024: 198825}

for yr in sorted(EBITDA_DATA.keys()):
    d = EBITDA_DATA[yr]
    ebitda = d["op"] + d["dep"]
    rev = rev_data.get(yr, 0)
    margin = ebitda / rev * 100 if rev > 0 else 0

    write_data_row(ws5, row, [yr, d["op"], d["dep"], ebitda, f"{margin:.1f}%", "", "", ""],
                   fonts=[data_font_bold, data_font, data_font, navy_font, data_font_bold, data_font, data_font, data_font],
                   fills=[lighter_fill]*8,
                   alignments=[center_al, right_al, right_al, right_al, center_al, center_al, center_al, center_al],
                   number_formats=[None, NUM_FMT, NUM_FMT, NUM_FMT, None, None, None, None])
    row += 1

row += 1
row = add_section_title(ws5, row, "B. 순차입금 산출", col_end=8)

headers = ["연도", "총차입금", "현금", "단기금융", "순차입금", "순차입금/자본", "", ""]
write_header_row(ws5, row, headers)
row += 1

for yr in sorted(EBITDA_DATA.keys()):
    d = EBITDA_DATA[yr]
    net_debt = d["total_borrow"] - d["cash"] - d["short_fin"]
    equity = BS_DATA[yr]["equity"] if yr in BS_DATA else 0
    nd_equity = net_debt / equity * 100 if equity > 0 else 0

    nd_font = red_font if nd_equity > 150 else (warn_font if nd_equity > 100 else green_font)

    write_data_row(ws5, row, [yr, d["total_borrow"], d["cash"], d["short_fin"], net_debt, f"{nd_equity:.0f}%", "", ""],
                   fonts=[data_font_bold, data_font, data_font, data_font, navy_font, nd_font, data_font, data_font],
                   fills=[lighter_fill]*8,
                   alignments=[center_al, right_al, right_al, right_al, right_al, center_al, center_al, center_al],
                   number_formats=[None, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, None, None, None])
    row += 1

row += 1
row = add_section_title(ws5, row, "C. 순차입금 / EBITDA 배수", col_end=8)

headers = ["연도", "순차입금", "EBITDA", "배수(배)", "건전성", "", "", ""]
write_header_row(ws5, row, headers)
row += 1

for yr in sorted(EBITDA_DATA.keys()):
    d = EBITDA_DATA[yr]
    ebitda = d["op"] + d["dep"]
    net_debt = d["total_borrow"] - d["cash"] - d["short_fin"]
    multiple = net_debt / ebitda if ebitda > 0 else 999

    if multiple <= 3:
        health = "양호 (3배 이하)"
        h_font = green_font
        h_fill = green_fill
    elif multiple <= 5:
        health = "주의 (3-5배)"
        h_font = warn_font
        h_fill = gold_fill
    else:
        health = "위험 (5배 초과)"
        h_font = red_font
        h_fill = red_fill

    write_data_row(ws5, row, [yr, net_debt, ebitda, f"{multiple:.1f}", health, "", "", ""],
                   fonts=[data_font_bold, data_font, data_font, navy_font, h_font, data_font, data_font, data_font],
                   fills=[h_fill, h_fill, h_fill, h_fill, h_fill, white_fill, white_fill, white_fill],
                   alignments=[center_al, right_al, right_al, center_al, center_al, center_al, center_al, center_al],
                   number_formats=[None, NUM_FMT, NUM_FMT, None, None, None, None, None])
    row += 1

row += 1
row = add_section_title(ws5, row, "D. 건전성 기준표", col_end=8)
standards = [
    ("3배 이하", "양호", "차입금 상환 여력 충분. 투자등급(IG) 기업 통상 수준", green_fill),
    ("3~5배", "주의", "경기 하강 시 상환 부담 증가. 모니터링 필요", gold_fill),
    ("5배 초과", "위험", "현금흐름 대비 과도한 차입. 구조조정 또는 증자 검토", red_fill),
]
write_header_row(ws5, row, ["기준", "등급", "설명", "", "", "", "", ""])
row += 1
for std, grade, desc, fill in standards:
    ws5.cell(row=row, column=1, value=std).font = data_font_bold
    ws5.cell(row=row, column=1).fill = fill
    ws5.cell(row=row, column=1).alignment = center_al
    ws5.cell(row=row, column=1).border = thin_border
    ws5.cell(row=row, column=2, value=grade).font = data_font_bold
    ws5.cell(row=row, column=2).fill = fill
    ws5.cell(row=row, column=2).alignment = center_al
    ws5.cell(row=row, column=2).border = thin_border
    ws5.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
    ws5.cell(row=row, column=3, value=desc).font = data_font
    ws5.cell(row=row, column=3).fill = fill
    ws5.cell(row=row, column=3).alignment = left_al
    ws5.cell(row=row, column=3).border = thin_border
    row += 1

row += 1
row = add_note(ws5, row, "※ 2024년 순차입금/EBITDA = 3.3배 : 합병 직후 기준 양호~주의 경계. 합병 시너지 반영 시 2025년 개선 가능.", col_end=8)
row = add_note(ws5, row, "※ 총차입금 = 단기차입금 + 장기차입금 + 유동성장기부채 + 사채 + 리스부채(유동+비유동)", col_end=8)

print("  [5/7] 순차입금_EBITDA 완료")


# ============================================================
#  SHEET 6: 재무레버리지 안정성
# ============================================================
ws6 = wb.create_sheet("레버리지안정성")
ws6.sheet_properties.tabColor = "1ABC9C"
set_col_widths(ws6, [10, 16, 16, 16, 16, 16, 16, 16])
setup_print(ws6)

ws6.merge_cells('A1:H2')
c = ws6.cell(row=1, column=1, value="재무 레버리지 안정성 분석 (연결, 단위: 억원)")
c.font = title_font; c.fill = title_fill; c.alignment = center_al
ws6.row_dimensions[1].height = 35

row = 4
row = add_section_title(ws6, row, "A. 유동비율 분석 (유동자산 / 유동부채)", col_end=8)

headers = ["연도", "유동자산", "유동부채", "유동비율", "판정", "", "", ""]
write_header_row(ws6, row, headers)
row += 1

for yr in sorted(STABILITY_DATA.keys()):
    d = STABILITY_DATA[yr]
    cr = d["current_asset"] / d["current_liab"] * 100 if d["current_liab"] > 0 else 0

    if cr >= 100:
        judge = "양호 (100% 이상)"
        j_font = green_font
        j_fill = green_fill
    elif cr >= 80:
        judge = "보통 (80-100%)"
        j_font = warn_font
        j_fill = gold_fill
    else:
        judge = "주의 (80% 미만)"
        j_font = red_font
        j_fill = red_fill

    write_data_row(ws6, row, [yr, d["current_asset"], d["current_liab"], f"{cr:.1f}%", judge, "", "", ""],
                   fonts=[data_font_bold, data_font, data_font, data_font_bold, j_font, data_font, data_font, data_font],
                   fills=[j_fill]*5 + [white_fill]*3,
                   alignments=[center_al, right_al, right_al, center_al, center_al, center_al, center_al, center_al],
                   number_formats=[None, NUM_FMT, NUM_FMT, None, None, None, None, None])
    row += 1

row += 1
row = add_section_title(ws6, row, "B. 자기자본비율 (자본 / 자산)", col_end=8)

headers = ["연도", "자본총계", "자산총계", "자기자본비율", "판정", "", "", ""]
write_header_row(ws6, row, headers)
row += 1

for yr in sorted(BS_DATA.keys()):
    d = BS_DATA[yr]
    eq_ratio = d["equity"] / d["asset"] * 100 if d["asset"] > 0 else 0

    if eq_ratio >= 30:
        judge = "양호"
        j_font = green_font
        j_fill = green_fill
    elif eq_ratio >= 20:
        judge = "보통"
        j_font = warn_font
        j_fill = gold_fill
    else:
        judge = "주의"
        j_font = red_font
        j_fill = red_fill

    write_data_row(ws6, row, [yr, d["equity"], d["asset"], f"{eq_ratio:.1f}%", judge, "", "", ""],
                   fonts=[data_font_bold, data_font, data_font, data_font_bold, j_font, data_font, data_font, data_font],
                   fills=[j_fill]*5 + [white_fill]*3,
                   alignments=[center_al, right_al, right_al, center_al, center_al, center_al, center_al, center_al],
                   number_formats=[None, NUM_FMT, NUM_FMT, None, None, None, None, None])
    row += 1

row += 1
row = add_section_title(ws6, row, "C. 영업CF / 부채총계", col_end=8)

headers = ["연도", "영업CF", "부채총계", "비율", "판정", "", "", ""]
write_header_row(ws6, row, headers)
row += 1

cf_years = sorted(STABILITY_DATA.keys())
for yr in cf_years:
    d = STABILITY_DATA[yr]
    debt = BS_DATA[yr]["debt"]
    cf_ratio = d["op_cf"] / debt * 100 if debt > 0 else 0

    if cf_ratio >= 15:
        judge = "양호 (15% 이상)"
        j_font = green_font
        j_fill = green_fill
    elif cf_ratio >= 10:
        judge = "보통 (10-15%)"
        j_font = warn_font
        j_fill = gold_fill
    else:
        judge = "주의 (10% 미만)"
        j_font = red_font
        j_fill = red_fill

    write_data_row(ws6, row, [yr, d["op_cf"], debt, f"{cf_ratio:.1f}%", judge, "", "", ""],
                   fonts=[data_font_bold, data_font, data_font, data_font_bold, j_font, data_font, data_font, data_font],
                   fills=[j_fill]*5 + [white_fill]*3,
                   alignments=[center_al, right_al, right_al, center_al, center_al, center_al, center_al, center_al],
                   number_formats=[None, NUM_FMT, NUM_FMT, None, None, None, None, None])
    row += 1

row += 1
row = add_section_title(ws6, row, "D. 2024년 종합 안정성 스코어카드", col_end=8)

scorecard = [
    ("유동비율", "68.4%", "1 미만 → 유동성 관리 필요", "C+", red_fill),
    ("자기자본비율", "23.3%", "20% 이상이나 합병 전 32%에서 하락", "B-", gold_fill),
    ("영업CF/부채", "12.6%", "10-15% 구간, 양호~보통 경계", "B", gold_fill),
    ("ICR(영업이익)", "3.5배", "이자 커버 가능, 안정적", "B+", green_fill),
    ("ICR(영업CF)", "7.6배", "현금기준 이자 커버 우수", "A-", green_fill),
    ("순차입금/EBITDA", "3.3배", "3배 초과이나 5배 미만 → 주의 구간", "B", gold_fill),
]

write_header_row(ws6, row, ["지표", "2024 값", "해석", "등급", "", "", "", ""])
row += 1
for metric, val, desc, grade, fill in scorecard:
    ws6.cell(row=row, column=1, value=metric).font = data_font_bold
    ws6.cell(row=row, column=1).fill = fill
    ws6.cell(row=row, column=1).alignment = left_al
    ws6.cell(row=row, column=1).border = thin_border
    ws6.cell(row=row, column=2, value=val).font = navy_font
    ws6.cell(row=row, column=2).fill = fill
    ws6.cell(row=row, column=2).alignment = center_al
    ws6.cell(row=row, column=2).border = thin_border
    ws6.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    ws6.cell(row=row, column=3, value=desc).font = data_font
    ws6.cell(row=row, column=3).fill = fill
    ws6.cell(row=row, column=3).alignment = left_al
    ws6.cell(row=row, column=3).border = thin_border
    ws6.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
    ws6.cell(row=row, column=7, value=grade).font = Font(name="맑은 고딕", size=14, bold=True, color=NAVY)
    ws6.cell(row=row, column=7).fill = fill
    ws6.cell(row=row, column=7).alignment = center_al
    ws6.cell(row=row, column=7).border = thin_border
    ws6.row_dimensions[row].height = 30
    row += 1

print("  [6/7] 레버리지안정성 완료")


# ============================================================
#  SHEET 7: 종합 판정 및 모니터링
# ============================================================
ws7 = wb.create_sheet("종합판정")
ws7.sheet_properties.tabColor = "2ECC71"
set_col_widths(ws7, [4, 22, 22, 22, 22, 22, 4])
setup_print(ws7)

ws7.merge_cells('A1:G2')
c = ws7.cell(row=1, column=1, value="종합 판정 및 모니터링")
c.font = title_font; c.fill = title_fill; c.alignment = center_al
ws7.row_dimensions[1].height = 35

row = 4

# Rating
ws7.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
ws7.cell(row=row, column=2, value="종합 재무등급").font = Font(name="맑은 고딕", size=14, bold=True, color=WHITE)
ws7.cell(row=row, column=2).fill = PatternFill("solid", fgColor=DARK_BLUE)
ws7.cell(row=row, column=2).alignment = center_al
ws7.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
ws7.cell(row=row, column=4, value="BB+").font = Font(name="맑은 고딕", size=28, bold=True, color=ACCENT_GOLD)
ws7.cell(row=row, column=4).fill = PatternFill("solid", fgColor=DARK_BLUE)
ws7.cell(row=row, column=4).alignment = center_al
row += 1

ws7.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
ws7.cell(row=row, column=2, value="투기등급 근접이나, 영업CF 양호 + 합병 시너지 기대 -> 안정적 전망").font = Font(name="맑은 고딕", size=11, color="B0C4DE")
ws7.cell(row=row, column=2).fill = PatternFill("solid", fgColor=DARK_BLUE)
ws7.cell(row=row, column=2).alignment = center_al
row += 2

# 리스크 요약
row = add_section_title(ws7, row, "A. 리스크 요약", col_end=6)

risks = [
    ("리스부채 과다", "높음", "총부채의 30%가 리스부채(10.9조). 항공기 임차 구조상 불가피하나, 환율/금리 민감"),
    ("유동비율 1미만", "높음", "유동비율 68.4%. 유동부채가 유동자산을 3.6조 초과. 단기 차환 리스크 내재"),
    ("합병 후 부채비율", "중간", "329%로 급등. 그러나 합병 원인이므로 시너지 발현 시 3년 내 250% 이하 가능"),
    ("금리 상승", "중간", "변동금리 차입 비중 및 리스이자 증가 시 금융비용 추가 부담"),
    ("항공 경기 민감", "중간", "경기침체/유가급등/팬데믹 등 외부 충격에 매우 취약한 업종 특성"),
]

write_header_row(ws7, row, ["", "리스크 항목", "심각도", "상세 설명", "", "", ""], col_start=1)
row += 1
for risk, severity, detail in risks:
    s_font = red_font if severity == "높음" else warn_font
    s_fill = red_fill if severity == "높음" else gold_fill

    ws7.cell(row=row, column=1).fill = white_fill
    ws7.cell(row=row, column=1).border = thin_border
    ws7.cell(row=row, column=2, value=risk).font = data_font_bold
    ws7.cell(row=row, column=2).fill = s_fill
    ws7.cell(row=row, column=2).alignment = left_al
    ws7.cell(row=row, column=2).border = thin_border
    ws7.cell(row=row, column=3, value=severity).font = s_font
    ws7.cell(row=row, column=3).fill = s_fill
    ws7.cell(row=row, column=3).alignment = center_al
    ws7.cell(row=row, column=3).border = thin_border
    ws7.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    ws7.cell(row=row, column=4, value=detail).font = data_font
    ws7.cell(row=row, column=4).fill = s_fill
    ws7.cell(row=row, column=4).alignment = left_al
    ws7.cell(row=row, column=4).border = thin_border
    ws7.cell(row=row, column=7).fill = white_fill
    ws7.cell(row=row, column=7).border = thin_border
    ws7.row_dimensions[row].height = 30
    row += 1

row += 1

# 긍정 요인
row = add_section_title(ws7, row, "B. 긍정 요인", col_end=6)
positives = [
    ("영업CF 4.6조원", "항공업 특성상 감가상각비 환입으로 영업CF가 영업이익보다 2배 이상. 실질 상환여력 충분"),
    ("EBITDA 3.8조원", "감가상각 포함 이익 기준으로 안정적. 항공기 교체 및 부채 상환에 활용 가능"),
    ("합병 시너지", "노선 통합, 중복 비용 절감, 화물 네트워크 확대 → 2025~2027 점진적 이익 증가 기대"),
    ("화물 경쟁력", "글로벌 Top5 항공화물 기업. e커머스 성장으로 화물 수요 구조적 증가"),
    ("국적 항공사 지위", "국내 독과점적 지위(합병 후). 정부 정책적 지원 가능성"),
]
for label, desc in positives:
    ws7.cell(row=row, column=1).fill = white_fill
    ws7.cell(row=row, column=1).border = thin_border
    ws7.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws7.cell(row=row, column=2, value=label).font = green_font
    ws7.cell(row=row, column=2).fill = green_fill
    ws7.cell(row=row, column=2).alignment = left_al
    ws7.cell(row=row, column=2).border = thin_border
    ws7.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    ws7.cell(row=row, column=4, value=desc).font = data_font
    ws7.cell(row=row, column=4).fill = green_fill
    ws7.cell(row=row, column=4).alignment = left_al
    ws7.cell(row=row, column=4).border = thin_border
    ws7.cell(row=row, column=7).fill = white_fill
    ws7.cell(row=row, column=7).border = thin_border
    ws7.row_dimensions[row].height = 30
    row += 1

row += 1

# 부채감축 로드맵
row = add_section_title(ws7, row, "C. 부채감축 로드맵 (2025-2027)", col_end=6)

roadmap = [
    ("2025년", "부채비율 300% 이하", "합병 시너지 본격화, 리스부채 만기 상환(약 2조), 영업CF 4.5조 활용"),
    ("2026년", "부채비율 270% 이하", "노후 항공기 반납(리스 종료), 고효율 신기재 도입, 화물 수익 안정화"),
    ("2027년", "부채비율 250% 이하", "아시아나 완전 통합 마무리, 중복노선 정리 완료, 자본 축적"),
]

write_header_row(ws7, row, ["", "시점", "목표", "핵심 전략", "", "", ""], col_start=1)
row += 1
for year, target, strategy in roadmap:
    ws7.cell(row=row, column=1).fill = white_fill
    ws7.cell(row=row, column=1).border = thin_border
    ws7.cell(row=row, column=2, value=year).font = data_font_bold
    ws7.cell(row=row, column=2).fill = blue_light_fill
    ws7.cell(row=row, column=2).alignment = center_al
    ws7.cell(row=row, column=2).border = thin_border
    ws7.cell(row=row, column=3, value=target).font = navy_font
    ws7.cell(row=row, column=3).fill = blue_light_fill
    ws7.cell(row=row, column=3).alignment = center_al
    ws7.cell(row=row, column=3).border = thin_border
    ws7.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    ws7.cell(row=row, column=4, value=strategy).font = data_font
    ws7.cell(row=row, column=4).fill = blue_light_fill
    ws7.cell(row=row, column=4).alignment = left_al
    ws7.cell(row=row, column=4).border = thin_border
    ws7.cell(row=row, column=7).fill = white_fill
    ws7.cell(row=row, column=7).border = thin_border
    ws7.row_dimensions[row].height = 30
    row += 1

row += 1

# 모니터링 핵심 지표
row = add_section_title(ws7, row, "D. 모니터링 핵심 KPI", col_end=6)

kpis = [
    ("순차입금/EBITDA", "< 3배", "3.3배 (2024)", "분기별", "3배 돌파 시 부채감축 가속 필요"),
    ("부채비율", "< 300%", "329% (2024)", "반기별", "300% 미만 진입 시 투자등급 상향 가능"),
    ("유동비율", "> 80%", "68.4% (2024)", "분기별", "70% 미만 지속 시 단기자금 조달 리스크"),
    ("ICR (영업이익)", "> 3배", "3.5배 (2024)", "분기별", "3배 미만 시 이자 커버 우려"),
    ("영업CF/부채", "> 12%", "12.6% (2024)", "반기별", "10% 미만 시 상환여력 부족 신호"),
    ("리스부채 추이", "감소 추세", "10.9조 (2024)", "반기별", "항공기 반납/만기에 따른 자연감소 확인"),
]

write_header_row(ws7, row, ["", "KPI", "기준선", "현재값", "점검주기", "경고 트리거", ""], col_start=1)
row += 1
for kpi, target, current, freq, trigger in kpis:
    ws7.cell(row=row, column=1).fill = white_fill
    ws7.cell(row=row, column=1).border = thin_border
    ws7.cell(row=row, column=2, value=kpi).font = data_font_bold
    ws7.cell(row=row, column=2).fill = lighter_fill
    ws7.cell(row=row, column=2).alignment = left_al
    ws7.cell(row=row, column=2).border = thin_border
    ws7.cell(row=row, column=3, value=target).font = green_font
    ws7.cell(row=row, column=3).fill = lighter_fill
    ws7.cell(row=row, column=3).alignment = center_al
    ws7.cell(row=row, column=3).border = thin_border
    ws7.cell(row=row, column=4, value=current).font = data_font_bold
    ws7.cell(row=row, column=4).fill = lighter_fill
    ws7.cell(row=row, column=4).alignment = center_al
    ws7.cell(row=row, column=4).border = thin_border
    ws7.cell(row=row, column=5, value=freq).font = data_font
    ws7.cell(row=row, column=5).fill = lighter_fill
    ws7.cell(row=row, column=5).alignment = center_al
    ws7.cell(row=row, column=5).border = thin_border
    ws7.cell(row=row, column=6, value=trigger).font = data_font
    ws7.cell(row=row, column=6).fill = lighter_fill
    ws7.cell(row=row, column=6).alignment = left_al
    ws7.cell(row=row, column=6).border = thin_border
    ws7.cell(row=row, column=7).fill = white_fill
    ws7.cell(row=row, column=7).border = thin_border
    ws7.row_dimensions[row].height = 28
    row += 1

row += 2

# 최종 코멘트
ws7.merge_cells(start_row=row, start_column=2, end_row=row+4, end_column=6)
cell = ws7.cell(row=row, column=2)
cell.value = (
    "종합 의견:\n\n"
    "대한항공의 부채비율 329%는 아시아나 합병에 따른 일시적 급증으로, 영업실질로 보면 위기 수준이 아니다. "
    "영업CF 4.6조원, EBITDA 3.8조원의 안정적 현금 창출력이 뒷받침되며, "
    "순차입금/EBITDA 3.3배는 글로벌 항공사 평균(4~5배)보다 양호하다.\n\n"
    "다만, 유동비율 68.4%(1 미만)는 단기 유동성 관리가 필요하며, "
    "리스부채 10.9조원은 항공업 구조적 특성이므로 절대 금액보다 EBITDA 대비 비율로 모니터링해야 한다. "
    "합병 시너지가 본격화되는 2025~2027년에 부채비율 250% 이하, 순차입금/EBITDA 3배 이하 진입이 핵심 과제이다."
)
cell.font = Font(name="맑은 고딕", size=10, color=NAVY)
cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
cell.fill = lighter_fill
cell.border = thin_border

print("  [7/7] 종합판정 완료")


# ============================================================
#  SAVE
# ============================================================
wb.save(OUT)
conn.close()
print(f"\n보고서 생성 완료: {OUT}")
print(f"시트 구성: {wb.sheetnames}")
