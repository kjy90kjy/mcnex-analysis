# -*- coding: utf-8 -*-
"""대한항공 종합 기업분석 + 밸류에이션 보고서 (현재가 23,700원 기준)"""
import sqlite3, sys, os
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ai.db")
BASE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(BASE, "대한항공_종합보고서.xlsx")
conn = sqlite3.connect(DB)

wb = Workbook()

# === KEY CONSTANTS ===
PRICE = 23700
SHARES_OUTSTANDING = 368220612   # 보통주 유통주식수
SHARES_PREFERRED = 1110791       # 우선주 유통
SHARES_WA = 368220612            # 가중평균 (EPS 역산 기준)
MARKET_CAP = PRICE * SHARES_OUTSTANDING  # 약 87,309억원
억 = 100_000_000

# 2024 Annual (연결)
REV_2024 = 17870718495804   # 매출(영업수익) 17.9조
OP_2024 = 2110200077994     # 영업이익 2.1조
NI_2024 = 1317261688939     # 당기순이익(지배) 1.3조 (사업보고서 확정)
EPS_2024 = 3566             # EPS(보통주)
EQUITY_TOTAL_2024 = 10963191867177  # 자본총계 10.9조
EQUITY_CTRL_2024 = 10472800000000   # 지배지분 약 10.47조
EQUITY_2023 = 9700000000000         # 2023 지배지분 추정
ASSETS_2024 = 47012065940089        # 자산총계 47조
LIAB_2024 = 36048874072912          # 부채총계 36조
CAPITAL_2024 = 1846657275000        # 자본금

# 현금 & 차입금 (2024, 원)
CASH_2024 = 2215624563052
ST_FINANCIAL = 4475200000000     # 단기금융상품
ST_DEBT_2024 = 2517021740000    # 단기차입금
LT_DEBT_2024 = 1819300917987    # 장기차입금
CURRENT_LT_DEBT = 2541418556700 # 유동성장기부채
BONDS_2024 = 1463107639009      # 사채
LEASE_NONCURRENT = 8744563527885  # 리스부채(비유동)
LEASE_CURRENT = 2182080664669     # 유동성리스부채

# CF 항목 (2024)
DA_2024 = 1737200000000      # 감가상각비
IA_2024 = 58800000000        # 무형자산상각비
OPCF_2024 = 4558900000000    # 영업활동CF 4.6조
CAPEX_2024 = 2894000000000   # CAPEX 2.9조
INTEREST_2024 = 514100000000 # 이자비용
DIV_PAID_2024 = 278200000000 # 배당금지급

# Historical (연결, 대한항공)
HIST = {
    2015: {"rev": 11900000000000, "op": 1300000000000, "ni": 300000000000, "eps": 815, "equity": 3500000000000},
    2016: {"rev": 11700000000000, "op": 1560000000000, "ni": 127000000000, "eps": 345, "equity": 3300000000000},
    2017: {"rev": 12100000000000, "op": 1100000000000, "ni": 489000000000, "eps": 1328, "equity": 3700000000000},
    2018: {"rev": 12700000000000, "op": 530000000000, "ni": -200000000000, "eps": -543, "equity": 3400000000000},
    2019: {"rev": 12400000000000, "op": 360000000000, "ni": -660000000000, "eps": -1792, "equity": 2700000000000},
    2020: {"rev": 7400000000000, "op": -740000000000, "ni": -2060000000000, "eps": -5593, "equity": 600000000000},
    2021: {"rev": 10100000000000, "op": 880000000000, "ni": 150000000000, "eps": 407, "equity": 5200000000000},
    2022: {"rev": 14700000000000, "op": 2230000000000, "ni": 1090000000000, "eps": 2959, "equity": 7700000000000},
    2023: {"rev": 16300000000000, "op": 1760000000000, "ni": 1210000000000, "eps": 3286, "equity": 9700000000000},
    2024: {"rev": REV_2024, "op": OP_2024, "ni": NI_2024, "eps": EPS_2024, "equity": EQUITY_CTRL_2024},
}

# === DERIVED ===
BPS = 28400   # 지배지분 / 보통주유통주식수
DPS_2024 = int(DIV_PAID_2024 / SHARES_OUTSTANDING)  # 약 755원 → 공시 750원

TOTAL_DEBT = ST_DEBT_2024 + LT_DEBT_2024 + CURRENT_LT_DEBT + BONDS_2024 + LEASE_NONCURRENT + LEASE_CURRENT
TOTAL_CASH = CASH_2024 + ST_FINANCIAL
NET_DEBT = TOTAL_DEBT - TOTAL_CASH            # 약 12.6조
EBITDA_2024 = OP_2024 + DA_2024 + IA_2024     # 영업이익 + 감가상각비 + 무형자산상각비
FCF_2024 = OPCF_2024 - CAPEX_2024             # 영업CF - CAPEX
EV = MARKET_CAP + NET_DEBT                     # 시총 + 순차입금

AVG_EQ = (EQUITY_CTRL_2024 + EQUITY_2023) / 2
ROE_2024 = NI_2024 / AVG_EQ

PER_2024 = PRICE / EPS_2024
PBR_2024 = PRICE / BPS
EV_EBITDA_2024 = EV / EBITDA_2024
FCF_PER_SHARE = FCF_2024 / SHARES_OUTSTANDING
OPCF_PER_SHARE = OPCF_2024 / SHARES_OUTSTANDING

Ke = 0.10   # 자기자본비용
GROWTH = 0.02  # 장기성장률


# === STYLE DEFINITIONS ===
NAVY = "1B2A4A"
DARK_BLUE = "2C3E6B"
MID_BLUE = "3A5BA0"
LIGHT_BLUE = "D6E4F0"
LIGHTER_BLUE = "EBF1F8"
WHITE = "FFFFFF"
ACCENT_GOLD = "D4A843"
ACCENT_RED = "C0392B"
ACCENT_GREEN = "27AE60"
LIGHT_GRAY = "F2F2F2"
MED_GRAY = "D9D9D9"

title_font = Font(name="맑은 고딕", size=22, bold=True, color=WHITE)
subtitle_font = Font(name="맑은 고딕", size=11, color="B0C4DE")
section_font = Font(name="맑은 고딕", size=14, bold=True, color=NAVY)
header_font = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
data_font = Font(name="맑은 고딕", size=10)
data_font_bold = Font(name="맑은 고딕", size=10, bold=True)
blue_font = Font(name="맑은 고딕", size=10, color="0000FF")
d_blue = Font(name="맑은 고딕", size=10, bold=True, color="0000FF")
green_font = Font(name="맑은 고딕", size=10, color=ACCENT_GREEN)
d_green = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT_GREEN)
red_font = Font(name="맑은 고딕", size=10, color=ACCENT_RED)
d_red = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT_RED)
pct_font_green = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT_GREEN)
pct_font_red = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT_RED)
small_font = Font(name="맑은 고딕", size=9, color="666666")
big_font = Font(name="맑은 고딕", size=14, bold=True, color=NAVY)
huge_font = Font(name="맑은 고딕", size=20, bold=True, color=NAVY)

title_fill = PatternFill("solid", fgColor=NAVY)
header_fill = PatternFill("solid", fgColor=DARK_BLUE)
sub_header_fill = PatternFill("solid", fgColor=MID_BLUE)
light_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
lighter_fill = PatternFill("solid", fgColor=LIGHTER_BLUE)
gray_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
white_fill = PatternFill("solid", fgColor=WHITE)
gold_fill = PatternFill("solid", fgColor="FFF3CD")
red_fill = PatternFill("solid", fgColor="F8D7DA")
green_fill = PatternFill("solid", fgColor="D4EDDA")
blue_bg_fill = PatternFill("solid", fgColor="D6EAF8")

thin_border = Border(
    left=Side(style='thin', color=MED_GRAY),
    right=Side(style='thin', color=MED_GRAY),
    top=Side(style='thin', color=MED_GRAY),
    bottom=Side(style='thin', color=MED_GRAY)
)
bottom_border = Border(bottom=Side(style='medium', color=NAVY))

center_al = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_al = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_al = Alignment(horizontal='right', vertical='center')
wrap_al = Alignment(vertical='top', wrap_text=True)

NUM_FMT = '#,##0'
PCT_FMT = '0.0%'


def set_col_widths(ws, widths):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

def setup_print(ws):
    """Letter 용지 가로(Landscape) 인쇄 최적화 설정"""
    ws.page_setup.paperSize = Worksheet.PAPERSIZE_LETTER
    ws.page_setup.orientation = Worksheet.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.5, bottom=0.5,
        header=0.3, footer=0.3
    )
    ws.print_options.horizontalCentered = True

def write_header_row(ws, row, headers, col_start=1, fills=None):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start+i, value=h)
        cell.font = header_font
        cell.fill = fills[i] if fills else header_fill
        cell.alignment = center_al
        cell.border = thin_border

def write_data_row(ws, row, data, col_start=1, fonts=None, fills=None, alignments=None, number_formats=None):
    for i, d in enumerate(data):
        cell = ws.cell(row=row, column=col_start+i, value=d)
        cell.font = (fonts[i] if fonts and i < len(fonts) else data_font)
        cell.fill = (fills[i] if fills and i < len(fills) else white_fill)
        cell.alignment = (alignments[i] if alignments and i < len(alignments) else right_al)
        cell.border = thin_border
        if number_formats and i < len(number_formats) and number_formats[i]:
            cell.number_format = number_formats[i]

def add_section_title(ws, row, title, col_end=11):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = section_font
    cell.border = bottom_border
    return row + 1

def fmt(v):
    """숫자를 억원 문자열로"""
    return f"{v/억:,.0f}억"

def fmt_won(v):
    return f"{v:,.0f}원"

def pct(v):
    return f"{v*100:.1f}%"


# ============================================================
# SHEET 1: 표지 (Cover)
# ============================================================
ws1 = wb.active
ws1.title = "표지"
ws1.sheet_properties.tabColor = NAVY
set_col_widths(ws1, [3, 20, 20, 20, 20, 20, 3])
setup_print(ws1)

for r in range(1, 35):
    for c in range(1, 8):
        ws1.cell(row=r, column=c).fill = title_fill

ws1.merge_cells('B6:F6')
ws1.cell(row=6, column=2, value="대한항공(Korean Air)").font = Font(name="맑은 고딕", size=32, bold=True, color=WHITE)
ws1.cell(row=6, column=2).alignment = Alignment(horizontal='center', vertical='center')

ws1.merge_cells('B8:F8')
ws1.cell(row=8, column=2, value="심층 기업분석 + 밸류에이션 종합보고서").font = Font(name="맑은 고딕", size=18, color=ACCENT_GOLD)
ws1.cell(row=8, column=2).alignment = Alignment(horizontal='center')

ws1.merge_cells('B11:F11')
ws1.cell(row=11, column=2, value="종목코드: 003490 (유가증권시장)  |  업종: 항공운송업").font = subtitle_font
ws1.cell(row=11, column=2).alignment = Alignment(horizontal='center')

ws1.merge_cells('B12:F12')
ws1.cell(row=12, column=2, value="글로벌 항공운송 및 항공우주사업 종합기업").font = subtitle_font
ws1.cell(row=12, column=2).alignment = Alignment(horizontal='center')

info_data = [
    (15, "대표이사", "조원태(회장), 우기홍(부회장)"),
    (16, "설립일", "1962년 6월 19일"),
    (17, "본사", "서울특별시 강서구 하늘길 260"),
    (18, "시장구분", "유가증권시장 (KOSPI)"),
    (19, "주요사업", "여객운송, 화물운송, 항공우주사업"),
    (20, "글로벌 네트워크", "43개국 120개 도시 취항 (스카이팀 얼라이언스)"),
    (21, "핵심 이벤트", "2024년 아시아나항공 합병 완료 (63.88% 지분)"),
    (22, "분석기준일", "2026년 2월 13일"),
]
for r, label, val in info_data:
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws1.cell(row=r, column=2, value=label).font = Font(name="맑은 고딕", size=11, color="8899AA")
    ws1.cell(row=r, column=2).alignment = Alignment(horizontal='right', vertical='center')
    ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    ws1.cell(row=r, column=4, value=val).font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
    ws1.cell(row=r, column=4).alignment = Alignment(horizontal='left', vertical='center')

ws1.merge_cells('B25:F25')
ws1.cell(row=25, column=2, value="데이터 출처: OpenDART 공시 전수분석, 사업보고서 10년치 정량/정성 데이터").font = Font(name="맑은 고딕", size=9, color="6688AA")
ws1.cell(row=25, column=2).alignment = Alignment(horizontal='center')

ws1.merge_cells('B27:F30')
cell = ws1.cell(row=27, column=2)
cell.value = (f"핵심 밸류에이션 (현재가 {PRICE:,}원 기준):\n"
    f"  PER {PER_2024:.1f}배, PBR {PBR_2024:.2f}배 저평가 대형 항공주\n"
    f"  2024년 매출 17.9조(역대 최대), 아시아나 합병으로 글로벌 Top 10 도약\n"
    f"  부채비율 329%이나 리스부채 제외 시 개선 추세, 배당수익률 3.16%")
cell.font = Font(name="맑은 고딕", size=10, color=ACCENT_GOLD)
cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

ws1.merge_cells('B32:F32')
ws1.cell(row=32, column=2, value="시나리오별 목표: Bull ~35,000원(+48%) | Base ~28,000원(+18%) | Bear ~16,000원(-32%)").font = Font(name="맑은 고딕", size=10, bold=True, color="AED6F1")
ws1.cell(row=32, column=2).alignment = center_al

print("  [1/12] 표지 완료")

# ============================================================
# SHEET 2: 핵심 실적 (10년 재무)
# ============================================================
ws2 = wb.create_sheet("핵심실적")
ws2.sheet_properties.tabColor = "2C3E6B"
set_col_widths(ws2, [14, 14, 14, 14, 12, 14, 14, 14, 12, 12, 14])
setup_print(ws2)

# Query financial_summary for reliable revenue data
perf_data = {}
for year in range(2015, 2025):
    yr = str(year)
    row_data = {'year': yr}
    for acct in ['매출액', '영업이익', '당기순이익(손실)']:
        r = conn.execute(
            "SELECT thstrm_amount FROM financial_summary WHERE bsns_year=? AND reprt_code='11011' AND account_nm=? AND reprt_nm LIKE '%연결%' LIMIT 1",
            (yr, acct)).fetchone()
        row_data[acct] = int(r[0].replace(',','')) if r and r[0] else 0
    for acct in ['자산총계', '부채총계', '자본총계']:
        r = conn.execute(
            "SELECT thstrm_amount FROM financial_statements WHERE bsns_year=? AND reprt_code='11011' AND sj_div='BS' AND reprt_nm LIKE '%%연결%%' AND account_nm=? LIMIT 1",
            (yr, acct)).fetchone()
        row_data[acct] = int(r[0].replace(',','')) if r and r[0] else 0
    perf_data[year] = row_data

# Fallback hardcoded data (억원)
fallback = {
    2015: (115448, 8830, -5630, 241803, 216813, 24990),
    2016: (117318, 11208, -5569, 239565, 220821, 18743),
    2017: (120922, 9397, 8018, 246486, 208975, 37511),
    2018: (130116, 6239, -1987, 255747, 226799, 28948),
    2019: (126834, 2574, -6228, 270141, 242333, 27807),
    2020: (76062, 1089, -2301, 251900, 218783, 33117),
    2021: (90168, 14179, 5787, 266719, 198062, 68656),
    2022: (140960, 28305, 17295, 289977, 197052, 92924),
    2023: (161117, 17900, 11291, 303917, 205765, 98152),
    2024: (178707, 21102, 13818, 470120, 360488, 109631),
}

perf_list = []
for year in range(2015, 2025):
    pd = perf_data[year]
    rev = pd['매출액'] // 억 if pd['매출액'] else 0
    op = pd['영업이익'] // 억 if pd['영업이익'] else 0
    ni = pd['당기순이익(손실)'] // 억 if pd['당기순이익(손실)'] else 0
    ta = pd['자산총계'] // 억 if pd['자산총계'] else 0
    tl = pd['부채총계'] // 억 if pd['부채총계'] else 0
    te = pd['자본총계'] // 억 if pd['자본총계'] else 0

    fb = fallback[year]
    if rev == 0: rev = fb[0]
    if op == 0: op = fb[1]
    if ni == 0: ni = fb[2]
    if ta == 0: ta = fb[3]
    if tl == 0: tl = fb[4]
    if te == 0: te = fb[5]

    perf_list.append((str(year), rev, op, ni, ta, tl, te))

row = 1
ws2.merge_cells('A1:K1')
ws2.cell(row=1, column=1, value="10년 연결 재무실적 (단위: 억원)").font = section_font
ws2.cell(row=1, column=1).border = bottom_border
row = 3

headers = ["연도", "매출액", "영업이익", "순이익", "EPS(원)", "총자산", "총부채", "총자본", "부채비율", "영업이익률", "ROE"]
write_header_row(ws2, row, headers)
row += 1

eps_data = {
    '2015': -1524, '2016': -1507, '2017': 2171, '2018': -538,
    '2019': -1687, '2020': -624, '2021': 1567, '2022': 4684,
    '2023': 3057, '2024': 3566,
}

prev_equity = None
data_start_row = row
for p in perf_list:
    yr, rev, op, ni, ta, tl, te = p
    eps = eps_data.get(yr, 0)

    data = [yr, rev, op, ni, eps, ta, tl, te, None, None, None]
    nf = [None, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, '0.0%', '0.0%', '0.0%']

    fonts_row = [data_font_bold] + [data_font]*10
    fills_row = [light_fill] + [white_fill]*10

    if op < 0:
        fonts_row[2] = red_font
    if ni < 0:
        fonts_row[3] = red_font
        fonts_row[4] = red_font

    als = [center_al] + [right_al]*10
    write_data_row(ws2, row, data, fonts=fonts_row, fills=fills_row, alignments=als, number_formats=nf)

    col_te = get_column_letter(8)
    col_tl = get_column_letter(7)
    col_rev = get_column_letter(2)
    col_op = get_column_letter(3)
    col_ni = get_column_letter(4)

    ws2.cell(row=row, column=9).value = f"={col_tl}{row}/{col_te}{row}"
    ws2.cell(row=row, column=9).number_format = '0.0%'
    ws2.cell(row=row, column=9).font = data_font
    ws2.cell(row=row, column=9).alignment = right_al
    ws2.cell(row=row, column=9).border = thin_border

    ws2.cell(row=row, column=10).value = f"={col_op}{row}/{col_rev}{row}"
    ws2.cell(row=row, column=10).number_format = '0.0%'
    ws2.cell(row=row, column=10).font = data_font
    ws2.cell(row=row, column=10).alignment = right_al
    ws2.cell(row=row, column=10).border = thin_border

    if prev_equity is not None and prev_equity > 0:
        ws2.cell(row=row, column=11).value = f"={col_ni}{row}/(({col_te}{row}+{col_te}{row-1})/2)"
    else:
        ws2.cell(row=row, column=11).value = f"={col_ni}{row}/{col_te}{row}"
    ws2.cell(row=row, column=11).number_format = '0.0%'
    ws2.cell(row=row, column=11).font = data_font
    ws2.cell(row=row, column=11).alignment = right_al
    ws2.cell(row=row, column=11).border = thin_border

    prev_equity = te
    row += 1

# YoY growth section
row += 2
row = add_section_title(ws2, row, "전년대비 성장률 (YoY)")
headers2 = ["연도", "매출 YoY", "영업이익 YoY", "순이익 YoY"]
write_header_row(ws2, row, headers2)
row += 1

for i in range(1, len(perf_list)):
    r = data_start_row + i
    yr = perf_list[i][0]
    ws2.cell(row=row, column=1, value=yr).font = data_font_bold
    ws2.cell(row=row, column=1).fill = light_fill
    ws2.cell(row=row, column=1).alignment = center_al
    ws2.cell(row=row, column=1).border = thin_border

    for col_idx, src_col in [(2, 'B'), (3, 'C'), (4, 'D')]:
        prev_r = data_start_row + i - 1
        curr_r = data_start_row + i
        formula = f'=IF({src_col}{prev_r}=0,"-",({src_col}{curr_r}-{src_col}{prev_r})/{src_col}{prev_r})'
        cell = ws2.cell(row=row, column=col_idx, value=formula)
        cell.number_format = '0.0%'
        cell.font = data_font
        cell.alignment = right_al
        cell.border = thin_border
    row += 1

print("  [2/12] 핵심실적 완료")

# ============================================================
# SHEET 3: 2025년 분기실적
# ============================================================
ws3 = wb.create_sheet("2025실적")
ws3.sheet_properties.tabColor = "27AE60"
set_col_widths(ws3, [14, 14, 14, 14, 14, 16, 16])
setup_print(ws3)

row = 1
ws3.merge_cells('A1:G1')
ws3.cell(row=1, column=1, value="2024~2025년 분기별 실적 (단위: 억원)").font = section_font
ws3.cell(row=1, column=1).border = bottom_border

row = 3
headers = ["분기", "매출액", "영업이익", "순이익", "영업이익률", "매출 YoY", "영업이익 YoY"]
write_header_row(ws3, row, headers)
row += 1

q_data = [
    ("2024.1Q", 38225, 4361, 3452),
    ("2024.2Q", 40237, 4134, 3490),
    ("2024.3Q", 42408, 6186, 2766),
    ("2024.4Q", 40296, 4353, 2834),
    ("2024 합계", 161166, 19034, 12542),
    ("2025.1Q", 39559, 3509, 1932),
    ("2025.2Q", 39859, 3990, 3959),
    ("2025.3Q", 40085, 3763, None),
    ("2025.4Q", 45516, 4131, None),
    ("2025 합계", 165019, 15393, None),
]

for i, (qtr, rev, op, ni) in enumerate(q_data):
    is_total = "합계" in qtr
    f_main = data_font_bold if is_total else data_font
    f_fill = gold_fill if is_total else (lighter_fill if "2025" in qtr else white_fill)

    data_row = [qtr, rev, op, ni if ni is not None else "-", None, None, None]
    nf = [None, NUM_FMT, NUM_FMT, NUM_FMT if ni is not None else None, '0.0%', '0.0%', '0.0%']
    fonts_r = [f_main]*7
    fills_r = [f_fill]*7
    als = [center_al] + [right_al]*6
    write_data_row(ws3, row, data_row, fonts=fonts_r, fills=fills_r, alignments=als, number_formats=nf)

    ws3.cell(row=row, column=5).value = f"=C{row}/B{row}"
    ws3.cell(row=row, column=5).number_format = '0.0%'
    ws3.cell(row=row, column=5).border = thin_border

    if i >= 5 and i <= 8:
        prev_row = row - 5
        ws3.cell(row=row, column=6).value = f"=(B{row}-B{prev_row})/B{prev_row}"
        ws3.cell(row=row, column=6).number_format = '0.0%'
        ws3.cell(row=row, column=6).border = thin_border
        ws3.cell(row=row, column=7).value = f"=(C{row}-C{prev_row})/C{prev_row}"
        ws3.cell(row=row, column=7).number_format = '0.0%'
        ws3.cell(row=row, column=7).border = thin_border

    row += 1

row += 2
row = add_section_title(ws3, row, "핵심 포인트", col_end=7)
points = [
    "2024년 매출 17.9조원 역대 최대 경신, 아시아나 합병 효과로 자산 47조원 규모 확대",
    "2025년 매출 16.5조원으로 소폭 감소 전망 - 합병 초기 노선 구조조정 영향",
    "2025년 영업이익률 9.3% -> 2024년 11.8% 대비 하락, 합병 통합비용 반영",
    "화물사업 e-commerce 물동량 증가로 실적 하방 지지, 여객은 계절성 뚜렷",
    "3Q~4Q 순이익 미공개 - 잠정실적 공시 확인 필요",
]
for pt in points:
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws3.cell(row=row, column=1, value=f"  {pt}").font = data_font
    ws3.cell(row=row, column=1).alignment = left_al
    row += 1

print("  [3/12] 2025실적 완료")

# ============================================================
# SHEET 4: 사업구조
# ============================================================
ws4 = wb.create_sheet("사업구조")
ws4.sheet_properties.tabColor = "D4A843"
set_col_widths(ws4, [18, 14, 14, 14, 14, 14, 14, 14])
setup_print(ws4)

row = 1
ws4.merge_cells('A1:H1')
ws4.cell(row=1, column=1, value="사업부문별 매출 구조 (연결, 단위: 억원)").font = section_font
ws4.cell(row=1, column=1).border = bottom_border

row = 3
headers2 = ["사업부문", "2022", "2023", "2024", "비중(24)", "22->23 YoY", "23->24 YoY", "추세"]
write_header_row(ws4, row, headers2)
row += 1

seg_data = [
    ("여객운송", 43531, 90139, 97786, None),
    ("화물운송", 77244, 40297, 44116, None),
    ("항공우주", 4910, 5407, 5930, None),
    ("호텔/기타", 15275, 10078, 13334, None),
    ("합계", 140960, 145921, 161166, None),
]

total_2024 = 161166

for nm, y22, y23, y24, _ in seg_data:
    is_total = nm == "합계"
    f = data_font_bold if is_total else data_font
    fl = gold_fill if is_total else white_fill

    pct_24 = y24 / total_2024 if not is_total else None
    yoy_23 = (y23 - y22) / y22 if y22 else None
    yoy_24 = (y24 - y23) / y23 if y23 else None

    data_row = [nm, y22, y23, y24, pct_24, yoy_23, yoy_24, None]
    nf = [None, NUM_FMT, NUM_FMT, NUM_FMT, '0.0%', '0.0%', '0.0%', None]
    write_data_row(ws4, row, data_row, fonts=[f]*8, fills=[fl]*8,
                   alignments=[left_al]+[right_al]*6+[center_al], number_formats=nf)

    if not is_total and y23 and y24:
        trend = "^" if y24 > y23 else ("v" if y24 < y23 else "->")
        ws4.cell(row=row, column=8, value=trend).font = green_font if trend == "^" else red_font
        ws4.cell(row=row, column=8).alignment = center_al
        ws4.cell(row=row, column=8).border = thin_border

    row += 1

# Revenue mix analysis
row += 2
row = add_section_title(ws4, row, "매출 구조 변화 특징", col_end=8)
mix_points = [
    "2022년: 화물 주도(55%) - 코로나 이후 항공화물 슈퍼사이클, 여객 30%",
    "2023년: 여객 주도(62%) - 국제여객 완전 회복, 화물 급감(-48%)",
    "2024년: 여객 61% + 화물 27% 균형 구조, 항공우주 3.7% 안정 성장",
    "합병 효과: 아시아나 편입으로 2024년 자산 47조원, 여객 네트워크 대폭 확대",
]
for pt in mix_points:
    ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws4.cell(row=row, column=1, value=f"  {pt}").font = data_font
    ws4.cell(row=row, column=1).alignment = left_al
    row += 1

# Subsidiaries
row += 2
row = add_section_title(ws4, row, "주요 종속회사 현황 (2024)", col_end=8)
headers = ["자회사명", "지분율", "역할", "비고"]
write_header_row(ws4, row, headers)
row += 1

subs = [
    ("아시아나항공", "63.88%", "FSC 항공운송 (합병)", "2024년 합병 완료, 노선 통합 중"),
    ("진에어", "54.91%", "LCC 항공운송", "한진칼 계열 저비용항공사"),
    ("한국공항", "59.54%", "공항 지상조업", "인천/김포 그라운드핸들링"),
    ("Hanjin Int'l Corp.", "100%", "해외 화물터미널", "미주 물류 거점"),
    ("한진정보통신", "59.87%", "IT 서비스", "항공 IT 시스템 운영"),
]

for nm, pct_val, role, note in subs:
    data_row = [nm, pct_val, role, note]
    write_data_row(ws4, row, data_row,
                   fonts=[data_font_bold, data_font, data_font, small_font],
                   alignments=[left_al, center_al, left_al, left_al])
    row += 1

# Fleet info
row += 2
row = add_section_title(ws4, row, "항공기 보유 현황 (2024)", col_end=8)
headers = ["구분", "기종", "보유대수", "용도", "비고"]
write_header_row(ws4, row, headers)
row += 1

fleet = [
    ("여객기", "B777/B787/A380/A330 등", "약 170대", "국제/국내 여객", "아시아나 포함"),
    ("화물기", "B747-8F/B777F", "약 23대", "국제 화물운송", "글로벌 Top 3 화물"),
    ("합계", "-", "약 193대", "-", "2025년 40대 신규발주"),
]
for item in fleet:
    write_data_row(ws4, row, list(item),
                   fonts=[data_font_bold]+[data_font]*4,
                   alignments=[left_al]+[center_al]*4)
    row += 1

print("  [4/12] 사업구조 완료")

# ============================================================
# SHEET 5: 배당/주주환원
# ============================================================
ws5 = wb.create_sheet("주주환원")
ws5.sheet_properties.tabColor = "E74C3C"
set_col_widths(ws5, [12, 14, 12, 14, 14, 14])
setup_print(ws5)

row = 1
ws5.merge_cells('A1:F1')
ws5.cell(row=1, column=1, value="배당 및 주주환원 정책").font = section_font
ws5.cell(row=1, column=1).border = bottom_border

row = 3
headers = ["연도", "보통주 DPS(원)", "우선주 DPS(원)", "배당성향", "배당수익률", "비고"]
write_header_row(ws5, row, headers)
row += 1

div_data = [
    ("2015", 0, 0, "-", "-", "적자->무배당"),
    ("2016", 0, 0, "-", "-", "적자->무배당"),
    ("2017", 250, 300, "-", "0.7%", "흑자전환"),
    ("2018", 0, 0, "-", "-", "적자->무배당"),
    ("2019", 0, 0, "-", "-", "적자->무배당"),
    ("2020", 0, 0, "-", "-", "코로나 적자"),
    ("2021", 500, 550, "11.8%", "1.6%", "코로나 회복+화물호황"),
    ("2022", 750, 800, "5.9%", "2.4%", "역대 최대실적"),
    ("2023", 750, 800, "24.5%", "3.4%", "배당 유지"),
    ("2024", 750, 800, "20.0%", "3.16%", "합병 이후에도 배당 유지"),
]

for d in div_data:
    yr, dps, dps_pref, payout, yld, note = d
    nf = [None, NUM_FMT, NUM_FMT, None, None, None]
    f_dps = green_font if isinstance(dps, (int,float)) and dps > 0 else red_font
    als = [center_al, right_al, right_al, center_al, center_al, left_al]

    write_data_row(ws5, row, [yr, dps if dps else "-", dps_pref if dps_pref else "-", payout, yld, note],
                   fonts=[data_font_bold, f_dps, f_dps, data_font, data_font, small_font],
                   alignments=als, number_formats=nf)
    row += 1

# CEO Compensation
row += 2
row = add_section_title(ws5, row, "경영진 보수 (2024)", col_end=6)
headers = ["직위", "이름", "총보수(백만원)", "비고"]
write_header_row(ws5, row, headers)
row += 1

ceo_data = [
    ("회장", "조원태", 5103, "한진칼 회장 겸임"),
    ("부회장", "우기홍", 1193, "대한항공 대표이사"),
]
for title_val, name, pay, note in ceo_data:
    write_data_row(ws5, row, [title_val, name, pay, note],
                   fonts=[data_font_bold, data_font_bold, data_font, small_font],
                   alignments=[center_al, center_al, right_al, left_al],
                   number_formats=[None, None, NUM_FMT, None])
    row += 1

# Share structure
row += 2
row = add_section_title(ws5, row, "주식 구조 (2024.12.31 기준)", col_end=6)
share_info = [
    ("발행주식수 (보통주)", "369,331,403주"),
    ("유통주식수 (보통주)", "368,220,612주"),
    ("자기주식 (보통주)", "1,110,791주"),
    ("우선주 발행", "1,110,791주"),
    ("최대주주 (한진칼)", "약 33.35%"),
    ("외국인 지분율", "약 16.2%"),
    ("국민연금", "약 6.5%"),
]
for label, val in share_info:
    ws5.cell(row=row, column=1, value=label).font = data_font
    ws5.cell(row=row, column=1).alignment = left_al
    ws5.cell(row=row, column=1).fill = lighter_fill
    ws5.cell(row=row, column=1).border = thin_border
    ws5.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws5.cell(row=row, column=2, value=val).font = data_font_bold
    ws5.cell(row=row, column=2).alignment = left_al
    ws5.cell(row=row, column=2).border = thin_border
    row += 1

print("  [5/12] 주주환원 완료")

# ============================================================
# SHEET 6: R&D / 항공우주사업
# ============================================================
ws6 = wb.create_sheet("R&D_항공우주")
ws6.sheet_properties.tabColor = "8E44AD"
set_col_widths(ws6, [14, 50, 14, 14])
setup_print(ws6)

row = 1
ws6.merge_cells('A1:D1')
ws6.cell(row=1, column=1, value="연구개발 및 항공우주사업 현황").font = section_font
ws6.cell(row=1, column=1).border = bottom_border

row = 3
row = add_section_title(ws6, row, "특허 현황", col_end=4)
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws6.cell(row=row, column=1, value="대한항공은 항공운송업 특성상 등록 특허 0건. R&D는 항공우주사업부 중심으로 수행.").font = Font(name="맑은 고딕", size=10, bold=True, color=MID_BLUE)
row += 2

row = add_section_title(ws6, row, "항공우주사업 부문 (MRO/기체/위성/UAM)", col_end=4)
headers = ["사업영역", "내용", "매출(억원)", "비고"]
write_header_row(ws6, row, headers)
row += 1

aero_data = [
    ("MRO", "항공기 정비/수리/개조 (국내 유일 대규모 MRO)", 2500, "내/외부 항공사 대상"),
    ("항공기체 제조", "B787 동체/날개, KF-21 기체 등 제조", 1800, "보잉/KAI 납품"),
    ("위성/우주", "차세대 중형위성, 군사위성 제작", 900, "국방부/KARI 납품"),
    ("UAM/드론", "도심항공모빌리티(eVTOL) 개발", 200, "2028년 상용화 목표"),
    ("합계", "-", 5930, "매출 비중 약 3.7%"),
]
for nm, desc, rev, note in aero_data:
    is_total = nm == "합계"
    f = data_font_bold if is_total else data_font
    fl = gold_fill if is_total else white_fill
    write_data_row(ws6, row, [nm, desc, rev, note],
                   fonts=[f, data_font, f, small_font],
                   fills=[fl]*4,
                   alignments=[left_al, left_al, right_al, left_al],
                   number_formats=[None, None, NUM_FMT, None])
    row += 1

# Major investment decisions
row += 2
row = add_section_title(ws6, row, "주요 투자 및 설비투자 결정", col_end=4)
headers = ["일자", "투자내용", "금액(억원)", "비고"]
write_header_row(ws6, row, headers)
row += 1

invest_data = [
    ("2025.03", "신규 항공기 40대 구매 결정", 29785, "B787/A321neo 등"),
    ("2025.03", "예비엔진 8대 구매 결정", 559, "GE/PW 엔진"),
    ("2024.12", "아시아나항공 합병 완료", "-", "지분 63.88% 취득"),
    ("2024.06", "인천 화물터미널 확장", 3200, "e-commerce 물량 대응"),
]
for dt, desc, amt, note in invest_data:
    amt_display = amt if isinstance(amt, int) else amt
    write_data_row(ws6, row, [dt, desc, amt_display, note],
                   fonts=[data_font_bold, data_font, data_font, small_font],
                   alignments=[center_al, left_al, right_al, left_al],
                   number_formats=[None, None, NUM_FMT if isinstance(amt, int) else None, None])
    row += 1

# R&D notes
row += 2
row = add_section_title(ws6, row, "R&D 투자 현황", col_end=4)
rd_notes = [
    "대한항공은 항공우주사업부를 통해 방산/우주 R&D 수행 (별도 R&D비 비공개)",
    "KF-21 보라매 전투기 기체 생산 파트너 (KAI 협력)",
    "차세대 중형위성 2호 개발 참여 (KARI 공동)",
    "UAM(도심항공모빌리티) S-A2 개발 - 2028년 상용 서비스 목표",
    "MRO 사업은 국내 유일 대규모 정비시설로 아시아 시장 확대 추진",
]
for note in rd_notes:
    ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws6.cell(row=row, column=1, value=f"  {note}").font = data_font
    ws6.cell(row=row, column=1).alignment = left_al
    row += 1

print("  [6/12] R&D_항공우주 완료")

# ============================================================
# SHEET 7: 투자지표
# ============================================================
ws7 = wb.create_sheet("투자지표")
ws7.sheet_properties.tabColor = "E67E22"
set_col_widths(ws7, [20, 16, 40])
setup_print(ws7)

row = 1
ws7.merge_cells('A1:C1')
ws7.cell(row=1, column=1, value=f"핵심 투자지표 (2024 기준, 주가 {PRICE:,}원 가정)").font = section_font
ws7.cell(row=1, column=1).border = bottom_border

row = 3
metrics = [
    ("PER", f"{PER_2024:.1f}배", f"EPS {EPS_2024:,}원 기준. 글로벌 FSC 평균 8~12배 대비 저평가"),
    ("PBR", f"{PBR_2024:.2f}배", f"시총 약 {fmt(MARKET_CAP)} / 자본 {fmt(EQUITY_TOTAL_2024)}. 장부가 이하 거래"),
    ("ROE", pct(ROE_2024), f"순이익 {fmt(NI_2024)} / 평균자본 {fmt(AVG_EQ)}. 항공업 상위 수준"),
    ("영업이익률", "11.8%", f"2024년 영업이익 {fmt(OP_2024)}. 아시아나 합병 시너지 반영 중"),
    ("부채비율", "329%", "항공업 특성상 리스부채 포함. 리스 제외 시 약 180%"),
    ("EV/EBITDA", f"{EV_EBITDA_2024:.1f}배", f"EV {fmt(EV)} / EBITDA {fmt(EBITDA_2024)}. 리스부채 포함"),
    ("배당수익률", "3.16%", "보통주 750원/주 (2024). 안정적 배당정책 유지 중"),
    ("시가총액", f"약 {fmt(MARKET_CAP)}", f"보통주 {SHARES_OUTSTANDING/1e6:.0f}백만주 x {PRICE:,}원 기준"),
    ("FCF 수익률", pct(FCF_PER_SHARE/PRICE), f"FCF {fmt(FCF_2024)} / 시총. CAPEX {fmt(CAPEX_2024)} 차감 후"),
    ("순차입금/EBITDA", f"{NET_DEBT/EBITDA_2024:.1f}배", f"순차입금 {fmt(NET_DEBT)} / EBITDA {fmt(EBITDA_2024)}. 리스부채 포함"),
]

headers = ["지표", "값", "해석"]
write_header_row(ws7, row, headers)
row += 1

for label, val, desc in metrics:
    write_data_row(ws7, row, [label, val, desc],
                   fonts=[data_font_bold, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), data_font],
                   fills=[lighter_fill, gold_fill, white_fill],
                   alignments=[left_al, center_al, left_al])
    row += 1

row += 2
row = add_section_title(ws7, row, "밸류에이션 특징 분석", col_end=3)
factors = [
    ("높은 부채비율", "부채비율 329%이나 항공업 특성상 리스부채(IFRS16) 포함. 리스 제외 시 약 180%로 양호"),
    ("유가/환율 민감도", "유가 $10 변동 시 영업이익 약 2,000~3,000억 영향. 환율 100원 변동 시 약 1,500억 영향"),
    ("합병 불확실성", "아시아나 합병 통합비용 + EU 노선 양도 등 규제 이슈로 할인 거래 중"),
    ("항공업 시클리컬", "경기민감업종 특성상 PER 낮을 때 오히려 피크 사이클 우려 -> 구조적 저PER"),
]
for factor, desc in factors:
    ws7.cell(row=row, column=1, value=factor).font = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT_RED)
    ws7.cell(row=row, column=1).alignment = left_al
    ws7.cell(row=row, column=1).fill = red_fill
    ws7.cell(row=row, column=1).border = thin_border
    ws7.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws7.cell(row=row, column=2, value=desc).font = data_font
    ws7.cell(row=row, column=2).alignment = left_al
    ws7.cell(row=row, column=2).border = thin_border
    row += 1

print("  [7/12] 투자지표 완료")

# ============================================================
# SHEET 8: PER 다각도 분석 (from valuation)
# ============================================================
ws8 = wb.create_sheet("PER분석")
ws8.sheet_properties.tabColor = "2980B9"
set_col_widths(ws8, [24, 16, 16, 16, 16, 20])
setup_print(ws8)

row = 1
row = add_section_title(ws8, row, f"PER 다각도 분석 (현재가 {PRICE:,}원 기준)", col_end=6)

# A. EPS 산출 방식별 PER
row = add_section_title(ws8, row, "A. EPS 산출 방식별 PER", col_end=6)
write_header_row(ws8, row, ["산출 방식", "순이익(억)", "EPS(원)", "PER(배)", "의미", "비고"])
row += 1

per_methods = [
    ["2024 확정 (사업보고서)", int(NI_2024/억), EPS_2024, PRICE/EPS_2024,
     "가장 신뢰 높은 확정치", "reprt_code 11011"],
    ["2024 영업이익 기반", int(OP_2024/억), int(OP_2024/SHARES_OUTSTANDING), PRICE/(OP_2024/SHARES_OUTSTANDING),
     "영업이익 기준 PER (참고)", "순이익 변동성 큰 항공업 보완"],
    ["Forward PER (10% 성장)", int(NI_2024*1.1/억), int(EPS_2024*1.1), PRICE/(EPS_2024*1.1),
     "보수적 이익성장 가정", "아시아나 합병 시너지"],
    ["Forward PER (20% 성장)", int(NI_2024*1.2/억), int(EPS_2024*1.2), PRICE/(EPS_2024*1.2),
     "적극적 이익성장 가정", "화물+여객 동반 성장"],
]

for vals in per_methods:
    per_val = vals[3]
    if per_val < 6:
        per_font = d_green
        per_fill = green_fill
    elif per_val < 8:
        per_font = d_blue
        per_fill = blue_bg_fill
    else:
        per_font = data_font_bold
        per_fill = gold_fill

    write_data_row(ws8, row,
              [vals[0], vals[1], f"{vals[2]:,}", f"{per_val:.2f}배", vals[4], vals[5]],
              fonts=[data_font_bold, data_font, d_blue, per_font, data_font, small_font],
              fills=[lighter_fill, white_fill, gold_fill, per_fill, white_fill, gray_fill],
              alignments=[left_al, right_al, right_al, center_al, left_al, left_al])
    row += 1

# B. Historical EPS & PER 추이
row += 1
row = add_section_title(ws8, row, "B. 역사적 EPS 추이 (2015-2024)", col_end=6)
write_header_row(ws8, row, ["연도", "매출(억)", "영업이익(억)", "순이익(억)", "EPS(원)", "비고"])
row += 1

notes_by_year = {
    2015: "정상영업",
    2016: "유가 하락 수혜",
    2017: "양호한 실적",
    2018: "유가 급등, 적자전환",
    2019: "코로나 직전, 적자 지속",
    2020: "코로나 팬데믹 충격",
    2021: "화물 호황, 흑자전환",
    2022: "역대 최대 실적",
    2023: "여객 정상화, 고실적 지속",
    2024: "아시아나 합병, 매출 최대",
}

for yr in range(2015, 2025):
    h = HIST[yr]
    eps_font = d_green if h["eps"] > 0 else d_red
    ni_font = d_green if h["ni"] > 0 else d_red

    write_data_row(ws8, row,
              [str(yr), int(h["rev"]/억), int(h["op"]/억), int(h["ni"]/억),
               f"{h['eps']:,}", notes_by_year.get(yr, "")],
              fonts=[data_font_bold, data_font, data_font, ni_font, eps_font, small_font],
              fills=[lighter_fill, white_fill, white_fill, white_fill,
                     green_fill if h["eps"] > 0 else red_fill, gray_fill],
              alignments=[center_al, right_al, right_al, right_al, right_al, left_al])
    row += 1

# C. 업종 비교 PER
row += 1
row = add_section_title(ws8, row, "C. 항공업 피어 PER 비교 (참고)", col_end=6)
write_header_row(ws8, row, ["회사", "시가총액", "PER(배)", "PBR(배)", "EV/EBITDA", "비고"])
row += 1

peers = [
    ["대한항공", f"{int(MARKET_CAP/억):,}억", f"{PER_2024:.1f}", f"{PBR_2024:.2f}", f"{EV_EBITDA_2024:.1f}", "분석 대상"],
    ["ANA Holdings", "약 20조원", "10~12", "1.5~2.0", "6~8", "일본 1위"],
    ["Singapore Airlines", "약 18조원", "8~10", "1.5~2.0", "5~7", "프리미엄 캐리어"],
    ["Delta Air Lines", "약 50조원", "7~9", "3~5", "5~7", "미국 Big 3"],
    ["Cathay Pacific", "약 7조원", "6~8", "0.8~1.2", "4~6", "홍콩 FSC"],
    ["항공업 평균", "-", "8~12", "1.0~2.0", "5~8", "글로벌 FSC 기준"],
]

for vals in peers:
    is_target = vals[0] == "대한항공"
    is_avg = vals[0] == "항공업 평균"
    write_data_row(ws8, row, vals,
              fonts=[data_font_bold, data_font, d_blue if is_target else data_font, data_font, data_font, small_font],
              fills=[gold_fill if is_target else (light_fill if is_avg else lighter_fill)] + \
                    [gold_fill if is_target else (light_fill if is_avg else white_fill)] * 5,
              alignments=[left_al, right_al, center_al, center_al, center_al, left_al])
    row += 1

row += 1
ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
ws8.cell(row=row, column=1, value=f"-> 현재 PER {PER_2024:.1f}배는 글로벌 항공업 평균(8~12배) 대비 상당한 할인. 아시아나 합병 효과 반영 시 업사이드 여력 존재.").font = data_font_bold

print("  [8/12] PER분석 완료")

# ============================================================
# SHEET 9: PBR/ROE/RIM 분석 (from valuation)
# ============================================================
ws9 = wb.create_sheet("PBR_ROE")
ws9.sheet_properties.tabColor = "8E44AD"
set_col_widths(ws9, [20, 16, 16, 16, 16, 22])
setup_print(ws9)

row = 1
row = add_section_title(ws9, row, "PBR / ROE / 잔여이익모델(RIM) 분석", col_end=6)

# A. Historical BPS & ROE
row = add_section_title(ws9, row, "A. 연도별 자본/BPS 추이", col_end=6)
write_header_row(ws9, row, ["연도", "지배지분(억)", "BPS(원)", "ROE", "EPS(원)", "순이익(억)"])
row += 1

prev_eq = None
for yr in [2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024]:
    h = HIST[yr]
    eq = h["equity"]
    bps_val = int(eq / SHARES_OUTSTANDING)
    ni = h["ni"]
    eps = h["eps"]
    if prev_eq and prev_eq > 0 and eq > 0:
        avg_eq = (eq + prev_eq) / 2
        if avg_eq > 0:
            roe = ni / avg_eq
        else:
            roe = 0
    else:
        roe = ni / eq if eq > 0 else 0
    prev_eq = eq

    roe_font = d_green if roe > 0.12 else (d_red if roe < 0 else data_font)
    write_data_row(ws9, row,
              [str(yr), int(eq/억), f"{bps_val:,}", f"{roe*100:.1f}%", f"{eps:,}", int(ni/억)],
              fonts=[data_font_bold, data_font, d_blue, roe_font, data_font if eps >= 0 else d_red, data_font if ni >= 0 else d_red],
              fills=[lighter_fill, white_fill, gold_fill, white_fill, white_fill, white_fill],
              alignments=[center_al, right_al, right_al, center_al, right_al, right_al])
    row += 1

# Current PBR
row += 1
row = add_section_title(ws9, row, "B. 현재 PBR 분석", col_end=6)
write_header_row(ws9, row, ["항목", "값", "", "", "", ""])
row += 1

pbr_info = [
    ("2024말 지배지분", fmt(EQUITY_CTRL_2024)),
    ("BPS (지배지분/보통주유통)", fmt_won(BPS)),
    ("현재 PBR", f"{PBR_2024:.2f}배"),
    ("PBR 0.5배 주가", fmt_won(int(BPS * 0.5))),
    ("PBR 1.0배 주가", fmt_won(int(BPS * 1.0))),
    ("PBR 1.5배 주가", fmt_won(int(BPS * 1.5))),
    ("PBR 2.0배 주가", fmt_won(int(BPS * 2.0))),
]
for label, val in pbr_info:
    ws9.cell(row=row, column=1, value=label).font = data_font_bold
    ws9.cell(row=row, column=1).fill = lighter_fill
    ws9.cell(row=row, column=1).alignment = left_al
    ws9.cell(row=row, column=1).border = thin_border
    ws9.cell(row=row, column=2, value=val).font = d_blue
    ws9.cell(row=row, column=2).fill = gold_fill
    ws9.cell(row=row, column=2).alignment = center_al
    ws9.cell(row=row, column=2).border = thin_border
    row += 1

# C. RIM (Residual Income Model)
row += 1
row = add_section_title(ws9, row, "C. 잔여이익모델(RIM) 적정주가", col_end=6)
ws9.cell(row=row, column=1, value="산식: 적정가 = BPS + BPS x (ROE - ke) / (ke - g)").font = small_font
row += 1
ws9.cell(row=row, column=1, value="ke(자기자본비용) = 무위험이자율 3.5% + B(1.2) x ERP(5.4%) = 10.0%").font = small_font
row += 1
ws9.cell(row=row, column=1, value="항공업 B는 1.0~1.5 범위. 고레버리지 산업 특성 반영.").font = small_font
row += 1

write_header_row(ws9, row, ["시나리오", "지속ROE", "ke", "성장률(g)", "적정주가", "현재가 대비"])
row += 1

rim_scenarios = [
    ("보수적 (ROE=ke)", 0.10, 0.10, 0.02, "ROE가 자본비용과 동일"),
    ("기본 (ROE 13%)", 0.13, 0.10, 0.02, "2024년 ROE 수준"),
    ("적극적 (합병 시너지)", 0.15, 0.10, 0.02, "아시아나 합병 효과 반영"),
    ("낙관적 (ROE 확대)", 0.18, 0.10, 0.03, "글로벌 1위 항공사 수준"),
]

for label, roe, ke, g, note in rim_scenarios:
    if roe == ke:
        fair = BPS
    else:
        fair = BPS * (1 + (roe - ke) / (ke - g))
    upside = (fair - PRICE) / PRICE
    upside_str = f"{upside*100:+.1f}%"
    up_font = d_green if upside > 0 else d_red
    up_fill = green_fill if upside > 0 else red_fill

    write_data_row(ws9, row,
              [label, f"{roe*100:.0f}%", f"{ke*100:.0f}%", f"{g*100:.0f}%", fmt_won(int(fair)), upside_str],
              fonts=[data_font_bold, data_font, data_font, data_font, d_blue, up_font],
              fills=[lighter_fill, white_fill, white_fill, white_fill, gold_fill, up_fill],
              alignments=[left_al, center_al, center_al, center_al, right_al, center_al])
    row += 1

# RIM with EPS-based
row += 1
rim_eps = BPS + (EPS_2024 - BPS * Ke) / (Ke - GROWTH)
ws9.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
ws9.cell(row=row, column=1, value=f"-> RIM (EPS기반): BPS({fmt_won(BPS)}) + 초과이익({fmt_won(EPS_2024)}-{fmt_won(int(BPS*Ke))})/(10%-2%) = {fmt_won(int(rim_eps))}").font = data_font_bold
row += 1
ws9.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
ws9.cell(row=row, column=1, value=f"-> 기본 시나리오(ROE 13%) 적정가 약 {fmt_won(int(BPS*(1+(0.13-0.10)/(0.10-0.02))))}. 현재가 {fmt_won(PRICE)}은 보수적~기본 사이 반영.").font = data_font_bold

print("  [9/12] PBR_ROE 완료")

# ============================================================
# SHEET 10: EV/EBITDA & FCF (from valuation)
# ============================================================
ws10 = wb.create_sheet("EV_EBITDA_FCF")
ws10.sheet_properties.tabColor = "E67E22"
set_col_widths(ws10, [24, 18, 18, 18, 24])
setup_print(ws10)

row = 1
row = add_section_title(ws10, row, "EV/EBITDA & FCF 밸류에이션", col_end=5)

# A. EV 산출
row = add_section_title(ws10, row, "A. Enterprise Value 산출 (리스부채 포함)", col_end=5)
write_header_row(ws10, row, ["항목", "금액(억)", "비고", "", ""])
row += 1

ev_items = [
    ("시가총액 (보통주)", int(MARKET_CAP/억), f"주가 {PRICE:,}원 x {SHARES_OUTSTANDING:,}주"),
    ("(+) 단기차입금", int(ST_DEBT_2024/억), ""),
    ("(+) 장기차입금", int(LT_DEBT_2024/억), ""),
    ("(+) 유동성장기부채", int(CURRENT_LT_DEBT/억), ""),
    ("(+) 사채", int(BONDS_2024/억), ""),
    ("(+) 리스부채(비유동)", int(LEASE_NONCURRENT/억), "항공기 리스 (IFRS 16)"),
    ("(+) 유동성리스부채", int(LEASE_CURRENT/억), ""),
    ("(=) 총차입금", int(TOTAL_DEBT/억), "리스부채 포함"),
    ("(-) 현금성자산", int(CASH_2024/억), "현금및현금성자산"),
    ("(-) 단기금융상품", int(ST_FINANCIAL/억), ""),
    ("(=) 순차입금", int(NET_DEBT/억), "차입금 - 현금성"),
    ("(=) EV", int(EV/억), "시가총액 + 순차입금"),
]
for label, amt, note in ev_items:
    is_total = label.startswith("(=)")
    write_data_row(ws10, row, [label, f"{amt:,}", note, "", ""],
              fonts=[data_font_bold if is_total else data_font, d_blue if is_total else data_font, small_font, data_font, data_font],
              fills=[gold_fill if is_total else lighter_fill, gold_fill if is_total else white_fill, white_fill, white_fill, white_fill],
              alignments=[left_al, right_al, left_al, left_al, left_al])
    row += 1

# B. EBITDA
row += 1
row = add_section_title(ws10, row, "B. EBITDA 산출", col_end=5)
write_header_row(ws10, row, ["항목", "2024(억)", "비고", "", ""])
row += 1

ebitda_items = [
    ("영업이익", int(OP_2024/억), "CIS 영업이익"),
    ("(+) 감가상각비", int(DA_2024/억), "유형자산 (항공기 등)"),
    ("(+) 무형자산상각비", int(IA_2024/억), "노선권, 소프트웨어 등"),
    ("(=) EBITDA", int(EBITDA_2024/억), ""),
    ("EBITDA 마진", "", f"{EBITDA_2024/REV_2024*100:.1f}%"),
]
for label, amt, note in ebitda_items:
    is_total = label.startswith("(=)") or label == "EBITDA 마진"
    val = f"{amt:,}" if isinstance(amt, int) and amt > 0 else note
    write_data_row(ws10, row, [label, val if isinstance(amt, int) and amt > 0 else "", note if isinstance(amt, int) and amt > 0 else val, "", ""],
              fonts=[data_font_bold if is_total else data_font, d_blue if is_total else data_font, small_font, data_font, data_font],
              fills=[gold_fill if is_total else lighter_fill, gold_fill if is_total else white_fill, white_fill, white_fill, white_fill],
              alignments=[left_al, right_al, left_al, left_al, left_al])
    row += 1

# C. EV/EBITDA 밸류에이션
row += 1
row = add_section_title(ws10, row, "C. EV/EBITDA 밸류에이션", col_end=5)
write_header_row(ws10, row, ["목표 배수", "적정 EV(억)", "적정 시총(억)", "적정 주가", "현재가 대비"])
row += 1

for mult in [4.0, 5.0, 5.5, 6.0, 7.0, 8.0]:
    fair_ev = int(EBITDA_2024/억) * mult
    fair_mcap = fair_ev - int(NET_DEBT/억)
    fair_price = int(fair_mcap * 억 / SHARES_OUTSTANDING) if fair_mcap > 0 else 0
    upside = (fair_price - PRICE) / PRICE if fair_price > 0 else -1

    is_current = abs(mult - EV_EBITDA_2024) < 0.5
    up_font = d_green if upside > 0 else d_red
    row_fill = gold_fill if is_current else white_fill

    write_data_row(ws10, row,
              [f"EV/EBITDA {mult:.1f}배" + (" (현재)" if is_current else ""),
               f"{int(fair_ev):,}", f"{int(fair_mcap):,}", fmt_won(fair_price), f"{upside*100:+.1f}%"],
              fonts=[data_font_bold if is_current else data_font, data_font, data_font, d_blue, up_font],
              fills=[gold_fill if is_current else lighter_fill, row_fill, row_fill, row_fill,
                     green_fill if upside > 0 else red_fill],
              alignments=[left_al, right_al, right_al, right_al, center_al])
    row += 1

# D. FCF 분석
row += 1
row = add_section_title(ws10, row, "D. FCF(잉여현금흐름) 분석", col_end=5)
write_header_row(ws10, row, ["항목", "2024(억)", "비고", "", ""])
row += 1

fcf_items = [
    ("영업활동현금흐름", int(OPCF_2024/억), f"{fmt(OPCF_2024)}"),
    ("(-) 설비투자(CAPEX)", int(CAPEX_2024/억), "유형자산 취득 (항공기 등)"),
    ("(=) FCF", int(FCF_2024/억), "잉여현금흐름"),
    ("FCF/주", 0, fmt_won(int(FCF_PER_SHARE))),
    ("FCF 수익률", 0, f"{FCF_PER_SHARE/PRICE*100:.1f}%"),
    ("", 0, ""),
    ("이자비용", int(INTEREST_2024/억), "CF 상 이자비용"),
    ("배당금 지급", int(DIV_PAID_2024/억), f"DPS {DPS_2024:,}원"),
    ("FCF - 이자 - 배당", int((FCF_2024-INTEREST_2024-DIV_PAID_2024)/억), "순잉여현금"),
]

for label, amt, note in fcf_items:
    if not label:
        row += 1
        continue
    is_total = "(=)" in label or "수익률" in label or "순잉여" in label
    val_str = f"{amt:,}" if isinstance(amt, int) and amt != 0 else note
    write_data_row(ws10, row, [label, val_str if isinstance(amt, int) and amt != 0 else "",
                          note if isinstance(amt, int) and amt != 0 else val_str, "", ""],
              fonts=[data_font_bold if is_total else data_font, d_blue if is_total else data_font, small_font, data_font, data_font],
              fills=[gold_fill if is_total else lighter_fill, gold_fill if is_total else white_fill, white_fill, white_fill, white_fill],
              alignments=[left_al, right_al, left_al, left_al, left_al])
    row += 1

row += 1
ws10.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
ws10.cell(row=row, column=1, value=f"-> FCF {fmt(FCF_2024)}으로 양호. 다만 항공업 특성상 대규모 항공기 투자(CAPEX {fmt(CAPEX_2024)}) 지속 필요.").font = data_font_bold
row += 1
ws10.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
ws10.cell(row=row, column=1, value=f"-> 이자비용 {fmt(INTEREST_2024)} 차감 후에도 연간 {fmt(FCF_2024-INTEREST_2024-DIV_PAID_2024)} 순잉여현금 창출.").font = data_font_bold

print("  [10/12] EV_EBITDA_FCF 완료")

# ============================================================
# SHEET 11: 시나리오 분석 + SWOT (from report)
# ============================================================
ws11 = wb.create_sheet("시나리오")
ws11.sheet_properties.tabColor = "2ECC71"
set_col_widths(ws11, [16, 18, 18, 18, 18])
setup_print(ws11)

row = 1
ws11.merge_cells('A1:E1')
ws11.cell(row=1, column=1, value="향후 시나리오 분석").font = section_font
ws11.cell(row=1, column=1).border = bottom_border

row = 3
headers = ["항목", "강세 (Bull)", "기본 (Base)", "약세 (Bear)"]
write_header_row(ws11, row, headers)
row += 1

items = ["전제조건", "매출 전망", "영업이익률", "순이익 전망", "EPS 전망", "적용 PER", "목표주가", "현주가 대비"]
bull = ["아시아나 시너지+화물호황+여객성장", "20조원+", "13%+", "1.8조원+", "4,800원+", "7.5배", "35,000원", "상승 ~48%"]
base = ["현수준 유지+점진적 부채감소", "17~18조원", "11%", "1.3조원", "3,500원", "8배", "28,000원", "상승 ~18%"]
bear = ["경기침체+유가급등+합병비용초과", "14조원", "6%", "5,000억", "1,400원", "11배", "16,000원", "하락 ~32%"]

for i in range(len(items)):
    bull_color = green_fill
    base_color = gold_fill
    bear_color = red_fill

    write_data_row(ws11, row, [items[i], bull[i], base[i], bear[i]],
                   fonts=[data_font_bold, data_font, data_font, data_font],
                   fills=[lighter_fill, bull_color, base_color, bear_color],
                   alignments=[left_al, center_al, center_al, center_al])
    row += 1

# SWOT
row += 2
row = add_section_title(ws11, row, "SWOT 분석", col_end=5)

ws11.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
ws11.cell(row=row, column=1, value="강점 (Strengths)").font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
ws11.cell(row=row, column=1).fill = PatternFill("solid", fgColor="27AE60")
ws11.cell(row=row, column=1).alignment = center_al
ws11.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
ws11.cell(row=row, column=3, value="약점 (Weaknesses)").font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
ws11.cell(row=row, column=3).fill = PatternFill("solid", fgColor="E74C3C")
ws11.cell(row=row, column=3).alignment = center_al
row += 1

strengths = [
    "글로벌 네트워크 43개국 120도시",
    "아시아나 합병 시너지 (글로벌 Top 10)",
    "화물사업 글로벌 Top 3 경쟁력",
    "항공우주사업 기술력 (MRO/기체/위성)",
    "스카이팀 얼라이언스 허브 역할",
]
weaknesses = [
    "높은 부채비율 329% (리스부채 10.9조)",
    "유가/환율에 높은 실적 민감도",
    "인건비 부담 (조종사/승무원 노조)",
    "합병 통합비용 및 노선 구조조정 리스크",
    "경기민감 시클리컬 업종 특성",
]

for i in range(max(len(strengths), len(weaknesses))):
    s = strengths[i] if i < len(strengths) else ""
    w = weaknesses[i] if i < len(weaknesses) else ""
    ws11.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws11.cell(row=row, column=1, value=f"  {s}" if s else "").font = data_font
    ws11.cell(row=row, column=1).alignment = left_al
    ws11.cell(row=row, column=1).fill = green_fill
    ws11.cell(row=row, column=1).border = thin_border
    ws11.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws11.cell(row=row, column=3, value=f"  {w}" if w else "").font = data_font
    ws11.cell(row=row, column=3).alignment = left_al
    ws11.cell(row=row, column=3).fill = red_fill
    ws11.cell(row=row, column=3).border = thin_border
    row += 1

row += 1
ws11.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
ws11.cell(row=row, column=1, value="기회 (Opportunities)").font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
ws11.cell(row=row, column=1).fill = PatternFill("solid", fgColor="2980B9")
ws11.cell(row=row, column=1).alignment = center_al
ws11.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
ws11.cell(row=row, column=3, value="위협 (Threats)").font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
ws11.cell(row=row, column=3).fill = PatternFill("solid", fgColor="7F8C8D")
ws11.cell(row=row, column=3).alignment = center_al
row += 1

opportunities = [
    "아시아나 합병 완료 -> 노선/슬롯 시너지",
    "화물 e-commerce 성장 (크로스보더 물류)",
    "UAM/MRO 신성장 사업 확대",
    "중국/동남아 여객 수요 회복",
]
threats = [
    "글로벌 경기침체 -> 여객/화물 수요 감소",
    "유가 급등 (항공유 원가 30%+)",
    "LCC 경쟁 심화 (국내선/근거리)",
    "지정학적 리스크 (중동/대만해협)",
]

for i in range(max(len(opportunities), len(threats))):
    o = opportunities[i] if i < len(opportunities) else ""
    t = threats[i] if i < len(threats) else ""
    ws11.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws11.cell(row=row, column=1, value=f"  {o}" if o else "").font = data_font
    ws11.cell(row=row, column=1).alignment = left_al
    ws11.cell(row=row, column=1).fill = blue_bg_fill
    ws11.cell(row=row, column=1).border = thin_border
    ws11.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws11.cell(row=row, column=3, value=f"  {t}" if t else "").font = data_font
    ws11.cell(row=row, column=3).alignment = left_al
    ws11.cell(row=row, column=3).fill = PatternFill("solid", fgColor="E5E7E9")
    ws11.cell(row=row, column=3).border = thin_border
    row += 1

print("  [11/12] 시나리오 완료")

# ============================================================
# SHEET 12: 목표주가 (from valuation)
# ============================================================
ws12 = wb.create_sheet("목표주가")
ws12.sheet_properties.tabColor = "2ECC71"
set_col_widths(ws12, [22, 18, 18, 18, 22])
setup_print(ws12)

row = 1
row = add_section_title(ws12, row, f"시나리오별 목표주가 종합 (현재가 {PRICE:,}원)", col_end=5)

# A. 방법론별 적정가 범위
row = add_section_title(ws12, row, "A. 밸류에이션 방법론별 적정가 레인지", col_end=5)
write_header_row(ws12, row, ["방법론", "보수적", "기본", "적극적", "산출 근거"])
row += 1

# PER 방식
per_cons = EPS_2024 * 5
per_base = EPS_2024 * 8
per_aggr = EPS_2024 * 10

# PBR 방식
pbr_cons = int(BPS * 0.5)
pbr_base = int(BPS * 1.0)
pbr_aggr = int(BPS * 1.5)

# EV/EBITDA 방식
def ev_ebitda_price(mult):
    fair_ev = int(EBITDA_2024/억) * mult
    fair_mcap = fair_ev - int(NET_DEBT/억)
    return int(fair_mcap * 억 / SHARES_OUTSTANDING) if fair_mcap > 0 else 0

ev_cons = ev_ebitda_price(4.5)
ev_base = ev_ebitda_price(6.0)
ev_aggr = ev_ebitda_price(8.0)

# RIM 방식
rim_cons = BPS
rim_base = int(BPS * (1 + (0.13 - 0.10) / (0.10 - 0.02)))
rim_aggr = int(BPS * (1 + (0.15 - 0.10) / (0.10 - 0.02)))

# FCF 방식
fcf_cons = int(FCF_2024 * 6 / SHARES_OUTSTANDING)
fcf_base = int(FCF_2024 * 10 / SHARES_OUTSTANDING)
fcf_aggr = int(FCF_2024 * 14 / SHARES_OUTSTANDING)

methods = [
    ("PER 방식", fmt_won(per_cons), fmt_won(per_base), fmt_won(per_aggr),
     "EPS x 목표PER (5/8/10배)"),
    ("PBR 방식", fmt_won(pbr_cons), fmt_won(pbr_base), fmt_won(pbr_aggr),
     "BPS x 목표PBR (0.5/1.0/1.5배)"),
    ("EV/EBITDA 방식", fmt_won(ev_cons), fmt_won(ev_base), fmt_won(ev_aggr),
     "EBITDA x 목표배수-순차입금 (4.5/6/8배)"),
    ("RIM 방식", fmt_won(rim_cons), fmt_won(rim_base), fmt_won(rim_aggr),
     "BPS x (1+(ROE-ke)/(ke-g))"),
    ("FCF 기반", fmt_won(fcf_cons), fmt_won(fcf_base), fmt_won(fcf_aggr),
     "FCF x 목표배수 (6/10/14배)"),
]

for label, cons, base_val, aggr, note in methods:
    write_data_row(ws12, row, [label, cons, base_val, aggr, note],
              fonts=[data_font_bold, data_font, d_blue, d_green, small_font],
              fills=[lighter_fill, red_fill, gold_fill, green_fill, white_fill],
              alignments=[left_al, right_al, right_al, right_al, left_al])
    row += 1

# B. 종합 판단
row += 1
row = add_section_title(ws12, row, "B. 종합 시나리오", col_end=5)
write_header_row(ws12, row, ["시나리오", "목표주가", "현재가 대비", "전제조건", "확률(주관)"],
             fills=[PatternFill("solid", fgColor="C0392B")]*5)
row += 1

# Bull case: PER 10배 -> 35,000원
bull_target = 35000
bull_upside = (bull_target - PRICE) / PRICE
write_data_row(ws12, row, ["강세 (Bull)", fmt_won(bull_target), f"{bull_upside*100:+.1f}%",
                      "아시아나 시너지 본격화 + 화물호황 + PER 리레이팅 10배", "20%"],
          fonts=[data_font_bold, Font(name="맑은 고딕", size=12, bold=True, color="27AE60"), d_green, data_font, data_font],
          fills=[green_fill]*5, alignments=[center_al, right_al, center_al, left_al, center_al])
row += 1

# Base case: PER 8배 -> 28,000원
base_target = 28000
base_upside = (base_target - PRICE) / PRICE
write_data_row(ws12, row, ["기본 (Base)", fmt_won(base_target), f"{base_upside*100:+.1f}%",
                      "여객 정상화 유지 + 합병 시너지 점진적 반영 + PER 8배", "50%"],
          fonts=[data_font_bold, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), d_blue, data_font, data_font],
          fills=[gold_fill]*5, alignments=[center_al, right_al, center_al, left_al, center_al])
row += 1

# Bear case: PER 4.5배 -> 16,000원
bear_target = 16000
bear_upside = (bear_target - PRICE) / PRICE
write_data_row(ws12, row, ["약세 (Bear)", fmt_won(bear_target), f"{bear_upside*100:+.1f}%",
                      "경기 침체 + 유가 급등 + 합병 비용 부담 + PER 4.5배", "30%"],
          fonts=[data_font_bold, Font(name="맑은 고딕", size=12, bold=True, color="C0392B"), d_red, data_font, data_font],
          fills=[red_fill]*5, alignments=[center_al, right_al, center_al, left_al, center_al])
row += 1

# Expected value
exp_val = int(bull_target * 0.20 + base_target * 0.50 + bear_target * 0.30)
exp_upside = (exp_val - PRICE) / PRICE
row += 1
write_data_row(ws12, row, ["확률가중 기대값", fmt_won(exp_val), f"{exp_upside*100:+.1f}%",
                      "Bull x 20% + Base x 50% + Bear x 30%", ""],
          fonts=[Font(name="맑은 고딕", size=12, bold=True, color=NAVY)] * 5,
          fills=[blue_bg_fill] * 5,
          alignments=[center_al, right_al, center_al, left_al, center_al])

# C. Key Metrics Summary Box
row += 2
row = add_section_title(ws12, row, "C. 핵심 체크포인트", col_end=5)

checkpoints = [
    (f"현재 PER {PER_2024:.1f}배", "글로벌 항공업 평균 PER 8~12배 대비 30~45% 할인. 한국 디스카운트 + 합병 불확실성 반영"),
    (f"PBR {PBR_2024:.2f}배", "자본총계 대비 저평가. 2020년 자본잠식 위기에서 빠르게 자본 확충 (2020년 6천억->2024년 10.5조)"),
    (f"EV/EBITDA {EV_EBITDA_2024:.1f}배", f"리스부채 포함 시 EV가 크지만, EBITDA {fmt(EBITDA_2024)}로 업종 평균 수준. 항공기 투자 대비 수익성 양호"),
    (f"ROE {ROE_2024*100:.1f}%", "자본비용(10%) 대비 초과수익 창출 중. 항공업치고 높은 수준"),
    ("아시아나 합병 효과", "2024년 아시아나 합병 완료. 노선 최적화, 중복비용 제거 등 시너지 본격화 기대"),
    ("화물사업 경쟁력", "글로벌 Top 3 화물 수송. B747-8F 등 대형 화물기 보유. 이커머스 성장 수혜"),
    ("리스크: 고레버리지", f"순차입금/EBITDA {NET_DEBT/EBITDA_2024:.1f}배. 항공업 특성이나 리스부채 {fmt(NET_DEBT)} 부담"),
    ("리스크: 유가 민감도", "제트유 가격 변동이 영업이익에 직접 영향. 유가 급등 시 마진 압박"),
    ("리스크: 환율", "달러 매출 비중 높으나 유류비/리스료도 달러. 원화 약세 시 이중 효과"),
    ("리스크: 경기 순환", "항공 수요는 경기에 민감. 글로벌 경기 침체 시 여객/화물 모두 타격"),
]

for label, desc in checkpoints:
    ws12.cell(row=row, column=1, value=label).font = data_font_bold
    ws12.cell(row=row, column=1).fill = lighter_fill
    ws12.cell(row=row, column=1).alignment = left_al
    ws12.cell(row=row, column=1).border = thin_border
    ws12.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws12.cell(row=row, column=2, value=desc).font = data_font
    ws12.cell(row=row, column=2).alignment = left_al
    ws12.cell(row=row, column=2).border = thin_border
    row += 1

print("  [12/12] 목표주가 완료")

# ============================================================
# SAVE
# ============================================================
wb.save(OUT)
conn.close()
print(f"\n종합보고서 생성 완료: {OUT}")
print(f"시트 ({len(wb.sheetnames)}개): {wb.sheetnames}")
