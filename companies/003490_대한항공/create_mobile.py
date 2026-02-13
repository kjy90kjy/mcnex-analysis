# -*- coding: utf-8 -*-
"""대한항공 모바일용 보고서 — 좁은 화면 최적화 (3열, 큰 글씨, 세로 스크롤)"""
import sqlite3, sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ai.db")
conn = sqlite3.connect(DB)

# === CONSTANTS ===
PRICE = 23700; SHARES = 369331403; SHARES_FLOAT = 368220612; SHARES_TREASURY = 1110791
억 = 100_000_000; MCAP = PRICE * SHARES
EPS24 = 3566; BPS24 = 28400; DPS24 = 750
REV24 = 178707 * 억; OP24 = 21102 * 억; NI24 = 13818 * 억
EQ24 = 109631 * 억; EQ23 = 98152 * 억; LIAB24 = 360488 * 억; TA24 = 470120 * 억
OPCF24 = 46000 * 억; CAPEX24 = 29000 * 억; FCF24 = OPCF24 - CAPEX24
EBITDA24 = OP24 + 15000 * 억  # 감가상각 약 1.5조 추정
NET_DEBT = 109000 * 억  # 순차입금 약 10.9조 (리스부채 포함)
EV = MCAP + NET_DEBT
ROE24 = NI24 / ((EQ24 + EQ23) / 2)

def fmt(v): return f"{v/억:,.0f}억"
def fw(v): return f"{v:,.0f}원"
def pct(v): return f"{v*100:.1f}%"

# === MOBILE STYLES ===
NAVY="1B2A4A"; DARK="2C3E6B"; W="FFFFFF"; GOLD_C="D4A843"; RED_C="C0392B"; GREEN_C="27AE60"

t1=Font(name="맑은 고딕",size=16,bold=True,color=W)
t2=Font(name="맑은 고딕",size=14,bold=True,color=NAVY)
hf_=Font(name="맑은 고딕",size=11,bold=True,color=W)
df_=Font(name="맑은 고딕",size=11)
db_=Font(name="맑은 고딕",size=11,bold=True)
bl_=Font(name="맑은 고딕",size=11,bold=True,color="0000FF")
gn_=Font(name="맑은 고딕",size=11,bold=True,color=GREEN_C)
rd_=Font(name="맑은 고딕",size=11,bold=True,color=RED_C)
sm_=Font(name="맑은 고딕",size=10,color="666666")
big_=Font(name="맑은 고딕",size=13,bold=True,color=NAVY)
huge_=Font(name="맑은 고딕",size=18,bold=True,color=NAVY)

tf_=PatternFill("solid",fgColor=NAVY)
hfl=PatternFill("solid",fgColor=DARK)
lf_=PatternFill("solid",fgColor="D6E4F0")
llf_=PatternFill("solid",fgColor="EBF1F8")
wf_=PatternFill("solid",fgColor=W)
gld_=PatternFill("solid",fgColor="FFF3CD")
gnf_=PatternFill("solid",fgColor="D4EDDA")
rdf_=PatternFill("solid",fgColor="F8D7DA")
blf_=PatternFill("solid",fgColor="D6EAF8")
grf_=PatternFill("solid",fgColor="F2F2F2")

ca_=Alignment(horizontal='center',vertical='center',wrap_text=True)
la_=Alignment(horizontal='left',vertical='center',wrap_text=True)
ra_=Alignment(horizontal='right',vertical='center',wrap_text=True)
tb_=Border(left=Side('thin',color="D0D0D0"),right=Side('thin',color="D0D0D0"),
           top=Side('thin',color="D0D0D0"),bottom=Side('thin',color="D0D0D0"))
bb_=Border(bottom=Side('medium',color=NAVY))

COL3 = [14, 18, 18]

def sw(ws, w):
    for i, v in enumerate(w, 1): ws.column_dimensions[get_column_letter(i)].width = v

def setup_print(ws):
    ws.page_setup.paperSize = Worksheet.PAPERSIZE_LETTER
    ws.page_setup.orientation = Worksheet.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.75, bottom=0.75,
        header=0.3, footer=0.3
    )
    ws.print_options.horizontalCentered = True

def mhdr(ws, r, vals):
    """모바일 헤더"""
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = hf_; c.fill = hfl; c.alignment = ca_; c.border = tb_

def mrow(ws, r, vals, fonts=None, fills=None, als=None, h=None):
    """모바일 데이터 행"""
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = fonts[i-1] if fonts else df_
        c.fill = fills[i-1] if fills else wf_
        c.alignment = als[i-1] if als else ca_
        c.border = tb_
    if h: ws.row_dimensions[r].height = h

def msec(ws, r, title, cols=3):
    """모바일 섹션 타이틀"""
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
    c = ws.cell(row=r, column=1, value=title)
    c.font = t2; c.border = bb_
    ws.row_dimensions[r].height = 28
    return r + 1

def minfo(ws, r, label, value, cols=3):
    """모바일 라벨-값 쌍"""
    ws.cell(row=r, column=1, value=label).font = db_
    ws.cell(row=r, column=1).fill = llf_; ws.cell(row=r, column=1).alignment = la_; ws.cell(row=r, column=1).border = tb_
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=cols)
    ws.cell(row=r, column=2, value=value).font = bl_
    ws.cell(row=r, column=2).fill = wf_; ws.cell(row=r, column=2).alignment = ra_; ws.cell(row=r, column=2).border = tb_
    ws.row_dimensions[r].height = 24
    return r + 1

wb = Workbook()

# ============================================================
# 단일 시트: 모든 내용을 세로로 길게 배치
# ============================================================
ws = wb.active
ws.title = "대한항공 분석"
ws.sheet_properties.tabColor = NAVY
sw(ws, COL3)
setup_print(ws)

row = 1

# ---- TITLE ----
ws.merge_cells('A1:C2')
c = ws.cell(row=1, column=1, value="대한항공(003490)\n모바일 요약 보고서")
c.font = t1; c.fill = tf_; c.alignment = ca_
ws.row_dimensions[1].height = 24; ws.row_dimensions[2].height = 24
row = 3
ws.merge_cells('A3:C3')
ws.cell(row=3, column=1, value=f"현재가 {PRICE:,}원 | 시총 {fmt(MCAP)} | 2026.02.13").font = Font(name="맑은 고딕", size=11, bold=True, color=DARK)
ws.cell(row=3, column=1).alignment = ca_; ws.row_dimensions[3].height = 22

# ==== 1. 기본정보 ====
row = 5
row = msec(ws, row, "기본 정보")
for lbl, val in [
    ("현재주가", fw(PRICE)), ("시가총액", fmt(MCAP)),
    ("유통시총", fmt(PRICE * SHARES_FLOAT)),
    ("발행주식수", f"{SHARES:,}주"), ("자기주식", f"{SHARES_TREASURY:,}주"),
    ("업종", "항공운송업 (FSC)"),
    ("주요사업", "여객운송, 화물운송, 항공우주사업"),
    ("대표이사", "조원태(회장), 우기홍(부회장)"),
    ("설립", "1962.06.19"),
    ("글로벌 네트워크", "43개국 120개 도시 (스카이팀)")]:
    row = minfo(ws, row, lbl, val)

# ==== 2. 핵심 밸류에이션 ====
row += 1
row = msec(ws, row, "핵심 밸류에이션")
mhdr(ws, row, ["지표", "값", "판정"]); row += 1

vals = [
    ("PER (2024)", f"{PRICE/EPS24:.1f}배", "저평가"),
    ("PBR", f"{PRICE/BPS24:.2f}배", "저평가"),
    ("EV/EBITDA", f"{EV/EBITDA24:.1f}배", "적정"),
    ("ROE", pct(ROE24), "양호"),
    ("영업이익률", pct(OP24/REV24), "양호"),
    ("배당수익률", pct(DPS24/PRICE), "양호"),
    ("FCF수익률", pct(FCF24/SHARES/PRICE), "양호"),
    ("부채비율", pct(LIAB24/EQ24), "주의"),
    ("리스제외 부채비율", "약 180%", "적정"),
]
for lbl, val, judge in vals:
    if "매우" in judge: jf, jfl = gn_, gnf_
    elif "저" in judge or "양호" in judge: jf, jfl = Font(name="맑은 고딕",size=11,bold=True,color="2E86C1"), blf_
    elif "적정" in judge: jf, jfl = db_, gld_
    else: jf, jfl = rd_, rdf_
    mrow(ws, row, [lbl, val, judge], fonts=[db_, bl_, jf], fills=[llf_, gld_, jfl], als=[la_, ca_, ca_], h=24)
    row += 1

# ==== 3. 2024 확정실적 ====
row += 1
row = msec(ws, row, "2024년 확정 실적")
for lbl, val in [
    ("매출액", f"{178707:,}억 (역대 최대)"), ("영업이익", f"{21102:,}억 (OPM 11.8%)"),
    ("순이익", f"{13818:,}억"), ("EPS", fw(EPS24)), ("BPS", fw(BPS24)),
    ("DPS", f"{fw(DPS24)} (보통주)"),
    ("총자산", f"{470120:,}억"), ("총부채", f"{360488:,}억"),
    ("총자본", f"{109631:,}억"), ("영업CF", f"약 46,000억"),
    ("FCF", f"약 17,000억"), ("EBITDA", f"약 36,102억")]:
    row = minfo(ws, row, lbl, val)

# ==== 4. 5개년 실적 ====
row += 1
row = msec(ws, row, "최근 5개년 실적 (억원)")
mhdr(ws, row, ["연도", "매출/영업이익", "순이익/EPS"]); row += 1
hist5 = [
    (2020, 76062, 1089, -2301, -624),
    (2021, 90168, 14179, 5787, 1567),
    (2022, 140960, 28305, 17295, 4684),
    (2023, 161117, 17900, 11291, 3057),
    (2024, 178707, 21102, 13818, 3566),
]
for yr, rev, op, ni, eps in hist5:
    nif = gn_ if ni > 0 else rd_
    mrow(ws, row, [str(yr), f"{rev:,} / {op:,}", f"{ni:,} / {eps:,}"],
         fonts=[db_, df_, nif], fills=[llf_, wf_, gnf_ if ni > 0 else rdf_],
         als=[ca_, ra_, ra_]); row += 1

# Key points
row += 1
for pt in ["2024 매출 17.9조 역대 최대 (아시아나 합병 효과)",
           "2022 영업이익 2.8조 피크 → 2024 2.1조 회복 추세",
           "2020 코로나 적자 → 2021~24 4년 연속 흑자"]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"• {pt}").font = df_; ws.cell(row=row, column=1).alignment = la_
    ws.row_dimensions[row].height = 22; row += 1

# ==== 5. 사업부문 ====
row += 1
row = msec(ws, row, "사업부문별 매출 (2024, 억원)")
mhdr(ws, row, ["사업부문", "매출액", "비중"]); row += 1
segments = [
    ("여객운송", 97786, "61%"),
    ("화물운송", 44116, "27%"),
    ("항공우주", 5930, "3.7%"),
    ("호텔/기타", 13334, "8.3%"),
    ("합계", 161166, "100%"),
]
for nm, rev, pct_val in segments:
    is_total = nm == "합계"
    fl = gld_ if is_total else wf_
    f = db_ if is_total else df_
    mrow(ws, row, [nm, f"{rev:,}", pct_val],
         fonts=[db_ if is_total else db_, f, f],
         fills=[llf_ if not is_total else gld_, fl, fl],
         als=[la_, ra_, ca_]); row += 1

# Subsidiaries
row += 1
for pt in ["아시아나항공: 63.88% (2024 합병 완료, 글로벌 Top 10 도약)",
           "진에어: 54.91% (LCC 저비용항공사)",
           "한국공항: 59.54% (인천/김포 지상조업)",
           "화물기: 약 23대 보유 (글로벌 화물 Top 3)"]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"• {pt}").font = df_; ws.cell(row=row, column=1).alignment = la_
    ws.row_dimensions[row].height = 22; row += 1

# ==== 6. 밸류에이션 - PER/PBR 목표가 ====
row += 1
row = msec(ws, row, "PER 기반 목표주가")
mhdr(ws, row, ["PER배수", "적정주가", "현재가대비"]); row += 1
for m in [6, 7, 8, 10, 12]:
    tp = EPS24 * m
    up = (tp - PRICE) / PRICE
    uf = gn_ if up > 0 else rd_; ufl = gnf_ if up > 0 else rdf_
    mrow(ws, row, [f"{m}배", fw(tp), f"{up*100:+.1f}%"],
         fonts=[db_, bl_, uf], fills=[llf_, gld_, ufl],
         als=[ca_, ra_, ca_]); row += 1

row += 1
row = msec(ws, row, "PBR 기반 목표주가")
mhdr(ws, row, ["PBR배수", "적정주가", "현재가대비"]); row += 1
for m_pct in [0.6, 0.8, 1.0, 1.2]:
    tp = int(BPS24 * m_pct)
    up = (tp - PRICE) / PRICE
    uf = gn_ if up > 0 else rd_; ufl = gnf_ if up > 0 else rdf_
    mrow(ws, row, [f"{m_pct:.1f}배", fw(tp), f"{up*100:+.1f}%"],
         fonts=[db_, bl_, uf], fills=[llf_, gld_, ufl],
         als=[ca_, ra_, ca_]); row += 1

# ==== 7. SWOT 분석 ====
row += 1
row = msec(ws, row, "SWOT 분석")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="강점 (S)").font = Font(name="맑은 고딕",size=12,bold=True,color=W)
ws.cell(row=row, column=1).fill = PatternFill("solid",fgColor=GREEN_C); ws.cell(row=row, column=1).alignment = ca_; row += 1
for s in ["글로벌 43개국 120도시 네트워크",
          "아시아나 합병 → 글로벌 Top 10 FSC",
          "화물사업 글로벌 Top 3 경쟁력",
          "항공우주사업 기술력 (MRO/기체/위성)",
          "스카이팀 얼라이언스 핵심 허브"]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"• {s}").font = df_; ws.cell(row=row, column=1).fill = gnf_
    ws.cell(row=row, column=1).alignment = la_; ws.cell(row=row, column=1).border = tb_; row += 1

ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="약점 (W)").font = Font(name="맑은 고딕",size=12,bold=True,color=W)
ws.cell(row=row, column=1).fill = PatternFill("solid",fgColor=RED_C); ws.cell(row=row, column=1).alignment = ca_; row += 1
for w_ in ["부채비율 329% (리스부채 10.9조 포함)",
           "유가/환율에 높은 실적 민감도",
           "인건비 부담 (조종사/승무원 노조)",
           "합병 통합비용 및 노선 구조조정 리스크",
           "경기민감 시클리컬 업종"]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"• {w_}").font = df_; ws.cell(row=row, column=1).fill = rdf_
    ws.cell(row=row, column=1).alignment = la_; ws.cell(row=row, column=1).border = tb_; row += 1

ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="기회 (O)").font = Font(name="맑은 고딕",size=12,bold=True,color=W)
ws.cell(row=row, column=1).fill = PatternFill("solid",fgColor="2980B9"); ws.cell(row=row, column=1).alignment = ca_; row += 1
for o_ in ["아시아나 합병 시너지 (노선/슬롯/비용절감)",
           "화물 e-commerce 성장 (크로스보더 물류)",
           "UAM/MRO 신성장 사업 확대",
           "중국/동남아 여객 수요 회복"]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"• {o_}").font = df_; ws.cell(row=row, column=1).fill = blf_
    ws.cell(row=row, column=1).alignment = la_; ws.cell(row=row, column=1).border = tb_; row += 1

ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="위협 (T)").font = Font(name="맑은 고딕",size=12,bold=True,color=W)
ws.cell(row=row, column=1).fill = PatternFill("solid",fgColor="7F8C8D"); ws.cell(row=row, column=1).alignment = ca_; row += 1
for t_ in ["글로벌 경기침체 → 여객/화물 감소",
           "유가 급등 (항공유 원가 30%+)",
           "LCC 경쟁 심화 (국내선/근거리)",
           "지정학적 리스크 (중동/대만해협)"]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"• {t_}").font = df_; ws.cell(row=row, column=1).fill = PatternFill("solid",fgColor="E5E7E9")
    ws.cell(row=row, column=1).alignment = la_; ws.cell(row=row, column=1).border = tb_; row += 1

# ==== 8. 시나리오별 목표주가 ====
row += 1
row = msec(ws, row, "시나리오별 목표주가")

bull_t = 35000; base_t = 28000; bear_t = 16000
exp_v = int(bull_t * 0.2 + base_t * 0.5 + bear_t * 0.3)

mhdr(ws, row, ["시나리오", "목표주가", "현재가대비"]); row += 1
mrow(ws, row, ["강세(Bull)", fw(bull_t), f"{(bull_t-PRICE)/PRICE*100:+.1f}%"],
     fonts=[db_, Font(name="맑은 고딕",size=12,bold=True,color=GREEN_C), gn_], fills=[gnf_]*3, h=28); row += 1
mrow(ws, row, ["기본(Base)", fw(base_t), f"{(base_t-PRICE)/PRICE*100:+.1f}%"],
     fonts=[db_, Font(name="맑은 고딕",size=12,bold=True,color=NAVY), bl_], fills=[gld_]*3, h=28); row += 1
mrow(ws, row, ["약세(Bear)", fw(bear_t), f"{(bear_t-PRICE)/PRICE*100:+.1f}%"],
     fonts=[db_, Font(name="맑은 고딕",size=12,bold=True,color=RED_C), rd_], fills=[rdf_]*3, h=28); row += 1
row += 1
mrow(ws, row, ["기대값", fw(exp_v), f"{(exp_v-PRICE)/PRICE*100:+.1f}%"],
     fonts=[Font(name="맑은 고딕",size=12,bold=True,color=NAVY)]*3, fills=[blf_]*3, h=28); row += 1

# 전제조건
row += 1
for sc, cond in [("Bull", "아시아나 시너지+화물호황+여객성장 → 매출20조+, OPM13%+, EPS4,800+, PER7.5배"),
                  ("Base", "현수준 유지+점진적 부채감소 → 매출17~18조, OPM11%, EPS3,500, PER8배"),
                  ("Bear", "경기침체+유가급등+합병비용초과 → 매출14조, OPM6%, EPS1,400, PER11배")]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"{sc}: {cond}").font = df_
    ws.cell(row=row, column=1).alignment = la_; ws.row_dimensions[row].height = 26; row += 1

# ==== 9. 배당 이력 ====
row += 1
row = msec(ws, row, "배당 이력 (보통주)")
mhdr(ws, row, ["연도", "DPS(원)", "배당률"]); row += 1
for yr, dps, yl in [("2017",250,"0.7%"),("2018",0,"무배당"),("2019",0,"무배당"),
                     ("2020",0,"무배당"),("2021",500,"1.6%"),("2022",750,"2.4%"),
                     ("2023",750,"3.4%"),("2024",750,f"{750/PRICE*100:.1f}%")]:
    f_dps = gn_ if (isinstance(dps, int) and dps > 0) else rd_
    mrow(ws, row, [yr, f"{dps:,}" if isinstance(dps, int) and dps > 0 else str(dps), yl],
         fonts=[db_, f_dps, df_], fills=[llf_, gld_ if (isinstance(dps, int) and dps > 0) else rdf_, wf_]); row += 1

# ==== 10. 10년 실적 요약 ====
row += 1
row = msec(ws, row, "10년 실적 요약 (억원)")
mhdr(ws, row, ["연도", "매출/영업이익", "순이익/EPS"]); row += 1
hist = [
    (2015, 115448, 8830, -5630, -1524),
    (2016, 117318, 11208, -5569, -1507),
    (2017, 120922, 9397, 8018, 2171),
    (2018, 130116, 6239, -1987, -538),
    (2019, 126834, 2574, -6228, -1687),
    (2020, 76062, 1089, -2301, -624),
    (2021, 90168, 14179, 5787, 1567),
    (2022, 140960, 28305, 17295, 4684),
    (2023, 161117, 17900, 11291, 3057),
    (2024, 178707, 21102, 13818, 3566),
]
for yr, rev, op, ni, eps in hist:
    nif = gn_ if ni > 0 else rd_
    mrow(ws, row, [str(yr), f"{rev:,} / {op:,}", f"{ni:,} / {eps:,}"],
         fonts=[db_, df_, nif], fills=[llf_, wf_, gnf_ if ni > 0 else rdf_],
         als=[ca_, ra_, ra_]); row += 1

# ==== 11. 핵심 모니터링 ====
row += 1
row = msec(ws, row, "핵심 모니터링 항목")
monitors = [
    ("1", "아시아나 합병 시너지", "노선 통합/슬롯 재배분/인력 구조조정. 통합비용 vs 시너지 규모"),
    ("2", "유가/환율 추이", "유가 $10 변동→영업이익 2~3천억. 환율 100원→약 1,500억 영향"),
    ("3", "분기실적 (여객/화물)", "RPK/FTK/탑승률 추이. e-commerce 화물 물동량 확인"),
    ("4", "부채감축 속도", "부채비율 329%→300% 이하 목표. 리스부채 vs 순차입금"),
    ("5", "배당정책/주주환원", "보통주 750원 유지 여부. 합병 부담으로 배당 축소 가능성"),
]
for rank, title, detail in monitors:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"{rank}. {title}").font = db_
    ws.cell(row=row, column=1).fill = gld_; ws.cell(row=row, column=1).alignment = la_; ws.cell(row=row, column=1).border = tb_
    ws.row_dimensions[row].height = 22; row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=detail).font = df_
    ws.cell(row=row, column=1).alignment = la_; ws.cell(row=row, column=1).border = tb_
    ws.row_dimensions[row].height = 36; row += 1

# Calendar
row += 1
row = msec(ws, row, "모니터링 캘린더")
mhdr(ws, row, ["시기", "이벤트", "중요도"]); row += 1
for t_, e_, imp in [("1월","연간 잠정실적 공시","★★★"),("2~3월","사업보고서/주총/배당확정","★★★"),
                     ("5월","1Q 잠정실적","★★☆"),("6~7월","하계 성수기 트래픽","★★★"),
                     ("8월","2Q 잠정실적","★★☆"),("11월","3Q 잠정실적","★★☆"),
                     ("수시","유가/환율 급변동","★★★"),("수시","항공기 발주/인도","★★☆")]:
    mrow(ws, row, [t_, e_, imp], fonts=[db_, df_, db_], fills=[llf_, wf_, gld_]); row += 1

# ==== Footer ====
row += 2
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="* 데이터: OpenDART 공시 전수분석 | 분석일: 2026.02.13").font = sm_
ws.cell(row=row, column=1).alignment = ca_
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="* 본 보고서는 투자 참고용이며 투자판단의 책임은 투자자에게 있습니다").font = sm_
ws.cell(row=row, column=1).alignment = ca_

# === SAVE ===
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "대한항공_모바일용.xlsx")
wb.save(OUT)
conn.close()
print(f"모바일용 보고서 생성: {OUT}")
print(f"총 {row}행, 3열 구성")
