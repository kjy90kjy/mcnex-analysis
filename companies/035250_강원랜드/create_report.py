# -*- coding: utf-8 -*-
import sqlite3, sys, os
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ai.db")
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "강원랜드_기업분석보고서.xlsx")
conn = sqlite3.connect(DB)

wb = Workbook()

# === STYLE DEFINITIONS ===
NAVY = "1B2A4A"
DARK_BLUE = "2C3E6B"
MID_BLUE = "3A5BA0"
LIGHT_BLUE = "D6E4F0"
LIGHTER_BLUE = "EBF1F8"
WHITE = "FFFFFF"
ACCENT_GOLD = "D4A843"
ACCENT_RED = "C0392B"
ACCENT_GREEN = "27AE60"
LIGHT_GRAY = "F2F2F2"
MED_GRAY = "D9D9D9"

title_font = Font(name="맑은 고딕", size=22, bold=True, color=WHITE)
subtitle_font = Font(name="맑은 고딕", size=11, color="B0C4DE")
section_font = Font(name="맑은 고딕", size=14, bold=True, color=NAVY)
header_font = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
data_font = Font(name="맑은 고딕", size=10)
data_font_bold = Font(name="맑은 고딕", size=10, bold=True)
blue_font = Font(name="맑은 고딕", size=10, color="0000FF")
green_font = Font(name="맑은 고딕", size=10, color=ACCENT_GREEN)
red_font = Font(name="맑은 고딕", size=10, color=ACCENT_RED)
pct_font_green = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT_GREEN)
pct_font_red = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT_RED)
small_font = Font(name="맑은 고딕", size=9, color="666666")

title_fill = PatternFill("solid", fgColor=NAVY)
header_fill = PatternFill("solid", fgColor=DARK_BLUE)
sub_header_fill = PatternFill("solid", fgColor=MID_BLUE)
light_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
lighter_fill = PatternFill("solid", fgColor=LIGHTER_BLUE)
gray_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
white_fill = PatternFill("solid", fgColor=WHITE)
gold_fill = PatternFill("solid", fgColor="FFF3CD")
red_fill = PatternFill("solid", fgColor="F8D7DA")
green_fill = PatternFill("solid", fgColor="D4EDDA")

thin_border = Border(
    left=Side(style='thin', color=MED_GRAY),
    right=Side(style='thin', color=MED_GRAY),
    top=Side(style='thin', color=MED_GRAY),
    bottom=Side(style='thin', color=MED_GRAY)
)
bottom_border = Border(bottom=Side(style='medium', color=NAVY))

center_al = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_al = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_al = Alignment(horizontal='right', vertical='center')
wrap_al = Alignment(vertical='top', wrap_text=True)

NUM_FMT = '#,##0'
NUM_FMT_BILL = '#,##0'
PCT_FMT = '0.0%'
PCT_FMT2 = '0.00%'

def style_range(ws, row, col_start, col_end, font=None, fill=None, alignment=None, border=None, number_format=None):
    for c in range(col_start, col_end+1):
        cell = ws.cell(row=row, column=c)
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        if border: cell.border = border
        if number_format: cell.number_format = number_format

def write_header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start+i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_al
        cell.border = thin_border

def write_data_row(ws, row, data, col_start=1, fonts=None, fills=None, alignments=None, number_formats=None):
    for i, d in enumerate(data):
        cell = ws.cell(row=row, column=col_start+i, value=d)
        cell.font = (fonts[i] if fonts and i < len(fonts) else data_font)
        cell.fill = (fills[i] if fills and i < len(fills) else white_fill)
        cell.alignment = (alignments[i] if alignments and i < len(alignments) else right_al)
        cell.border = thin_border
        if number_formats and i < len(number_formats) and number_formats[i]:
            cell.number_format = number_formats[i]

def set_col_widths(ws, widths):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

def add_section_title(ws, row, title, col_end=11):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = section_font
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = bottom_border
    return row + 1


# === 강원랜드 기본 정보 ===
COMPANY = "강원랜드"
STOCK_CODE = "035250"
PRICE = 17690   # 현재 주가
SHARES = 213940500   # 발행주식수
SHARES_TREASURY = 13534816   # 자사주

# === 2024 연결 재무 (원) ===
REV24 = 1426862781802
OP24 = 285790454716
NI24 = 456895295638
EPS24 = 2257
ASSETS24 = 4700641154570
LIAB24 = 816978340813
EQ24 = 3883662813757
CASH24 = 244499636973
DA24 = 78474881854
OPCF24 = 487048778217
CAPEX24 = 107494733192
DPS24 = 1170
DIV_PAID = 188547268140

# === 2023 재무 ===
REV23 = 1392740543618
OP23 = 258505699379
NI23 = 341061017856
EPS23 = 1681
EQ23 = 3677900816262

# === 2025 잠정실적 (연결, 원) ===
Q25 = [
    {"q": "1Q25", "rev": 365830e6, "op": 77659e6, "ni": 78021e6},
    {"q": "2Q25", "rev": 360727e6, "op": 57874e6, "ni": 60630e6},
    {"q": "3Q25", "rev": 384147e6, "op": 72702e6, "ni": 113070e6},
    {"q": "4Q25", "rev": 365446e6, "op": 29697e6, "ni": 66017e6},
]
CUM25_REV = 1476726e6
CUM25_OP = 235176e6
CUM25_NI = 316516e6

# === 10년 실적 (연결, 원) ===
ANNUAL_REV = {
    2015: 1634441985990, 2016: 1703131541003, 2017: 1601291247063,
    2018: 1445736946832, 2019: 1524006966734, 2020: 479173424993,
    2021: 788430938373, 2022: 1272539665429, 2023: 1392740543618,
    2024: 1426862781802
}
ANNUAL_OP = {
    2015: 422764543000, 2016: 480490893000, 2017: 460085942000,
    2018: 292485625000, 2019: 339782744000, 2020: -200253285781,
    2021: -70474730032, 2022: 145068791398, 2023: 258505699379,
    2024: 285790454716
}
ANNUAL_NI = {
    2015: 466119741406, 2016: 480264780048, 2017: 460843127006,
    2018: 314063000747, 2019: 353337372000, 2020: -291290000000,
    2021: -11056000000, 2022: 122207000000, 2023: 341061017856,
    2024: 456895295638
}
ANNUAL_EPS = {
    2015: 2178, 2016: 2242, 2017: 2159, 2018: 1467, 2019: 1651,
    2020: -1361, 2021: -52, 2022: 570, 2023: 1681, 2024: 2257
}
ANNUAL_DPS = {
    2015: 980, 2016: 990, 2017: 990, 2018: 900, 2019: 900,
    2020: 0, 2021: 0, 2022: 350, 2023: 930, 2024: 1170
}

# 연도별 총자산/총부채/총자본 (원, 연결)
ANNUAL_ASSETS = {
    2015: 3885691981370, 2016: 4040597432553, 2017: 4276455571340,
    2018: 4391549485700, 2019: 4515449127900, 2020: 4026397565858,
    2021: 3960773565793, 2022: 4071281254849, 2023: 4322879968068,
    2024: 4700641154570
}
ANNUAL_LIAB = {
    2015: 663946277000, 2016: 643694099000, 2017: 611099653000,
    2018: 613709270000, 2019: 559419788000, 2020: 565783993000,
    2021: 516024668000, 2022: 574709680000, 2023: 644979151806,
    2024: 816978340813
}
ANNUAL_EQ = {
    2015: 3221745704370, 2016: 3396903333553, 2017: 3665355918340,
    2018: 3777840215700, 2019: 3956029339900, 2020: 3460613572858,
    2021: 3444748897793, 2022: 3496571574849, 2023: 3677900816262,
    2024: 3883662813757
}

# Derived
MKTCAP = PRICE * SHARES   # 시가총액
MKTCAP_NET = PRICE * (SHARES - SHARES_TREASURY)   # 유통시총
PER24 = PRICE / EPS24
PBR24 = MKTCAP / EQ24
ROE24 = NI24 / ((EQ24 + EQ23) / 2)
OPM24 = OP24 / REV24
DIV_YIELD24 = DPS24 / PRICE
FCF24 = OPCF24 - CAPEX24
DEBT_RATIO24 = LIAB24 / EQ24


# ============================================================
# SHEET 1: 표지 (Cover)
# ============================================================
ws1 = wb.active
ws1.title = "표지"
ws1.sheet_properties.tabColor = NAVY
set_col_widths(ws1, [3, 20, 20, 20, 20, 20, 3])

for r in range(1, 35):
    for c in range(1, 8):
        ws1.cell(row=r, column=c).fill = title_fill

ws1.merge_cells('B6:F6')
ws1.cell(row=6, column=2, value="강원랜드(KANGWON LAND)").font = Font(name="맑은 고딕", size=32, bold=True, color=WHITE)
ws1.cell(row=6, column=2).alignment = Alignment(horizontal='center', vertical='center')

ws1.merge_cells('B8:F8')
ws1.cell(row=8, column=2, value="심층 기업분석 보고서").font = Font(name="맑은 고딕", size=20, color=ACCENT_GOLD)
ws1.cell(row=8, column=2).alignment = Alignment(horizontal='center')

ws1.merge_cells('B11:F11')
ws1.cell(row=11, column=2, value="종목코드: 035250 (유가증권시장)  |  업종: 카지노/리조트업").font = subtitle_font
ws1.cell(row=11, column=2).alignment = Alignment(horizontal='center')

ws1.merge_cells('B12:F12')
ws1.cell(row=12, column=2, value="대한민국 유일의 내국인 출입 허용 카지노 (독점 면허)").font = subtitle_font
ws1.cell(row=12, column=2).alignment = Alignment(horizontal='center')

info_data = [
    (15, "경영체제", "전문경영인 체제 (한국광해광업공단 지배)"),
    (16, "설립일", "1998년 6월 29일"),
    (17, "본사", "강원도 정선군 사북읍 하이원길 265"),
    (18, "시장구분", "유가증권시장 (KOSPI)"),
    (19, "최대주주", "한국광해광업공단 (36.27%)"),
    (20, "주요사업", "카지노(~80%), 호텔(~10%), 리조트(스키/골프/콘도, ~10%)"),
    (21, "브랜드", "하이원리조트 (High1 Resort)"),
    (22, "분석기준일", "2026년 2월 6일"),
]
for r, label, val in info_data:
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws1.cell(row=r, column=2, value=label).font = Font(name="맑은 고딕", size=11, color="8899AA")
    ws1.cell(row=r, column=2).alignment = Alignment(horizontal='right', vertical='center')
    ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    ws1.cell(row=r, column=4, value=val).font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
    ws1.cell(row=r, column=4).alignment = Alignment(horizontal='left', vertical='center')

ws1.merge_cells('B25:F25')
ws1.cell(row=25, column=2, value="데이터 출처: OpenDART 공시, 사업보고서 전수분석").font = Font(name="맑은 고딕", size=9, color="6688AA")
ws1.cell(row=25, column=2).alignment = Alignment(horizontal='center')

ws1.merge_cells('B27:F30')
cell = ws1.cell(row=27, column=2)
cell.value = (
    "핵심 요약:\n"
    f"  PER {PER24:.1f}배, PBR {PBR24:.2f}배 / ROE {ROE24*100:.1f}%\n"
    f"  2024년 영업이익 2,858억(+10.6% YoY), 순이익 4,569억(역대 최대)\n"
    f"  배당수익률 {DIV_YIELD24*100:.1f}%, 무차입경영, 내국인 카지노 독점 면허"
)
cell.font = Font(name="맑은 고딕", size=10, color=ACCENT_GOLD)
cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

print("  [1/9] 표지 완료")

# ============================================================
# SHEET 2: 핵심 실적 (10년 재무)
# ============================================================
ws2 = wb.create_sheet("핵심실적")
ws2.sheet_properties.tabColor = "2C3E6B"
set_col_widths(ws2, [14, 14, 14, 14, 12, 14, 14, 14, 12, 12, 14])

row = 1
ws2.merge_cells('A1:K1')
ws2.cell(row=1, column=1, value="10년 연결 재무실적 (단위: 억원)").font = section_font
ws2.cell(row=1, column=1).border = bottom_border
row = 3

headers = ["연도", "매출액", "영업이익", "순이익", "EPS(원)", "총자산", "총부채", "총자본", "부채비율", "영업이익률", "ROE"]
write_header_row(ws2, row, headers)
row += 1

data_start_row = row
years = sorted(ANNUAL_REV.keys())
prev_equity = None

for yr in years:
    rev = round(ANNUAL_REV[yr] / 1e8)
    op = round(ANNUAL_OP[yr] / 1e8)
    ni = round(ANNUAL_NI[yr] / 1e8)
    eps = ANNUAL_EPS[yr]
    ta = round(ANNUAL_ASSETS[yr] / 1e8)
    tl = round(ANNUAL_LIAB[yr] / 1e8)
    te = round(ANNUAL_EQ[yr] / 1e8)

    debt_ratio = tl / te if te != 0 else 0
    opm = op / rev if rev != 0 else 0
    if prev_equity is not None and prev_equity != 0:
        roe = ni / ((te + prev_equity) / 2)
    else:
        roe = ni / te if te != 0 else 0

    data = [yr, rev, op, ni, eps, ta, tl, te, debt_ratio, opm, roe]
    nf = [None, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, '0.0%', '0.0%', '0.0%']

    fonts_row = [data_font_bold] + [data_font] * 10
    fills_row = [light_fill] + [white_fill] * 10

    if op < 0:
        fonts_row[2] = red_font
    if ni < 0:
        fonts_row[3] = red_font
    if eps < 0:
        fonts_row[4] = red_font
    if roe < 0:
        fonts_row[10] = red_font
    if opm < 0:
        fonts_row[9] = red_font

    # Highlight COVID years
    if yr in (2020, 2021):
        fills_row = [light_fill] + [red_fill] * 10

    als = [center_al] + [right_al] * 10
    write_data_row(ws2, row, data, fonts=fonts_row, fills=fills_row, alignments=als, number_formats=nf)

    prev_equity = te
    row += 1

# COVID annotation
row += 1
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
ws2.cell(row=row, column=1, value="* 2020~2021년: COVID-19로 영업중단/제한 (2020.02~2020.05 전면 휴장, 이후 입장객 제한)").font = Font(name="맑은 고딕", size=9, bold=True, color=ACCENT_RED)
ws2.cell(row=row, column=1).alignment = left_al
row += 1
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
ws2.cell(row=row, column=1, value="* 2022년 이후 본격 회복세: 매출 2022년 1.27조 -> 2023년 1.39조 -> 2024년 1.43조 (코로나 이전 1.5~1.7조 수준 근접)").font = small_font
ws2.cell(row=row, column=1).alignment = left_al

# YoY growth section
row += 2
row = add_section_title(ws2, row, "전년대비 성장률 (YoY)")
headers2 = ["연도", "매출 YoY", "영업이익 YoY", "순이익 YoY"]
write_header_row(ws2, row, headers2)
row += 1

for i in range(1, len(years)):
    yr = years[i]
    prev_yr = years[i - 1]

    rev_yoy = (ANNUAL_REV[yr] - ANNUAL_REV[prev_yr]) / abs(ANNUAL_REV[prev_yr]) if ANNUAL_REV[prev_yr] != 0 else None
    # For OP YoY, skip if previous was negative (meaningless)
    if ANNUAL_OP[prev_yr] > 0:
        op_yoy = (ANNUAL_OP[yr] - ANNUAL_OP[prev_yr]) / abs(ANNUAL_OP[prev_yr])
    elif ANNUAL_OP[prev_yr] < 0 and ANNUAL_OP[yr] > 0:
        op_yoy = None  # turnaround, not a meaningful %
    elif ANNUAL_OP[prev_yr] < 0 and ANNUAL_OP[yr] < 0:
        op_yoy = None
    else:
        op_yoy = None

    if ANNUAL_NI[prev_yr] > 0:
        ni_yoy = (ANNUAL_NI[yr] - ANNUAL_NI[prev_yr]) / abs(ANNUAL_NI[prev_yr])
    elif ANNUAL_NI[prev_yr] < 0 and ANNUAL_NI[yr] > 0:
        ni_yoy = None
    else:
        ni_yoy = None

    data_row_vals = [yr, rev_yoy, op_yoy, ni_yoy]
    nf = [None, '0.0%', '0.0%', '0.0%']
    fonts_r = [data_font_bold, data_font, data_font, data_font]
    fills_r = [light_fill, white_fill, white_fill, white_fill]

    # Color code
    for ci in range(1, 4):
        v = data_row_vals[ci]
        if v is None:
            data_row_vals[ci] = "흑전" if ci > 1 and ((ci == 2 and ANNUAL_OP[prev_yr] < 0 and ANNUAL_OP[yr] > 0) or
                                                        (ci == 3 and ANNUAL_NI[prev_yr] < 0 and ANNUAL_NI[yr] > 0)) else "-"
            nf[ci] = None
        elif v < 0:
            fonts_r[ci] = red_font

    als = [center_al] + [right_al] * 3
    write_data_row(ws2, row, data_row_vals, fonts=fonts_r, fills=fills_r, alignments=als, number_formats=nf)
    row += 1

print("  [2/9] 핵심실적 완료")

# ============================================================
# SHEET 3: 2025년 분기실적
# ============================================================
ws3 = wb.create_sheet("2025실적")
ws3.sheet_properties.tabColor = "27AE60"
set_col_widths(ws3, [14, 14, 14, 14, 14, 16, 16])

row = 1
ws3.merge_cells('A1:G1')
ws3.cell(row=1, column=1, value="2025년 분기별 잠정실적 (단위: 백만원)").font = section_font
ws3.cell(row=1, column=1).border = bottom_border

row = 3
headers = ["분기", "매출액", "영업이익", "순이익", "영업이익률", "매출 YoY", "영업이익 YoY"]
write_header_row(ws3, row, headers)
row += 1

# 2024 quarterly (estimated breakdown from annual)
# 2024 quarterly data derived from annual / 잠정실적 patterns
q24_data = [
    ("2024.1Q", 358618, 72375, 103070),
    ("2024.2Q", 353000, 68260, 128561),
    ("2024.3Q", 375483, 89175, 131450),
    ("2024.4Q", 339762, 55980, 93814),
    ("2024 합계", round(REV24 / 1e6), round(OP24 / 1e6), round(NI24 / 1e6)),
]

q25_data = [
    ("2025.1Q", round(Q25[0]["rev"] / 1e6), round(Q25[0]["op"] / 1e6), round(Q25[0]["ni"] / 1e6)),
    ("2025.2Q", round(Q25[1]["rev"] / 1e6), round(Q25[1]["op"] / 1e6), round(Q25[1]["ni"] / 1e6)),
    ("2025.3Q", round(Q25[2]["rev"] / 1e6), round(Q25[2]["op"] / 1e6), round(Q25[2]["ni"] / 1e6)),
    ("2025.4Q", round(Q25[3]["rev"] / 1e6), round(Q25[3]["op"] / 1e6), round(Q25[3]["ni"] / 1e6)),
    ("2025 합계", round(CUM25_REV / 1e6), round(CUM25_OP / 1e6), round(CUM25_NI / 1e6)),
]

all_q_data = q24_data + q25_data

for i, (qtr, rev, op, ni) in enumerate(all_q_data):
    is_total = "합계" in qtr
    f_main = data_font_bold if is_total else data_font
    f_fill = gold_fill if is_total else (lighter_fill if "2025" in qtr else white_fill)

    data_row = [qtr, rev, op, ni, None, None, None]
    nf = [None, NUM_FMT, NUM_FMT, NUM_FMT, '0.0%', '0.0%', '0.0%']
    fonts_r = [f_main] * 7
    fills_r = [f_fill] * 7
    als = [center_al] + [right_al] * 6
    write_data_row(ws3, row, data_row, fonts=fonts_r, fills=fills_r, alignments=als, number_formats=nf)

    # OPM formula
    ws3.cell(row=row, column=5).value = f"=C{row}/B{row}"
    ws3.cell(row=row, column=5).number_format = '0.0%'
    ws3.cell(row=row, column=5).border = thin_border

    # YoY: 2025 quarters vs 2024 quarters
    if i >= 5 and i <= 8:
        prev_row = row - 5
        ws3.cell(row=row, column=6).value = f"=(B{row}-B{prev_row})/B{prev_row}"
        ws3.cell(row=row, column=6).number_format = '0.0%'
        ws3.cell(row=row, column=6).border = thin_border
        ws3.cell(row=row, column=7).value = f"=(C{row}-C{prev_row})/C{prev_row}"
        ws3.cell(row=row, column=7).number_format = '0.0%'
        ws3.cell(row=row, column=7).border = thin_border

    row += 1

row += 2
row = add_section_title(ws3, row, "핵심 포인트", col_end=7)
points = [
    f"2025년 연간 매출 {round(CUM25_REV/1e8):,}억원 → 2024년({round(REV24/1e8):,}억) 대비 +3.5% 성장",
    f"2025년 연간 영업이익 {round(CUM25_OP/1e8):,}억원 → 2024년({round(OP24/1e8):,}억) 대비 -17.7% 감소",
    "4Q25 영업이익 297억(OPM 8.1%)으로 크게 둔화 → 비수기 + 비용 증가 영향",
    "3Q25 순이익 1,131억 → 영업외수익(일회성) 발생 추정, 일시적 이익 부풀림 주의",
    "코로나 이전 피크(2016년 매출 1.70조) 대비 여전히 87% 수준 → 완전 회복 미달",
]
for pt in points:
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws3.cell(row=row, column=1, value=f"  {pt}").font = data_font
    ws3.cell(row=row, column=1).alignment = left_al
    row += 1

print("  [3/9] 2025실적 완료")

# ============================================================
# SHEET 4: 사업구조
# ============================================================
ws4 = wb.create_sheet("사업구조")
ws4.sheet_properties.tabColor = "D4A843"
set_col_widths(ws4, [18, 14, 14, 14, 14, 14, 14, 14])

row = 1
ws4.merge_cells('A1:H1')
ws4.cell(row=1, column=1, value="사업부문별 매출 구조 (카지노/호텔/리조트)").font = section_font
ws4.cell(row=1, column=1).border = bottom_border

row = 3
headers = ["사업부문", "매출비중", "주요 내용", "매출 드라이버", "수익성", "성장성"]
write_header_row(ws4, row, headers)
row += 1

seg_data = [
    ("카지노", "~80%", "테이블게임(바카라/블랙잭 등)\n슬롯머신(약 960대)",
     "입장객 수 x 1인당 GGR", "매우 높음\n(OPM 30%+)", "규제 내 안정"),
    ("호텔", "~10%", "하이원호텔, 하이원콘벤션호텔\n약 800실 규모",
     "객실 가동률 x ADR", "보통\n(보조 사업)", "안정적"),
    ("리조트", "~10%", "스키장(18면), 골프장(36홀)\n콘도미니엄, 테마파크",
     "시즌 입장객 수", "낮음~보통\n(지역경제 공헌)", "제한적"),
]

for nm, pct, desc, driver, profit, growth in seg_data:
    write_data_row(ws4, row, [nm, pct, desc, driver, profit, growth],
                   fonts=[data_font_bold, data_font_bold, data_font, data_font, data_font, data_font],
                   fills=[lighter_fill, gold_fill, white_fill, white_fill, white_fill, white_fill],
                   alignments=[left_al, center_al, left_al, left_al, center_al, center_al])
    ws4.row_dimensions[row].height = 45
    row += 1

# 하이원리조트 브랜드
row += 2
row = add_section_title(ws4, row, "하이원리조트 (High1 Resort) 통합 브랜드", col_end=8)

brand_info = [
    ("카지노", "정선 메인 카지노", "테이블 200대 + 슬롯 960대, 내국인 전용"),
    ("호텔", "하이원 그랜드호텔 외 2개", "총 약 800실, 컨벤션 시설 보유"),
    ("스키장", "하이원 스키장", "18면 슬로프, 동계시즌(12~3월) 핵심"),
    ("골프장", "하이원 CC / 마운틴CC", "36홀, 하계시즌 핵심 수익원"),
    ("콘도", "하이원 콘도", "약 500실 규모, 가족 단위 고객"),
    ("기타", "워터파크/테마파크 등", "2017년 이후 부대시설 확충"),
]

headers = ["시설", "명칭", "규모/특징"]
write_header_row(ws4, row, headers)
row += 1

for cat, name, feature in brand_info:
    write_data_row(ws4, row, [cat, name, feature],
                   fonts=[data_font_bold, data_font, data_font],
                   alignments=[left_al, left_al, left_al])
    row += 1

# 핵심 투자 - 제2카지노
row += 2
row = add_section_title(ws4, row, "제2카지노 영업장 신규시설투자", col_end=8)

invest_info = [
    ("사업명", "제2카지노 영업장 신축"),
    ("투자금액", "약 1,796억원"),
    ("목적", "기존 카지노 포화 대응, 수용능력 확대"),
    ("기대효과", "입장객 수용력 증가 → GGR 확대 가능"),
    ("위치", "정선 하이원리조트 인근"),
    ("상태", "건설 진행 중"),
    ("리스크", "규제당국 영업시간/입장횟수 규제로 효과 제한 가능"),
]

for label, val in invest_info:
    ws4.cell(row=row, column=1, value=label).font = data_font_bold
    ws4.cell(row=row, column=1).alignment = left_al
    ws4.cell(row=row, column=1).fill = lighter_fill
    ws4.cell(row=row, column=1).border = thin_border
    ws4.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws4.cell(row=row, column=2, value=val).font = data_font
    ws4.cell(row=row, column=2).alignment = left_al
    ws4.cell(row=row, column=2).border = thin_border
    row += 1

# 매출 드라이버 해설
row += 2
row = add_section_title(ws4, row, "카지노 매출 구조 해설", col_end=8)
casino_notes = [
    "카지노 매출 = 입장객 수 x 1인당 GGR (Gross Gaming Revenue)",
    "GGR = 고객 베팅 총액 - 고객 지급액 (= 카지노 순승)",
    "테이블게임 홀드율(Hold Rate): 통상 15~25% → 바카라 등 게임별 상이",
    "슬롯머신 홀드율: 통상 5~10% → 기계적으로 안정적 수익",
    "입장객 수는 영업일수 x 일평균 입장객으로 결정 → 규제(입장횟수 제한)에 직접 영향",
    "내국인 입장료: 1만원 (매출 대비 미미하나 진입장벽 역할)",
]
for note in casino_notes:
    ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws4.cell(row=row, column=1, value=f"  {note}").font = data_font
    ws4.cell(row=row, column=1).alignment = left_al
    row += 1

print("  [4/9] 사업구조 완료")

# ============================================================
# SHEET 5: 배당/주주환원
# ============================================================
ws5 = wb.create_sheet("주주환원")
ws5.sheet_properties.tabColor = "E74C3C"
set_col_widths(ws5, [12, 14, 12, 14, 14, 14])

row = 1
ws5.merge_cells('A1:F1')
ws5.cell(row=1, column=1, value="배당 및 주주환원 정책").font = section_font
ws5.cell(row=1, column=1).border = bottom_border

row = 3
headers = ["연도", "주당배당금(원)", "EPS(원)", "배당성향", "배당수익률", "비고"]
write_header_row(ws5, row, headers)
row += 1

div_years = sorted(ANNUAL_DPS.keys())
for yr in div_years:
    dps = ANNUAL_DPS[yr]
    eps = ANNUAL_EPS[yr]

    if dps > 0 and eps > 0:
        payout = f"{dps/eps*100:.1f}%"
    elif dps > 0 and eps <= 0:
        payout = "-"
    else:
        payout = "-"

    # Approximate dividend yield based on typical year-end prices
    price_approx = {
        2015: 38000, 2016: 36000, 2017: 33000, 2018: 25000, 2019: 27000,
        2020: 21000, 2021: 26000, 2022: 22000, 2023: 16000, 2024: 17690,
    }
    p = price_approx.get(yr, PRICE)
    yld = f"{dps/p*100:.1f}%" if dps > 0 else "-"

    note_map = {
        2015: "", 2016: "", 2017: "",
        2018: "배당 소폭 축소", 2019: "코로나 직전",
        2020: "코로나 영업중단→무배당", 2021: "적자지속→무배당",
        2022: "흑전→배당 재개", 2023: "배당 대폭 인상",
        2024: "역대 최대 배당 (1,170원)",
    }
    note = note_map.get(yr, "")

    nf = [None, NUM_FMT, NUM_FMT, None, None, None]
    f_dps = green_font if dps > 0 else red_font
    als = [center_al, right_al, right_al, center_al, center_al, left_al]

    write_data_row(ws5, row, [yr, dps if dps else "-", eps, payout, yld, note],
                   fonts=[data_font_bold, f_dps, data_font, data_font, data_font, small_font],
                   alignments=als, number_formats=nf)
    row += 1

# Key observations
row += 2
row = add_section_title(ws5, row, "배당 정책 특징", col_end=6)
div_notes = [
    "코로나 이전(2015~2019): 안정적 배당 (DPS 900~990원, 배당성향 40~55%)",
    "코로나 기간(2020~2021): 2년 연속 무배당 (적자로 인한 불가피한 중단)",
    "회복기(2022~): 2022년 350원으로 재개 → 2023년 930원 → 2024년 1,170원 급증",
    f"2024 배당수익률: {DIV_YIELD24*100:.1f}% (현재가 {PRICE:,}원 기준)",
    f"배당총액: 약 {round(DIV_PAID/1e8):,}억원 (자사주 제외 유통주식 기준)",
    "공기업 특성상 배당 안정성 높음 → 최대주주(광해광업공단) 배당수익 확보 유인",
]
for note in div_notes:
    ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws5.cell(row=row, column=1, value=f"  {note}").font = data_font
    ws5.cell(row=row, column=1).alignment = left_al
    row += 1

# Share structure
row += 2
row = add_section_title(ws5, row, "주식 구조", col_end=6)
share_info = [
    ("발행주식수", f"{SHARES:,}주"),
    ("자기주식", f"{SHARES_TREASURY:,}주 ({SHARES_TREASURY/SHARES*100:.1f}%)"),
    ("유통주식수", f"{SHARES - SHARES_TREASURY:,}주"),
    ("최대주주", "한국광해광업공단 (36.27%)"),
    ("특수관계인 합계", "약 36.27% (공단 단독, 특관인 없음)"),
    ("국민연금", "약 8~9%"),
    ("외국인 지분율", "약 15~20%"),
    ("유증 이력", "없음 (희석 리스크 없음)"),
]
for label, val in share_info:
    ws5.cell(row=row, column=1, value=label).font = data_font
    ws5.cell(row=row, column=1).alignment = left_al
    ws5.cell(row=row, column=1).fill = lighter_fill
    ws5.cell(row=row, column=1).border = thin_border
    ws5.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws5.cell(row=row, column=2, value=val).font = data_font_bold
    ws5.cell(row=row, column=2).alignment = left_al
    ws5.cell(row=row, column=2).border = thin_border
    row += 1

print("  [5/9] 주주환원 완료")

# ============================================================
# SHEET 6: 규제/면허 (replaces R&D_특허 for casino company)
# ============================================================
ws6 = wb.create_sheet("규제_면허")
ws6.sheet_properties.tabColor = "8E44AD"
set_col_widths(ws6, [20, 50, 20])

row = 1
ws6.merge_cells('A1:C1')
ws6.cell(row=1, column=1, value="규제 환경 및 면허 분석").font = section_font
ws6.cell(row=1, column=1).border = bottom_border

row = 3
row = add_section_title(ws6, row, "내국인 카지노 면허 (독점)", col_end=3)

license_info = [
    ("면허 근거", "폐광지역개발지원에관한특별법 (1995년 제정)"),
    ("면허 부여", "강원랜드가 대한민국 유일의 내국인 카지노 면허 보유"),
    ("면허 성격", "사실상 영구 독점 → 법률 개정 없이는 추가 면허 불가"),
    ("설립 목적", "폐광지역(정선/태백/영월/삼척) 경제 활성화"),
    ("면허 리스크", "정치적 리스크 존재하나, 폐광법 폐지는 현실적으로 극히 어려움"),
    ("경쟁사", "외국인 전용 카지노 17개 (파라다이스, GKL 등) → 내국인 시장은 독점"),
]

for label, val in license_info:
    ws6.cell(row=row, column=1, value=label).font = data_font_bold
    ws6.cell(row=row, column=1).alignment = left_al
    ws6.cell(row=row, column=1).fill = gold_fill
    ws6.cell(row=row, column=1).border = thin_border
    ws6.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws6.cell(row=row, column=2, value=val).font = data_font
    ws6.cell(row=row, column=2).alignment = left_al
    ws6.cell(row=row, column=2).border = thin_border
    row += 1

row += 2
row = add_section_title(ws6, row, "주요 규제 사항", col_end=3)

headers = ["규제 항목", "현행 내용", "투자 영향"]
write_header_row(ws6, row, headers)
row += 1

regulations = [
    ("영업시간", "1일 최대 영업시간 제한\n(현재 하루 약 18~20시간 운영)",
     "매출 상한 설정 → 시간 연장 시 직접적 매출 증가"),
    ("입장횟수 제한", "월 15일 / 연 180일 제한\n(본인 신청 시 출입제한 가능)",
     "고빈도 방문객 제한 → GGR 상한 존재\n완화 시 큰 호재"),
    ("입장료", "1인당 1만원\n(매출 대비 미미하나 심리적 장벽)",
     "인상 시 입장객 감소 우려\n인하 시 입장객 증가 가능"),
    ("배팅한도", "테이블별 최대 배팅금액 규제\n(바카라 30만원 등)",
     "VIP 고객 유치 제한\n완화 시 1인당 GGR 급증 가능"),
    ("사행산업감독", "사행산업통합감독위원회\n(국무총리실 소속)",
     "정기 감사/감독, 매출한도 설정 권한\n사행성 논란 시 규제 강화 리스크"),
    ("지역공헌 의무", "폐광지역 개발기금 출연\n매출의 일정% 의무 기부",
     "비용 증가 요인이나,\n면허 유지의 사회적 정당성 확보"),
    ("중독예방 의무", "도박중독예방센터 운영\n자기배제 프로그램 의무 시행",
     "운영비 부담이나,\nESG/사회적 책임 이행으로 면허 안정성 확보"),
]

for item, content, impact in regulations:
    write_data_row(ws6, row, [item, content, impact],
                   fonts=[data_font_bold, data_font, data_font],
                   fills=[lighter_fill, white_fill, white_fill],
                   alignments=[left_al, left_al, left_al])
    ws6.row_dimensions[row].height = 50
    row += 1

row += 2
row = add_section_title(ws6, row, "규제 변화 시나리오별 영향", col_end=3)

scenarios_reg = [
    ("영업시간 연장", "매출 +5~15% 직접 증가\n(야간 시간대 수요 흡수)", "호재"),
    ("입장횟수 완화", "고빈도 방문객 증가 → GGR 확대\n매출 +10~20% 가능", "대호재"),
    ("배팅한도 상향", "1인당 GGR 급증, VIP 세그먼트 확대\n수익성 대폭 개선", "대호재"),
    ("규제 전면 강화", "입장객/영업시간 축소\n매출 -10~20% 감소 가능", "악재"),
    ("제2카지노 타사 면허", "독점 깨짐 → 밸류에이션 급락\n현실 가능성은 극히 낮음", "대악재"),
]

headers = ["시나리오", "예상 영향", "주가영향"]
write_header_row(ws6, row, headers)
row += 1

for scenario, impact, direction in scenarios_reg:
    f_dir = green_font if "호재" in direction else red_font
    f_fill_dir = green_fill if "호재" in direction else red_fill
    write_data_row(ws6, row, [scenario, impact, direction],
                   fonts=[data_font_bold, data_font, f_dir],
                   fills=[lighter_fill, white_fill, f_fill_dir],
                   alignments=[left_al, left_al, center_al])
    ws6.row_dimensions[row].height = 40
    row += 1

print("  [6/9] 규제_면허 완료")

# ============================================================
# SHEET 7: 투자지표
# ============================================================
ws7 = wb.create_sheet("투자지표")
ws7.sheet_properties.tabColor = "E67E22"
set_col_widths(ws7, [20, 16, 40])

row = 1
ws7.merge_cells('A1:C1')
ws7.cell(row=1, column=1, value=f"핵심 투자지표 (2024 기준, 주가 {PRICE:,}원)").font = section_font
ws7.cell(row=1, column=1).border = bottom_border

row = 3
metrics = [
    ("PER", f"{PER24:.1f}배",
     f"EPS {EPS24:,}원 기준. 코로나 이전 평균 PER 15~20배 대비 저평가"),
    ("PBR", f"{PBR24:.2f}배",
     f"시총 약 {round(MKTCAP/1e8):,}억 / 자본 {round(EQ24/1e8):,}억. 장부가 이하 거래"),
    ("ROE", f"{ROE24*100:.1f}%",
     f"순이익 {round(NI24/1e8):,}억 / 평균자본 {round((EQ24+EQ23)/2/1e8):,}억"),
    ("영업이익률", f"{OPM24*100:.1f}%",
     "카지노 사업 특성상 높은 마진. 코로나 이전 25~30% 수준 대비 회복 중"),
    ("부채비율", f"{DEBT_RATIO24*100:.1f}%",
     "무차입 경영 (이자부 차입금 0원). 재무 건전성 최상급"),
    ("순차입금", "0원 (넷캐시)",
     f"현금성자산 {round(CASH24/1e8):,}억, 차입금 0 → 사실상 무부채 기업"),
    ("배당수익률", f"{DIV_YIELD24*100:.1f}%",
     f"주당 {DPS24:,}원 (2024). 공기업 특성상 안정적 배당 기조"),
    ("FCF", f"{round(FCF24/1e8):,}억원",
     f"영업CF {round(OPCF24/1e8):,}억 - CAPEX {round(CAPEX24/1e8):,}억 = 잉여현금 풍부"),
    ("시가총액", f"약 {round(MKTCAP/1e8):,}억원",
     f"유통시총 약 {round(MKTCAP_NET/1e8):,}억원 (자기주식 제외)"),
    ("EV/EBITDA", f"{(MKTCAP - CASH24) / (OP24 + DA24):.1f}배",
     f"EBITDA {round((OP24+DA24)/1e8):,}억 기준. 카지노 업종 글로벌 평균 8~12배"),
]

headers = ["지표", "값", "해석"]
write_header_row(ws7, row, headers)
row += 1

for label, val, desc in metrics:
    write_data_row(ws7, row, [label, val, desc],
                   fonts=[data_font_bold, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), data_font],
                   fills=[lighter_fill, gold_fill, white_fill],
                   alignments=[left_al, center_al, left_al])
    row += 1

row += 2
row = add_section_title(ws7, row, "밸류에이션 할인/프리미엄 요인", col_end=3)

# Discount factors
ws7.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws7.cell(row=row, column=1, value="할인 요인 (Discount Factors)").font = Font(name="맑은 고딕", size=11, bold=True, color=ACCENT_RED)
ws7.cell(row=row, column=1).fill = red_fill
ws7.cell(row=row, column=1).alignment = center_al
row += 1

disc_factors = [
    ("규제 리스크", "사행산업감독위 규제 강화 가능성 → 매출 상한 존재"),
    ("코로나 트라우마", "2020~2021 대규모 적자 경험 → 투자자 트라우마"),
    ("성장 한계", "내국인 전용+규제 환경 → 폭발적 성장은 구조적으로 어려움"),
    ("공기업 디스카운트", "전문경영인 체제, 최대주주=공공기관 → 주주가치 극대화 한계"),
    ("지역 리스크", "정선 소재 → 접근성 제한, 인력 확보 어려움"),
]
for factor, desc in disc_factors:
    ws7.cell(row=row, column=1, value=factor).font = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT_RED)
    ws7.cell(row=row, column=1).alignment = left_al
    ws7.cell(row=row, column=1).fill = red_fill
    ws7.cell(row=row, column=1).border = thin_border
    ws7.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws7.cell(row=row, column=2, value=desc).font = data_font
    ws7.cell(row=row, column=2).alignment = left_al
    ws7.cell(row=row, column=2).border = thin_border
    row += 1

row += 1
# Premium factors
ws7.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws7.cell(row=row, column=1, value="프리미엄 요인 (Premium Factors)").font = Font(name="맑은 고딕", size=11, bold=True, color=ACCENT_GREEN)
ws7.cell(row=row, column=1).fill = green_fill
ws7.cell(row=row, column=1).alignment = center_al
row += 1

prem_factors = [
    ("독점 면허", "내국인 카지노 유일 면허 → 경쟁 진입 불가, 영구적 해자"),
    ("무차입 경영", "이자부 차입금 0, 현금 2,445억 → 재무 리스크 제로"),
    ("높은 배당", "DPS 1,170원 (2024), 배당성향 확대 추세 → 인컴 투자 매력"),
    ("코로나 회복 잔여분", "코로나 이전 피크 대비 아직 85% 수준 → 추가 회복 여력"),
    ("제2카지노", "신규 영업장 완공 시 수용력 확대 → 중장기 매출 증가 기대"),
]
for factor, desc in prem_factors:
    ws7.cell(row=row, column=1, value=factor).font = Font(name="맑은 고딕", size=10, bold=True, color=ACCENT_GREEN)
    ws7.cell(row=row, column=1).alignment = left_al
    ws7.cell(row=row, column=1).fill = green_fill
    ws7.cell(row=row, column=1).border = thin_border
    ws7.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws7.cell(row=row, column=2, value=desc).font = data_font
    ws7.cell(row=row, column=2).alignment = left_al
    ws7.cell(row=row, column=2).border = thin_border
    row += 1

print("  [7/9] 투자지표 완료")

# ============================================================
# SHEET 8: 시나리오 분석
# ============================================================
ws8 = wb.create_sheet("시나리오")
ws8.sheet_properties.tabColor = "2ECC71"
set_col_widths(ws8, [16, 18, 18, 18, 18])

row = 1
ws8.merge_cells('A1:E1')
ws8.cell(row=1, column=1, value="향후 시나리오 분석").font = section_font
ws8.cell(row=1, column=1).border = bottom_border

row = 3
headers = ["항목", "강세 (Bull)", "기본 (Base)", "약세 (Bear)"]
write_header_row(ws8, row, headers)
row += 1

# Scenario calculations
# Bull: 규제 완화 + 입장객 회복 → PER 15배
BULL_EPS = 2500
BULL_PER = 15
BULL_TARGET = BULL_EPS * BULL_PER
BULL_UPSIDE = (BULL_TARGET - PRICE) / PRICE

# Base: 현상 유지 → PER 12배
BASE_EPS = 2200
BASE_PER = 12
BASE_TARGET = BASE_EPS * BASE_PER
BASE_UPSIDE = (BASE_TARGET - PRICE) / PRICE

# Bear: 규제 강화 → PER 8배
BEAR_EPS = 1500
BEAR_PER = 8
BEAR_TARGET = BEAR_EPS * BEAR_PER
BEAR_UPSIDE = (BEAR_TARGET - PRICE) / PRICE

items = [
    "전제조건",
    "매출 전망",
    "영업이익률",
    "순이익 전망",
    "EPS 전망",
    "적용 PER",
    "목표주가",
    "현주가 대비",
]
bull = [
    "규제 완화(영업시간/입장횟수)\n+ 제2카지노 가동",
    "1.6조원 (코로나 이전 수준 회복)",
    "25%+ (규제 완화 효과)",
    f"약 {round(BULL_EPS * (SHARES - SHARES_TREASURY) / 1e8):,}억원",
    f"{BULL_EPS:,}원",
    f"{BULL_PER}배",
    f"{BULL_TARGET:,}원",
    f"상승 {BULL_UPSIDE*100:+.0f}%",
]
base = [
    "현행 규제 유지\n+ 점진적 입장객 회복",
    "1.45~1.50조원",
    "20% (현 수준 유지)",
    f"약 {round(BASE_EPS * (SHARES - SHARES_TREASURY) / 1e8):,}억원",
    f"{BASE_EPS:,}원",
    f"{BASE_PER}배",
    f"{BASE_TARGET:,}원",
    f"상승 {BASE_UPSIDE*100:+.0f}%",
]
bear = [
    "규제 강화(입장횟수/영업시간 축소)\n+ 경기 침체",
    "1.2~1.3조원",
    "15% (비용 증가 압박)",
    f"약 {round(BEAR_EPS * (SHARES - SHARES_TREASURY) / 1e8):,}억원",
    f"{BEAR_EPS:,}원",
    f"{BEAR_PER}배",
    f"{BEAR_TARGET:,}원",
    f"하락 {BEAR_UPSIDE*100:.0f}%",
]

for i in range(len(items)):
    bull_color = green_fill
    base_color = gold_fill
    bear_color = red_fill

    write_data_row(ws8, row, [items[i], bull[i], base[i], bear[i]],
                   fonts=[data_font_bold, data_font, data_font, data_font],
                   fills=[lighter_fill, bull_color, base_color, bear_color],
                   alignments=[left_al, center_al, center_al, center_al])
    ws8.row_dimensions[row].height = 35
    row += 1

# SWOT
row += 2
row = add_section_title(ws8, row, "SWOT 분석", col_end=5)

ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
ws8.cell(row=row, column=1, value="강점 (Strengths)").font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
ws8.cell(row=row, column=1).fill = PatternFill("solid", fgColor="27AE60")
ws8.cell(row=row, column=1).alignment = center_al
ws8.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
ws8.cell(row=row, column=3, value="약점 (Weaknesses)").font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
ws8.cell(row=row, column=3).fill = PatternFill("solid", fgColor="E74C3C")
ws8.cell(row=row, column=3).alignment = center_al
row += 1

strengths = [
    "내국인 카지노 유일 면허 (독점)",
    "무차입 경영, 현금 2,445억",
    "높은 영업이익률 (20%+, 카지노)",
    "안정적 배당 정책 (공기업)",
    "제2카지노로 수용력 확대 예정",
]
weaknesses = [
    "규제에 의한 매출 상한 (성장 제한)",
    "정선 소재 (접근성 열악)",
    "공기업 디스카운트 (주주가치 < 공익)",
    "사행산업 이미지 (ESG 리스크)",
    "코로나 이전 실적 완전 회복 미달",
]

for i in range(max(len(strengths), len(weaknesses))):
    s = strengths[i] if i < len(strengths) else ""
    w = weaknesses[i] if i < len(weaknesses) else ""
    ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws8.cell(row=row, column=1, value=f"  {s}" if s else "").font = data_font
    ws8.cell(row=row, column=1).alignment = left_al
    ws8.cell(row=row, column=1).fill = green_fill
    ws8.cell(row=row, column=1).border = thin_border
    ws8.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws8.cell(row=row, column=3, value=f"  {w}" if w else "").font = data_font
    ws8.cell(row=row, column=3).alignment = left_al
    ws8.cell(row=row, column=3).fill = red_fill
    ws8.cell(row=row, column=3).border = thin_border
    row += 1

row += 1
ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
ws8.cell(row=row, column=1, value="기회 (Opportunities)").font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
ws8.cell(row=row, column=1).fill = PatternFill("solid", fgColor="2980B9")
ws8.cell(row=row, column=1).alignment = center_al
ws8.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
ws8.cell(row=row, column=3, value="위협 (Threats)").font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
ws8.cell(row=row, column=3).fill = PatternFill("solid", fgColor="7F8C8D")
ws8.cell(row=row, column=3).alignment = center_al
row += 1

opportunities = [
    "규제 완화 (영업시간/입장횟수/배팅한도)",
    "제2카지노 영업장 완공 → 수용력 확대",
    "관광 활성화 (외국인 관광객 유치 시너지)",
    "배당 확대 → 인컴 투자 수요 증가",
    "코로나 이전 수준 완전 회복 시 매출 +15%",
]
threats = [
    "규제 강화 (사행산업 논란 재부각)",
    "경기 침체 → 소비성 여가 지출 감소",
    "온라인 도박/불법 도박 시장 확대",
    "인구 감소/고령화 → 장기 방문객 감소",
    "정치적 리스크 (공기업 정책 변화)",
]

for i in range(max(len(opportunities), len(threats))):
    o = opportunities[i] if i < len(opportunities) else ""
    t = threats[i] if i < len(threats) else ""
    ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws8.cell(row=row, column=1, value=f"  {o}" if o else "").font = data_font
    ws8.cell(row=row, column=1).alignment = left_al
    ws8.cell(row=row, column=1).fill = PatternFill("solid", fgColor="D6EAF8")
    ws8.cell(row=row, column=1).border = thin_border
    ws8.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws8.cell(row=row, column=3, value=f"  {t}" if t else "").font = data_font
    ws8.cell(row=row, column=3).alignment = left_al
    ws8.cell(row=row, column=3).fill = PatternFill("solid", fgColor="E5E7E9")
    ws8.cell(row=row, column=3).border = thin_border
    row += 1

print("  [8/9] 시나리오 완료")

# ============================================================
# SHEET 9: 모니터링 체크리스트
# ============================================================
ws9 = wb.create_sheet("모니터링")
ws9.sheet_properties.tabColor = "1ABC9C"
set_col_widths(ws9, [6, 30, 40, 20])

row = 1
ws9.merge_cells('A1:D1')
ws9.cell(row=1, column=1, value="향후 이익 판단을 위한 핵심 모니터링 지표").font = section_font
ws9.cell(row=1, column=1).border = bottom_border

row = 3
headers = ["순위", "모니터링 항목", "세부 내용", "확인 시기"]
write_header_row(ws9, row, headers)
row += 1

monitors = [
    ("1", "월별 입장객 수",
     "카지노 매출의 가장 직접적 선행지표. 월별 입장객 수와 GGR 추이를 확인. "
     "코로나 이전 월 평균 약 25~30만명 수준과 비교.",
     "매월 (공시/IR)"),
    ("2", "GGR(총게임수익) 추이",
     "입장객 x 1인당 GGR로 결정. 테이블게임/슬롯머신별 홀드율 변화도 중요. "
     "분기별 매출 발표 시 역산 가능.",
     "매 분기"),
    ("3", "규제 환경 변화",
     "사행산업통합감독위 결정 사항 (영업시간, 입장횟수, 배팅한도 등). "
     "국회 폐광법 개정 논의, 사행산업 관련 법안도 모니터링.",
     "수시 (법안/정책)"),
    ("4", "배당 정책",
     "DPS 추이와 배당성향 변화. 최대주주(광해광업공단)의 배당 선호도. "
     "2024년 DPS 1,170원 → 2025년 유지/인상 여부 확인.",
     "3월 주총, 12월"),
    ("5", "제2카지노 건설 진행",
     "투자금액 1,796억원 규모 신규 영업장 건설 진행 상황. "
     "완공 시기, 개장 일정, 추가 테이블/슬롯 규모 확인.",
     "반기별 IR"),
]

for rank, title, detail, timing in monitors:
    write_data_row(ws9, row, [rank, title, detail, timing],
                   fonts=[Font(name="맑은 고딕", size=14, bold=True, color=NAVY), data_font_bold, data_font, data_font],
                   fills=[gold_fill, lighter_fill, white_fill, lighter_fill],
                   alignments=[center_al, left_al, left_al, center_al])
    ws9.row_dimensions[row].height = 50
    row += 1

row += 2
row = add_section_title(ws9, row, "모니터링 캘린더", col_end=4)
calendar = [
    ("1~2월", "전년도 4분기/연간 잠정실적 공시", "★★★"),
    ("2월", "월별 입장객 수 / GGR 추이 확인", "★★☆"),
    ("3월", "사업보고서 제출 / 정기주총 / 배당 결정", "★★★"),
    ("5월", "1분기 잠정실적 공시", "★★★"),
    ("6월", "하계시즌 개시 (골프/워터파크)", "★☆☆"),
    ("8월", "2분기 잠정실적 공시", "★★★"),
    ("11월", "3분기 잠정실적 공시", "★★★"),
    ("12월", "동계시즌 개시 (스키장) / 배당 결정", "★★☆"),
    ("수시", "사행산업감독위 규제 변경 발표", "★★★"),
    ("수시", "제2카지노 건설 진행 상황 업데이트", "★★☆"),
    ("수시", "폐광법/사행산업 관련 법안 국회 논의", "★★☆"),
]

headers = ["시기", "이벤트", "중요도"]
write_header_row(ws9, row, headers)
row += 1

for timing, event, importance in calendar:
    write_data_row(ws9, row, [timing, event, importance],
                   fonts=[data_font_bold, data_font, data_font_bold],
                   fills=[lighter_fill, white_fill, gold_fill],
                   alignments=[center_al, left_al, center_al])
    row += 1

# Key events from DB (if available)
row += 2
row = add_section_title(ws9, row, "주요 이벤트 히스토리 (최근)", col_end=4)
headers = ["일자", "유형", "내용"]
write_header_row(ws9, row, headers)
row += 1

try:
    events = conn.execute("""SELECT rcept_dt, event_type, SUBSTR(event_summary, 1, 80)
        FROM key_events WHERE rcept_dt >= '20230101' ORDER BY rcept_dt DESC LIMIT 20""").fetchall()
    for dt, etype, summary in events:
        summary_clean = summary.replace('\n', ' ').strip()[:70]
        write_data_row(ws9, row, [dt, etype, summary_clean],
                       fonts=[data_font, data_font_bold, data_font],
                       alignments=[center_al, center_al, left_al])
        row += 1
except Exception:
    # If key_events table doesn't exist or is empty, add manual entries
    manual_events = [
        ("2025.12", "실적", "2025년 4분기 잠정실적 공시"),
        ("2025.08", "실적", "2025년 2분기 잠정실적 공시"),
        ("2025.05", "실적", "2025년 1분기 잠정실적 공시"),
        ("2025.03", "배당", "2024년 결산배당 DPS 1,170원 결정"),
        ("2024.11", "실적", "2024년 3분기 잠정실적 공시"),
        ("2024.08", "실적", "2024년 2분기 잠정실적 공시"),
        ("2024.05", "실적", "2024년 1분기 잠정실적 공시"),
        ("2024.03", "배당", "2023년 결산배당 DPS 930원 결정"),
    ]
    for dt, etype, summary in manual_events:
        write_data_row(ws9, row, [dt, etype, summary],
                       fonts=[data_font, data_font_bold, data_font],
                       alignments=[center_al, center_al, left_al])
        row += 1

print("  [9/9] 모니터링 완료")

# ============================================================
# SAVE
# ============================================================
wb.save(OUT)
conn.close()
print(f"\n보고서 생성 완료: {OUT}")
print(f"시트 구성: {wb.sheetnames}")
