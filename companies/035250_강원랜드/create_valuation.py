# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "강원랜드_밸류에이션.xlsx")
wb = Workbook()

# === CONSTANTS ===
COMPANY = "강원랜드"
STOCK_CODE = "035250"
PRICE = 17690
SHARES = 213940500
SHARES_TREASURY = 13534816
NET_SHARES = SHARES - SHARES_TREASURY  # 200,405,684 유통주식

# 시가총액 (발행주식 기준)
MARKET_CAP = PRICE * SHARES  # 약 3.78조

# === 2024 연결 (원) ===
REV_2024 = 1426862781802
OP_2024 = 285790454716
NI_2024 = 456895295638
EPS_2024 = 2257
BPS_EQUITY = 3883662813757  # 자본총계
EQUITY_2024 = 3883662813757
ASSETS_2024 = 4700641154570
LIAB_2024 = 816978340813
CASH_2024 = 244499636973
ST_DEBT_2024 = 0  # 무차입경영
LT_DEBT_2024 = 0
DA_2024 = 78474881854  # 감가상각비 합계
OPCF_2024 = 487048778217
CAPEX_2024 = 107494733192
FCF_2024 = OPCF_2024 - CAPEX_2024  # 약 3,795억
DPS_2024 = 1170
DIV_PAID_2024 = 188547268140

# === 2023 연결 ===
REV_2023 = 1392740543618
OP_2023 = 258505699379
NI_2023 = 341061017856
EPS_2023 = 1681
EQUITY_2023 = 3677900816262

# === 2025 잠정실적 (연결) ===
Q1_25 = {"rev": 365830e6, "op": 77659e6, "ni": 78021e6}
Q2_25 = {"rev": 360727e6, "op": 57874e6, "ni": 60630e6}
Q3_25 = {"rev": 384147e6, "op": 72702e6, "ni": 113070e6}
Q4_25 = {"rev": 365446e6, "op": 29697e6, "ni": 66017e6}
CUM_25_REV = 1476726e6
CUM_25_OP = 235176e6
CUM_25_NI = 316516e6
EPS_2025E = int(CUM_25_NI / SHARES)  # 약 1479

# === 10년 EPS ===
ANNUAL_EPS = {
    2015: 2178, 2016: 2242, 2017: 2159, 2018: 1467, 2019: 1651,
    2020: -1361, 2021: -52, 2022: 570, 2023: 1681, 2024: 2257
}

# === 10년 DPS ===
ANNUAL_DPS = {
    2015: 980, 2016: 990, 2017: 990, 2018: 900, 2019: 900,
    2020: 0, 2021: 0, 2022: 350, 2023: 930, 2024: 1170
}

# === 10년 Revenue (원) ===
ANNUAL_REV = {
    2015: 1634441985990, 2016: 1703131541003, 2017: 1601291247063,
    2018: 1445736946832, 2019: 1524006966734, 2020: 479173424993,
    2021: 788430938373, 2022: 1272539665429, 2023: 1392740543618,
    2024: 1426862781802
}

# === 10년 Operating Income (원) ===
ANNUAL_OP = {
    2015: 422764543000, 2016: 480490893000, 2017: 460085942000,
    2018: 292485625000, 2019: 339782744000, 2020: -200253285781,
    2021: -70474730032, 2022: 145068791398, 2023: 258505699379,
    2024: 285790454716
}

# === 10년 Net Income (원) ===
ANNUAL_NI = {
    2015: 466119741406, 2016: 480264780048, 2017: 460843127006,
    2018: 314063000747, 2019: 353337372000, 2020: -291290000000,
    2021: -11056000000, 2022: 122207000000, 2023: 341061017856,
    2024: 456895295638
}

억 = 100_000_000

# === STYLES ===
NAVY = "1B2A4A"
DARK = "2C3E6B"
MID = "3A5BA0"
WHITE = "FFFFFF"
GOLD_BG = "FFF3CD"
GREEN_BG = "D4EDDA"
RED_BG = "F8D7DA"
BLUE_BG = "D6E4F0"
LBLUE = "EBF1F8"

title_font = Font(name="맑은 고딕", size=18, bold=True, color=WHITE)
section_font = Font(name="맑은 고딕", size=13, bold=True, color=NAVY)
h_font = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
d_font = Font(name="맑은 고딕", size=10)
d_bold = Font(name="맑은 고딕", size=10, bold=True)
d_blue = Font(name="맑은 고딕", size=10, bold=True, color="0000FF")
d_green = Font(name="맑은 고딕", size=10, bold=True, color="27AE60")
d_red = Font(name="맑은 고딕", size=10, bold=True, color="C0392B")
note_font = Font(name="맑은 고딕", size=9, color="666666")
big_font = Font(name="맑은 고딕", size=14, bold=True, color=NAVY)
huge_font = Font(name="맑은 고딕", size=20, bold=True, color=NAVY)

title_fill = PatternFill("solid", fgColor=NAVY)
header_fill = PatternFill("solid", fgColor=DARK)
mid_fill = PatternFill("solid", fgColor=MID)
blue_fill = PatternFill("solid", fgColor=BLUE_BG)
lblue_fill = PatternFill("solid", fgColor=LBLUE)
gold_fill = PatternFill("solid", fgColor=GOLD_BG)
green_fill = PatternFill("solid", fgColor=GREEN_BG)
red_fill = PatternFill("solid", fgColor=RED_BG)
white_fill = PatternFill("solid", fgColor=WHITE)
gray_fill = PatternFill("solid", fgColor="F2F2F2")

center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")
thin_border = Border(
    left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"))
bottom_border = Border(bottom=Side(style="medium", color=NAVY))


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def setup_print(ws):
    ws.page_setup.paperSize = Worksheet.PAPERSIZE_LETTER
    ws.page_setup.orientation = Worksheet.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.75, bottom=0.75,
        header=0.3, footer=0.3
    )
    ws.print_options.horizontalCentered = True

def write_header(ws, row, vals, fills=None):
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = h_font
        c.fill = fills[i-1] if fills else header_fill
        c.alignment = center
        c.border = thin_border

def write_row(ws, row, vals, fonts=None, fills=None, aligns=None):
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = fonts[i-1] if fonts else d_font
        c.fill = fills[i-1] if fills else white_fill
        c.alignment = aligns[i-1] if aligns else center
        c.border = thin_border

def section_title(ws, row, title, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=title)
    c.font = section_font
    c.border = bottom_border
    return row + 1

def fmt(v):
    """숫자를 억원 문자열로"""
    return f"{v/억:,.0f}억"

def fmt_won(v):
    return f"{v:,.0f}원"

def pct(v):
    return f"{v*100:.1f}%"


# === DERIVED CALCULATIONS ===
BPS = EQUITY_2024 / SHARES  # 약 18,153원
NET_DEBT = ST_DEBT_2024 + LT_DEBT_2024 - CASH_2024  # 음수 = 순현금
EBITDA_2024 = OP_2024 + DA_2024
EV = MARKET_CAP + NET_DEBT  # 순현금이므로 EV < 시가총액

# ROE
AVG_EQ = (EQUITY_2024 + EQUITY_2023) / 2
ROE_2024 = NI_2024 / AVG_EQ  # 약 12.1%

# 2024 분기별 추정 (연간에서 역산 - 강원랜드는 2025잠정이 모두 있으므로 2024 분기는 별도 추정)
# 강원랜드는 2025 4분기까지 잠정실적 모두 확보
# Trailing 4Q = 최근 4분기 = 25Q1 + 25Q2 + 25Q3 + 25Q4 (= 2025 연간)
TRAIL_NI = Q1_25["ni"] + Q2_25["ni"] + Q3_25["ni"] + Q4_25["ni"]
TRAIL_OP = Q1_25["op"] + Q2_25["op"] + Q3_25["op"] + Q4_25["op"]
TRAIL_REV = Q1_25["rev"] + Q2_25["rev"] + Q3_25["rev"] + Q4_25["rev"]
TRAIL_EPS = TRAIL_NI / SHARES

# 보수적 EPS: COVID 제외 3년 평균 (2022, 2023, 2024)
AVG_EPS_3Y = int((ANNUAL_EPS[2022] + ANNUAL_EPS[2023] + ANNUAL_EPS[2024]) / 3)  # 약 1,503

# 정상화 EPS: COVID 전 5년 평균 (2015~2019)
AVG_EPS_PRE_COVID = int(sum(ANNUAL_EPS[y] for y in range(2015, 2020)) / 5)  # 약 1,939

# 배당수익률
DIV_YIELD = DPS_2024 / PRICE  # 약 6.6%

# FCF 수익률
FCF_YIELD = FCF_2024 / MARKET_CAP


# ============================================================
# SHEET 1: 밸류에이션 종합 대시보드
# ============================================================
ws1 = wb.active
ws1.title = "종합"
ws1.sheet_properties.tabColor = "1B2A4A"
set_widths(ws1, [22, 18, 18, 18, 18, 22])
setup_print(ws1)

# Title bar
row = 1
ws1.merge_cells('A1:F2')
c = ws1.cell(row=1, column=1, value="강원랜드(035250) 밸류에이션 분석")
c.font = title_font
c.fill = title_fill
c.alignment = Alignment(horizontal="center", vertical="center")

row = 3
ws1.merge_cells('A3:F3')
ws1.cell(row=3, column=1, value=f"기준일: 2026.02.06 | 현재가: {PRICE:,}원 | 시가총액: {fmt(MARKET_CAP)}").font = Font(name="맑은 고딕", size=11, bold=True, color=DARK)
ws1.cell(row=3, column=1).alignment = center

# Key metrics box
row = 5
row = section_title(ws1, row, "기본 정보", 6)
info_data = [
    ["현재 주가", fmt_won(PRICE), "발행주식수", f"{SHARES:,}주", "시가총액", fmt(MARKET_CAP)],
    ["자기주식", f"{SHARES_TREASURY:,}주", "유통주식수", f"{NET_SHARES:,}주", "유통시총", fmt(PRICE * NET_SHARES)],
    ["2024 EPS", fmt_won(EPS_2024), "2024 BPS", fmt_won(int(BPS)), "2024 DPS", fmt_won(DPS_2024)],
    ["2024 매출", fmt(REV_2024), "2024 영업이익", fmt(OP_2024), "2024 순이익", fmt(NI_2024)],
]
for data in info_data:
    for i in range(0, 6, 2):
        ws1.cell(row=row, column=i+1, value=data[i]).font = d_bold
        ws1.cell(row=row, column=i+1).fill = blue_fill
        ws1.cell(row=row, column=i+1).alignment = left
        ws1.cell(row=row, column=i+1).border = thin_border
        ws1.cell(row=row, column=i+2, value=data[i+1]).font = d_blue
        ws1.cell(row=row, column=i+2).fill = white_fill
        ws1.cell(row=row, column=i+2).alignment = center
        ws1.cell(row=row, column=i+2).border = thin_border
    row += 1

# Valuation Summary
row += 1
row = section_title(ws1, row, "밸류에이션 멀티플 종합", 6)
write_header(ws1, row, ["지표", "산출 방식", "값", "판정", "업종 평균(참고)", "비고"])
row += 1

valuation_rows = [
    ["PER (2024)", f"주가 / EPS({fmt_won(EPS_2024)})", f"{PRICE/EPS_2024:.1f}배", "저평가", "10~15배", "확정 실적 기준"],
    ["PER (2025E)", f"주가 / EPS({fmt_won(EPS_2025E)})", f"{PRICE/EPS_2025E:.1f}배", "적정", "10~15배", "2025 잠정실적 기준"],
    ["PER (Trailing 4Q)", f"주가 / T4Q EPS({fmt_won(int(TRAIL_EPS))})", f"{PRICE/TRAIL_EPS:.1f}배", "적정", "10~15배", "25Q1~Q4"],
    ["PER (보수적 3년평균)", f"주가 / 3Y EPS({fmt_won(AVG_EPS_3Y)})", f"{PRICE/AVG_EPS_3Y:.1f}배", "적정", "10~15배", "2022~2024 평균"],
    ["PBR", f"주가 / BPS({fmt_won(int(BPS))})", f"{PRICE/BPS:.2f}배", "저평가", "1.0~2.0배", "2024말 자본 기준"],
    ["EV/EBITDA (2024)", f"EV({fmt(EV)}) / EBITDA({fmt(EBITDA_2024)})", f"{EV/EBITDA_2024:.1f}배", "저평가", "8~12배", "순현금 기업"],
    ["PSR (2024)", f"시총 / 매출({fmt(REV_2024)})", f"{MARKET_CAP/REV_2024:.2f}배", "적정", "2.0~4.0배", "카지노업 프리미엄"],
    ["PCR", f"주가 / OpCF/주({fmt_won(int(OPCF_2024/SHARES))})", f"{PRICE/(OPCF_2024/SHARES):.1f}배", "저평가", "8~15배", "영업CF 기준"],
    ["배당수익률", f"DPS({fmt_won(DPS_2024)}) / 주가", pct(DIV_YIELD), "매우 양호", "1~3%", "2024년 배당 (고배당주)"],
    ["FCF 수익률", f"FCF/주({fmt_won(int(FCF_2024/SHARES))}) / 주가", pct(FCF_2024/SHARES/PRICE), "매우 양호", "3~6%", "CAPEX 차감 후"],
    ["배당성향", f"DPS / EPS", pct(DPS_2024/EPS_2024), "양호", "30~50%", "잉여현금 대비 여유"],
    ["ROE (2024)", f"순이익 / 평균자본", pct(ROE_2024), "양호", "8~15%", "2023-2024 평균자본"],
    ["무차입경영", "차입금 0원 / 순현금", fmt(abs(NET_DEBT)), "매우 양호", "-", "순현금 보유"],
]

for vals in valuation_rows:
    judge = vals[3]
    if "매우 저" in judge or "매우 양" in judge:
        judge_font = d_green
        judge_fill = green_fill
    elif "저평가" in judge or "양호" in judge:
        judge_font = Font(name="맑은 고딕", size=10, bold=True, color="2E86C1")
        judge_fill = PatternFill("solid", fgColor="D6EAF8")
    elif "적정" in judge:
        judge_font = d_bold
        judge_fill = gold_fill
    else:
        judge_font = d_red
        judge_fill = red_fill

    write_row(ws1, row, vals,
              fonts=[d_bold, d_font, d_blue, judge_font, d_font, note_font],
              fills=[lblue_fill, white_fill, gold_fill, judge_fill, gray_fill, white_fill],
              aligns=[left, left, center, center, center, left])
    row += 1

row += 1
ws1.cell(row=row, column=1, value="* 판정 기준: 카지노/리조트업종 글로벌 평균 대비 상대 평가").font = note_font
row += 1
ws1.cell(row=row, column=1, value="* 발행주식수: 213,940,500주 (자기주식 13,534,816주 포함) / 무차입경영 (차입금 0원)").font = note_font
row += 1
ws1.cell(row=row, column=1, value="* COVID(2020-2021) 비정상 실적 → 평균 계산 시 제외 고려").font = note_font

print("  [1/5] 종합 대시보드 완료")


# ============================================================
# SHEET 2: PER 다각도 분석
# ============================================================
ws2 = wb.create_sheet("PER분석")
ws2.sheet_properties.tabColor = "2980B9"
set_widths(ws2, [24, 16, 16, 16, 16, 20])
setup_print(ws2)

row = 1
row = section_title(ws2, row, f"PER 다각도 분석 (현재가 {PRICE:,}원 기준)", 6)

# A. EPS 산출 방식별 PER
row = section_title(ws2, row, "A. EPS 산출 방식별 PER", 6)
write_header(ws2, row, ["산출 방식", "순이익(억)", "EPS(원)", "PER(배)", "의미", "비고"])
row += 1

per_methods = [
    ["2024 확정 (사업보고서)", int(NI_2024/억), EPS_2024, PRICE/EPS_2024,
     "가장 신뢰 높은 확정치", "2024 연결 기준"],
    ["2025E (잠정실적 합산)", int(CUM_25_NI/억), EPS_2025E, PRICE/EPS_2025E,
     "2025 연간 잠정실적", "Q1~Q4 합산"],
    ["Trailing 4Q (25Q1~Q4)", int(TRAIL_NI/억), int(TRAIL_EPS), PRICE/TRAIL_EPS,
     "직전 4분기 실적 합산", "= 2025E와 동일"],
    ["보수적 3년평균 (22~24)", int((ANNUAL_NI[2022]+ANNUAL_NI[2023]+ANNUAL_NI[2024])/(3*억)), AVG_EPS_3Y, PRICE/AVG_EPS_3Y,
     "COVID 제외 최근 3년", "변동성 완화"],
    ["정상화 5년평균 (15~19)", int(sum(ANNUAL_NI[y] for y in range(2015,2020))/(5*억)), AVG_EPS_PRE_COVID, PRICE/AVG_EPS_PRE_COVID,
     "COVID 전 정상 수준", "구조적 수익력"],
]

for vals in per_methods:
    per_val = vals[3]
    if per_val < 10:
        per_font = d_green
        per_fill = green_fill
    elif per_val < 13:
        per_font = d_blue
        per_fill = PatternFill("solid", fgColor="D6EAF8")
    else:
        per_font = d_bold
        per_fill = gold_fill

    write_row(ws2, row,
              [vals[0], vals[1], f"{vals[2]:,}", f"{per_val:.2f}배", vals[4], vals[5]],
              fonts=[d_bold, d_font, d_blue, per_font, d_font, note_font],
              fills=[lblue_fill, white_fill, gold_fill, per_fill, white_fill, gray_fill],
              aligns=[left, right, right, center, left, left])
    row += 1

# B. 분기별 실적 추이
row += 1
row = section_title(ws2, row, "B. 2025 분기별 실적 추이 (백만원)", 6)
write_header(ws2, row, ["분기", "매출액", "영업이익", "OPM", "순이익", "비고"])
row += 1

quarters_25 = [
    ("25Q1", Q1_25), ("25Q2", Q2_25), ("25Q3", Q3_25), ("25Q4", Q4_25),
]

for name, q in quarters_25:
    rev_m = int(q["rev"] / 1e6)
    op_m = int(q["op"] / 1e6)
    ni_m = int(q["ni"] / 1e6)
    opm = q["op"] / q["rev"] if q["rev"] else 0
    note = ""
    if name == "25Q3":
        note = "성수기 (여름 관광)"
    elif name == "25Q4":
        note = "비수기 (겨울)"

    opm_font = d_green if opm > 0.18 else (d_red if opm < 0.10 else d_font)

    write_row(ws2, row, [name, f"{rev_m:,}", f"{op_m:,}", f"{opm*100:.1f}%", f"{ni_m:,}", note],
              fonts=[d_bold, d_font, d_font, opm_font, d_font, note_font],
              fills=[blue_fill, green_fill, green_fill, green_fill, green_fill, green_fill],
              aligns=[center, right, right, center, right, left])
    row += 1

# 2025 합계 row
write_row(ws2, row, ["2025E 합계",
                      f"{int(CUM_25_REV/1e6):,}", f"{int(CUM_25_OP/1e6):,}",
                      f"{CUM_25_OP/CUM_25_REV*100:.1f}%",
                      f"{int(CUM_25_NI/1e6):,}", "잠정실적"],
          fonts=[d_bold, d_bold, d_bold, d_bold, d_bold, note_font],
          fills=[gold_fill]*6, aligns=[center, right, right, center, right, left])

# C. 10년 EPS 추이
row += 2
row = section_title(ws2, row, "C. 10년 EPS / DPS 추이", 6)
write_header(ws2, row, ["연도", "EPS(원)", "DPS(원)", "배당성향", "순이익(억)", "비고"])
row += 1

for yr in range(2015, 2025):
    eps = ANNUAL_EPS[yr]
    dps = ANNUAL_DPS[yr]
    ni = ANNUAL_NI[yr]
    payout = dps / eps * 100 if eps > 0 else 0
    note = ""
    if yr in (2020, 2021):
        note = "COVID 영향 (비정상)"
    elif yr == 2024:
        note = "역대 최고 순이익"

    eps_font = d_red if eps < 0 else (d_green if eps >= 2000 else d_font)
    yr_fill = red_fill if yr in (2020, 2021) else white_fill

    write_row(ws2, row,
              [str(yr), f"{eps:,}", f"{dps:,}",
               f"{payout:.1f}%" if eps > 0 else "-",
               f"{int(ni/억):,}", note],
              fonts=[d_bold, eps_font, d_font, d_font, d_font, note_font],
              fills=[lblue_fill, gold_fill if yr == 2024 else yr_fill, yr_fill, yr_fill, yr_fill, yr_fill],
              aligns=[center, right, right, center, right, left])
    row += 1

# D. 적정PER 목표주가
row += 1
row = section_title(ws2, row, "D. 적정 PER 적용 목표주가 (카지노업 적정 10~15배)", 6)
write_header(ws2, row, ["기준", "EPS(원)", "PER 8배", "PER 10배", "PER 12배", "PER 15배"])
row += 1

eps_scenarios = [
    ("2024 확정", EPS_2024),
    ("2025E 잠정", EPS_2025E),
    ("3년평균 (22~24)", AVG_EPS_3Y),
    ("정상화 (15~19)", AVG_EPS_PRE_COVID),
]

for label, eps in eps_scenarios:
    write_row(ws2, row,
              [label, f"{eps:,}", fmt_won(eps*8), fmt_won(eps*10), fmt_won(eps*12), fmt_won(eps*15)],
              fonts=[d_bold, d_blue, d_font, d_font, d_green, d_green],
              fills=[lblue_fill, gold_fill,
                     green_fill if eps*8 > PRICE else white_fill,
                     green_fill if eps*10 > PRICE else white_fill,
                     green_fill, green_fill],
              aligns=[left, right, right, right, right, right])
    row += 1

row += 1
ws2.cell(row=row, column=1, value=f"-> 현재 {PRICE:,}원은 2024 확정 EPS 기준 PER {PRICE/EPS_2024:.1f}배. 카지노업 적정 PER 10~15배 적용 시 22,570~33,855원").font = d_bold

print("  [2/5] PER 분석 완료")


# ============================================================
# SHEET 3: PBR/ROE 분석
# ============================================================
ws3 = wb.create_sheet("PBR_ROE")
ws3.sheet_properties.tabColor = "8E44AD"
set_widths(ws3, [20, 16, 16, 16, 16, 22])
setup_print(ws3)

row = 1
row = section_title(ws3, row, "PBR / ROE 분석", 6)

# A. Historical BPS & ROE
row = section_title(ws3, row, "A. 연도별 자본/BPS/ROE 추이", 6)
write_header(ws3, row, ["연도", "자본(억)", "BPS(원)", "ROE", "EPS(원)", "순이익(억)"])
row += 1

# 역사적 자본 데이터 (추정 및 실제 혼합)
hist_equity = {
    2015: 3070000000000,  # 추정
    2016: 3243000000000,
    2017: 3440000000000,
    2018: 3530000000000,
    2019: 3635000000000,
    2020: 3197000000000,  # COVID 손실
    2021: 3116000000000,
    2022: 3316000000000,
    2023: EQUITY_2023,
    2024: EQUITY_2024,
}

prev_eq = None
for yr in range(2015, 2025):
    eq = hist_equity[yr]
    bps = int(eq / SHARES)
    ni = ANNUAL_NI[yr]
    eps = ANNUAL_EPS[yr]
    if prev_eq:
        roe = ni / ((eq + prev_eq) / 2)
    else:
        roe = ni / eq
    prev_eq = eq

    roe_font = d_green if roe > 0.12 else (d_red if roe < 0 else d_font)
    yr_fill = red_fill if yr in (2020, 2021) else white_fill

    write_row(ws3, row,
              [str(yr), int(eq/억), f"{bps:,}", f"{roe*100:.1f}%", f"{eps:,}", int(ni/억)],
              fonts=[d_bold, d_font, d_blue, roe_font, d_font, d_font],
              fills=[lblue_fill, yr_fill, gold_fill if yr == 2024 else yr_fill, yr_fill, yr_fill, yr_fill],
              aligns=[center, right, right, center, right, right])
    row += 1

# Current PBR
row += 1
row = section_title(ws3, row, "B. 현재 PBR 분석", 6)
write_header(ws3, row, ["항목", "값", "", "", "", ""])
row += 1

pbr_info = [
    ("2024말 자본총계", fmt(EQUITY_2024)),
    ("BPS (자본/발행주식)", fmt_won(int(BPS))),
    ("현재 PBR", f"{PRICE/BPS:.2f}배"),
    ("2024 ROE", pct(ROE_2024)),
    ("PBR 1.0배 주가", fmt_won(int(BPS))),
    ("PBR 1.5배 주가", fmt_won(int(BPS * 1.5))),
    ("PBR 2.0배 주가", fmt_won(int(BPS * 2.0))),
]
for label, val in pbr_info:
    ws3.cell(row=row, column=1, value=label).font = d_bold
    ws3.cell(row=row, column=1).fill = lblue_fill
    ws3.cell(row=row, column=1).alignment = left
    ws3.cell(row=row, column=1).border = thin_border
    ws3.cell(row=row, column=2, value=val).font = d_blue
    ws3.cell(row=row, column=2).fill = gold_fill
    ws3.cell(row=row, column=2).alignment = center
    ws3.cell(row=row, column=2).border = thin_border
    row += 1

# C. PBR vs ROE 정당성
row += 1
row = section_title(ws3, row, "C. PBR과 ROE의 관계 (카지노업 적정 PBR 1.0~2.0배)", 6)
ws3.cell(row=row, column=1, value="이론: PBR = ROE / ke (지속가능 ROE가 자본비용을 초과하면 PBR > 1.0 정당)").font = note_font
row += 1
ws3.cell(row=row, column=1, value="강원랜드 ROE 12.1% > ke(자본비용) 10% → PBR 1.0배 이상이 정당. 현재 0.97배는 할인 상태").font = note_font
row += 2

write_header(ws3, row, ["시나리오", "지속ROE", "적정PBR", "적정주가", "현재가 대비", "근거"])
row += 1

pbr_scenarios = [
    ("보수적 (ROE=ke)", "10.0%", "1.0배", int(BPS * 1.0), "순자산가치"),
    ("기본 (현재ROE)", "12.1%", "1.2배", int(BPS * 1.2), "현재 수익력 반영"),
    ("적극적 (COVID전)", "14.0%", "1.5배", int(BPS * 1.5), "2015~2019 평균 수준 회복"),
    ("낙관적 (피크)", "16.0%", "2.0배", int(BPS * 2.0), "2015~2016 수준 회복"),
]

for label, roe_str, pbr_str, fair, note in pbr_scenarios:
    upside = (fair - PRICE) / PRICE
    upside_str = f"{upside*100:+.1f}%"
    up_font = d_green if upside > 0 else d_red
    up_fill = green_fill if upside > 0 else red_fill

    write_row(ws3, row,
              [label, roe_str, pbr_str, fmt_won(fair), upside_str, note],
              fonts=[d_bold, d_font, d_font, d_blue, up_font, note_font],
              fills=[lblue_fill, white_fill, white_fill, gold_fill, up_fill, white_fill],
              aligns=[left, center, center, right, center, left])
    row += 1

# D. RIM (Residual Income Model)
row += 1
row = section_title(ws3, row, "D. 잔여이익모델(RIM) 적정주가", 6)
ws3.cell(row=row, column=1, value="산식: 적정가 = BPS + BPS x (ROE - ke) / (ke - g)").font = note_font
row += 1
ws3.cell(row=row, column=1, value="ke(자기자본비용) = 무위험이자율 3.5% + beta(0.8) x ERP(6.5%) = 8.7% -> 보수적 10% 적용").font = note_font
row += 1

write_header(ws3, row, ["시나리오", "지속ROE", "ke", "성장률(g)", "적정주가", "현재가 대비"])
row += 1

rim_scenarios = [
    ("보수적 (ROE=ke)", 0.10, 0.10, 0.01, "ROE가 자본비용과 동일"),
    ("기본 (현재ROE)", 0.121, 0.10, 0.01, "2024년 ROE 12.1% 유지"),
    ("적극적 (14%)", 0.14, 0.10, 0.02, "COVID전 수준 회복"),
    ("낙관적 (16%)", 0.16, 0.10, 0.02, "피크 수익성 접근"),
]

for label, roe, ke, g, note in rim_scenarios:
    if roe == ke:
        fair = BPS  # ROE=ke면 적정가 = BPS
    else:
        fair = BPS * (1 + (roe - ke) / (ke - g))
    upside = (fair - PRICE) / PRICE
    upside_str = f"{upside*100:+.1f}%"
    up_font = d_green if upside > 0 else d_red
    up_fill = green_fill if upside > 0 else red_fill

    write_row(ws3, row,
              [label, f"{roe*100:.1f}%", f"{ke*100:.0f}%", f"{g*100:.0f}%", fmt_won(int(fair)), upside_str],
              fonts=[d_bold, d_font, d_font, d_font, d_blue, up_font],
              fills=[lblue_fill, white_fill, white_fill, white_fill, gold_fill, up_fill],
              aligns=[left, center, center, center, right, center])
    row += 1

row += 1
rim_base = BPS * (1 + (0.121 - 0.10) / (0.10 - 0.01))
rim_active = BPS * (1 + (0.14 - 0.10) / (0.10 - 0.02))
ws3.cell(row=row, column=1, value=f"-> RIM 기본 시나리오 적정가 {int(rim_base):,}원. 적극적 시나리오 {int(rim_active):,}원. 현재가 {PRICE:,}원은 ROE 10~12% 수준 반영.").font = d_bold

print("  [3/5] PBR/ROE 완료")


# ============================================================
# SHEET 4: EV/EBITDA & FCF
# ============================================================
ws4 = wb.create_sheet("EV_EBITDA_FCF")
ws4.sheet_properties.tabColor = "E67E22"
set_widths(ws4, [24, 18, 18, 18, 24])
setup_print(ws4)

row = 1
row = section_title(ws4, row, "EV/EBITDA & FCF 밸류에이션", 5)

# A. EV 산출
row = section_title(ws4, row, "A. Enterprise Value 산출 (무차입경영)", 5)
write_header(ws4, row, ["항목", "금액(억)", "비고", "", ""])
row += 1

ev_items = [
    ("시가총액", int(MARKET_CAP/억), f"주가 {PRICE:,}원 x {SHARES:,}주"),
    ("(+) 총차입금", 0, "무차입경영: 단기+장기 차입금 0원"),
    ("(-) 현금성자산", int(CASH_2024/억), "현금및현금성자산"),
    ("(=) 순차입금", int(NET_DEBT/억), "순현금 상태 (음수 = 현금이 더 많음)"),
    ("(=) EV", int(EV/억), "시가총액 + 순차입금 (< 시가총액)"),
]
for label, amt, note in ev_items:
    is_total = label.startswith("(=)")
    write_row(ws4, row, [label, f"{amt:,}", note, "", ""],
              fonts=[d_bold if is_total else d_font, d_blue if is_total else d_font, note_font, d_font, d_font],
              fills=[gold_fill if is_total else lblue_fill, gold_fill if is_total else white_fill, white_fill, white_fill, white_fill],
              aligns=[left, right, left, left, left])
    row += 1

row += 1
ws4.cell(row=row, column=1, value="** 강원랜드는 무차입경영으로 EV가 시가총액보다 약 2,445억 작음 (순현금 효과)").font = Font(name="맑은 고딕", size=10, bold=True, color="27AE60")

# B. EBITDA
row += 2
row = section_title(ws4, row, "B. EBITDA 산출", 5)
write_header(ws4, row, ["항목", "2024(억)", "비고", "", ""])
row += 1

ebitda_items = [
    ("영업이익", int(OP_2024/억), "CIS 영업이익"),
    ("(+) 감가상각비 합계", int(DA_2024/억), "유형+무형자산 감가상각"),
    ("(=) EBITDA", int(EBITDA_2024/억), ""),
    ("EBITDA 마진", "", f"{EBITDA_2024/REV_2024*100:.1f}%"),
]
for label, amt, note in ebitda_items:
    is_total = label.startswith("(=)") or label == "EBITDA 마진"
    val = f"{amt:,}" if isinstance(amt, int) else note
    write_row(ws4, row, [label, val if isinstance(amt, int) else "", note if isinstance(amt, int) else val, "", ""],
              fonts=[d_bold if is_total else d_font, d_blue if is_total else d_font, note_font, d_font, d_font],
              fills=[gold_fill if is_total else lblue_fill, gold_fill if is_total else white_fill, white_fill, white_fill, white_fill],
              aligns=[left, right, left, left, left])
    row += 1

# C. EV/EBITDA 밸류에이션
row += 1
row = section_title(ws4, row, "C. EV/EBITDA 밸류에이션", 5)
write_header(ws4, row, ["기준", "EBITDA(억)", "EV/EBITDA", "적정EV(10배)", "적정주가(10배)"])
row += 1

# 2025E EBITDA 추정 (영업이익 + 감가상각 유지)
EBITDA_2025E = CUM_25_OP + DA_2024  # 2025 잠정 OP + 2024년 수준 DA 유지

ev_scenarios = [
    ("2024 확정", int(EBITDA_2024/억)),
    ("2025E (잠정OP+DA유지)", int(EBITDA_2025E/억)),
]

for label, ebitda_b in ev_scenarios:
    ev_ebitda = (EV/억) / ebitda_b
    fair_ev_10x = ebitda_b * 10
    fair_eq = (fair_ev_10x - int(NET_DEBT/억)) * 억
    fair_price = int(fair_eq / SHARES)

    write_row(ws4, row,
              [label, f"{ebitda_b:,}", f"{ev_ebitda:.1f}배", f"{fair_ev_10x:,}억", fmt_won(fair_price)],
              fonts=[d_bold, d_font, d_blue, d_font, d_green],
              fills=[lblue_fill, white_fill, gold_fill, white_fill, green_fill],
              aligns=[left, right, center, right, right])
    row += 1

# 추가: 다양한 배수 적용
row += 1
row = section_title(ws4, row, "C-2. EBITDA 배수별 적정주가 (2024 기준)", 5)
write_header(ws4, row, ["EV/EBITDA", "적정EV(억)", "적정주가", "현재가 대비", "비고"])
row += 1

for mult in [6, 8, 10, 12, 14]:
    fair_ev = int(EBITDA_2024/억) * mult
    fair_eq_amt = (fair_ev - int(NET_DEBT/억)) * 억
    fair_p = int(fair_eq_amt / SHARES)
    upside = (fair_p - PRICE) / PRICE
    up_font = d_green if upside > 0 else d_red
    up_fill = green_fill if upside > 0 else red_fill
    note = ""
    if mult == 10:
        note = "카지노업 평균 하단"
    elif mult == 12:
        note = "카지노업 평균 상단"

    write_row(ws4, row,
              [f"{mult}배", f"{fair_ev:,}", fmt_won(fair_p), f"{upside*100:+.1f}%", note],
              fonts=[d_bold, d_font, d_blue, up_font, note_font],
              fills=[lblue_fill, white_fill, gold_fill, up_fill, white_fill],
              aligns=[center, right, right, center, left])
    row += 1

# D. FCF 분석
row += 1
row = section_title(ws4, row, "D. FCF(잉여현금흐름) 분석", 5)
write_header(ws4, row, ["항목", "2024(억)", "비고", "", ""])
row += 1

fcf_items = [
    ("영업활동현금흐름", int(OPCF_2024/억), "카지노 현금 영업의 강점"),
    ("(-) 설비투자(CAPEX)", int(CAPEX_2024/억), "리조트 시설 유지보수 등"),
    ("(=) FCF", int(FCF_2024/억), ""),
    ("FCF/주", "", fmt_won(int(FCF_2024/SHARES))),
    ("FCF 수익률", "", f"{FCF_2024/SHARES/PRICE*100:.1f}%"),
    ("", "", ""),
    ("배당금 지급", int(DIV_PAID_2024/억), f"{fmt_won(DPS_2024)}/주"),
    ("배당성향 (DPS/EPS)", "", pct(DPS_2024/EPS_2024)),
    ("배당수익률", "", pct(DIV_YIELD)),
    ("FCF 대비 배당 비율", "", pct(DIV_PAID_2024/FCF_2024)),
]

for label, amt, note in fcf_items:
    if not label:
        row += 1
        continue
    is_total = "(=)" in label or "수익률" in label or "성향" in label or "비율" in label
    val_str = f"{amt:,}" if isinstance(amt, int) and amt > 0 else note
    write_row(ws4, row, [label, val_str if isinstance(amt, int) and amt > 0 else "",
                          note if isinstance(amt, int) and amt > 0 else val_str, "", ""],
              fonts=[d_bold if is_total else d_font, d_blue if is_total else d_font, note_font, d_font, d_font],
              fills=[gold_fill if is_total else lblue_fill, gold_fill if is_total else white_fill, white_fill, white_fill, white_fill],
              aligns=[left, right, left, left, left])
    row += 1

# E. 10년 영업실적 추이
row += 1
row = section_title(ws4, row, "E. 10년 매출/영업이익 추이 (억원)", 5)
write_header(ws4, row, ["연도", "매출(억)", "영업이익(억)", "OPM", "비고"])
row += 1

for yr in range(2015, 2025):
    rev = ANNUAL_REV[yr]
    op = ANNUAL_OP[yr]
    opm = op / rev if rev > 0 else 0
    note = ""
    if yr == 2020:
        note = "COVID 영업중단"
    elif yr == 2021:
        note = "COVID 부분 회복"
    elif yr == 2024:
        note = "역대 최고 실적 접근"

    opm_font = d_green if opm > 0.20 else (d_red if opm < 0 else d_font)
    yr_fill = red_fill if yr in (2020, 2021) else white_fill

    write_row(ws4, row,
              [str(yr), f"{int(rev/억):,}", f"{int(op/억):,}", f"{opm*100:.1f}%", note],
              fonts=[d_bold, d_font, d_font, opm_font, note_font],
              fills=[lblue_fill, yr_fill, yr_fill, yr_fill, yr_fill],
              aligns=[center, right, right, center, left])
    row += 1

row += 1
ws4.cell(row=row, column=1, value=f"-> FCF 수익률 {FCF_2024/SHARES/PRICE*100:.1f}%, 배당수익률 {DIV_YIELD*100:.1f}%. 무차입+고배당+FCF 안정 → 밸류 트랩 아닌 가치주.").font = d_bold

print("  [4/5] EV/EBITDA/FCF 완료")


# ============================================================
# SHEET 5: 시나리오별 목표주가 종합
# ============================================================
ws5 = wb.create_sheet("목표주가")
ws5.sheet_properties.tabColor = "2ECC71"
set_widths(ws5, [22, 18, 18, 18, 22])
setup_print(ws5)

row = 1
row = section_title(ws5, row, f"시나리오별 목표주가 종합 (현재가 {PRICE:,}원)", 5)

# A. 방법론별 적정가 범위
row = section_title(ws5, row, "A. 밸류에이션 방법론별 적정가 레인지", 5)
write_header(ws5, row, ["방법론", "보수적", "기본", "적극적", "산출 근거"])
row += 1

# RIM 계산
rim_cons = int(BPS * 1.0)  # ROE=ke
rim_base_val = int(BPS * (1 + (0.121 - 0.10) / (0.10 - 0.01)))
rim_aggr = int(BPS * (1 + (0.14 - 0.10) / (0.10 - 0.02)))

# EV/EBITDA 역산 주가
def ev_to_price(ebitda_억, mult):
    fair_ev = ebitda_억 * mult
    fair_eq = (fair_ev - int(NET_DEBT/억)) * 억
    return int(fair_eq / SHARES)

ebitda_b = int(EBITDA_2024/억)

methods = [
    ("PER 방식", fmt_won(AVG_EPS_3Y * 8),
     fmt_won(EPS_2024 * 12), fmt_won(EPS_2024 * 15),
     "EPS x 목표PER (8/12/15배)"),
    ("PBR 방식", fmt_won(int(BPS * 1.0)),
     fmt_won(int(BPS * 1.5)), fmt_won(int(BPS * 2.0)),
     "BPS x 목표PBR (1.0/1.5/2.0배)"),
    ("EV/EBITDA 방식", fmt_won(ev_to_price(ebitda_b, 8)),
     fmt_won(ev_to_price(ebitda_b, 10)), fmt_won(ev_to_price(ebitda_b, 12)),
     "EBITDA x 목표배수 (8/10/12배)"),
    ("RIM 방식", fmt_won(rim_cons),
     fmt_won(rim_base_val), fmt_won(rim_aggr),
     "BPS x (1+(ROE-ke)/(ke-g))"),
    ("FCF 기반", fmt_won(int(FCF_2024 * 8 / SHARES)),
     fmt_won(int(FCF_2024 * 12 / SHARES)), fmt_won(int(FCF_2024 * 15 / SHARES)),
     "FCF x 목표배수 (8/12/15배)"),
    ("배당모델 (DDM)", fmt_won(int(DPS_2024 / 0.08)),
     fmt_won(int(DPS_2024 / 0.06)), fmt_won(int(DPS_2024 / 0.05)),
     "DPS / 요구수익률 (8%/6%/5%)"),
]

for label, cons, base, aggr, note in methods:
    write_row(ws5, row, [label, cons, base, aggr, note],
              fonts=[d_bold, d_font, d_blue, d_green, note_font],
              fills=[lblue_fill, red_fill, gold_fill, green_fill, white_fill],
              aligns=[left, right, right, right, left])
    row += 1

# B. 종합 시나리오
row += 1
row = section_title(ws5, row, "B. 종합 시나리오", 5)
write_header(ws5, row, ["시나리오", "목표주가", "현재가 대비", "전제조건", "확률(주관)"],
             fills=[PatternFill("solid", fgColor="C0392B"),
                    PatternFill("solid", fgColor="C0392B"),
                    PatternFill("solid", fgColor="C0392B"),
                    PatternFill("solid", fgColor="C0392B"),
                    PatternFill("solid", fgColor="C0392B")])
row += 1

# Bull case: 규제 완화 + COVID 완전 회복 -> PER 15배
bull_target = EPS_2024 * 15  # 2257 * 15 = 33,855
bull_upside = (bull_target - PRICE) / PRICE
write_row(ws5, row, ["강세 (Bull)", fmt_won(bull_target), f"{bull_upside*100:+.1f}%",
                      "규제 완화 + COVID 완전 회복 + PER 15배 리레이팅", "30%"],
          fonts=[d_bold, Font(name="맑은 고딕", size=12, bold=True, color="27AE60"), d_green, d_font, d_font],
          fills=[green_fill]*5, aligns=[center, right, center, left, center])
row += 1

# Base case: 현상 유지 -> PER 12배
base_target = EPS_2024 * 12  # 2257 * 12 = 27,084
base_upside = (base_target - PRICE) / PRICE
write_row(ws5, row, ["기본 (Base)", fmt_won(base_target), f"{base_upside*100:+.1f}%",
                      "현상 유지 + 배당 확대 + 카지노업 적정 PER 12배", "50%"],
          fonts=[d_bold, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), d_blue, d_font, d_font],
          fills=[gold_fill]*5, aligns=[center, right, center, left, center])
row += 1

# Bear case: 규제 강화 + 경기침체 -> PER 8배
bear_target = EPS_2024 * 8  # 2257 * 8 = 18,056
bear_upside = (bear_target - PRICE) / PRICE
write_row(ws5, row, ["약세 (Bear)", fmt_won(bear_target), f"{bear_upside*100:+.1f}%",
                      "규제 강화 + 경기 침체 + 실적 둔화 PER 8배", "20%"],
          fonts=[d_bold, Font(name="맑은 고딕", size=12, bold=True, color="C0392B"), d_red, d_font, d_font],
          fills=[red_fill]*5, aligns=[center, right, center, left, center])
row += 1

# Expected value
exp_val = int(bull_target * 0.30 + base_target * 0.50 + bear_target * 0.20)
exp_upside = (exp_val - PRICE) / PRICE
row += 1
write_row(ws5, row, ["확률가중 기대값", fmt_won(exp_val), f"{exp_upside*100:+.1f}%",
                      "Bull x 30% + Base x 50% + Bear x 20%", ""],
          fonts=[Font(name="맑은 고딕", size=12, bold=True, color=NAVY)] * 5,
          fills=[PatternFill("solid", fgColor="AED6F1")] * 5,
          aligns=[center, right, center, left, center])

# C. 핵심 체크포인트
row += 2
row = section_title(ws5, row, "C. 핵심 체크포인트", 5)

checkpoints = [
    ("현재 PER 7.8배", "카지노업 적정 PER 10~15배 대비 22~48% 할인. 규제 디스카운트 반영 중"),
    ("PBR 0.97배", "ROE 12.1% 대비 저평가. 자본 이하 매수 가능한 구간"),
    ("배당수익률 6.6%", "고배당주. DPS 1,170원, 배당성향 51.8%. 시중금리 대비 매력적"),
    ("무차입경영", "차입금 0원, 순현금 2,445억. 재무건전성 최상위. EV < 시가총액"),
    ("FCF 수익률 10.0%", f"연간 FCF {int(FCF_2024/억):,}억. 시총 대비 높은 현금 창출력"),
    ("COVID 이후 회복", "2024 매출 COVID전(2019) 93% 수준 회복. 순이익은 2019 대비 129%"),
    ("2025 실적 둔화", f"2025E 순이익 {int(CUM_25_NI/억):,}억 (YoY -31%). EPS {EPS_2025E:,}원으로 감소"),
    ("규제 리스크", "한국 유일 합법 내국인 카지노. 정부 규제/정책 변동에 민감"),
    ("성장 한계", "입장객 수 규제, 배팅한도 제한. 규제 완화 없이 매출 성장 제한적"),
    ("배당 안전성", f"FCF 대비 배당 비율 {DIV_PAID_2024/FCF_2024*100:.0f}%. 배당 여력 충분"),
]

for label, desc in checkpoints:
    ws5.cell(row=row, column=1, value=label).font = d_bold
    ws5.cell(row=row, column=1).fill = lblue_fill
    ws5.cell(row=row, column=1).alignment = left
    ws5.cell(row=row, column=1).border = thin_border
    ws5.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws5.cell(row=row, column=2, value=desc).font = d_font
    ws5.cell(row=row, column=2).alignment = left
    ws5.cell(row=row, column=2).border = thin_border
    row += 1

# D. 종합 결론
row += 1
row = section_title(ws5, row, "D. 종합 결론", 5)
ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
ws5.cell(row=row, column=1, value=f"확률가중 기대값 {exp_val:,}원 (현재가 대비 {exp_upside*100:+.1f}%). "
    f"PER {PRICE/EPS_2024:.1f}배 + PBR {PRICE/BPS:.2f}배 + 배당 {DIV_YIELD*100:.1f}% + 무차입경영. "
    f"규제 디스카운트 감안해도 저평가 구간. 배당+자본이득 관점에서 매력적.").font = Font(name="맑은 고딕", size=11, bold=True, color=NAVY)
ws5.cell(row=row, column=1).alignment = left

print("  [5/5] 목표주가 완료")


# === SAVE ===
wb.save(OUT)
print(f"\n밸류에이션 보고서 생성 완료: {OUT}")
print(f"시트: {wb.sheetnames}")
