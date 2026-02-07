# -*- coding: utf-8 -*-
"""강원랜드 2025년 이익 역성장 분석 보고서"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

BASE = os.path.dirname(os.path.abspath(__file__))
억 = 100_000_000

# ============================================================
#  CONSTANTS
# ============================================================
COMPANY = "강원랜드"
PRICE = 17690
SHARES = 213940500

# === 연간 실적 (연결, 원) ===
ANNUAL = {
    2015: {"rev":1634441985990,"op":422764543000,"ni":466119741406,"eps":2178,"dps":980},
    2016: {"rev":1703131541003,"op":480490893000,"ni":480264780048,"eps":2242,"dps":990},
    2017: {"rev":1601291247063,"op":460085942000,"ni":460843127006,"eps":2159,"dps":990},
    2018: {"rev":1445736946832,"op":292485625000,"ni":314063000747,"eps":1467,"dps":900},
    2019: {"rev":1524006966734,"op":339782744000,"ni":353337372000,"eps":1651,"dps":900},
    2020: {"rev":479173424993, "op":-200253285781,"ni":-291290000000,"eps":-1361,"dps":0},
    2021: {"rev":788430938373, "op":-70474730032, "ni":-11056000000, "eps":-52,  "dps":0},
    2022: {"rev":1272539665429,"op":145068791398,"ni":122207000000,"eps":570,  "dps":350},
    2023: {"rev":1392740543618,"op":258505699379,"ni":341061017856,"eps":1681, "dps":930},
    2024: {"rev":1426862781802,"op":285790454716,"ni":456895295638,"eps":2257, "dps":1170},
}

# === 2025 잠정실적 (연결) ===
CUM25 = {"rev":1476726e6, "op":235176e6, "ni":316516e6}
CUM25_EPS = int(CUM25["ni"] / SHARES)

# === 분기별 실적 (별도/연결 혼재, 잠정실적 기준) ===
Q24 = [
    {"q":"1Q24","rev":368163,"op":75780,"opm":20.6},
    {"q":"2Q24","rev":337719,"op":73416,"opm":21.7},
    {"q":"3Q24","rev":375541,"op":93039,"opm":24.8},
    {"q":"4Q24","rev":344338,"op":41769,"opm":12.1},
]
Q25 = [
    {"q":"1Q25","rev":365830,"op":77659,"opm":21.2},
    {"q":"2Q25","rev":360727,"op":57874,"opm":16.0},
    {"q":"3Q25","rev":383662,"op":71880,"opm":18.7},
    {"q":"4Q25","rev":365446,"op":29697,"opm":8.1},
]

# === P&L 구조 (연결 사업보고서, 억원) ===
PNL = {
    2022: {"rev":12707,"cogs":9218,"gp":3489,"sga":1313,"op":2176,
           "other_inc":46,"other_exp":229,"fin_inc":280,"fin_exp":598,
           "pbt":1679,"tax":522,"ni":1156,
           "fin_eval":None,"fin_disp":None},
    2023: {"rev":13886,"cogs":9794,"gp":4092,"sga":1269,"op":2823,
           "other_inc":61,"other_exp":272,"fin_inc":2159,"fin_exp":77,
           "pbt":4572,"tax":1164,"ni":3409,
           "fin_eval":1756,"fin_disp":78},
    2024: {"rev":14269,"cogs":10080,"gp":4189,"sga":1331,"op":2858,
           "other_inc":1291,"other_exp":262,"fin_inc":1789,"fin_exp":28,
           "pbt":5653,"tax":1085,"ni":4569,
           "fin_eval":1313,"fin_disp":159},
}

# === 반기 비교 (연결, 억원) ===
H1 = {
    "24": {"rev":3380,"cogs":2343,"gp":1037,"sga":297,"op":739,
           "other_inc":765,"other_exp":25,"fin_inc":406,"fin_exp":7,
           "pbt":1884,"ni":1610},
    "25": {"rev":3608,"cogs":2683,"gp":925,"sga":340,"op":585,
           "other_inc":7,"other_exp":202,"fin_inc":432,"fin_exp":7,
           "pbt":817,"ni":628},
}

# === 퇴직급여/사외적립 (CF, 억원) ===
RETIRE = {
    2019: {"cost":216,"paid":73, "fund":-222},
    2020: {"cost":550,"paid":76, "fund":-641},
    2021: {"cost":291,"paid":100,"fund":-124},
    2022: {"cost":268,"paid":145,"fund":-309},
    2023: {"cost":167,"paid":269,"fund":-216},
    2024: {"cost":173,"paid":223,"fund":202},
}

# === 금융자산 (BS, 2024 억원) ===
FIN_ASSETS = {"현금":2445, "유동금융자산":9327, "비유동금융자산":19683}

# ============================================================
#  STYLE (동일 패턴)
# ============================================================
NAVY="1B2A4A"; DARK="2C3E6B"; MID="3A5BA0"; LB="D6E4F0"; LLB="EBF1F8"; W="FFFFFF"
GOLD_C="D4A843"; RED_C="C0392B"; GREEN_C="27AE60"; GRAY_C="F2F2F2"

title_font=Font(name="맑은 고딕",size=22,bold=True,color=W)
sub_font=Font(name="맑은 고딕",size=11,color="B0C4DE")
sec_font=Font(name="맑은 고딕",size=14,bold=True,color=NAVY)
sub2_font=Font(name="맑은 고딕",size=12,bold=True,color=DARK)
hdr_font=Font(name="맑은 고딕",size=10,bold=True,color=W)
df=Font(name="맑은 고딕",size=10)
db=Font(name="맑은 고딕",size=10,bold=True)
d_blue=Font(name="맑은 고딕",size=10,bold=True,color="0000FF")
d_grn=Font(name="맑은 고딕",size=10,bold=True,color=GREEN_C)
d_red=Font(name="맑은 고딕",size=10,bold=True,color=RED_C)
d_navy=Font(name="맑은 고딕",size=12,bold=True,color=NAVY)
sm=Font(name="맑은 고딕",size=9,color="666666")
warn_font=Font(name="맑은 고딕",size=10,bold=True,color="B8860B")

tf=PatternFill("solid",fgColor=NAVY)
hf=PatternFill("solid",fgColor=DARK)
mf=PatternFill("solid",fgColor=MID)
lf=PatternFill("solid",fgColor=LB)
llf=PatternFill("solid",fgColor=LLB)
gf=PatternFill("solid",fgColor=GRAY_C)
wf=PatternFill("solid",fgColor=W)
gld=PatternFill("solid",fgColor="FFF3CD")
rdf=PatternFill("solid",fgColor="F8D7DA")
gnf=PatternFill("solid",fgColor="D4EDDA")
blf=PatternFill("solid",fgColor="D6EAF8")

ca=Alignment(horizontal='center',vertical='center',wrap_text=True)
la=Alignment(horizontal='left',vertical='center',wrap_text=True)
ra=Alignment(horizontal='right',vertical='center')
tb=Border(left=Side('thin',color="D9D9D9"),right=Side('thin',color="D9D9D9"),
          top=Side('thin',color="D9D9D9"),bottom=Side('thin',color="D9D9D9"))
bb=Border(bottom=Side('medium',color=NAVY))
NF='#,##0'; PF='0.0%'; PF1='0.0%'

def sw(ws,w):
    for i,v in enumerate(w,1): ws.column_dimensions[get_column_letter(i)].width=v
def setup_print(ws):
    ws.page_setup.paperSize = Worksheet.PAPERSIZE_LETTER
    ws.page_setup.orientation = Worksheet.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.75, bottom=0.75,
        header=0.3, footer=0.3
    )
    ws.print_options.horizontalCentered = True
def wh(ws,r,h,fills=None):
    for i,v in enumerate(h,1):
        c=ws.cell(row=r,column=i,value=v); c.font=hdr_font; c.fill=fills[i-1] if fills else hf; c.alignment=ca; c.border=tb
def wr(ws,r,d,fonts=None,fills=None,als=None,nfs=None):
    for i,v in enumerate(d,1):
        c=ws.cell(row=r,column=i,value=v)
        c.font=fonts[i-1] if fonts else df; c.fill=fills[i-1] if fills else wf
        c.alignment=als[i-1] if als else ca; c.border=tb
        if nfs and i<=len(nfs) and nfs[i-1]: c.number_format=nfs[i-1]
def st(ws,r,t,ce=8):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ce)
    c=ws.cell(row=r,column=1,value=t); c.font=sec_font; c.border=bb; return r+1
def st2(ws,r,t,ce=8):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ce)
    c=ws.cell(row=r,column=1,value=t); c.font=sub2_font; c.border=bb; return r+1
def note(ws,r,t,ce=8):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ce)
    c=ws.cell(row=r,column=1,value=t); c.font=sm; c.alignment=la; return r+1
def bullet(ws,r,t,ce=8,font=None):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ce)
    c=ws.cell(row=r,column=1,value=f"  {t}"); c.font=font or df; c.alignment=la; return r+1

wb = Workbook()

# ============================================================
#  SHEET 1: 요약 대시보드
# ============================================================
ws = wb.active; ws.title = "요약"
sw(ws,[4,22,14,14,14,14,14,14])
setup_print(ws)

# Title banner
ws.merge_cells("A1:H2")
c=ws.cell(row=1,column=1,value=f"{COMPANY} 2025년 이익 역성장 분석"); c.font=title_font; c.fill=tf; c.alignment=ca
ws.merge_cells("A3:H3")
c=ws.cell(row=3,column=1,value="매출은 성장했는데 이익은 왜 줄었나?  |  구조적 문제인가, 일시적 문제인가?"); c.font=sub_font; c.fill=tf; c.alignment=ca
ws.row_dimensions[1].height=35; ws.row_dimensions[2].height=10

r = 5
r = st(ws,r,"A. 2025년 실적 개요")
wh(ws,r,["","항목","2024","2025","YoY","성격","",""])
r += 1
rows_data = [
    ["","매출",f"{14269:,}",f"{14767:,}","+3.5%","정상 성장","",""],
    ["","영업이익",f"{2858:,}",f"{2352:,}","-17.7%","구조적 우려","",""],
    ["","OPM","20.0%","15.9%","-4.1p","구조적 우려","",""],
    ["","순이익",f"{4569:,}",f"{3165:,}","-30.7%","대부분 일시적","",""],
    ["","EPS",f"{2257:,}",f"{CUM25_EPS:,}","","","",""],
]
tag_fills = {"정상 성장":gnf, "구조적 우려":rdf, "대부분 일시적":gld}
for row in rows_data:
    tag = row[5]
    fls = [wf]*8
    fts = [df]*8
    if tag in tag_fills:
        fls[5] = tag_fills[tag]
        fts[5] = warn_font if tag=="구조적 우려" else db
    fts[1] = db
    wr(ws,r,row,fonts=fts,fills=fls)
    r += 1
note(ws,r,"※ 단위: 억원 (연결 기준). OPM = 영업이익률. 2025년은 잠정실적 기준.")
r += 2

# 핵심 진단
r = st(ws,r,"B. 핵심 진단: 이익 역성장의 두 가지 층위")
r += 1
# Box 1 - 순이익
ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
c=ws.cell(row=r,column=2,value="순이익 -30.7%: 대부분 일시적"); c.font=Font(name="맑은 고딕",size=12,bold=True,color=W); c.fill=PatternFill("solid",fgColor="2E86C1"); c.alignment=ca
ws.merge_cells(start_row=r,start_column=5,end_row=r,end_column=7)
c=ws.cell(row=r,column=5,value="영업이익 -17.7%: 구조적 우려"); c.font=Font(name="맑은 고딕",size=12,bold=True,color=W); c.fill=PatternFill("solid",fgColor=RED_C); c.alignment=ca
r += 1
ws.merge_cells(start_row=r,start_column=2,end_row=r+3,end_column=4)
c=ws.cell(row=r,column=2,value="2024년 기타수익 1,291억 (전년 61억)\n→ 일회성 소멸이 NI 역성장의 절반+\n\n정상화 시 2024 NI ≈ 3,500억\n→ 2025 NI 3,165억은 약 -10% 수준")
c.font=df; c.fill=blf; c.alignment=Alignment(horizontal='left',vertical='top',wrap_text=True)
ws.merge_cells(start_row=r,start_column=5,end_row=r+3,end_column=7)
c=ws.cell(row=r,column=5,value="매출 +6.7% vs 매출원가 +14.5%\n→ 원가가 매출의 2배 속도로 증가\n\n원가율: 69.3% → 74.3% (+5.0p)\n인건비 절감 구조적으로 불가")
c.font=df; c.fill=rdf; c.alignment=Alignment(horizontal='left',vertical='top',wrap_text=True)
r += 5

# 요인 분류표
r = st(ws,r,"C. 역성장 요인 분류")
wh(ws,r,["","요인","영향 규모","일시/구조","걱정 수준","개선 가능성","",""])
r += 1
factors = [
    ["","기타수익 1,291억 소멸","NI -1,000억+","일시적","낮음","해당없음","",""],
    ["","금융자산평가 변동","NI ±수백억","일시적","낮음","해당없음","",""],
    ["","OPM 악화 (20%→16%)","OP -506억","구조적","높음","매우 낮음","",""],
    ["","매출 천장 (규제)","장기 마진 축소","구조적","높음","제2카지노 (27~28년)","",""],
    ["","퇴직급여 (연173억)","경상적 비용","경상적","낮음","해당없음","",""],
]
for frow in factors:
    fls = [wf]*8
    fts = [df]*8
    fts[1] = db
    if frow[3] == "구조적":
        fls = [rdf]*8
        fts[4] = d_red
    elif frow[3] == "일시적":
        fls = [gnf]*8
        fts[4] = d_grn
    wr(ws,r,frow,fonts=fts,fills=fls)
    r += 1

# ============================================================
#  SHEET 2: 비용구조 분석
# ============================================================
ws2 = wb.create_sheet("비용구조")
sw(ws2,[4,20,14,14,14,14,14,14])
setup_print(ws2)

ws2.merge_cells("A1:H2")
c=ws2.cell(row=1,column=1,value="비용구조 분석: 매출원가·판관비 추이"); c.font=title_font; c.fill=tf; c.alignment=ca
ws2.row_dimensions[1].height=35

r = 4
r = st(ws2,r,"A. 연간 P&L 구조 (연결, 억원)")
wh(ws2,r,["","항목","2022","2023","2024","22→23","23→24",""])
r += 1
for label, key in [("매출","rev"),("매출원가","cogs"),("매출총이익","gp"),
                     ("판관비","sga"),("영업이익","op")]:
    v22 = PNL[2022][key]; v23 = PNL[2023][key]; v24 = PNL[2024][key]
    chg1 = f"{(v23-v22)/abs(v22)*100:+.1f}%" if v22 else ""
    chg2 = f"{(v24-v23)/abs(v23)*100:+.1f}%" if v23 else ""
    fts = [df,db,df,df,df,df,df,df]
    fls = [wf]*8
    if label == "영업이익":
        fls = [lf]*8; fts[1] = d_navy
    wr(ws2,r,["",label,v22,v23,v24,chg1,chg2,""],fonts=fts,fills=fls,nfs=[None,None,NF,NF,NF,None,None,None])
    r += 1
r += 1

# 마진율 추이
r = st(ws2,r,"B. 마진율 추이")
wh(ws2,r,["","지표","2022","2023","2024","2025","변화","판단"])
r += 1
margin_data = [
    ["","원가율","72.5%","70.5%","70.6%","74.3%(H1)","↑ 악화","구조적 비용 증가"],
    ["","GPM","27.5%","29.5%","29.4%","25.6%(H1)","↓ 악화","인건비 상승 반영"],
    ["","판관비율","10.3%","9.1%","9.3%","9.4%(H1)","→ 소폭 상승",""],
    ["","OPM","17.1%","20.3%","20.0%","15.9%","↓ 급락","핵심 우려"],
]
for mrow in margin_data:
    fts = [df]*8; fts[1] = db
    fls = [wf]*8
    if "악화" in mrow[6] or "급락" in mrow[6]:
        fls = [rdf]*8; fts[6] = d_red
    wr(ws2,r,mrow,fonts=fts,fills=fls)
    r += 1
r += 1

# 반기 비교 상세
r = st(ws2,r,"C. 반기 비교 상세 (연결, 억원)")
wh(ws2,r,["","항목","2024 H1","2025 H1","변화(억)","변화(%)","비중(24H1)","비중(25H1)"])
r += 1
h_items = [
    ("매출","rev",None),("매출원가","cogs","rev"),("매출총이익","gp","rev"),
    ("판관비","sga","rev"),("영업이익","op","rev"),
    ("기타수익","other_inc",None),("기타비용","other_exp",None),
    ("금융수익","fin_inc",None),("세전이익","pbt",None),("순이익","ni",None),
]
for label, key, pct_base in h_items:
    v24 = H1["24"][key]; v25 = H1["25"][key]
    diff = v25 - v24
    chg_pct = f"{diff/abs(v24)*100:+.1f}%" if v24 else ""
    r24 = f"{v24/H1['24']['rev']*100:.1f}%" if pct_base else ""
    r25 = f"{v25/H1['25']['rev']*100:.1f}%" if pct_base else ""
    fts = [df]*8; fts[1] = db
    fls = [wf]*8
    if label in ("영업이익","순이익"):
        fls = [lf]*8; fts[1] = d_navy
    if label == "기타수익":
        fls = [gld]*8
    wr(ws2,r,["",label,v24,v25,diff,chg_pct,r24,r25],fonts=fts,fills=fls,nfs=[None,None,NF,NF,"+#,##0;-#,##0",None,None,None])
    r += 1
r += 1
note(ws2,r,"핵심: 매출 +228억(+6.7%)인데 매출원가 +340억(+14.5%) → 원가가 매출의 2배 이상 속도로 증가")
r += 1
note(ws2,r,"기타수익: 2024H1 765억 → 2025H1 7억 (일회성 소멸). 기타비용: 25억 → 202억 (일회성 비용 발생)")
r += 2

# 분기별 OPM
r = st(ws2,r,"D. 분기별 영업이익률(OPM) 비교")
wh(ws2,r,["","분기","24매출(백만)","24OP(백만)","24 OPM","25매출(백만)","25OP(백만)","25 OPM"])
r += 1
for i in range(4):
    q24 = Q24[i]; q25 = Q25[i]
    opm_chg = q25["opm"] - q24["opm"]
    fls = [wf]*8
    if opm_chg < -4: fls = [rdf]*8
    elif opm_chg < -1: fls = [gld]*8
    wr(ws2,r,["",f"Q{i+1}",q24["rev"],q24["op"],q24["opm"]/100,q25["rev"],q25["op"],q25["opm"]/100],
       fills=fls,nfs=[None,None,NF,NF,PF,NF,NF,PF])
    r += 1
note(ws2,r+1,"Q1은 전년 수준 유지, Q2부터 급격 악화. 특히 Q4는 OPM 8.1%로 비수기 효과 + 비용 증가 동시 작용")

# ============================================================
#  SHEET 3: 일시적 요인 (기타수익·금융수익)
# ============================================================
ws3 = wb.create_sheet("일시적요인")
sw(ws3,[4,22,14,14,14,14,14,14])
setup_print(ws3)

ws3.merge_cells("A1:H2")
c=ws3.cell(row=1,column=1,value="일시적 요인: 기타수익·금융자산평가이익"); c.font=title_font; c.fill=tf; c.alignment=ca
ws3.row_dimensions[1].height=35

r = 4
r = st(ws3,r,"A. 기타수익 추이 (연결, 억원)")
wh(ws3,r,["","항목","2022","2023","2024","2025H1","판단",""])
r += 1
oi_data = [
    ["","기타수익",46,61,1291,7,"2024 비정상",""],
    ["","기타비용",229,272,262,202,"2025 비용 발생",""],
    ["","기타순손익",-183,-211,1029,-195,"",""],
]
for orow in oi_data:
    fls = [wf]*8; fts = [df]*8; fts[1] = db
    if orow[1] == "기타순손익": fls = [lf]*8; fts[1] = d_navy
    if orow[1] == "기타수익": fls = [gld]*8
    wr(ws3,r,orow,fonts=fts,fills=fls,nfs=[None,None,NF,NF,NF,NF,None,None])
    r += 1
r += 1
r = bullet(ws3,r,"• 2024년 기타수익 1,291억원은 전년(61억) 대비 21배 증가한 비정상적 일회성 항목")
r = bullet(ws3,r,"• 2025년 상반기 기타수익은 7억원에 불과 → 완전 소멸 확인")
r = bullet(ws3,r,"• 세후 영향 추정: 약 1,000억+ → 순이익 역성장(-30.7%)의 절반 이상 설명")
r = bullet(ws3,r,"• 기타수익 제거 시 2024 정상화 NI ≈ 3,500억대 → 2025 NI 3,165억은 약 -10% 수준")
r += 1

r = st(ws3,r,"B. 금융자산·금융수익 분석")
wh(ws3,r,["","항목","2022","2023","2024","변화","비고",""])
r += 1
fi_data = [
    ["","금융수익 합계",280,2159,1789,"-370","시장 변동",""],
    ["","  금융자산평가이익","",1756,1313,"-443","비현금/변동성",""],
    ["","  금융자산처분이익","",78,159,"+81","",""],
    ["","  이자수익","",324,318,"-6","안정적",""],
    ["","금융원가",598,77,28,"-49","개선",""],
]
for frow in fi_data:
    fts = [df]*8; fts[1] = db if not frow[1].startswith("  ") else df
    fls = [wf]*8
    if frow[1] == "금융수익 합계": fls = [lf]*8
    wr(ws3,r,frow,fonts=fts,fills=fls,nfs=[None,None,NF,NF,NF,None,None,None])
    r += 1
r += 1

r = st(ws3,r,"C. 금융자산 보유 현황 (2024말, 억원)")
wh(ws3,r,["","자산 유형","금액","시총 대비","비고","","",""])
r += 1
mcap_b = PRICE * SHARES / 억
for name, val in FIN_ASSETS.items():
    pct_v = val / mcap_b * 100
    wr(ws3,r,["",name,val,f"{pct_v:.1f}%","","","",""],
       fonts=[df,db,df,df,df,df,df,df],nfs=[None,None,NF,None,None,None,None,None])
    r += 1
total = sum(FIN_ASSETS.values())
wr(ws3,r,["","합계",total,f"{total/mcap_b*100:.1f}%","시총의 83%","","",""],
   fonts=[df,d_navy,d_navy,d_navy,d_red,df,df,df],fills=[lf]*8,nfs=[None,None,NF,None,None,None,None,None])
r += 2
r = bullet(ws3,r,"• 강원랜드는 금융자산 3.1조원 보유 (시가총액 3.78조의 83%)")
r = bullet(ws3,r,"• 금융자산평가이익이 연 1,300~1,700억씩 발생 → 시장 상황에 따라 변동")
r = bullet(ws3,r,"• 이는 카지노 영업과 무관한 금융투자 수익 → 핵심이익이 아님")

# ============================================================
#  SHEET 4: 퇴직급여 분석
# ============================================================
ws4 = wb.create_sheet("퇴직급여")
sw(ws4,[4,22,14,14,14,14,14,14])
setup_print(ws4)

ws4.merge_cells("A1:H2")
c=ws4.cell(row=1,column=1,value="퇴직급여 분석: 경상적 비용, 금액 자체는 소규모"); c.font=title_font; c.fill=tf; c.alignment=ca
ws4.row_dimensions[1].height=35

r = 4
r = st(ws4,r,"A. 퇴직급여·퇴직금지급·사외적립 추이 (억원)")
wh(ws4,r,["","연도","퇴직급여(비용)","퇴직금지급(현금)","사외적립자산","비고","",""])
r += 1
for yr in sorted(RETIRE.keys()):
    d = RETIRE[yr]
    memo = ""
    fls = [wf]*8; fts = [df]*8
    fts[1] = db
    if yr == 2020:
        memo = "COVID 구조조정 추정"
        fls = [rdf]*8; fts[5] = d_red
    elif yr == 2024:
        memo = "사외적립 환수 (+202억)"
        fls = [gld]*8; fts[5] = warn_font
    wr(ws4,r,["",yr,d["cost"],d["paid"],d["fund"],memo,"",""],fonts=fts,fills=fls,nfs=[None,None,NF,NF,"+#,##0;-#,##0",None,None,None])
    r += 1
r += 1

r = st(ws4,r,"B. 판단")
r = bullet(ws4,r,"• 퇴직급여(비용): 연 167~173억으로 안정적. 매출 1.4조 대비 1.2%에 불과 → 큰 문제 아님")
r = bullet(ws4,r,"• 2020년 550억은 COVID 희망퇴직/구조조정 관련 일시적 비용 → 반복 없음")
r = bullet(ws4,r,"• 퇴직금지급(현금유출)은 145→269→223억으로 변동 있으나, 실제 퇴직자 수에 따른 것")
r += 1
r = bullet(ws4,r,"• 2024년 사외적립자산 환수(+202억): 기존에 과적립된 퇴직연금을 돌려받음",font=warn_font)
r = bullet(ws4,r,"  → 순확정급여자산: 1,161억(2023) → 713억(2024)으로 감소")
r = bullet(ws4,r,"  → 현금 확보 목적으로 판단, 제2카지노 투자 등에 활용 가능")
r += 1
r = bullet(ws4,r,"결론: 퇴직급여는 경상적이나 이익 역성장의 핵심 원인이 아님. 연 170억 수준 지속 예상.",font=d_navy)

# ============================================================
#  SHEET 5: 인건비 구조
# ============================================================
ws5 = wb.create_sheet("인건비구조")
sw(ws5,[4,28,14,14,14,14,14,14])
setup_print(ws5)

ws5.merge_cells("A1:H2")
c=ws5.cell(row=1,column=1,value="인건비 구조: 절감 가능성 분석"); c.font=title_font; c.fill=tf; c.alignment=ca
ws5.row_dimensions[1].height=35

r = 4
r = st(ws5,r,"A. 카지노업 원가 구조의 특수성")
r = bullet(ws5,r,"• 매출원가의 99%가 '용역 제공 원가' (제품 판매 원가는 23~77억에 불과)")
r = bullet(ws5,r,"• 용역 원가 = 딜러 인건비 + 보안인력 + 서비스인력 + 시설운영비")
r = bullet(ws5,r,"• 카지노 사업은 극도로 노동집약적: 자동화 대체가 구조적으로 어려움")
r = bullet(ws5,r,"• 정선 지방 입지 → 직원 유치/유지를 위한 추가 복리후생비 필요")
r += 1

r = st(ws5,r,"B. 인건비 절감이 불가능한 5가지 이유")
reasons = [
    ("1. 공기업 DNA",
     "최대주주 한국광해광업공단(36.27%), 정부 영향하의 전문경영인 체제\n구조조정/인원감축이 사실상 불가능 (정치적 리스크 매우 높음)"),
    ("2. 설립 목적 = 지역고용",
     "폐광지역개발지원특별법으로 설립, 정선군/태백시 최대 고용처\n인력 감축 = 설립 취지에 정면 위배 → 감독기관/정치권 반대"),
    ("3. 호봉제 + 연공서열",
     "공기업형 임금체계: 근속연수 증가 시 자동 인상\n성과급 체계로의 전환이 노조/이해관계자 저항으로 어려움"),
    ("4. 카지노 자동화 한계",
     "딜러 업무는 대면 서비스 본질 → 무인화 불가\n보안/감시 인력은 사행산업감독위 규제 요건으로 의무적 배치"),
    ("5. 인플레이션 직격",
     "최저임금 인상 → 하위 직급 자동 인상\n물가/에너지 상승 → 복리후생비, 식대, 교통비 등 간접비도 동반 상승"),
]
for title, desc in reasons:
    ws5.merge_cells(start_row=r,start_column=2,end_row=r,end_column=8)
    c=ws5.cell(row=r,column=2,value=title); c.font=db; c.fill=rdf; c.alignment=la; c.border=tb
    r += 1
    ws5.merge_cells(start_row=r,start_column=2,end_row=r+1,end_column=8)
    c=ws5.cell(row=r,column=2,value=desc); c.font=df; c.alignment=Alignment(horizontal='left',vertical='top',wrap_text=True); c.border=tb
    r += 3

r = st(ws5,r,"C. 인건비 vs 매출 성장 가위(Scissors) 구조")
wh(ws5,r,["","항목","추정 성장률","근거","장기 전망","","",""])
r += 1
scissors = [
    ["","매출 성장","연 2~4%","규제(입장제한/영업시간)로 상한 존재","제2카지노 완공(27~28년) 전까지 제한적","","",""],
    ["","인건비 상승","연 5~8%","최저임금+호봉제+물가 연동","구조적으로 지속. 절감 수단 없음","","",""],
    ["","결과: OPM","점진적 하락","비용이 매출보다 빠르게 증가","Pre-COVID 22% → 현재 16% → 추가 하락 가능","","",""],
]
for srow in scissors:
    fts = [df]*8; fts[1] = db; fls = [wf]*8
    if "결과" in srow[1]: fls = [rdf]*8; fts = [d_red]*8; fts[1] = d_red
    wr(ws5,r,srow,fonts=fts,fills=fls)
    r += 1
r += 1
r = bullet(ws5,r,"결론: 인건비 절감은 구조적으로 불가능. 매출 성장 < 비용 증가 구도가 지속될 가능성 높음.",font=d_red)
r = bullet(ws5,r,"유일한 해결책: 제2카지노영업장 완공(1,796억 투자)으로 매출 천장을 높이는 것.",font=d_navy)

# ============================================================
#  SHEET 6: 매출 천장 분석
# ============================================================
ws6 = wb.create_sheet("매출천장")
sw(ws6,[4,14,14,14,14,14,14,14])
setup_print(ws6)

ws6.merge_cells("A1:H2")
c=ws6.cell(row=1,column=1,value="매출 천장: 규제가 만든 성장의 한계"); c.font=title_font; c.fill=tf; c.alignment=ca
ws6.row_dimensions[1].height=35

r = 4
r = st(ws6,r,"A. 10년 매출 추이 (연결, 억원)")
wh(ws6,r,["","연도","매출","YoY","영업이익","OPM","비고",""])
r += 1
prev_rev = None
for yr in range(2015,2025):
    d = ANNUAL[yr]
    rev_b = d["rev"]/억; op_b = d["op"]/억
    opm = d["op"]/d["rev"]*100 if d["rev"] else 0
    yoy = f"{(d['rev']-prev_rev)/prev_rev*100:+.1f}%" if prev_rev and prev_rev > 0 else ""
    memo = ""
    fls = [wf]*8; fts = [df]*8; fts[1] = db
    if yr in (2020,2021):
        memo = "COVID" if yr==2020 else "COVID 회복 중"
        fls = [rdf]*8; fts[5] = d_red
    elif yr == 2017:
        memo = "Pre-COVID 피크"
        fls = [gnf]*8
    wr(ws6,r,["",yr,round(rev_b),yoy,round(op_b),f"{opm:.1f}%",memo,""],fonts=fts,fills=fls,nfs=[None,None,NF,None,NF,None,None,None])
    prev_rev = d["rev"]
    r += 1
# 2025
wr(ws6,r,["",2025,round(CUM25["rev"]/억),"+3.5%",round(CUM25["op"]/억),"15.9%","잠정실적",""],
   fonts=[df,db,d_blue,df,df,d_red,sm,df],fills=[blf]*8,nfs=[None,None,NF,None,NF,None,None,None])
r += 2

r = bullet(ws6,r,"• Pre-COVID 피크(2017): 매출 16,013억 → 2025년에도 아직 미회복 (14,767억)")
r = bullet(ws6,r,"• 7년이 지나도 2017년 매출을 넘지 못하는 구조",font=d_red)
r += 1

r = st(ws6,r,"B. 매출 천장의 원인: 규제")
reg_items = [
    ("영업시간 제한","주 120시간 이내 (1일 최대 20시간)","매출 상한 직접 결정"),
    ("입장횟수 제한","월 15회 이내","상습 이용자 제한 → 고가치 고객 이탈"),
    ("입장료","9,000원 (2023년 인상)","심리적 진입장벽"),
    ("배팅한도","테이블당 최대 30만원","1인당 GGR 상한"),
    ("총량규제","사행산업 매출총량 관리","산업 전체 성장 제한"),
]
wh(ws6,r,["","규제 항목","현황","매출 영향","","","",""])
r += 1
for item in reg_items:
    wr(ws6,r,["",item[0],item[1],item[2],"","","",""],fonts=[df,db,df,df,df,df,df,df])
    r += 1
r += 1

r = st(ws6,r,"C. 제2카지노영업장 (유일한 돌파구)")
r = bullet(ws6,r,"• 투자 규모: 1,796억원 (자기자본 대비 4.9%)")
r = bullet(ws6,r,"• 공시일: 2024.11.28 (신규시설투자 자율공시)")
r = bullet(ws6,r,"• 예상 완공: 2027~2028년")
r = bullet(ws6,r,"• 기대 효과: 테이블/슬롯 증설 → 수용인원 증가 → 매출 천장 상향")
r = bullet(ws6,r,"• 리스크: 감가상각비 연 100~150억 추가 부담, 규제가 동시 강화될 가능성",font=d_red)

# ============================================================
#  SHEET 7: 종합 판단
# ============================================================
ws7 = wb.create_sheet("종합판단")
sw(ws7,[4,28,14,14,14,14,14,14])
setup_print(ws7)

ws7.merge_cells("A1:H2")
c=ws7.cell(row=1,column=1,value="종합 판단: 해자는 건재하나 효율화 유인 부재"); c.font=title_font; c.fill=tf; c.alignment=ca
ws7.row_dimensions[1].height=35

r = 4
r = st(ws7,r,"A. 버핏 관점 핵심 판단")
r += 1
ws7.merge_cells(start_row=r,start_column=2,end_row=r+4,end_column=8)
c=ws7.cell(row=r,column=2,value=(
    "강원랜드의 경제적 해자(규제 독점)는 건재하다. 내국인 카지노 면허는 여전히 유일하며, "
    "경쟁자 진입 가능성은 극히 낮다.\n\n"
    "그러나 그 해자가 역설적으로 비용 구조를 개선할 유인도 차단하고 있다. "
    "독점이라 경쟁은 없지만, 공기업이라 효율화도 안 되는 양날의 검.\n\n"
    "OPM이 Pre-COVID 22%에서 16%로 내려온 것은 '독점 프리미엄의 침식'이다. "
    "규제 완화나 제2카지노 없이는 반등이 어렵다."
))
c.font=Font(name="맑은 고딕",size=11); c.alignment=Alignment(horizontal='left',vertical='top',wrap_text=True)
c.fill=llf; c.border=tb
r += 6

r = st(ws7,r,"B. 구조적 리스크 vs 일시적 노이즈")
wh(ws7,r,["","구분","요인","영향","지속성","투자 시사점","",""])
r += 1
judge_data = [
    ["","일시적","기타수익 소멸","NI -1,000억","1년","2024가 비정상, 2025가 정상","",""],
    ["","일시적","금융자산평가","NI ±수백억","매년 변동","핵심이익 아님, 무시 가능","",""],
    ["","구조적","인건비 상승","OP -500억/년","영구","절감 불가, 매년 악화","",""],
    ["","구조적","매출 천장","성장 제한","27~28년까지","제2카지노 완공 전까지 지속","",""],
    ["","경상적","퇴직급여","연 170억","매년","금액 소규모, 무시 가능","",""],
]
for jrow in judge_data:
    fts = [df]*8; fts[1] = db; fls = [wf]*8
    if jrow[1] == "구조적": fls = [rdf]*8; fts[1] = d_red
    elif jrow[1] == "일시적": fls = [gnf]*8; fts[1] = d_grn
    wr(ws7,r,jrow,fonts=fts,fills=fls)
    r += 1
r += 1

r = st(ws7,r,"C. 시나리오별 OPM 전망")
wh(ws7,r,["","시나리오","OPM 전망","전제 조건","확률","비고","",""])
r += 1
scen = [
    ["","Best","18~20%","규제완화 + 제2카지노 + 비용관리","15%","2027년 이후","",""],
    ["","Base","14~16%","현상유지, 인건비 계속 상승","55%","현재 수준 유지","",""],
    ["","Worst","10~12%","규제강화 + 인건비 급등","30%","영업시간 추가 축소 시","",""],
]
for srow in scen:
    fts = [df]*8; fts[1] = db; fls = [wf]*8
    if srow[1]=="Best": fls = [gnf]*8
    elif srow[1]=="Worst": fls = [rdf]*8
    else: fls = [blf]*8
    wr(ws7,r,srow,fonts=fts,fills=fls)
    r += 1
r += 1

r = st(ws7,r,"D. 모니터링 포인트")
monitor = [
    "매출원가율 추이: 74%(H1) → 연간 확정치 확인 (70% 이하 회복 여부)",
    "분기별 OPM: Q2~Q4 악화 추세가 반전되는지",
    "제2카지노 건설 진행 상황: 착공/허가/완공 일정",
    "사행산업 규제 변화: 영업시간/입장횟수/배팅한도 조정 여부",
    "인건비 관련: 단체협약, 임금인상률, 정원 변동",
    "기타수익/기타비용: 일회성 항목 규모 (정상화 수준 확인)",
]
for m in monitor:
    r = bullet(ws7,r,f"• {m}")
r += 1

r = st(ws7,r,"E. 최종 결론")
r += 1
ws7.merge_cells(start_row=r,start_column=2,end_row=r+3,end_column=8)
c=ws7.cell(row=r,column=2,value=(
    "NI -30.7%의 절반 이상은 일시적이다 (2024년 기타수익 소멸). "
    "그러나 OPM 4p 하락과 매출 성장 한계는 구조적 리스크이며, "
    "인건비 절감이 불가능한 공기업 구조에서 이 가위(scissors) 효과는 당분간 지속될 것이다.\n\n"
    "투자 관점에서 강원랜드는 '해자는 튼튼하지만 성안의 비용이 계속 올라가는 성' — "
    "배당(수익률 6.6%)과 순현금(3.1조 금융자산)이 버텨주는 한 급락 리스크는 낮으나, "
    "마진 개선 없이는 주가 재평가도 어렵다."
))
c.font=Font(name="맑은 고딕",size=11,bold=True); c.alignment=Alignment(horizontal='left',vertical='top',wrap_text=True)
c.fill=gld; c.border=tb

# ============================================================
#  SAVE
# ============================================================
out = os.path.join(BASE, f"{COMPANY}_이익역성장분석.xlsx")
wb.save(out)
print(f"저장 완료: {out}")
print(f"시트: {[ws.title for ws in wb.worksheets]}")
