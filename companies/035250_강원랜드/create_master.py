"""강원랜드 마스터 보고서 생성 - 5개 보고서를 하나의 파일로 통합 (모바일용 제외)

종합보고서(12시트) + 투자구루분석(7시트) + 이익역성장분석(7시트) = 26시트
기업분석보고서/밸류에이션은 종합보고서에 이미 포함되므로 제외.
"""
import copy
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

DIR = os.path.dirname(os.path.abspath(__file__))

# 합칠 파일 순서 (종합보고서 → 구루분석 → 이익역성장)
SOURCES = [
    ("강원랜드_종합보고서.xlsx", None),           # 전체 시트
    ("강원랜드_투자구루분석.xlsx", None),          # 전체 시트
    ("강원랜드_이익역성장분석.xlsx", None),        # 전체 시트
]

OUTPUT = os.path.join(DIR, "강원랜드_마스터보고서.xlsx")


def copy_sheet(src_ws, dst_wb, title):
    """src_ws의 내용을 dst_wb에 새 시트로 복사 (값, 스타일, 병합, 너비, 높이)"""
    dst_ws = dst_wb.create_sheet(title=title)

    # 1) 셀 값 + 스타일 복사
    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column)
            dst_cell.value = cell.value

            # 스타일 복사
            if cell.has_style:
                dst_cell.font = copy.copy(cell.font)
                dst_cell.fill = copy.copy(cell.fill)
                dst_cell.border = copy.copy(cell.border)
                dst_cell.alignment = copy.copy(cell.alignment)
                dst_cell.number_format = cell.number_format
                dst_cell.protection = copy.copy(cell.protection)

    # 2) 병합 셀 복사
    for merged in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged))

    # 3) 열 너비 복사
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
        if dim.hidden:
            dst_ws.column_dimensions[col_letter].hidden = True

    # 4) 행 높이 복사
    for row_num, dim in src_ws.row_dimensions.items():
        if dim.height:
            dst_ws.row_dimensions[row_num].height = dim.height
        if dim.hidden:
            dst_ws.row_dimensions[row_num].hidden = True

    # 5) 시트 속성 복사
    dst_ws.sheet_properties = copy.copy(src_ws.sheet_properties)
    if src_ws.print_options:
        dst_ws.print_options = copy.copy(src_ws.print_options)

    return dst_ws


def main():
    # 기본 워크북 생성 (빈 시트 하나 생성됨)
    master_wb = load_workbook(os.path.join(DIR, SOURCES[0][0]))

    # 나머지 파일들의 시트 추가
    for i, (fname, sheet_filter) in enumerate(SOURCES):
        if i == 0:
            continue  # 첫 번째 파일은 이미 로드됨

        fpath = os.path.join(DIR, fname)
        if not os.path.exists(fpath):
            print(f"  [SKIP] {fname} 파일 없음")
            continue

        src_wb = load_workbook(fpath)
        sheets = sheet_filter if sheet_filter else src_wb.sheetnames

        for sname in sheets:
            if sname not in src_wb.sheetnames:
                print(f"  [SKIP] {fname} > {sname} 시트 없음")
                continue
            src_ws = src_wb[sname]

            # 시트명 중복 방지
            dst_title = sname
            if dst_title in master_wb.sheetnames:
                # 출처 접미사 추가
                suffix = fname.replace("강원랜드_", "").replace(".xlsx", "")
                dst_title = f"{sname}_{suffix}"

            print(f"  + {dst_title}")
            copy_sheet(src_ws, master_wb, dst_title)

        src_wb.close()

    # 저장
    master_wb.save(OUTPUT)
    master_wb.close()
    print(f"\n>>> {os.path.basename(OUTPUT)} 저장 완료 ({len(load_workbook(OUTPUT, read_only=True).sheetnames)}시트)")


if __name__ == "__main__":
    print("=" * 60)
    print("  강원랜드 마스터 보고서 생성")
    print("  종합보고서 + 투자구루분석 + 이익역성장분석")
    print("=" * 60)
    main()
