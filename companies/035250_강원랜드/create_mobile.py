# -*- coding: utf-8 -*-
"""강원랜드 모바일용 보고서 — 좁은 화면 최적화 (3열, 큰 글씨, 세로 스크롤)"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# === CONSTANTS ===
COMPANY = "강원랜드"
STOCK_CODE = "035250"
PRICE = 17690
SHARES = 213940500
SHARES_TREASURY = 13534816
SHARES_OUT = SHARES - SHARES_TREASURY  # 유통주식수 200,405,684
억 = 100_000_000
MCAP = PRICE * SHARES  # 시가총액

# === 2024 연결 (원) ===
REV24 = 1426862781802
OP24 = 285790454716
NI24 = 456895295638
EPS24 = 2257
EQ24 = 3883662813757
EQ23 = 3677900816262
ASSETS24 = 4700641154570
LIAB24 = 816978340813
CASH24 = 244499636973
ST_DEBT = 0; LT_DEBT = 0  # 무차입경영
DA24 = 78474881854
OPCF24 = 487048778217
CAPEX24 = 107494733192
DPS24 = 1170
DIV_PAID = 188547268140

# 2023
REV23 = 1392740543618; OP23 = 258505699379; NI23 = 341061017856; EPS23 = 1681

# === 2025 잠정실적 ===
Q25 = [
    {"q": "1Q25", "rev": 365830e6, "op": 77659e6, "ni": 78021e6},
    {"q": "2Q25", "rev": 360727e6, "op": 57874e6, "ni": 60630e6},
    {"q": "3Q25", "rev": 384147e6, "op": 72702e6, "ni": 113070e6},
    {"q": "4Q25", "rev": 365446e6, "op": 29697e6, "ni": 66017e6},
]
CUM25_REV = 1476726e6; CUM25_OP = 235176e6; CUM25_NI = 316516e6

# === 10년 실적 ===
ANNUAL_REV = {2015: 1634441985990, 2016: 1703131541003, 2017: 1601291247063, 2018: 1445736946832,
              2019: 1524006966734, 2020: 479173424993, 2021: 788430938373, 2022: 1272539665429,
              2023: 1392740543618, 2024: 1426862781802}
ANNUAL_OP = {2015: 422764543000, 2016: 480490893000, 2017: 460085942000, 2018: 292485625000,
             2019: 339782744000, 2020: -200253285781, 2021: -70474730032, 2022: 145068791398,
             2023: 258505699379, 2024: 285790454716}
ANNUAL_NI = {2015: 466119741406, 2016: 480264780048, 2017: 460843127006, 2018: 314063000747,
             2019: 353337372000, 2020: -291290000000, 2021: -11056000000, 2022: 122207000000,
             2023: 341061017856, 2024: 456895295638}
ANNUAL_EPS = {2015: 2178, 2016: 2242, 2017: 2159, 2018: 1467, 2019: 1651,
              2020: -1361, 2021: -52, 2022: 570, 2023: 1681, 2024: 2257}
ANNUAL_DPS = {2015: 980, 2016: 990, 2017: 990, 2018: 900, 2019: 900,
              2020: 0, 2021: 0, 2022: 350, 2023: 930, 2024: 1170}

# === 파생 계산 ===
BPS = EQ24 / SHARES
NET_DEBT = ST_DEBT + LT_DEBT - CASH24   # 무차입이므로 음수(Net Cash)
EBITDA24 = OP24 + DA24
FCF24 = OPCF24 - CAPEX24
EV = MCAP + NET_DEBT                     # Net Cash이므로 EV < 시총
ROE24 = NI24 / ((EQ24 + EQ23) / 2)

# Trailing 4Q (4Q24 + 1Q25~3Q25 없으므로 2024 연간 그대로 사용, 2025 잠정 사용)
# 4Q24는 연간 - (1Q~3Q)24를 역산하기 어려우므로 2025 연간 잠정 기준 EPS 계산
EPS25E = int(CUM25_NI / SHARES)  # 2025 잠정 전체 EPS
TRAIL_EPS = EPS24  # 가장 최근 확정 연간 기준

def fmt(v): return f"{v / 억:,.0f}억"
def fw(v): return f"{v:,.0f}원"
def pct(v): return f"{v * 100:.1f}%"

# === MOBILE STYLES ===
NAVY = "1B2A4A"; DARK = "2C3E6B"; W = "FFFFFF"; GOLD_C = "D4A843"; RED_C = "C0392B"; GREEN_C = "27AE60"

t1 = Font(name="맑은 고딕", size=16, bold=True, color=W)       # title
t2 = Font(name="맑은 고딕", size=14, bold=True, color=NAVY)     # section
hf_ = Font(name="맑은 고딕", size=11, bold=True, color=W)       # header
df_ = Font(name="맑은 고딕", size=11)                            # data
db_ = Font(name="맑은 고딕", size=11, bold=True)                 # data bold
bl_ = Font(name="맑은 고딕", size=11, bold=True, color="0000FF") # blue
gn_ = Font(name="맑은 고딕", size=11, bold=True, color=GREEN_C)
rd_ = Font(name="맑은 고딕", size=11, bold=True, color=RED_C)
sm_ = Font(name="맑은 고딕", size=10, color="666666")
big_ = Font(name="맑은 고딕", size=13, bold=True, color=NAVY)
huge_ = Font(name="맑은 고딕", size=18, bold=True, color=NAVY)

tf_ = PatternFill("solid", fgColor=NAVY)
hfl = PatternFill("solid", fgColor=DARK)
lf_ = PatternFill("solid", fgColor="D6E4F0")
llf_ = PatternFill("solid", fgColor="EBF1F8")
wf_ = PatternFill("solid", fgColor=W)
gld_ = PatternFill("solid", fgColor="FFF3CD")
gnf_ = PatternFill("solid", fgColor="D4EDDA")
rdf_ = PatternFill("solid", fgColor="F8D7DA")
blf_ = PatternFill("solid", fgColor="D6EAF8")
grf_ = PatternFill("solid", fgColor="F2F2F2")

ca_ = Alignment(horizontal='center', vertical='center', wrap_text=True)
la_ = Alignment(horizontal='left', vertical='center', wrap_text=True)
ra_ = Alignment(horizontal='right', vertical='center', wrap_text=True)
tb_ = Border(left=Side('thin', color="D0D0D0"), right=Side('thin', color="D0D0D0"),
             top=Side('thin', color="D0D0D0"), bottom=Side('thin', color="D0D0D0"))
bb_ = Border(bottom=Side('medium', color=NAVY))

COL3 = [14, 18, 18]

def sw(ws, w):
    for i, v in enumerate(w, 1): ws.column_dimensions[get_column_letter(i)].width = v

def mhdr(ws, r, vals):
    """모바일 헤더"""
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = hf_; c.fill = hfl; c.alignment = ca_; c.border = tb_

def mrow(ws, r, vals, fonts=None, fills=None, als=None, h=None):
    """모바일 데이터 행"""
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = fonts[i - 1] if fonts else df_
        c.fill = fills[i - 1] if fills else wf_
        c.alignment = als[i - 1] if als else ca_
        c.border = tb_
    if h: ws.row_dimensions[r].height = h

def msec(ws, r, title, cols=3):
    """모바일 섹션 타이틀"""
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
    c = ws.cell(row=r, column=1, value=title)
    c.font = t2; c.border = bb_
    ws.row_dimensions[r].height = 28
    return r + 1

def minfo(ws, r, label, value, cols=3):
    """모바일 라벨-값 쌍"""
    ws.cell(row=r, column=1, value=label).font = db_
    ws.cell(row=r, column=1).fill = llf_; ws.cell(row=r, column=1).alignment = la_; ws.cell(row=r, column=1).border = tb_
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=cols)
    ws.cell(row=r, column=2, value=value).font = bl_
    ws.cell(row=r, column=2).fill = wf_; ws.cell(row=r, column=2).alignment = ra_; ws.cell(row=r, column=2).border = tb_
    ws.row_dimensions[r].height = 24
    return r + 1

wb = Workbook()

# ============================================================
# 단일 시트: 모든 내용을 세로로 길게 배치
# ============================================================
ws = wb.active
ws.title = "강원랜드 분석"
ws.sheet_properties.tabColor = NAVY
sw(ws, COL3)

row = 1

# ---- TITLE ----
ws.merge_cells('A1:C2')
c = ws.cell(row=1, column=1, value="강원랜드(035250)\n종합 기업분석 보고서")
c.font = t1; c.fill = tf_; c.alignment = ca_
ws.row_dimensions[1].height = 24; ws.row_dimensions[2].height = 24
row = 3
ws.merge_cells('A3:C3')
ws.cell(row=3, column=1, value=f"현재가 {PRICE:,}원 | 시총 {fmt(MCAP)} | 2026.02.06").font = Font(name="맑은 고딕", size=11, bold=True, color=DARK)
ws.cell(row=3, column=1).alignment = ca_; ws.row_dimensions[3].height = 22

# ==== 1. 기본정보 ====
row = 5
row = msec(ws, row, "기본 정보")
for lbl, val in [
    ("현재주가", fw(PRICE)), ("시가총액", fmt(MCAP)),
    ("유통시총", fmt(PRICE * SHARES_OUT)),
    ("발행주식수", f"{SHARES:,}주"), ("자기주식", f"{SHARES_TREASURY:,}주"),
    ("업종", "카지노 / 리조트 (내국인 카지노 유일)"),
    ("최대주주", "한국광해광업공단 36.27%"),
    ("소재지", "강원도 정선군 사북읍"),
    ("설립", "1998.06.29 (폐광지역 경제진흥)")]:
    row = minfo(ws, row, lbl, val)

# ==== 2. 핵심 밸류에이션 ====
row += 1
row = msec(ws, row, "핵심 밸류에이션")
mhdr(ws, row, ["지표", "값", "판정"]); row += 1

PER24 = PRICE / EPS24
PBR = PRICE / BPS
EV_EBITDA = EV / EBITDA24
DIV_YIELD = DPS24 / PRICE
FCF_YIELD = (FCF24 / SHARES) / PRICE
DEBT_RATIO = LIAB24 / EQ24

vals = [
    ("PER (2024)", f"{PER24:.1f}배", "저평가" if PER24 < 10 else "적정"),
    ("PER (25E 잠정)", f"{PRICE / EPS25E:.1f}배" if EPS25E > 0 else "N/A", "적정"),
    ("PBR", f"{PBR:.2f}배", "저평가" if PBR < 1.0 else "적정"),
    ("EV/EBITDA", f"{EV_EBITDA:.1f}배", "저평가" if EV_EBITDA < 8 else "적정"),
    ("배당수익률", pct(DIV_YIELD), "매우양호" if DIV_YIELD > 0.05 else "양호"),
    ("FCF수익률", pct(FCF_YIELD), "양호"),
    ("ROE", pct(ROE24), "양호" if ROE24 > 0.10 else "적정"),
    ("OPM", pct(OP24 / REV24), "양호"),
    ("부채비율", pct(DEBT_RATIO), "매우건전"),
    ("차입금", "0원 (무차입경영)", "매우건전"),
]
for lbl, val, judge in vals:
    if "매우" in judge: jf, jfl = gn_, gnf_
    elif "저" in judge or "양호" in judge or "건전" in judge: jf, jfl = Font(name="맑은 고딕", size=11, bold=True, color="2E86C1"), blf_
    elif "적정" in judge: jf, jfl = db_, gld_
    else: jf, jfl = rd_, rdf_
    mrow(ws, row, [lbl, val, judge], fonts=[db_, bl_, jf], fills=[llf_, gld_, jfl], als=[la_, ca_, ca_], h=24)
    row += 1

# ==== 3. 2024 확정실적 ====
row += 1
row = msec(ws, row, "2024년 확정 실적")
REV_YOY = (REV24 - REV23) / REV23 * 100
OP_YOY = (OP24 - OP23) / OP23 * 100
NI_YOY = (NI24 - NI23) / NI23 * 100
for lbl, val in [
    ("매출액", f"{fmt(REV24)} (YoY {REV_YOY:+.1f}%)"),
    ("영업이익", f"{fmt(OP24)} (OPM {OP24 / REV24 * 100:.1f}%, YoY {OP_YOY:+.1f}%)"),
    ("순이익", f"{fmt(NI24)} (YoY {NI_YOY:+.1f}%)"),
    ("EPS", fw(EPS24)), ("BPS", fw(int(BPS))),
    ("DPS", f"{fw(DPS24)} (전년 930원, YoY +25.8%)"),
    ("총자산", fmt(ASSETS24)), ("총부채", fmt(LIAB24)),
    ("총자본", fmt(EQ24)),
    ("영업CF", fmt(OPCF24)),
    ("CAPEX", fmt(CAPEX24)),
    ("FCF", fmt(FCF24)),
    ("EBITDA", fmt(EBITDA24)),
    ("감가상각", fmt(DA24)),
    ("차입금", "0원 (무차입경영)"),
    ("현금성자산", fmt(CASH24))]:
    row = minfo(ws, row, lbl, val)

# ==== 4. 2025 분기실적 ====
row += 1
row = msec(ws, row, "2025년 분기별 잠정실적 (백만원)")
mhdr(ws, row, ["분기", "매출", "영업이익"]); row += 1
for q in Q25:
    mrow(ws, row, [q["q"], f"{int(q['rev'] / 1e6):,}", f"{int(q['op'] / 1e6):,}"],
         fonts=[db_, df_, df_], fills=[blf_, gnf_, gnf_], als=[ca_, ra_, ra_]); row += 1
# 누적
mrow(ws, row, ["25누계", f"{int(CUM25_REV / 1e6):,}", f"{int(CUM25_OP / 1e6):,}"],
     fonts=[db_] * 3, fills=[gld_] * 3, als=[ca_, ra_, ra_]); row += 1
# 순이익 행도 추가
row += 1
mhdr(ws, row, ["분기", "순이익", "비고"]); row += 1
for q in Q25:
    note = ""
    if q["q"] == "3Q25": note = "일회성 이익 포함"
    if q["q"] == "4Q25": note = "계절적 비수기"
    mrow(ws, row, [q["q"], f"{int(q['ni'] / 1e6):,}", note],
         fonts=[db_, df_, sm_], fills=[blf_, gnf_, wf_], als=[ca_, ra_, la_]); row += 1
mrow(ws, row, ["25누계NI", f"{int(CUM25_NI / 1e6):,}", f"EPS {EPS25E:,}원"],
     fonts=[db_] * 3, fills=[gld_] * 3, als=[ca_, ra_, ra_]); row += 1

# Key points
row += 1
for pt in [f"2025 잠정 매출 {fmt(CUM25_REV)} (YoY {(CUM25_REV - REV24) / REV24 * 100:+.1f}%)",
           f"2025 잠정 영업이익 {fmt(CUM25_OP)} (YoY {(CUM25_OP - OP24) / OP24 * 100:+.1f}%)",
           "3Q25 순이익 1,131억 (일회성 이익 반영 추정)",
           "4Q25 영업이익 297억으로 분기 최저 (계절적 비수기)"]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"• {pt}").font = df_; ws.cell(row=row, column=1).alignment = la_
    ws.row_dimensions[row].height = 22; row += 1

# ==== 5. PER 분석 ====
row += 1
row = msec(ws, row, "PER 다각도 분석")
mhdr(ws, row, ["기준", "EPS", "PER"]); row += 1
for lbl, eps_v in [("2024 확정", EPS24), ("2025 잠정", EPS25E),
                    ("보수적(25E*0.9)", int(EPS25E * 0.9))]:
    per = PRICE / eps_v if eps_v > 0 else 0
    pf = gn_ if per < 8 else (bl_ if per < 12 else db_)
    pfl = gnf_ if per < 8 else (blf_ if per < 12 else gld_)
    mrow(ws, row, [lbl, fw(eps_v), f"{per:.1f}배"], fonts=[db_, bl_, pf], fills=[llf_, gld_, pfl]); row += 1

# 카지노 적정 PER
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="카지노업 적정 PER: 10~15배 (글로벌 평균)").font = sm_
ws.cell(row=row, column=1).alignment = la_; row += 1

# Target PER
row += 1
row = msec(ws, row, "목표PER별 적정주가")
mhdr(ws, row, ["PER배수", "2024 EPS", "25E EPS"]); row += 1
for m in [8, 10, 12, 15]:
    v24 = EPS24 * m; v25 = EPS25E * m
    mrow(ws, row, [f"{m}배", fw(v24), fw(v25)],
         fonts=[db_, df_, df_],
         fills=[llf_, gnf_ if v24 > PRICE else wf_, gnf_ if v25 > PRICE else wf_],
         als=[ca_, ra_, ra_]); row += 1

# ==== 6. EV/EBITDA ====
row += 1
row = msec(ws, row, "EV/EBITDA & FCF")
for lbl, val in [
    ("시가총액", fmt(MCAP)),
    ("Net Cash (무차입)", fmt(abs(int(NET_DEBT)))),
    ("EV (시총-Net Cash)", fmt(EV)),
    ("", ""),
    ("EBITDA (2024)", fmt(EBITDA24)),
    ("EV/EBITDA", f"{EV_EBITDA:.1f}배"),
    ("", ""),
    ("영업CF", fmt(OPCF24)),
    ("CAPEX", fmt(CAPEX24)),
    ("FCF", fmt(FCF24)),
    ("FCF/주", fw(int(FCF24 / SHARES))),
    ("FCF 수익률", pct(FCF_YIELD)),
    ("", ""),
    ("배당지급", f"{fmt(DIV_PAID)} ({fw(DPS24)}/주)"),
    ("배당성향(24)", pct(DIV_PAID / NI24)),
    ("배당수익률", pct(DIV_YIELD))]:
    if not lbl: row += 1; continue
    row = minfo(ws, row, lbl, val)

# ==== 7. RIM 적정주가 ====
row += 1
row = msec(ws, row, "잔여이익모델(RIM) 적정주가")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="BPS x (1+(ROE-ke)/(ke-g)) | ke=10%").font = sm_
ws.cell(row=row, column=1).alignment = la_; row += 1
mhdr(ws, row, ["시나리오", "적정주가", "현재가대비"]); row += 1
for lbl, roe, ke, g in [("보수적(ROE=ke)", 0.10, 0.10, 0.02),
                          ("기본(ROE12%)", 0.12, 0.10, 0.02),
                          ("적극(ROE15%)", 0.15, 0.10, 0.02),
                          ("낙관(ROE18%)", 0.18, 0.10, 0.03)]:
    fv = BPS * (1 + (roe - ke) / (ke - g)); up = (fv - PRICE) / PRICE
    uf = gn_ if up > 0 else rd_; ufl = gnf_ if up > 0 else rdf_
    mrow(ws, row, [lbl, fw(int(fv)), f"{up * 100:+.1f}%"], fonts=[db_, bl_, uf], fills=[llf_, gld_, ufl]); row += 1

# ==== 8. 시나리오별 목표주가 ====
row += 1
row = msec(ws, row, "시나리오별 목표주가")
# Bull: PER 15배 x 2024 EPS, Base: PER 12배, Bear: PER 8배
bull_t = EPS24 * 15; base_t = EPS24 * 12; bear_t = EPS24 * 8
exp_v = int(bull_t * 0.2 + base_t * 0.5 + bear_t * 0.3)

mhdr(ws, row, ["시나리오", "목표주가", "현재가대비"]); row += 1
mrow(ws, row, ["강세(Bull)", fw(bull_t), f"{(bull_t - PRICE) / PRICE * 100:+.1f}%"],
     fonts=[db_, Font(name="맑은 고딕", size=12, bold=True, color=GREEN_C), gn_], fills=[gnf_] * 3, h=28); row += 1
mrow(ws, row, ["기본(Base)", fw(base_t), f"{(base_t - PRICE) / PRICE * 100:+.1f}%"],
     fonts=[db_, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), bl_], fills=[gld_] * 3, h=28); row += 1
mrow(ws, row, ["약세(Bear)", fw(bear_t), f"{(bear_t - PRICE) / PRICE * 100:+.1f}%"],
     fonts=[db_, Font(name="맑은 고딕", size=12, bold=True, color=RED_C), rd_], fills=[rdf_] * 3, h=28); row += 1
row += 1
mrow(ws, row, ["기대값", fw(exp_v), f"{(exp_v - PRICE) / PRICE * 100:+.1f}%"],
     fonts=[Font(name="맑은 고딕", size=12, bold=True, color=NAVY)] * 3, fills=[blf_] * 3, h=28); row += 1

# 전제조건
row += 1
for sc, cond in [("Bull", "PER15배: 규제완화+입장객증가+제2카지노기대감+배당확대"),
                  ("Base", "PER12배: 현행 유지, 안정적 배당, 점진적 입장객 회복"),
                  ("Bear", "PER8배: 규제강화+경기침체+온라인도박 합법화 우려")]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"{sc}: {cond}").font = df_
    ws.cell(row=row, column=1).alignment = la_; ws.row_dimensions[row].height = 22; row += 1

# ==== 9. 배당 분석 ====
row += 1
row = msec(ws, row, "배당 분석")
mhdr(ws, row, ["연도", "DPS(원)", "배당수익률"]); row += 1
for yr in range(2015, 2025):
    dps = ANNUAL_DPS[yr]
    yl = f"{dps / PRICE * 100:.1f}%" if dps > 0 else "0%"
    nif = gn_ if dps >= 900 else (rd_ if dps == 0 else df_)
    mrow(ws, row, [str(yr), f"{dps:,}", yl],
         fonts=[db_, nif, df_], fills=[llf_, gnf_ if dps >= 900 else (rdf_ if dps == 0 else wf_), wf_]); row += 1

# 배당 요약
row += 1
for pt in [f"2024 DPS {DPS24:,}원 = 배당수익률 {DIV_YIELD * 100:.1f}% (고배당)",
           f"배당성향(2024) {DIV_PAID / NI24 * 100:.1f}% — 순이익 대비 안정적",
           "COVID(2020-2021) 무배당 후 2022년 복원, 매년 증가 추세",
           "카지노업 특성상 설비투자 낮고 현금창출력 높아 배당여력 충분"]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"• {pt}").font = df_; ws.cell(row=row, column=1).alignment = la_
    ws.row_dimensions[row].height = 22; row += 1

# ==== 10. 사업구조 ====
row += 1
row = msec(ws, row, "사업구조 (카지노/리조트)")
for lbl, val in [
    ("카지노사업", "매출 약 80% (핵심 수익원)"),
    ("호텔사업", "매출 약 10% (하이원호텔, 콘도)"),
    ("리조트사업", "매출 약 10% (스키장, 골프장 등)"),
    ("", ""),
    ("매출 드라이버", "입장객 x Drop(투입금액) x Hold Rate"),
    ("GGR(총게임수익)", "카지노 매출 = GGR"),
    ("제2카지노영업장", "투자액 1,796억 (영업 개시 기대)"),
    ("", ""),
    ("직원수(추정)", "약 3,800~4,000명"),
    ("영업시간 규제", "1일 최대 18시간"),
    ("입장료", "내국인 1만원"),
    ("입장제한", "월 15회, 지정 출입제한 등")]:
    if not lbl: row += 1; continue
    row = minfo(ws, row, lbl, val)

# ==== 11. 규제/면허 ====
row += 1
row = msec(ws, row, "규제 및 면허 (카지노 핵심)")
for lbl, val in [
    ("면허현황", "내국인 카지노 유일 합법 면허"),
    ("근거법", "폐광지역개발지원에관한특별법"),
    ("감독기관", "사행산업통합감독위원회"),
    ("영업시간", "1일 최대 18시간 (규제)"),
    ("입장제한", "월 15회 / 1회 24시간 이내"),
    ("출입제한", "본인/가족 신청 출입차단 가능"),
    ("사행산업기금", "매출의 일정비율 납부 의무"),
    ("지역공헌", "폐광지역 경제진흥 법적 의무"),
    ("규제리스크", "정부 정책에 따라 영업시간/입장 변동"),
    ("면허가치", "진입장벽 절대적 — 신규 면허 가능성 극히 낮음")]:
    row = minfo(ws, row, lbl, val)

# ==== 12. SWOT ====
row += 1
row = msec(ws, row, "SWOT 분석")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="강점 (S)").font = Font(name="맑은 고딕", size=12, bold=True, color=W)
ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=GREEN_C); ws.cell(row=row, column=1).alignment = ca_; row += 1
for s in ["규제독점: 국내 유일 내국인 카지노 면허",
          "무차입경영: 차입금 0원, 순현금 보유",
          "고마진: 영업이익률 20%+ (카지노 특성)",
          "고배당: DPS 1,170원, 배당수익률 6.6%",
          "안정적 현금흐름: 영업CF 4,870억, FCF 3,796억"]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"• {s}").font = df_; ws.cell(row=row, column=1).fill = gnf_
    ws.cell(row=row, column=1).alignment = la_; ws.cell(row=row, column=1).border = tb_; row += 1

ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="약점 (W)").font = Font(name="맑은 고딕", size=12, bold=True, color=W)
ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=RED_C); ws.cell(row=row, column=1).alignment = ca_; row += 1
for w_ in ["전문경영인 부재: 정부 영향력(낙하산 인사 우려)",
           "단일 사업장: 강원도 정선 1곳에 집중",
           "지방 입지: 서울 접근성 낮음 (약 3시간)",
           "성장 한계: 입장객/영업시간 규제로 매출 천장",
           "ESG 리스크: 도박중독 사회적 비용"]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"• {w_}").font = df_; ws.cell(row=row, column=1).fill = rdf_
    ws.cell(row=row, column=1).alignment = la_; ws.cell(row=row, column=1).border = tb_; row += 1

ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="기회 (O)").font = Font(name="맑은 고딕", size=12, bold=True, color=W)
ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="2980B9"); ws.cell(row=row, column=1).alignment = ca_; row += 1
for o_ in ["COVID후 입장객 본격 회복 (2024 코로나 이전 수준 근접)",
           "영업시간 규제 완화 가능성 (현 18시간→확대)",
           "제2카지노영업장 완공 (수용인원 확대)",
           "관광수요 증가 (복합리조트, K-관광)",
           "배당정책 확대 여지 (FCF 대비 배당비율 낮음)"]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"• {o_}").font = df_; ws.cell(row=row, column=1).fill = blf_
    ws.cell(row=row, column=1).alignment = la_; ws.cell(row=row, column=1).border = tb_; row += 1

ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="위협 (T)").font = Font(name="맑은 고딕", size=12, bold=True, color=W)
ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="7F8C8D"); ws.cell(row=row, column=1).alignment = ca_; row += 1
for t_ in ["온라인 도박 합법화 시 독점 가치 훼손",
           "규제 강화: 입장제한 횟수 축소, 영업시간 단축",
           "경기침체: 여가소비 감소 → 입장객/GGR 하락",
           "인구감소: 장기 잠재 고객층 축소",
           "정치적 리스크: 정권 교체 시 경영진/정책 변동"]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"• {t_}").font = df_; ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="E5E7E9")
    ws.cell(row=row, column=1).alignment = la_; ws.cell(row=row, column=1).border = tb_; row += 1

# ==== 13. 10년 실적 요약 ====
row += 1
row = msec(ws, row, "10년 실적 요약 (억원)")
mhdr(ws, row, ["연도", "매출/영업이익", "순이익/EPS"]); row += 1
for yr in range(2015, 2025):
    rev_a = int(ANNUAL_REV[yr] / 억)
    op_a = int(ANNUAL_OP[yr] / 억)
    ni_a = int(ANNUAL_NI[yr] / 억)
    eps_a = ANNUAL_EPS[yr]
    nif = gn_ if ni_a > 0 else rd_
    note = ""
    if yr == 2020: note = " *COVID"
    if yr == 2021: note = " *COVID"
    mrow(ws, row, [f"{yr}{note}", f"{rev_a:,} / {op_a:,}", f"{ni_a:,} / {eps_a:,}"],
         fonts=[db_, df_, nif],
         fills=[llf_, wf_, gnf_ if ni_a > 2000 else (rdf_ if ni_a < 0 else wf_)],
         als=[ca_, ra_, ra_]); row += 1

# COVID 주석
row += 1
for pt in ["2020-2021: COVID-19로 영업정지/축소 → 대규모 적자",
           "2022: 영업 정상화 개시, 2023-2024: 완전 회복",
           "2015-2019 평균 OPM 25%+ → 본래 고마진 사업",
           "2024 순이익 4,569억 = 역대 최고 근접 수준"]:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"• {pt}").font = sm_; ws.cell(row=row, column=1).alignment = la_
    ws.row_dimensions[row].height = 20; row += 1

# ==== 14. 모니터링 ====
row += 1
row = msec(ws, row, "핵심 모니터링 항목")
monitors = [
    ("1", "월별 입장객 수", "카지노 매출의 직접 드라이버. 월별 입장객 공시 모니터링 필수"),
    ("2", "GGR(총게임수익) 추이", "입장객 x Drop x Hold Rate. 고객 1인당 소비액 변화"),
    ("3", "영업시간 규제 변동", "현행 1일 18시간. 확대 시 매출 즉시 증가, 축소 시 직접 타격"),
    ("4", "배당정책 / DPS", "고배당 매력 핵심. 배당성향, DPS 증가율 추적"),
    ("5", "제2카지노영업장 건설", "투자 1,796억. 완공 시 수용인원 확대 → 매출 Jump-Up 가능"),
    ("6", "IR / 경영진 교체", "정부 영향력 하 경영진 교체 시 전략 변동 리스크"),
]
for rank, title, detail in monitors:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=f"{rank}. {title}").font = db_
    ws.cell(row=row, column=1).fill = gld_; ws.cell(row=row, column=1).alignment = la_; ws.cell(row=row, column=1).border = tb_
    ws.row_dimensions[row].height = 22; row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=detail).font = df_
    ws.cell(row=row, column=1).alignment = la_; ws.cell(row=row, column=1).border = tb_
    ws.row_dimensions[row].height = 36; row += 1

# 모니터링 캘린더
row += 1
row = msec(ws, row, "모니터링 캘린더")
mhdr(ws, row, ["시기", "이벤트", "중요도"]); row += 1
for t_, e_, imp in [("매월", "입장객 수 공시", "★★★"),
                     ("2월", "4Q+연간 실적 공시", "★★★"),
                     ("3월", "사업보고서/주총/배당확정", "★★★"),
                     ("5월", "1Q 실적", "★★☆"),
                     ("8월", "2Q+반기 실적", "★★☆"),
                     ("11월", "3Q 실적", "★★☆"),
                     ("수시", "영업시간 규제 변경", "★★★"),
                     ("수시", "제2카지노 건설 진행", "★★☆"),
                     ("수시", "경영진 인사/정부 정책", "★★★")]:
    mrow(ws, row, [t_, e_, imp], fonts=[db_, df_, db_], fills=[llf_, wf_, gld_]); row += 1

# ==== Footer ====
row += 2
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="* 데이터: OpenDART 공시 전수분석 + 잠정실적 | 분석일: 2026.02.06").font = sm_
ws.cell(row=row, column=1).alignment = ca_
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value="* 본 보고서는 투자 참고용이며 투자판단의 책임은 투자자에게 있습니다").font = sm_
ws.cell(row=row, column=1).alignment = ca_

# === SAVE ===
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "강원랜드_모바일용.xlsx")
wb.save(OUT)
print(f"모바일용 보고서 생성: {OUT}")
print(f"총 {row}행, 3열 구성")
