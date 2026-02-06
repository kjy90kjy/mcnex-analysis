# -*- coding: utf-8 -*-
"""강원랜드 종합 기업분석 + 밸류에이션 보고서 (현재가 17,690원 기준)"""
import sqlite3, sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ai.db")
BASE = os.path.dirname(os.path.abspath(__file__))
conn = sqlite3.connect(DB)

# === KEY CONSTANTS ===
COMPANY = "강원랜드"
STOCK_CODE = "035250"
PRICE = 17690
SHARES = 213940500
SHARES_TREASURY = 13534816
SHARES_OUT = SHARES - SHARES_TREASURY  # 유통주식수
억 = 100_000_000
MCAP = PRICE * SHARES  # ~3.78조

# === 2024 연결 (원) ===
REV24 = 1426862781802
OP24 = 285790454716
NI24 = 456895295638
EPS24 = 2257
EQ24 = 3883662813757
EQ23 = 3677900816262
ASSETS24 = 4700641154570
LIAB24 = 816978340813
CASH24 = 244499636973
ST_DEBT = 0; LT_DEBT = 0  # 무차입경영
DA24 = 78474881854
OPCF24 = 487048778217
CAPEX24 = 107494733192
DPS24 = 1170
DIV_PAID = 188547268140

# 2023
REV23 = 1392740543618; OP23 = 258505699379; NI23 = 341061017856; EPS23 = 1681

# === 2025 잠정실적 (4분기 모두 있음) ===
Q24 = [
    {"q":"1Q24","rev":367419e6,"op":78917e6,"ni":116972e6},  # 2024 분기 추정 (연간/4 근사)
    {"q":"2Q24","rev":353281e6,"op":71129e6,"ni":113724e6},
    {"q":"3Q24","rev":371735e6,"op":74744e6,"ni":124461e6},
    {"q":"4Q24","rev":REV24-367419e6-353281e6-371735e6,
     "op":OP24-78917e6-71129e6-74744e6,
     "ni":NI24-116972e6-113724e6-124461e6},
]
Q25 = [
    {"q":"1Q25","rev":365830e6,"op":77659e6,"ni":78021e6},
    {"q":"2Q25","rev":360727e6,"op":57874e6,"ni":60630e6},
    {"q":"3Q25","rev":384147e6,"op":72702e6,"ni":113070e6},
    {"q":"4Q25","rev":365446e6,"op":29697e6,"ni":66017e6},
]
CUM25_REV = 1476726e6; CUM25_OP = 235176e6; CUM25_NI = 316516e6

# === 10년 실적 ===
ANNUAL_REV = {2015:1634441985990,2016:1703131541003,2017:1601291247063,2018:1445736946832,2019:1524006966734,2020:479173424993,2021:788430938373,2022:1272539665429,2023:1392740543618,2024:1426862781802}
ANNUAL_OP = {2015:422764543000,2016:480490893000,2017:460085942000,2018:292485625000,2019:339782744000,2020:-200253285781,2021:-70474730032,2022:145068791398,2023:258505699379,2024:285790454716}
ANNUAL_NI = {2015:466119741406,2016:480264780048,2017:460843127006,2018:314063000747,2019:353337372000,2020:-291290000000,2021:-11056000000,2022:122207000000,2023:341061017856,2024:456895295638}
ANNUAL_EPS = {2015:2178,2016:2242,2017:2159,2018:1467,2019:1651,2020:-1361,2021:-52,2022:570,2023:1681,2024:2257}
ANNUAL_DPS = {2015:980,2016:990,2017:990,2018:900,2019:900,2020:0,2021:0,2022:350,2023:930,2024:1170}

# Derived
BPS = EQ24 / SHARES  # ~18,153
NET_DEBT = ST_DEBT + LT_DEBT - CASH24  # 음수 (Net Cash)
EBITDA24 = OP24 + DA24
FCF24 = OPCF24 - CAPEX24
EV = MCAP + NET_DEBT  # 무차입이므로 EV < 시가총액
AVG_EQ = (EQ24 + EQ23) / 2
ROE24 = NI24 / AVG_EQ

# Trailing (2025 전체 4분기 확정)
TRAIL_NI = CUM25_NI  # 2025 전체
TRAIL_OP = CUM25_OP
TRAIL_REV = CUM25_REV
TRAIL_EPS = TRAIL_NI / SHARES  # 2025 전체 EPS

# 2025 연환산 (4분기 모두 있으므로 = 연간 확정)
E25_ANN_EPS = CUM25_NI / SHARES
# 보수적: 2025 하반기 = 상반기와 동일 가정 (이미 확정이므로 실제 = CUM)
E25_CONS_EPS = CUM25_NI / SHARES

# === STYLE ===
NAVY="1B2A4A"; DARK="2C3E6B"; MID="3A5BA0"; LB="D6E4F0"; LLB="EBF1F8"; W="FFFFFF"
GOLD_C="D4A843"; RED_C="C0392B"; GREEN_C="27AE60"; GRAY_C="F2F2F2"

title_font=Font(name="맑은 고딕",size=22,bold=True,color=W)
sub_font=Font(name="맑은 고딕",size=11,color="B0C4DE")
sec_font=Font(name="맑은 고딕",size=14,bold=True,color=NAVY)
hdr_font=Font(name="맑은 고딕",size=10,bold=True,color=W)
df=Font(name="맑은 고딕",size=10)
db=Font(name="맑은 고딕",size=10,bold=True)
d_blue=Font(name="맑은 고딕",size=10,bold=True,color="0000FF")
d_grn=Font(name="맑은 고딕",size=10,bold=True,color=GREEN_C)
d_red=Font(name="맑은 고딕",size=10,bold=True,color=RED_C)
d_navy=Font(name="맑은 고딕",size=12,bold=True,color=NAVY)
sm=Font(name="맑은 고딕",size=9,color="666666")

tf=PatternFill("solid",fgColor=NAVY)
hf=PatternFill("solid",fgColor=DARK)
mf=PatternFill("solid",fgColor=MID)
lf=PatternFill("solid",fgColor=LB)
llf=PatternFill("solid",fgColor=LLB)
gf=PatternFill("solid",fgColor=GRAY_C)
wf=PatternFill("solid",fgColor=W)
gld=PatternFill("solid",fgColor="FFF3CD")
rdf=PatternFill("solid",fgColor="F8D7DA")
gnf=PatternFill("solid",fgColor="D4EDDA")
blf=PatternFill("solid",fgColor="D6EAF8")

ca=Alignment(horizontal='center',vertical='center',wrap_text=True)
la=Alignment(horizontal='left',vertical='center',wrap_text=True)
ra=Alignment(horizontal='right',vertical='center')
tb=Border(left=Side('thin',color="D9D9D9"),right=Side('thin',color="D9D9D9"),
          top=Side('thin',color="D9D9D9"),bottom=Side('thin',color="D9D9D9"))
bb=Border(bottom=Side('medium',color=NAVY))
NF='#,##0'; PF='0.0%'

def sw(ws,w):
    for i,v in enumerate(w,1): ws.column_dimensions[get_column_letter(i)].width=v
def wh(ws,r,h,fills=None):
    for i,v in enumerate(h,1):
        c=ws.cell(row=r,column=i,value=v); c.font=hdr_font; c.fill=fills[i-1] if fills else hf; c.alignment=ca; c.border=tb
def wr(ws,r,d,fonts=None,fills=None,als=None,nfs=None):
    for i,v in enumerate(d,1):
        c=ws.cell(row=r,column=i,value=v)
        c.font=fonts[i-1] if fonts else df; c.fill=fills[i-1] if fills else wf
        c.alignment=als[i-1] if als else ca; c.border=tb
        if nfs and i<=len(nfs) and nfs[i-1]: c.number_format=nfs[i-1]
def st(ws,r,t,ce=11):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ce)
    c=ws.cell(row=r,column=1,value=t); c.font=sec_font; c.border=bb; return r+1
def fmt(v): return f"{v/억:,.0f}억"
def fw(v): return f"{v:,.0f}원"
def pct(v): return f"{v*100:.1f}%"

wb = Workbook()

# ============================================================
# SHEET 1: 표지
# ============================================================
ws = wb.active; ws.title="표지"; ws.sheet_properties.tabColor=NAVY
sw(ws,[3,20,20,20,20,20,3])
for r in range(1,35):
    for c in range(1,8): ws.cell(row=r,column=c).fill=tf
ws.merge_cells('B6:F6')
ws.cell(row=6,column=2,value="강원랜드(KANGWON LAND)").font=Font(name="맑은 고딕",size=32,bold=True,color=W)
ws.cell(row=6,column=2).alignment=Alignment(horizontal='center',vertical='center')
ws.merge_cells('B8:F8')
ws.cell(row=8,column=2,value="심층 기업분석 + 밸류에이션 종합보고서").font=Font(name="맑은 고딕",size=18,color=GOLD_C)
ws.cell(row=8,column=2).alignment=ca
ws.merge_cells('B11:F11')
ws.cell(row=11,column=2,value="035250 (유가증권)  |  카지노/리조트업  |  국내 유일 내국인 카지노").font=sub_font
ws.cell(row=11,column=2).alignment=ca

DIV_YIELD = DPS24/PRICE
info=[
    (14,"현재주가",f"{PRICE:,}원 (2026.02.06 기준)"),
    (15,"시가총액",f"{fmt(MCAP)} (유통시총 {fmt(PRICE*SHARES_OUT)})"),
    (16,"대표이사 / 설립일","문태곤 / 1998.06.29"),
    (17,"본사","강원도 정선군 사북읍 하이원길 265"),
    (18,"최대주주","한국광해광업공단 (36.27%)"),
    (19,"사업영역","카지노(~80%) + 호텔(~10%) + 리조트(스키/골프/콘도, ~10%)"),
    (20,"분석기준일","2026년 2월 6일 | 데이터: OpenDART 전수분석"),
]
for r,lbl,val in info:
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
    ws.cell(row=r,column=2,value=lbl).font=Font(name="맑은 고딕",size=11,color="8899AA")
    ws.cell(row=r,column=2).alignment=Alignment(horizontal='right',vertical='center')
    ws.merge_cells(start_row=r,start_column=4,end_row=r,end_column=6)
    ws.cell(row=r,column=4,value=val).font=Font(name="맑은 고딕",size=11,bold=True,color=W)
    ws.cell(row=r,column=4).alignment=la
ws.merge_cells('B23:F26')
c=ws.cell(row=23,column=2)
c.value=(f"핵심 밸류에이션 (현재가 {PRICE:,}원 기준):\n"
    f"  PER {PRICE/EPS24:.1f}배(2024) / {PRICE/TRAIL_EPS:.1f}배(2025E) -- 카지노업 PER 10~15배 대비 저평가\n"
    f"  PBR {PRICE/BPS:.2f}배 | EV/EBITDA {EV/EBITDA24:.1f}배 | 배당수익률 {DIV_YIELD*100:.1f}% (고배당주)\n"
    f"  무차입경영 (차입금 0원) | Net Cash {fmt(CASH24)} | ROE {ROE24*100:.1f}%")
c.font=Font(name="맑은 고딕",size=10,color=GOLD_C)
c.alignment=Alignment(horizontal='left',vertical='top',wrap_text=True)
ws.merge_cells('B28:F28')
ws.cell(row=28,column=2,value="시나리오별 목표: Bull ~33,800원(+91%) | Base ~26,600원(+50%) | Bear ~11,800원(-33%)").font=Font(name="맑은 고딕",size=10,bold=True,color="AED6F1")
ws.cell(row=28,column=2).alignment=ca
print("  [1/12] 표지")

# ============================================================
# SHEET 2: 핵심실적 (10년 연결)
# ============================================================
ws2=wb.create_sheet("핵심실적"); ws2.sheet_properties.tabColor=DARK
sw(ws2,[14,14,14,14,12,14,14,14,12,12,14])

row=1; ws2.merge_cells('A1:K1')
ws2.cell(row=1,column=1,value="10년 연결 재무실적 (단위: 억원)").font=sec_font
ws2.cell(row=1,column=1).border=bb; row=3
wh(ws2,row,["연도","매출액","영업이익","순이익","EPS(원)","DPS(원)","배당성향","비고","부채비율","영업이익률","ROE"])
row+=1
start_row=row
years=sorted(ANNUAL_REV.keys())
prev_eq_val=None
for yr in years:
    rev=int(ANNUAL_REV[yr])//억; op=int(ANNUAL_OP[yr])//억; ni=int(ANNUAL_NI[yr])//억
    eps=ANNUAL_EPS[yr]; dps=ANNUAL_DPS[yr]
    payout=f"{dps/eps*100:.1f}%" if eps>0 and dps>0 else "-"
    note=""
    if yr==2020: note="COVID 셧다운"
    elif yr==2021: note="COVID 부분회복"
    elif yr==2022: note="위드코로나"
    elif yr==2024: note="역대최고 순이익"
    fts=[db]+[df]*10; fls=[lf]+[wf]*10
    if op<0: fts[2]=d_red
    if ni<0: fts[3]=d_red
    if yr in [2020,2021]: fls=[lf]+[rdf]*7+[wf]*3  # COVID 기간 빨간 배경
    wr(ws2,row,[yr,rev,op,ni,eps,dps,payout,note,None,None,None],
       fonts=fts,fills=fls,als=[ca]+[ra]*6+[la]+[ra]*3,
       nfs=[None]+[NF]*3+[NF,NF,None,None,PF,PF,PF])
    # 부채비율, 영업이익률, ROE는 수식 대신 직접 계산
    opm=ANNUAL_OP[yr]/ANNUAL_REV[yr] if ANNUAL_REV[yr]!=0 else 0
    ws2.cell(row=row,column=10,value=opm).number_format=PF
    ws2.cell(row=row,column=10).font=df; ws2.cell(row=row,column=10).alignment=ra; ws2.cell(row=row,column=10).border=tb
    row+=1

row+=1
ws2.merge_cells(start_row=row,start_column=1,end_row=row,end_column=11)
ws2.cell(row=row,column=1,value="* 2020~2021년 COVID-19로 카지노 영업 중단/제한. 비정상 기간으로 밸류에이션 참고 시 제외 권장.").font=Font(name="맑은 고딕",size=9,bold=True,color=RED_C)
ws2.cell(row=row,column=1).alignment=la
row+=2; st(ws2,row,"전년대비 성장률 (YoY)"); row+=1
wh(ws2,row,["연도","매출 YoY","영업이익 YoY","순이익 YoY"]); row+=1
prev_data=None
for yr in years:
    if prev_data is not None:
        prev_rev,prev_op,prev_ni=prev_data
        r_yoy=(ANNUAL_REV[yr]-prev_rev)/abs(prev_rev) if prev_rev!=0 else None
        o_yoy=(ANNUAL_OP[yr]-prev_op)/abs(prev_op) if prev_op!=0 else None
        n_yoy=(ANNUAL_NI[yr]-prev_ni)/abs(prev_ni) if prev_ni!=0 else None
        wr(ws2,row,[yr,r_yoy,o_yoy,n_yoy],
           fonts=[db,df,df,df],fills=[lf,wf,wf,wf],als=[ca,ra,ra,ra],nfs=[None,PF,PF,PF])
        row+=1
    prev_data=(ANNUAL_REV[yr],ANNUAL_OP[yr],ANNUAL_NI[yr])
print("  [2/12] 핵심실적")

# ============================================================
# SHEET 3: 2025실적 (4분기 잠정실적 모두)
# ============================================================
ws3=wb.create_sheet("2025실적"); ws3.sheet_properties.tabColor=GREEN_C
sw(ws3,[14,14,14,14,14,16,16])
row=1; ws3.merge_cells('A1:G1')
ws3.cell(row=1,column=1,value="2025년 분기별 잠정실적 (단위: 백만원)").font=sec_font
ws3.cell(row=1,column=1).border=bb; row=3
wh(ws3,row,["분기","매출액","영업이익","순이익","영업이익률","매출 YoY","영업이익 YoY"]); row+=1

# 2024 분기 데이터
q24_data=[
    ("24Q1",int(Q24[0]["rev"]/1e6),int(Q24[0]["op"]/1e6),int(Q24[0]["ni"]/1e6)),
    ("24Q2",int(Q24[1]["rev"]/1e6),int(Q24[1]["op"]/1e6),int(Q24[1]["ni"]/1e6)),
    ("24Q3",int(Q24[2]["rev"]/1e6),int(Q24[2]["op"]/1e6),int(Q24[2]["ni"]/1e6)),
    ("24Q4",int(Q24[3]["rev"]/1e6),int(Q24[3]["op"]/1e6),int(Q24[3]["ni"]/1e6)),
    ("24합계",int(REV24/1e6),int(OP24/1e6),int(NI24/1e6)),
]
q25_data=[
    ("25Q1",int(Q25[0]["rev"]/1e6),int(Q25[0]["op"]/1e6),int(Q25[0]["ni"]/1e6)),
    ("25Q2",int(Q25[1]["rev"]/1e6),int(Q25[1]["op"]/1e6),int(Q25[1]["ni"]/1e6)),
    ("25Q3",int(Q25[2]["rev"]/1e6),int(Q25[2]["op"]/1e6),int(Q25[2]["ni"]/1e6)),
    ("25Q4",int(Q25[3]["rev"]/1e6),int(Q25[3]["op"]/1e6),int(Q25[3]["ni"]/1e6)),
    ("25합계",int(CUM25_REV/1e6),int(CUM25_OP/1e6),int(CUM25_NI/1e6)),
]
qd=q24_data+q25_data
q24_start_row=row
for i,(q,rv,op,ni) in enumerate(qd):
    tot="합계" in q; is25="25" in q
    fl=gld if tot else (gnf if is25 else wf); fn=db if tot else df
    wr(ws3,row,[q,rv,op,ni,None,None,None],fonts=[fn]*7,fills=[fl]*7,als=[ca]+[ra]*6,nfs=[None,NF,NF,NF,PF,PF,PF])
    ws3.cell(row=row,column=5,value=op/rv if rv!=0 else 0).number_format=PF
    ws3.cell(row=row,column=5).border=tb; ws3.cell(row=row,column=5).font=df
    # YoY for 25Q1~25Q4 vs 24Q1~24Q4
    if 5<=i<=8:  # 25Q1~25Q4
        pr_row=q24_start_row+(i-5)
        ws3.cell(row=row,column=6,value=f"=(B{row}-B{pr_row})/B{pr_row}").number_format=PF
        ws3.cell(row=row,column=6).border=tb; ws3.cell(row=row,column=6).font=df
        ws3.cell(row=row,column=7,value=f"=(C{row}-C{pr_row})/C{pr_row}").number_format=PF
        ws3.cell(row=row,column=7).border=tb; ws3.cell(row=row,column=7).font=df
    row+=1
row+=2; st(ws3,row,"핵심 포인트",7); row+=1
for pt in ["2025 연간 매출 1.477조 (YoY +3.5%) -- 매출은 소폭 성장",
           "2025 연간 영업이익 2,352억 (YoY -17.7%) -- 수익성 하락이 핵심 이슈",
           "4Q25 영업이익 297억으로 급락 (OPM 8.1%) -- 계절적 요인 or 비용 증가 확인 필요",
           "3Q25 순이익 1,131억으로 양호 -- 영업외수익(금융수익 등) 기여 추정",
           "COVID 이전 수준(매출 1.5~1.7조) 대비 아직 미회복. 구조적 성장 한계 vs 규제 완화 기대"]:
    ws3.merge_cells(start_row=row,start_column=1,end_row=row,end_column=7)
    ws3.cell(row=row,column=1,value=f"  {pt}").font=df; ws3.cell(row=row,column=1).alignment=la; row+=1
print("  [3/12] 2025실적")

# ============================================================
# SHEET 4: 사업구조
# ============================================================
ws4=wb.create_sheet("사업구조"); ws4.sheet_properties.tabColor=GOLD_C
sw(ws4,[18,14,14,14,14,14,14,14])
row=1; ws4.merge_cells('A1:H1')
ws4.cell(row=1,column=1,value="사업부문별 매출 구조 (연결, 단위: 억원)").font=sec_font
ws4.cell(row=1,column=1).border=bb; row=3

st(ws4,row,"A. 매출 구성 (2024 기준 추정)",8); row+=1
wh(ws4,row,["사업부문","내용","매출비중","매출(억)","핵심 드라이버","수익성","특성","비고"]); row+=1
for nm,desc,pct_v,rev_v,driver,margin,char,note in [
    ("카지노","테이블게임+슬롯머신","~80%",f"{int(REV24*0.8/억):,}","입장객수 x 1인당 GGR","高","독점면허",
     "내국인 전용 카지노"),
    ("호텔","하이원호텔+컨벤션","~10%",f"{int(REV24*0.1/억):,}","투숙률, 객단가","中","카지노 연계",
     "카지노 방문객 연계"),
    ("리조트","스키/골프/콘도","~10%",f"{int(REV24*0.1/억):,}","방문객수, 시즌","中~低","계절성",
     "동절기 스키, 하절기 골프"),
]:
    wr(ws4,row,[nm,desc,pct_v,rev_v,driver,margin,char,note],
       fonts=[db,df,d_blue,df,df,df,df,sm],fills=[llf,wf,gld,wf,wf,wf,wf,gf],
       als=[la,la,ca,ra,la,ca,ca,la]); row+=1

row+=2; st(ws4,row,"B. 매출 드라이버 분석",8); row+=1
wh(ws4,row,["핵심 변수","설명","2024 상황","향후 전망"]); row+=1
for var,desc,now,outlook in [
    ("입장객수","카지노 매출의 핵심 볼륨 지표","COVID 이전 대비 ~90% 회복","규제완화(입장횟수) 시 상승여력"),
    ("GGR(총게임수익)","1인당 게임순수익금 (drop x win rate)","Table+Slot 합산 개선 추세","테이블게임 비중 증가 시 GGR 상승"),
    ("영업시간","현재 주간영업(08~06시)","하루 22시간","24시간 영업 전환 기대감"),
    ("입장료","1회 9,000원 (2024 기준)","입장료 인상 이슈","인상 시 입장객 감소 vs 객단가 상승"),
    ("입장횟수","월 15회 제한","사행산업감독위 규제","완화 시 최대 수혜"),
]:
    wr(ws4,row,[var,desc,now,outlook],fonts=[db,df,df,df],fills=[llf,wf,wf,blf],als=[la,la,la,la]); row+=1

row+=2; st(ws4,row,"C. 주요 투자 프로젝트",8); row+=1
wh(ws4,row,["프로젝트","투자금액","일정","기대효과"]); row+=1
for proj,amt,sched,eff in [
    ("제2카지노영업장","1,796억원","건설 진행 중","수용능력 확대, 매출 성장 드라이버"),
    ("하이원 리조트 리뉴얼","수백억","지속 투자","리조트 경쟁력 강화"),
    ("스마트카지노","수십억","순차 도입","무인 슬롯, 디지털 전환"),
]:
    wr(ws4,row,[proj,amt,sched,eff],fonts=[db,d_blue,df,df],fills=[llf,gld,wf,gnf],als=[la,ra,la,la]); row+=1
print("  [4/12] 사업구조")

# ============================================================
# SHEET 5: 주주환원
# ============================================================
ws5=wb.create_sheet("주주환원"); ws5.sheet_properties.tabColor="E74C3C"
sw(ws5,[12,14,12,14,14,14])
row=1; ws5.merge_cells('A1:F1')
ws5.cell(row=1,column=1,value="배당 및 주주환원 정책").font=sec_font; ws5.cell(row=1,column=1).border=bb; row=3
wh(ws5,row,["연도","주당배당(원)","EPS(원)","배당성향","배당수익률","비고"]); row+=1
for yr in sorted(ANNUAL_EPS.keys()):
    eps_v=ANNUAL_EPS[yr]; dps_v=ANNUAL_DPS[yr]
    po=f"{dps_v/eps_v*100:.1f}%" if eps_v>0 and dps_v>0 else "-"
    yl=f"{dps_v/PRICE*100:.2f}%" if dps_v>0 else "-"
    note=""
    if yr==2020: note="COVID, 무배당"
    elif yr==2021: note="COVID, 무배당"
    elif yr==2024: note="대폭 인상"
    elif yr==2023: note="배당 재개 본격화"
    wr(ws5,row,[yr,dps_v if dps_v else "-",eps_v,po,yl,note],
       fonts=[db,d_grn if dps_v>0 else d_red,df,df,df,sm],als=[ca,ra,ra,ca,ca,la])
    row+=1

row+=2; st(ws5,row,"배당수익률 분석",6); row+=1
ws5.merge_cells(start_row=row,start_column=1,end_row=row,end_column=6)
ws5.cell(row=row,column=1,value=f"현재 배당수익률 {DPS24/PRICE*100:.1f}% -- 코스피 평균(~2%) 대비 고배당. 안정적 현금흐름 기반 배당 증가 추세.").font=db
ws5.cell(row=row,column=1).alignment=la; row+=2

st(ws5,row,f"주식 구조 (2024.12.31) -- 현재가 {PRICE:,}원 기준",6); row+=1
for lbl,val in [("발행주식수(보통주)",f"{SHARES:,}주"),("자기주식",f"{SHARES_TREASURY:,}주 ({SHARES_TREASURY/SHARES*100:.1f}%)"),
                ("유통주식수",f"{SHARES_OUT:,}주"),("시가총액",fmt(MCAP)),
                ("유통시총",fmt(PRICE*SHARES_OUT)),
                ("최대주주(한국광해광업공단)","77,633,489주 (36.27%)"),
                ("유증 이력","없음 (설립 이후 유상증자 없음)")]:
    ws5.cell(row=row,column=1,value=lbl).font=df; ws5.cell(row=row,column=1).alignment=la
    ws5.cell(row=row,column=1).fill=llf; ws5.cell(row=row,column=1).border=tb
    ws5.merge_cells(start_row=row,start_column=2,end_row=row,end_column=3)
    ws5.cell(row=row,column=2,value=val).font=db; ws5.cell(row=row,column=2).alignment=la; ws5.cell(row=row,column=2).border=tb
    row+=1
print("  [5/12] 주주환원")

# ============================================================
# SHEET 6: 규제_면허 (replaces R&D_특허)
# ============================================================
ws6=wb.create_sheet("규제_면허"); ws6.sheet_properties.tabColor="8E44AD"
sw(ws6,[18,50,18,18])
row=1; ws6.merge_cells('A1:D1')
ws6.cell(row=1,column=1,value="규제 환경 및 면허 현황").font=sec_font; ws6.cell(row=1,column=1).border=bb

row=3; st(ws6,row,"A. 카지노 면허 (핵심 경쟁우위)",4); row+=1
wh(ws6,row,["항목","내용","영향","비고"]); row+=1
for item,content,impact,note in [
    ("내국인 카지노 면허","폐광지역 개발 지원에 관한 특별법에 의한 유일한 면허",
     "절대적 진입장벽","법률 개정 없이 신규 면허 불가"),
    ("면허 기한","반영구적 (특별법에 근거)","사업 지속성 보장","정치적 리스크 매우 낮음"),
    ("면허 취소 사유","법령 위반, 공익 저해 등 극단적 경우만","실질적 취소 가능성 극히 낮음",""),
    ("외국인 카지노(참고)","외국인 전용 17개소 (파라다이스, GKL 등)","직접 경쟁 아님","고객층 완전 분리"),
]:
    wr(ws6,row,[item,content,impact,note],fonts=[db,df,df,sm],fills=[llf,wf,blf,gf],als=[la,la,la,la]); row+=1

row+=1; st(ws6,row,"B. 영업 규제 현황",4); row+=1
wh(ws6,row,["규제 항목","현행 기준","사업 영향","완화 가능성"]); row+=1
for reg,curr,impact,relax in [
    ("영업시간","08:00~익일06:00 (22시간)","매출 상한 존재","24시간 전환 논의 중"),
    ("입장횟수","월 15회 제한","고빈도 고객 매출 제한","완화 시 매출 10~15% 증가 추정"),
    ("입장료","1회 9,000원","저가 고객 유입 제한","인상 논의 (입장객 감소 리스크)"),
    ("베팅한도","테이블별 상이 (최대 30만원)","VIP 고객 유치 제한","조정 가능성 있음"),
    ("신분증 확인","신분증 의무 제시","미성년자/출입금지자 차단","규제 유지"),
    ("출입제한제도","본인/가족 신청 출입차단","매출 감소 요인","사회적 책임으로 유지"),
]:
    relax_font=d_grn if "완화" in relax or "전환" in relax else df
    wr(ws6,row,[reg,curr,impact,relax],fonts=[db,df,df,relax_font],fills=[llf,wf,wf,wf],als=[la,la,la,la]); row+=1

row+=1; st(ws6,row,"C. 감독 체계 및 사회적 의무",4); row+=1
wh(ws6,row,["항목","내용","비고",""]); row+=1
for item,content,note in [
    ("감독기관","사행산업통합감독위원회","문화체육관광부 산하"),
    ("지역공헌 의무","카지노 매출의 일정비율 지역발전기금 출연","정선군/강원도 발전"),
    ("도박중독 대응","한국도박문제관리센터 연계, 자체 프로그램 운영","CSR 의무"),
    ("공기업적 성격","최대주주 한국광해광업공단(정부산하기관)","경영 자율성 제한 요인"),
    ("특별법 근거","폐광지역개발지원에관한특별법","2025년 일몰 연장 완료"),
]:
    wr(ws6,row,[item,content,note,""],fonts=[db,df,sm,df],fills=[llf,wf,gf,wf],als=[la,la,la,la]); row+=1

row+=1
ws6.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
ws6.cell(row=row,column=1,value="* 핵심: 내국인 카지노 면허는 대체 불가능한 독점 자산. 규제 완화/강화가 주가의 최대 변수.").font=Font(name="맑은 고딕",size=10,bold=True,color=MID)
ws6.cell(row=row,column=1).alignment=la
print("  [6/12] 규제_면허")

# ============================================================
# SHEET 7: 밸류에이션 종합
# ============================================================
ws7=wb.create_sheet("밸류에이션"); ws7.sheet_properties.tabColor="E67E22"
sw(ws7,[22,18,18,18,18,22])
row=1; ws7.merge_cells('A1:F2')
c=ws7.cell(row=1,column=1,value=f"강원랜드 밸류에이션 종합 (현재가 {PRICE:,}원)")
c.font=Font(name="맑은 고딕",size=18,bold=True,color=W); c.fill=tf; c.alignment=ca
row=4
# Key info
for d in [["현재 주가",fw(PRICE),"발행주식수",f"{SHARES:,}주","시가총액",fmt(MCAP)],
          ["2024 EPS",fw(EPS24),"2024 BPS",fw(int(BPS)),"2024 DPS",fw(DPS24)],
          ["2025E EPS",fw(int(TRAIL_EPS)),"ROE (2024)",pct(ROE24),"부채비율",pct(LIAB24/EQ24)]]:
    for i in range(0,6,2):
        ws7.cell(row=row,column=i+1,value=d[i]).font=db; ws7.cell(row=row,column=i+1).fill=blf
        ws7.cell(row=row,column=i+1).alignment=la; ws7.cell(row=row,column=i+1).border=tb
        ws7.cell(row=row,column=i+2,value=d[i+1]).font=d_blue; ws7.cell(row=row,column=i+2).fill=wf
        ws7.cell(row=row,column=i+2).alignment=ca; ws7.cell(row=row,column=i+2).border=tb
    row+=1
row+=1; st(ws7,row,"멀티플 종합",6); row+=1
wh(ws7,row,["지표","산출 방식","값","판정","카지노업 참고","비고"]); row+=1

E25_EPS = int(CUM25_NI / SHARES)
PER_24 = PRICE/EPS24
PER_25 = PRICE/E25_EPS

vals_list = [
    ["PER (2024 확정)",f"주가/EPS({fw(EPS24)})",f"{PER_24:.1f}배","저평가","10~15배","확정 사업보고서"],
    ["PER (2025E)",f"주가/EPS({fw(E25_EPS)})",f"{PER_25:.1f}배","적정~고평가","10~15배","2025 잠정실적 전체"],
    ["PBR",f"주가/BPS({fw(int(BPS))})",f"{PRICE/BPS:.2f}배","적정","0.8~1.5배","무차입 고ROE"],
    ["EV/EBITDA",f"EV({fmt(EV)})/EBITDA({fmt(EBITDA24)})",f"{EV/EBITDA24:.1f}배","저평가","8~12배","무차입→EV<시총"],
    ["PSR (2024)",f"시총/매출({fmt(REV24)})",f"{MCAP/REV24:.2f}배","적정","2~4배","카지노 고마진"],
    ["배당수익률",f"DPS({fw(DPS24)})/주가",pct(DPS24/PRICE),"매우양호","2~4%","고배당주"],
    ["FCF 수익률",f"FCF({fw(int(FCF24/SHARES))})/주가",pct(FCF24/SHARES/PRICE),"양호","3~6%",""],
    ["ROE",f"순이익/평균자본",pct(ROE24),"양호","8~15%","무차입 순수 ROE"],
    ["Net Cash/시총",f"순현금({fmt(CASH24)})/시총",pct(CASH24/MCAP),"양호","-","무차입경영 프리미엄"],
]
for v in vals_list:
    j=v[3]
    if "매우" in j: jf,jfl=d_grn,gnf
    elif "저평가" in j or "양호" in j: jf,jfl=Font(name="맑은 고딕",size=10,bold=True,color="2E86C1"),blf
    elif "적정" in j: jf,jfl=db,gld
    else: jf,jfl=d_red,rdf
    wr(ws7,row,v,fonts=[db,df,d_blue,jf,df,sm],fills=[llf,wf,gld,jfl,gf,wf],als=[la,la,ca,ca,ca,la]); row+=1
print("  [7/12] 밸류에이션 종합")

# ============================================================
# SHEET 8: PER 상세
# ============================================================
ws8=wb.create_sheet("PER분석"); ws8.sheet_properties.tabColor="2980B9"
sw(ws8,[24,16,16,16,16,20])
row=1; st(ws8,row,f"PER 다각도 분석 (현재가 {PRICE:,}원)",6); row+=1
st(ws8,row,"A. EPS 산출 방식별 PER",6); row+=1
wh(ws8,row,["산출 방식","순이익(억)","EPS(원)","PER(배)","의미","비고"]); row+=1
for lbl,ni_v,eps_v,desc,note in [
    ("2024 확정",int(NI24/억),EPS24,"확정치 기준","사업보고서"),
    ("2025 잠정 (전체)",int(CUM25_NI/억),int(CUM25_NI/SHARES),"2025년 전체","4분기 모두 확정"),
    ("2024-2025 평균",int((NI24+CUM25_NI)/2/억),int((NI24+CUM25_NI)/2/SHARES),"2개년 평균","변동성 완화"),
    ("COVID제외 5년평균",int((ANNUAL_NI[2022]+ANNUAL_NI[2023]+ANNUAL_NI[2024]+CUM25_NI)/4/억),
     int((ANNUAL_NI[2022]+ANNUAL_NI[2023]+ANNUAL_NI[2024]+CUM25_NI)/4/SHARES),
     "정상화 후 평균","2022~2025"),
]:
    per=PRICE/eps_v if eps_v>0 else 0
    pf2=d_grn if per<10 else (d_blue if per<12 else db)
    pfl=gnf if per<10 else (blf if per<12 else gld)
    wr(ws8,row,[lbl,ni_v,f"{eps_v:,}",f"{per:.2f}배" if per>0 else "N/A",desc,note],
       fonts=[db,df,d_blue,pf2,df,sm],fills=[llf,wf,gld,pfl,wf,gf],als=[la,ra,ra,ca,la,la]); row+=1

# B. Quarterly
row+=1; st(ws8,row,"B. 분기별 실적 추이 (백만원)",6); row+=1
wh(ws8,row,["분기","매출액","영업이익","OPM","순이익","YoY 순이익"]); row+=1
qs=[("24Q1",Q24[0]),("24Q2",Q24[1]),("24Q3",Q24[2]),("24Q4",Q24[3]),
    ("25Q1",Q25[0]),("25Q2",Q25[1]),("25Q3",Q25[2]),("25Q4",Q25[3])]
pq={"25Q1":Q24[0],"25Q2":Q24[1],"25Q3":Q24[2],"25Q4":Q24[3]}
for n,q in qs:
    rm=int(q["rev"]/1e6); om=int(q["op"]/1e6); nm_=int(q["ni"]/1e6)
    opm=q["op"]/q["rev"] if q["rev"] else 0
    yoy=""
    if n in pq:
        p=pq[n]
        if p["ni"]!=0: yoy=f"{(q['ni']-p['ni'])/abs(p['ni'])*100:+.1f}%"
    is25=n.startswith("25"); fl2=gnf if is25 else wf
    wr(ws8,row,[n,f"{rm:,}",f"{om:,}",f"{opm*100:.1f}%",f"{nm_:,}",yoy],
       fonts=[db,df,df,d_grn if opm>0.15 else(d_red if opm<0.10 else df),df,db],
       fills=[blf if is25 else llf]+[fl2]*5,als=[ca,ra,ra,ca,ra,ca]); row+=1
# Annual sum
wr(ws8,row,["2025 연간",f"{int(CUM25_REV/1e6):,}",f"{int(CUM25_OP/1e6):,}",
            f"{CUM25_OP/CUM25_REV*100:.1f}%",f"{int(CUM25_NI/1e6):,}",""],
   fonts=[db]*6,fills=[gld]*6,als=[ca,ra,ra,ca,ra,ca]); row+=1

# C. Target PER
row+=1; st(ws8,row,"C. 목표PER별 적정주가 (카지노업 PER 10~15배)",6); row+=1
wh(ws8,row,["기준","EPS(원)","PER 8배","PER 10배","PER 12배","PER 15배"]); row+=1
for lbl,eps in [("2024 확정",EPS24),("2025E 잠정",int(CUM25_NI/SHARES)),
                ("2개년 평균",int((NI24+CUM25_NI)/2/SHARES)),("정상화 5년평균",int((ANNUAL_NI[2022]+ANNUAL_NI[2023]+ANNUAL_NI[2024]+CUM25_NI)/4/SHARES))]:
    wr(ws8,row,[lbl,f"{eps:,}",fw(eps*8),fw(eps*10),fw(eps*12),fw(eps*15)],
       fonts=[db,d_blue,df,df,d_grn,d_grn],fills=[llf,gld,wf,gnf if eps*10>PRICE else wf,gnf,gnf],
       als=[la,ra,ra,ra,ra,ra]); row+=1
row+=1
ws8.merge_cells(start_row=row,start_column=1,end_row=row,end_column=6)
ws8.cell(row=row,column=1,value=f"-> 현재 {PRICE:,}원 = 2024 EPS 기준 PER {PRICE/EPS24:.1f}배. 2025E 기준 PER {PRICE/(CUM25_NI/SHARES):.1f}배. 카지노업 PER 10~15배 참고.").font=db
print("  [8/12] PER분석")

# ============================================================
# SHEET 9: PBR/ROE
# ============================================================
ws9=wb.create_sheet("PBR_ROE"); ws9.sheet_properties.tabColor="8E44AD"
sw(ws9,[20,16,16,16,16,22])
row=1; st(ws9,row,"PBR / ROE / 잔여이익모델(RIM)",6); row+=1

# Historical
st(ws9,row,"A. 연도별 BPS/ROE 추이",6); row+=1
wh(ws9,row,["연도","자본(억)","BPS(원)","ROE","EPS(원)","순이익(억)"]); row+=1

# 자본 데이터 (추정)
EQ_HIST = {
    2019: 3_326_000_000_000,  # 추정
    2020: 3_035_000_000_000,  # COVID 적자로 자본 감소
    2021: 3_024_000_000_000,
    2022: 3_312_000_000_000,
    2023: EQ23,
    2024: EQ24,
}
peq=None
for yr in [2019,2020,2021,2022,2023,2024]:
    ni=ANNUAL_NI[yr]; eps=ANNUAL_EPS[yr]; eq=EQ_HIST[yr]
    bps=int(eq/SHARES)
    roe=ni/((eq+peq)/2) if peq else ni/eq; peq=eq
    rf=d_grn if roe>0.10 else(d_red if roe<0 else df)
    ni_억=int(ni/억)
    note_fill=rdf if yr in [2020,2021] else wf
    wr(ws9,row,[str(yr),int(eq/억),f"{bps:,}",f"{roe*100:.1f}%",f"{eps:,}",ni_억],
       fonts=[db,df,d_blue,rf,df,d_red if ni<0 else df],
       fills=[llf if yr not in [2020,2021] else rdf,note_fill,gld,note_fill,note_fill,note_fill],
       als=[ca,ra,ra,ca,ra,ra]); row+=1

# PBR
row+=1; st(ws9,row,"B. 현재 PBR",6); row+=1
for lbl,val in [(f"BPS ({fw(int(BPS))})",f"PBR = {PRICE/BPS:.2f}배"),
                ("PBR 0.8배 주가",fw(int(BPS*0.8))),("PBR 1.0배 주가",fw(int(BPS))),
                ("PBR 1.2배 주가",fw(int(BPS*1.2))),("PBR 1.5배 주가",fw(int(BPS*1.5)))]:
    ws9.cell(row=row,column=1,value=lbl).font=db; ws9.cell(row=row,column=1).fill=llf
    ws9.cell(row=row,column=1).alignment=la; ws9.cell(row=row,column=1).border=tb
    ws9.cell(row=row,column=2,value=val).font=d_blue; ws9.cell(row=row,column=2).fill=gld
    ws9.cell(row=row,column=2).alignment=ca; ws9.cell(row=row,column=2).border=tb; row+=1

# RIM
row+=1; st(ws9,row,"C. 잔여이익모델(RIM) 적정주가",6); row+=1
ws9.cell(row=row,column=1,value="산식: BPS x (1 + (ROE-ke)/(ke-g)) | ke=9% (무위험3.5%+beta0.8xERP6.5%)").font=sm; row+=1
# 강원랜드는 beta가 낮음 (독점 공공재적 성격) → ke=9%
wh(ws9,row,["시나리오","지속ROE","ke","성장률(g)","적정주가","현재가 대비"]); row+=1
for lbl,roe,ke,g in [("보수적(정상화ROE)",0.08,0.09,0.01),("기본(2024수준)",0.121,0.09,0.015),
                      ("적극적(규제완화)",0.15,0.09,0.02),("낙관적(최적화)",0.18,0.09,0.02)]:
    fv=BPS*(1+(roe-ke)/(ke-g)); up=(fv-PRICE)/PRICE
    uf=d_grn if up>0 else d_red; ufl=gnf if up>0 else rdf
    wr(ws9,row,[lbl,f"{roe*100:.1f}%",f"{ke*100:.0f}%",f"{g*100:.1f}%",fw(int(fv)),f"{up*100:+.1f}%"],
       fonts=[db,df,df,df,d_blue,uf],fills=[llf,wf,wf,wf,gld,ufl],als=[la,ca,ca,ca,ra,ca]); row+=1
row+=1
ws9.merge_cells(start_row=row,start_column=1,end_row=row,end_column=6)
rim_base=BPS*(1+(0.121-0.09)/(0.09-0.015))
rim_agg=BPS*(1+(0.15-0.09)/(0.09-0.02))
ws9.cell(row=row,column=1,value=f"-> 기본(ROE12.1%) 적정가 ~{fw(int(rim_base))}. 적극적(15%) ~{fw(int(rim_agg))}. 현재 PBR {PRICE/BPS:.2f}배.").font=db
print("  [9/12] PBR/ROE")

# ============================================================
# SHEET 10: EV/EBITDA
# ============================================================
ws10=wb.create_sheet("EV_EBITDA"); ws10.sheet_properties.tabColor="E67E22"
sw(ws10,[24,18,18,18,24])
row=1; st(ws10,row,"EV/EBITDA & FCF 밸류에이션",5); row+=1

# EV
st(ws10,row,"A. Enterprise Value (무차입경영)",5); row+=1
wh(ws10,row,["항목","금액(억)","비고","",""]); row+=1
for lbl,amt,note in [("시가총액",int(MCAP/억),f"주가{PRICE:,}x{SHARES:,}주"),
                      ("(+) 총차입금",0,"단기 0 + 장기 0 (무차입경영)"),
                      ("(-) 현금",int(CASH24/억),"현금및현금성자산"),
                      ("(=) 순차입금",int(NET_DEBT/억),"음수 = Net Cash position"),
                      ("(=) EV",int(EV/억),"시총+순차입금 (시총보다 작음)")]:
    tot="(=)" in lbl
    amt_display=f"{amt:,}" if isinstance(amt,int) else str(amt)
    wr(ws10,row,[lbl,amt_display,note,"",""],fonts=[db if tot else df,d_blue if tot else df,sm,df,df],
       fills=[gld if tot else llf,gld if tot else wf,wf,wf,wf],als=[la,ra,la,la,la]); row+=1

row+=1
ws10.merge_cells(start_row=row,start_column=1,end_row=row,end_column=5)
ws10.cell(row=row,column=1,value="* 무차입경영: 차입금 0원, 현금 2,445억 보유. EV가 시가총액보다 낮아 EV/EBITDA 배수 유리.").font=Font(name="맑은 고딕",size=9,bold=True,color="2E86C1")
ws10.cell(row=row,column=1).alignment=la; row+=2

# EBITDA
st(ws10,row,"B. EBITDA",5); row+=1
wh(ws10,row,["항목","2024(억)","비고","",""]); row+=1
for lbl,amt,note in [("영업이익",int(OP24/억),""),
                      ("(+) 감가상각비",int(DA24/억),"유형+무형자산 상각"),
                      ("(=) EBITDA",int(EBITDA24/억),f"마진 {EBITDA24/REV24*100:.1f}%")]:
    tot="(=)" in lbl
    wr(ws10,row,[lbl,f"{amt:,}",note,"",""],fonts=[db if tot else df,d_blue if tot else df,sm,df,df],
       fills=[gld if tot else llf,gld if tot else wf,wf,wf,wf],als=[la,ra,la,la,la]); row+=1

# EV/EBITDA scenarios
row+=1; st(ws10,row,"C. EV/EBITDA 적정주가",5); row+=1
wh(ws10,row,["기준","EBITDA(억)","현재배수","적정EV(10배)","적정주가(10배)"]); row+=1
ebitda_24=int(EBITDA24/억)
ebitda_25=int((CUM25_OP+DA24)/억)  # 2025 OP + 2024 DA 근사
for lbl,eb in [("2024 확정",ebitda_24),("2025E (OP+DA)",ebitda_25)]:
    ev_eb=(EV/억)/eb; fev10=eb*10; fp=int((fev10-int(NET_DEBT/억))*억/SHARES)
    wr(ws10,row,[lbl,f"{eb:,}",f"{ev_eb:.1f}배",f"{fev10:,}억",fw(fp)],
       fonts=[db,df,d_blue,df,d_grn],fills=[llf,wf,gld,wf,gnf],als=[la,ra,ca,ra,ra]); row+=1

# FCF
row+=1; st(ws10,row,"D. FCF & 주주환원",5); row+=1
wh(ws10,row,["항목","2024(억)","비고","",""]); row+=1
for lbl,amt,note in [("영업활동CF",int(OPCF24/억),""),
                      ("(-) CAPEX",int(CAPEX24/억),"유형자산취득 (제2카지노 포함)"),
                      ("(=) FCF",int(FCF24/억),""),
                      ("FCF/주","",fw(int(FCF24/SHARES))),
                      ("FCF 수익률","",pct(FCF24/SHARES/PRICE)),
                      ("","",""),
                      ("배당금 지급",int(DIV_PAID/억),f"{fw(DPS24)}/주"),
                      ("자기주식(보유)",f"{SHARES_TREASURY:,}주",f"취득가 기준 ~{int(SHARES_TREASURY*15000/억):,}억 추정"),
                      ("총 배당환원율","",pct(DIV_PAID/MCAP))]:
    if not lbl: row+=1; continue
    tot="(=)" in lbl or "수익률" in lbl or "환원율" in lbl
    v2=f"{amt:,}" if isinstance(amt,int) and amt>0 else (str(amt) if amt else note)
    v3=note if isinstance(amt,int) and amt>0 else ("" if not amt else "")
    wr(ws10,row,[lbl,v2,v3,"",""],fonts=[db if tot else df,d_blue if tot else df,sm,df,df],
       fills=[gld if tot else llf,gld if tot else wf,wf,wf,wf],als=[la,ra,la,la,la]); row+=1
row+=1
ws10.merge_cells(start_row=row,start_column=1,end_row=row,end_column=5)
ws10.cell(row=row,column=1,value=f"-> FCF수익률 {FCF24/SHARES/PRICE*100:.1f}%, 배당수익률 {DPS24/PRICE*100:.1f}%. 무차입+고배당+안정적 FCF 창출 기업.").font=db
print("  [10/12] EV/EBITDA")

# ============================================================
# SHEET 11: 시나리오
# ============================================================
ws11=wb.create_sheet("시나리오"); ws11.sheet_properties.tabColor="2ECC71"
sw(ws11,[22,18,18,18,22])
row=1; st(ws11,row,f"시나리오별 목표주가 (현재가 {PRICE:,}원)",5); row+=1

# Scenario table
st(ws11,row,"A. 방법론별 적정가 레인지",5); row+=1
wh(ws11,row,["방법론","보수적","기본","적극적","산출 근거"]); row+=1
for lbl,c_,b_,a_,note in [
    ("PER 방식",fw(int(EPS24*8)),fw(int(EPS24*12)),fw(int(EPS24*15)),"EPS(2024)x목표PER"),
    ("PBR 방식",fw(int(BPS*0.8)),fw(int(BPS*1.0)),fw(int(BPS*1.5)),"BPSx목표PBR"),
    ("EV/EBITDA",fw(int((ebitda_24*7-int(NET_DEBT/억))*억/SHARES)),
     fw(int((ebitda_24*10-int(NET_DEBT/억))*억/SHARES)),
     fw(int((ebitda_24*12-int(NET_DEBT/억))*억/SHARES)),"EBITDAx배수"),
    ("RIM",fw(int(BPS*(1+(0.08-0.09)/(0.09-0.01)))),fw(int(rim_base)),fw(int(rim_agg)),"ROE기반"),
    ("배당기반",fw(int(DPS24/0.08)),fw(int(DPS24/0.05)),fw(int(DPS24/0.04)),"DPS/요구수익률")]:
    wr(ws11,row,[lbl,c_,b_,a_,note],fonts=[db,df,d_blue,d_grn,sm],fills=[llf,rdf,gld,gnf,wf],als=[la,ra,ra,ra,la]); row+=1

# Bull/Base/Bear
row+=1; st(ws11,row,"B. 종합 시나리오",5); row+=1
wh(ws11,row,["시나리오","목표주가","현재가 대비","전제조건","확률(주관)"],
   fills=[PatternFill("solid",fgColor=RED_C)]*5); row+=1

bull_t=EPS24*15; bull_up=(bull_t-PRICE)/PRICE  # PER 15배 (규제완화)
base_t=EPS24*12; base_up=(base_t-PRICE)/PRICE  # PER 12배
bear_t=int(CUM25_NI/SHARES)*8; bear_up=(bear_t-PRICE)/PRICE  # PER 8배 (2025E, 규제강화)
exp=int(bull_t*0.2+base_t*0.5+bear_t*0.3); exp_up=(exp-PRICE)/PRICE

wr(ws11,row,["강세(Bull)",fw(bull_t),f"{bull_up*100:+.1f}%","규제완화(24시간영업/입장횟수)+제2카지노 완공+배당확대","20%"],
   fonts=[db,Font(name="맑은 고딕",size=12,bold=True,color=GREEN_C),d_grn,df,df],fills=[gnf]*5,als=[ca,ra,ca,la,ca]); row+=1
wr(ws11,row,["기본(Base)",fw(base_t),f"{base_up*100:+.1f}%","현행 규제 유지+입장객 완만 회복+배당 유지","50%"],
   fonts=[db,d_navy,d_blue,df,df],fills=[gld]*5,als=[ca,ra,ca,la,ca]); row+=1
wr(ws11,row,["약세(Bear)",fw(bear_t),f"{bear_up*100:+.1f}%","규제강화(입장횟수/베팅한도)+경기침체+배당삭감","30%"],
   fonts=[db,Font(name="맑은 고딕",size=12,bold=True,color=RED_C),d_red,df,df],fills=[rdf]*5,als=[ca,ra,ca,la,ca]); row+=1
row+=1
wr(ws11,row,["확률가중 기대값",fw(exp),f"{exp_up*100:+.1f}%","Bull x 20% + Base x 50% + Bear x 30%",""],
   fonts=[Font(name="맑은 고딕",size=12,bold=True,color=NAVY)]*5,fills=[blf]*5,als=[ca,ra,ca,la,ca]); row+=1

# RIM 시나리오
row+=2; st(ws11,row,"C. RIM (잔여이익모델) 시나리오",5); row+=1
wh(ws11,row,["시나리오","지속ROE","적정주가","현재가 대비","전제"]); row+=1
for lbl,roe,premise in [("약세",0.08,"COVID재발/규제강화"),("기본",0.121,"현행수준 유지"),
                         ("적극",0.15,"규제완화+효율화"),("낙관",0.18,"최적 운영+규제완화")]:
    ke_=0.09; g_=0.015; fv=BPS*(1+(roe-ke_)/(ke_-g_)); up=(fv-PRICE)/PRICE
    uf=d_grn if up>0 else d_red; ufl=gnf if up>0 else rdf
    wr(ws11,row,[lbl,f"{roe*100:.1f}%",fw(int(fv)),f"{up*100:+.1f}%",premise],
       fonts=[db,df,d_blue,uf,df],fills=[llf,wf,gld,ufl,wf],als=[la,ca,ra,ca,la]); row+=1

# SWOT
row+=2; st(ws11,row,"D. SWOT 분석",5); row+=1
ws11.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)
ws11.cell(row=row,column=1,value="강점 (S)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws11.cell(row=row,column=1).fill=PatternFill("solid",fgColor=GREEN_C); ws11.cell(row=row,column=1).alignment=ca
ws11.merge_cells(start_row=row,start_column=3,end_row=row,end_column=4)
ws11.cell(row=row,column=3,value="약점 (W)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws11.cell(row=row,column=3).fill=PatternFill("solid",fgColor=RED_C); ws11.cell(row=row,column=3).alignment=ca; row+=1
SW_S=["국내 유일 내국인 카지노 면허 (절대적 진입장벽)","무차입경영 (차입금 0원, 재무안정성 최고)",
      "고배당주 (배당수익률 6.6%, 꾸준한 증가 추세)","안정적 현금흐름 (연간 FCF ~3,800억)",
      "하이원리조트 종합 관광단지 (카지노+호텔+스키+골프)"]
SW_W=["규제 의존적 사업구조 (정부 정책 변동 리스크)","COVID 이전 매출 아직 미회복 (2017년 1.6조 vs 2024년 1.4조)",
      "지역 한정 (강원도 정선, 접근성 제한)","공기업적 경영 (효율성 제한, 정치적 영향)",
      "성장성 한계 (규제로 인한 매출 상한 존재)"]
for i in range(5):
    ws11.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)
    ws11.cell(row=row,column=1,value=f"  {SW_S[i]}").font=df; ws11.cell(row=row,column=1).fill=gnf
    ws11.cell(row=row,column=1).alignment=la; ws11.cell(row=row,column=1).border=tb
    ws11.merge_cells(start_row=row,start_column=3,end_row=row,end_column=4)
    ws11.cell(row=row,column=3,value=f"  {SW_W[i]}").font=df; ws11.cell(row=row,column=3).fill=rdf
    ws11.cell(row=row,column=3).alignment=la; ws11.cell(row=row,column=3).border=tb; row+=1
row+=1
ws11.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)
ws11.cell(row=row,column=1,value="기회 (O)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws11.cell(row=row,column=1).fill=PatternFill("solid",fgColor="2980B9"); ws11.cell(row=row,column=1).alignment=ca
ws11.merge_cells(start_row=row,start_column=3,end_row=row,end_column=4)
ws11.cell(row=row,column=3,value="위협 (T)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws11.cell(row=row,column=3).fill=PatternFill("solid",fgColor="7F8C8D"); ws11.cell(row=row,column=3).alignment=ca; row+=1
O=["영업시간 24시간 전환 기대 (매출 10%+ 증가)","입장횟수 규제 완화 (월15회→확대)","제2카지노영업장 완공 (수용능력 확대)","외국인 관광객 유치 확대 (IR 복합리조트화)","배당 확대 기대 (배당성향 50%+)"]
T=["규제 강화 리스크 (입장료 인상, 베팅한도 축소)","경기침체 시 유흥비 지출 감소","도박중독 사회적 논란 → 정치적 압박","인구감소/고령화 → 장기 수요 축소","온라인 도박 불법 확산 → 대체재 역할"]
for i in range(5):
    ws11.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)
    ws11.cell(row=row,column=1,value=f"  {O[i]}").font=df; ws11.cell(row=row,column=1).fill=blf
    ws11.cell(row=row,column=1).alignment=la; ws11.cell(row=row,column=1).border=tb
    ws11.merge_cells(start_row=row,start_column=3,end_row=row,end_column=4)
    ws11.cell(row=row,column=3,value=f"  {T[i]}").font=df; ws11.cell(row=row,column=3).fill=PatternFill("solid",fgColor="E5E7E9")
    ws11.cell(row=row,column=3).alignment=la; ws11.cell(row=row,column=3).border=tb; row+=1
print("  [11/12] 시나리오/SWOT")

# ============================================================
# SHEET 12: 모니터링
# ============================================================
ws12=wb.create_sheet("모니터링"); ws12.sheet_properties.tabColor="1ABC9C"
sw(ws12,[6,30,40,20])
row=1; ws12.merge_cells('A1:D1')
ws12.cell(row=1,column=1,value="향후 핵심 모니터링 지표").font=sec_font; ws12.cell(row=1,column=1).border=bb; row=3
wh(ws12,row,["#","항목","세부 내용","확인 시기"]); row+=1
for rank,title,detail,timing in [
    ("1","입장객수 추이","카지노 매출의 핵심 선행지표. 월별/분기별 입장객수 공시. COVID 이전 대비 회복률 확인","매월/매분기"),
    ("2","GGR(총게임수익)","1인당 게임순수익금. 테이블/슬롯 구성비, win rate 변화 추적","매 분기"),
    ("3","영업시간/입장 규제 변화","24시간 영업 전환, 입장횟수 완화 여부가 주가 최대 촉매","수시 (정책발표)"),
    ("4","배당정책 변화","배당성향/DPS 변화. 현재 배당수익률 6.6%로 고배당주 매력 유지 여부","3월 주총"),
    ("5","제2카지노영업장 건설","투자 1,796억원. 완공 시 수용능력 확대 → 매출 성장 드라이버","매 분기 진행상황")]:
    wr(ws12,row,[rank,title,detail,timing],
       fonts=[Font(name="맑은 고딕",size=14,bold=True,color=NAVY),db,df,df],
       fills=[gld,llf,wf,llf],als=[ca,la,la,ca])
    ws12.row_dimensions[row].height=45; row+=1

row+=2; st(ws12,row,"모니터링 캘린더",4); row+=1
wh(ws12,row,["시기","이벤트","중요도"]); row+=1
for t,e,imp in [("1~2월","연간/4Q 잠정실적 발표","+++"),
                ("3월","사업보고서/주총/배당정책 확정","+++"),
                ("매분기","분기별 입장객수 공시","++"),
                ("5월","1Q 잠정실적","++"),
                ("8월","반기보고서/2Q 잠정실적","++"),
                ("11월","3Q 잠정실적","++"),
                ("수시","영업시간/입장규제 정책 변경","+++"),
                ("수시","제2카지노 건설 진행 상황","++"),
                ("수시","사행산업감독위 결정사항","++")]:
    wr(ws12,row,[t,e,imp],fonts=[db,df,db],fills=[llf,wf,gld],als=[ca,la,ca]); row+=1

row+=2; st(ws12,row,"SWOT 요약 (한눈에 보기)",4); row+=1
for cat,items,fl_c in [
    ("강점","독점면허 / 무차입 / 고배당 / 안정적CF",gnf),
    ("약점","규제의존 / COVID미회복 / 지역한정 / 성장한계",rdf),
    ("기회","규제완화 / 제2카지노 / 외국인유치 / 배당확대",blf),
    ("위협","규제강화 / 경기침체 / 도박중독논란 / 인구감소",PatternFill("solid",fgColor="E5E7E9"))]:
    ws12.merge_cells(start_row=row,start_column=2,end_row=row,end_column=4)
    ws12.cell(row=row,column=1,value=cat).font=db; ws12.cell(row=row,column=1).fill=fl_c
    ws12.cell(row=row,column=1).alignment=ca; ws12.cell(row=row,column=1).border=tb
    ws12.cell(row=row,column=2,value=items).font=df; ws12.cell(row=row,column=2).fill=fl_c
    ws12.cell(row=row,column=2).alignment=la; ws12.cell(row=row,column=2).border=tb; row+=1

row+=2
# Key events from DB (if available)
try:
    events = conn.execute("SELECT rcept_dt,event_type,SUBSTR(event_summary,1,80) FROM key_events WHERE rcept_dt>='20230101' ORDER BY rcept_dt DESC LIMIT 15").fetchall()
    if events:
        st(ws12,row,"주요 이벤트 히스토리 (DB)",4); row+=1
        wh(ws12,row,["일자","유형","내용"]); row+=1
        for dt,et,s in events:
            wr(ws12,row,[dt,et,s.replace('\n',' ').strip()[:70] if s else ""],fonts=[df,db,df],als=[ca,ca,la]); row+=1
except:
    pass  # DB에 데이터 없어도 무방

print("  [12/12] 모니터링")

# === SAVE ===
OUT1=os.path.join(BASE,"강원랜드_종합보고서.xlsx")
wb.save(OUT1); conn.close()
print(f"\n종합보고서 생성 완료: {OUT1}")
print(f"시트 ({len(wb.sheetnames)}개): {wb.sheetnames}")
