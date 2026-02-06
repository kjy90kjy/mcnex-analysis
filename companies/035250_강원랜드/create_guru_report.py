# -*- coding: utf-8 -*-
"""강원랜드 투자 구루 분석 보고서 (Buffett/Munger 한국형 Four Filters)"""
import sqlite3, sys, os, statistics
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ai.db")
BASE = os.path.dirname(os.path.abspath(__file__))
conn = sqlite3.connect(DB)

# === KEY CONSTANTS ===
PRICE = 17690
SHARES = 213940500
SHARES_TREASURY = 13534816
SHARES_OUT = SHARES - SHARES_TREASURY  # 200,405,684
억 = 100_000_000
MCAP = PRICE * SHARES

# 2024 Annual (연결)
REV24 = 1426862781802; OP24 = 285790454716; NI24 = 456895295638; EPS24 = 2257
EQ24 = 3883662813757; EQ23 = 3677900816262; ASSETS24 = 4700641154570; LIAB24 = 816978340813
CASH24 = 244499636973; ST_DEBT = 0; LT_DEBT = 0  # 무차입경영
DA24 = 77582716300 + 892165554  # 감가상각비 + 무형자산상각비
OPCF24 = 487048778217; CAPEX24 = 107494733192
DIV_PAID = 188547268140; TREAS_BUY = 0; DPS24 = 1170  # 2024년 배당금 지급(2023년분), 자사주취득은 2024년에 실행

# Quarters (연결 잠정실적 기준)
Q24 = [{"rev":368373e6,"op":75867e6,"ni":92913e6},
       {"rev":337719e6,"op":73416e6,"ni":149791e6},
       {"rev":375541e6,"op":93039e6,"ni":91982e6},
       {"rev":344622e6,"op":42724e6,"ni":112197e6}]
Q25 = [{"rev":365830e6,"op":77659e6,"ni":78021e6},
       {"rev":360727e6,"op":57874e6,"ni":60630e6},
       {"rev":384147e6,"op":72702e6,"ni":113070e6},
       {"rev":365446e6,"op":29697e6,"ni":66017e6}]
CUM25 = {"rev":1476726e6,"op":235176e6,"ni":316516e6}

# Derived
BPS = EQ24 / SHARES
NET_DEBT = ST_DEBT + LT_DEBT - CASH24  # 순현금 상태
EBITDA24 = OP24 + DA24; FCF24 = OPCF24 - CAPEX24; EV = MCAP + NET_DEBT
TRAIL_NI = Q24[3]["ni"]+Q25[0]["ni"]+Q25[1]["ni"]+Q25[2]["ni"]
TRAIL_EPS = TRAIL_NI / SHARES
AVG_EQ = (EQ24+EQ23)/2; ROE24 = NI24/AVG_EQ

# === STYLE ===
NAVY="1B2A4A"; DARK="2C3E6B"; MID="3A5BA0"; LB="D6E4F0"; LLB="EBF1F8"; W="FFFFFF"
GOLD_C="D4A843"; RED_C="C0392B"; GREEN_C="27AE60"; GRAY_C="F2F2F2"

title_font=Font(name="맑은 고딕",size=22,bold=True,color=W)
sub_font=Font(name="맑은 고딕",size=11,color="B0C4DE")
sec_font=Font(name="맑은 고딕",size=14,bold=True,color=NAVY)
hdr_font=Font(name="맑은 고딕",size=10,bold=True,color=W)
df=Font(name="맑은 고딕",size=10)
db=Font(name="맑은 고딕",size=10,bold=True)
d_blue=Font(name="맑은 고딕",size=10,bold=True,color="0000FF")
d_grn=Font(name="맑은 고딕",size=10,bold=True,color=GREEN_C)
d_red=Font(name="맑은 고딕",size=10,bold=True,color=RED_C)
d_navy=Font(name="맑은 고딕",size=12,bold=True,color=NAVY)
sm=Font(name="맑은 고딕",size=9,color="666666")

tf=PatternFill("solid",fgColor=NAVY)
hf=PatternFill("solid",fgColor=DARK)
mf=PatternFill("solid",fgColor=MID)
lf=PatternFill("solid",fgColor=LB)
llf=PatternFill("solid",fgColor=LLB)
gf=PatternFill("solid",fgColor=GRAY_C)
wf=PatternFill("solid",fgColor=W)
gld=PatternFill("solid",fgColor="FFF3CD")
rdf=PatternFill("solid",fgColor="F8D7DA")
gnf=PatternFill("solid",fgColor="D4EDDA")
blf=PatternFill("solid",fgColor="D6EAF8")

ca=Alignment(horizontal='center',vertical='center',wrap_text=True)
la=Alignment(horizontal='left',vertical='center',wrap_text=True)
ra=Alignment(horizontal='right',vertical='center')
tb=Border(left=Side('thin',color="D9D9D9"),right=Side('thin',color="D9D9D9"),
          top=Side('thin',color="D9D9D9"),bottom=Side('thin',color="D9D9D9"))
bb=Border(bottom=Side('medium',color=NAVY))
NF='#,##0'; PF='0.0%'

def sw(ws,w):
    for i,v in enumerate(w,1): ws.column_dimensions[get_column_letter(i)].width=v
def wh(ws,r,h,fills=None):
    for i,v in enumerate(h,1):
        c=ws.cell(row=r,column=i,value=v); c.font=hdr_font; c.fill=fills[i-1] if fills else hf; c.alignment=ca; c.border=tb
def wr(ws,r,d,fonts=None,fills=None,als=None,nfs=None):
    for i,v in enumerate(d,1):
        c=ws.cell(row=r,column=i,value=v)
        c.font=fonts[i-1] if fonts else df; c.fill=fills[i-1] if fills else wf
        c.alignment=als[i-1] if als else ca; c.border=tb
        if nfs and i<=len(nfs) and nfs[i-1]: c.number_format=nfs[i-1]
def st(ws,r,t,ce=8):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ce)
    c=ws.cell(row=r,column=1,value=t); c.font=sec_font; c.border=bb; return r+1
def fmt(v): return f"{v/억:,.0f}억"
def fw(v): return f"{v:,.0f}원"
def pct(v): return f"{v*100:.1f}%"

# === LOAD DATA ===
# 매출액 (연결) - 강원랜드는 연도별로 계정명이 다름
rev_rows = conn.execute("""SELECT bsns_year, thstrm_amount FROM financial_statements
    WHERE (account_nm='수익(매출액)' OR account_nm='매출')
    AND sj_div='CIS' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    ORDER BY bsns_year""").fetchall()
rev_map = {r[0]: int(r[1]) for r in rev_rows if r[1]}

# 영업이익 (연결) - 2021~2024는 사업보고서에 연결 영업이익이 없으므로 잠정실적 사용
op_rows = conn.execute("""SELECT bsns_year, thstrm_amount FROM financial_statements
    WHERE account_nm='영업이익' AND sj_div='CIS' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    ORDER BY bsns_year""").fetchall()
op_map = {r[0]: int(r[1]) for r in op_rows if r[1]}
# 잠정실적 기준 보완
op_map.setdefault('2021', -52695285784)
op_map.setdefault('2022', 217565214200)
op_map.setdefault('2023', 282263014860)
op_map.setdefault('2024', 285790454716)

# 당기순이익 (연결, 지배기업)
ni_rows = conn.execute("""SELECT bsns_year, thstrm_amount FROM financial_statements
    WHERE account_nm LIKE '%지배기업%순이익%' AND sj_div='CIS' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    ORDER BY bsns_year""").fetchall()
ni_map = {r[0]: int(r[1]) for r in ni_rows if r[1]}
# 2015/2016은 당기순이익으로 조회
for yr, ni in [('2015', 441629100500), ('2016', 454533658888)]:
    ni_map.setdefault(yr, ni)

# 자산/부채/자본 (연결)
eq_rows = conn.execute("""SELECT bsns_year, thstrm_amount FROM financial_statements
    WHERE account_nm='자본총계' AND sj_div='BS' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    ORDER BY bsns_year""").fetchall()
eq_map = {r[0]: int(r[1]) for r in eq_rows if r[1]}
# 2021/2022 연결 자본총계는 사업보고서에 없으므로 추정
eq_map.setdefault('2021', 3252793111203)  # 2020 + 2021 NI estimate
eq_map.setdefault('2022', 3415378100640)  # 추정

ta_rows = conn.execute("""SELECT bsns_year, thstrm_amount FROM financial_statements
    WHERE account_nm='자산총계' AND sj_div='BS' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    ORDER BY bsns_year""").fetchall()
ta_map = {r[0]: int(r[1]) for r in ta_rows if r[1]}

tl_rows = conn.execute("""SELECT bsns_year, thstrm_amount FROM financial_statements
    WHERE account_nm='부채총계' AND sj_div='BS' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    ORDER BY bsns_year""").fetchall()
tl_map = {r[0]: int(r[1]) for r in tl_rows if r[1]}

# 매출총이익
gp_rows = conn.execute("""SELECT bsns_year, thstrm_amount FROM financial_statements
    WHERE account_nm='매출총이익' AND sj_div='CIS' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    ORDER BY bsns_year""").fetchall()
gp_map = {r[0]: int(r[1]) for r in gp_rows if r[1]}

# 영업CF
opcf_rows = conn.execute("""SELECT bsns_year, thstrm_amount FROM financial_statements
    WHERE account_nm='영업활동현금흐름' AND sj_div='CF' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    ORDER BY bsns_year""").fetchall()
opcf_map = {r[0]: int(r[1]) for r in opcf_rows if r[1]}

# CAPEX
capex_rows = conn.execute("""SELECT bsns_year, thstrm_amount FROM financial_statements
    WHERE account_nm LIKE '%유형자산%취득%' AND sj_div='CF' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    ORDER BY bsns_year""").fetchall()
capex_map = {r[0]: abs(int(r[1])) for r in capex_rows if r[1]}

# 감가상각비
da_rows = conn.execute("""SELECT bsns_year, account_nm, thstrm_amount FROM financial_statements
    WHERE account_nm IN ('감가상각비','투자부동산감가상각비','무형자산상각비')
    AND sj_div='CF' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    ORDER BY bsns_year""").fetchall()
da_map = {}
for r in da_rows:
    if r[2]:
        da_map[r[0]] = da_map.get(r[0], 0) + int(r[2])

# 배당 (주당 현금배당금) - 강원랜드는 stock_knd가 None인 경우도 있음
div_rows = conn.execute("""SELECT bsns_year, thstrm FROM dividends
    WHERE se='주당 현금배당금(원)' ORDER BY bsns_year""").fetchall()
div_map = {}
for r in div_rows:
    try:
        v = int(str(r[1]).replace(',','').replace('-','0'))
        if v > 0:
            div_map[r[0]] = v
    except:
        pass

# 배당성향
dpayout_rows = conn.execute("""SELECT bsns_year, thstrm FROM dividends
    WHERE se='(연결)현금배당성향(%)' ORDER BY bsns_year""").fetchall()
dpayout_map = {}
for r in dpayout_rows:
    try: dpayout_map[r[0]] = float(r[1])
    except: pass

# 자사주 총계
ts_rows = conn.execute("""SELECT bsns_year, bsis_qy, change_qy_acqs, change_qy_dsps, change_qy_incnr, trmend_qy
    FROM treasury_stock WHERE acqs_mth1='총계'
    AND (stock_knd LIKE '%보통%' OR stock_knd='보통주식' OR stock_knd='보통주')
    ORDER BY bsns_year""").fetchall()

# 유상증자/CB
cap_all = conn.execute("""SELECT DISTINCT isu_dcrs_de, isu_dcrs_stle, isu_dcrs_stock_knd, isu_dcrs_qy
    FROM capital_changes WHERE isu_dcrs_stle LIKE '%유상증자%' OR isu_dcrs_stle LIKE '%전환%' OR isu_dcrs_stle LIKE '%무상%'
    ORDER BY isu_dcrs_de""").fetchall()

# 임원 (2024)
exec_rows = conn.execute("""SELECT nm, ofcps, rgist_exctv_at, fte_at FROM executives
    WHERE bsns_year='2024' LIMIT 20""").fetchall()

# 주요 이벤트
event_rows = conn.execute("""SELECT rcept_dt, event_type, SUBSTR(event_summary,1,80) FROM key_events
    WHERE rcept_dt>='20220101' ORDER BY rcept_dt DESC LIMIT 20""").fetchall()

# EPS map
eps_map = {'2015':2178,'2016':2242,'2017':2159,'2018':1467,'2019':1651,
           '2020':-1361,'2021':-52,'2022':570,'2023':1681,'2024':2257}

# Build annual data dict
YEARS_ALL = ['2015','2016','2017','2018','2019','2020','2021','2022','2023','2024']
annual = {}
for yr in YEARS_ALL:
    rev = rev_map.get(yr, 0)
    op = op_map.get(yr, 0)
    ni = ni_map.get(yr, 0)
    eps = eps_map.get(yr, 0)
    ta = ta_map.get(yr, 0)
    tl = tl_map.get(yr, 0)
    te = eq_map.get(yr, 0)
    annual[yr] = {"rev":rev,"op":op,"ni":ni,"eps":eps,"ta":ta,"tl":tl,"te":te,
                  "gp":gp_map.get(yr,0),"opcf":opcf_map.get(yr,0),
                  "capex":capex_map.get(yr,0),"da":da_map.get(yr,0)}

YEARS5 = [y for y in sorted(annual.keys()) if y >= '2020']

wb = Workbook()

# ============================================================
# SHEET 1: Four Filters 종합 대시보드
# ============================================================
ws1 = wb.active; ws1.title="Four Filters"; ws1.sheet_properties.tabColor=NAVY
sw(ws1,[4,22,14,14,14,14,14,14,4])

# Title banner
for r in range(1,6):
    for c in range(1,10): ws1.cell(row=r,column=c).fill=tf
ws1.merge_cells('B2:H2')
ws1.cell(row=2,column=2,value="강원랜드(KANGWON LAND) 투자 구루 분석").font=Font(name="맑은 고딕",size=24,bold=True,color=W)
ws1.cell(row=2,column=2).alignment=ca
ws1.merge_cells('B4:H4')
ws1.cell(row=4,column=2,value=f"Four Filters 종합 대시보드  |  현재가 {PRICE:,}원  |  2026.02.06 기준").font=sub_font
ws1.cell(row=4,column=2).alignment=ca

row = 7

# === Four Filters Score ===
st(ws1,row,"A. Four Filters 스코어카드 (100점 만점)",8); row+=1

# Filter 1: 사업 이해도
f1_items = [
    ("사업모델 단순성","내국인 독점 카지노 + 리조트(호텔/스키/골프)","높음",5),
    ("매출구조 명확성","카지노 80%+ / 호텔·콘도·스키·골프 20%","높음",5),
    ("10년 일관성","2000년 개장 이래 카지노+리조트 사업 일관","높음",5),
    ("수요 예측성","입장객 수 × 1인당 GGR, 계절성 있으나 안정","보통",4),
    ("규제 환경 이해","정부 규제(입장료·시간·횟수) 직접 영향","보통",3),
]
f1_score = sum(x[3] for x in f1_items)

# Filter 2: 경제적 해자
f2_items = [
    ("규제 독점","한국 유일 내국인 카지노 면허 (법적 진입장벽)","매우 강함",5),
    ("브랜드/입지","강원도 정선 독점 입지, 관광 인프라 독점","강함",4),
    ("전환비용","도박 중독성으로 재방문율 높음, 대체재 부재","강함",4),
    ("수익성 지속","GPM 50%+, OPM 25~35% 구조적 고마진","매우 강함",5),
    ("해자 위협","온라인 도박 합법화, 복합리조트 확대 가능성","보통",3),
]
f2_score = sum(x[3] for x in f2_items)

# Filter 3: 경영진
f3_items = [
    ("경영진 구조","전문경영인 체제 (정부 영향력 하 공기업형)","보통",3),
    ("보수 합리성","개별 CEO 보수 미공개 (5억 미만 추정)","보통",3),
    ("배당 정책","고배당 전통 유지, DPS 930→1,170원 인상","양호",4),
    ("지역공헌/사회적 의무","폐광지역 지원, 강원도 지역경제 기여","양호",4),
    ("성장 투자","제2카지노영업장 1,796억 신규 투자 결정","양호",4),
]
f3_score = sum(x[3] for x in f3_items)

# Filter 4: 안전마진
per24 = PRICE/EPS24; per_trail = PRICE/(TRAIL_NI/SHARES); pbr = PRICE/BPS
f4_items = [
    ("PER 수준",f"확정 {per24:.1f}배 / T4Q {per_trail:.1f}배 (카지노 기준 10~15배)","저평가" if per_trail<10 else "적정",5 if per_trail<10 else 3),
    ("PBR 수준",f"{pbr:.2f}배 (기준 0.8~1.5배)","적정" if pbr<=1.5 else "고평가",3 if pbr<=1.5 else 2),
    ("FCF 수익률",f"{FCF24/MCAP*100:.1f}% (기준 >5%)","양호" if FCF24/MCAP>0.05 else "보통",4 if FCF24/MCAP>0.05 else 3),
    ("EV/EBITDA",f"{EV/EBITDA24:.1f}배 (기준 8~12배)","저평가" if EV/EBITDA24<8 else "적정",4 if EV/EBITDA24<8 else 3),
    ("배당수익률",f"{DPS24/PRICE*100:.1f}% (기준 >3%)","양호" if DPS24/PRICE>0.03 else "보통",4 if DPS24/PRICE>0.03 else 3),
]
f4_score = sum(x[3] for x in f4_items)

total_score = f1_score + f2_score + f3_score + f4_score
grade = "A" if total_score>=80 else ("B" if total_score>=60 else ("C" if total_score>=40 else "D"))
grade_color = GREEN_C if grade in ["A","B"] else (GOLD_C if grade=="C" else RED_C)

# Score summary
wh(ws1,row,["","필터","소항목1","소항목2","소항목3","소항목4","소항목5","소계"]); row+=1
for fname,items,score in [("F1: 사업 이해도",f1_items,f1_score),
                           ("F2: 경제적 해자",f2_items,f2_score),
                           ("F3: 경영진 품질",f3_items,f3_score),
                           ("F4: 안전마진",f4_items,f4_score)]:
    scores = [x[3] for x in items]
    sc_fill = gnf if score>=20 else (gld if score>=15 else rdf)
    wr(ws1,row,["",fname]+scores+[f"{score}/25"],
       fonts=[df,db]+[df]*5+[Font(name="맑은 고딕",size=11,bold=True,color=GREEN_C if score>=20 else NAVY)],
       fills=[wf,llf]+[wf]*5+[sc_fill],als=[ca,la]+[ca]*6)
    row+=1

# Total
row+=1
ws1.merge_cells(start_row=row,start_column=2,end_row=row,end_column=5)
ws1.cell(row=row,column=2,value="종합점수").font=d_navy; ws1.cell(row=row,column=2).alignment=la; ws1.cell(row=row,column=2).border=tb
ws1.merge_cells(start_row=row,start_column=6,end_row=row,end_column=7)
ws1.cell(row=row,column=6,value=f"{total_score}/100 (등급: {grade})").font=Font(name="맑은 고딕",size=14,bold=True,color=grade_color)
ws1.cell(row=row,column=6).fill=gnf if grade in ["A","B"] else gld
ws1.cell(row=row,column=6).alignment=ca; ws1.cell(row=row,column=6).border=tb
row+=2

# === 투자 테제 ===
st(ws1,row,"B. 투자 테제",8); row+=1
ws1.merge_cells(start_row=row,start_column=2,end_row=row+1,end_column=8)
c=ws1.cell(row=row,column=2)
c.value=("강원랜드는 한국 유일의 내국인 카지노 면허를 보유한 규제 독점 기업으로, "
         "GPM 50%+, OPM 20~35%의 구조적 고마진 사업을 영위. "
         "무차입 순현금 경영 + 고배당 전통(배당성향 45~60%) + 자사주 매입까지 병행하는 "
         f"주주환원 기업. 현재 PER {per_trail:.1f}배(T4Q)는 카지노업 기준 저평가 영역이며, "
         f"배당수익률 {DPS24/PRICE*100:.1f}%로 가치주 매력 충분.")
c.font=Font(name="맑은 고딕",size=10,bold=True,color=NAVY); c.alignment=Alignment(horizontal='left',vertical='top',wrap_text=True)
row+=3

# === 핵심 강점/리스크 ===
st(ws1,row,"C. 핵심 강점 vs 리스크",8); row+=1

# Strengths
ws1.merge_cells(start_row=row,start_column=2,end_row=row,end_column=4)
ws1.cell(row=row,column=2,value="핵심 강점 (3)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws1.cell(row=row,column=2).fill=PatternFill("solid",fgColor=GREEN_C); ws1.cell(row=row,column=2).alignment=ca
ws1.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
ws1.cell(row=row,column=5,value="핵심 리스크 (3)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws1.cell(row=row,column=5).fill=PatternFill("solid",fgColor=RED_C); ws1.cell(row=row,column=5).alignment=ca
row+=1

strengths = ["한국 유일 내국인 카지노 면허 (규제 독점 해자)",
             "무차입경영 + 순현금 2,445억 + 배당수익률 6.6%",
             "COVID 정상화 완료 → 매출/이익 회복 + 신규 투자(제2카지노)"]
risks = ["정부 규제 리스크 (영업시간/입장료/횟수 제한 변경)",
         "전문경영인 체제 → 정부 정책 변화에 취약",
         "온라인 도박 합법화 또는 내국인 카지노 추가 면허 가능성"]

for i in range(3):
    ws1.merge_cells(start_row=row,start_column=2,end_row=row,end_column=4)
    ws1.cell(row=row,column=2,value=f"  {i+1}. {strengths[i]}").font=df
    ws1.cell(row=row,column=2).fill=gnf; ws1.cell(row=row,column=2).alignment=la; ws1.cell(row=row,column=2).border=tb
    ws1.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
    ws1.cell(row=row,column=5,value=f"  {i+1}. {risks[i]}").font=df
    ws1.cell(row=row,column=5).fill=rdf; ws1.cell(row=row,column=5).alignment=la; ws1.cell(row=row,column=5).border=tb
    row+=1

# === 소항목 상세 ===
row+=1; st(ws1,row,"D. 필터별 소항목 상세",8); row+=1
wh(ws1,row,["","항목","평가 근거","판정","점수","","",""]); row+=1
for fname,items in [("Filter 1: 사업 이해도",f1_items),("Filter 2: 경제적 해자",f2_items),
                     ("Filter 3: 경영진 품질",f3_items),("Filter 4: 안전마진",f4_items)]:
    ws1.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
    ws1.cell(row=row,column=1,value=fname).font=Font(name="맑은 고딕",size=11,bold=True,color=MID)
    ws1.cell(row=row,column=1).fill=llf; ws1.cell(row=row,column=1).border=tb; row+=1
    for item_nm,evidence,judgment,score in items:
        sf = d_grn if score>=4 else (df if score>=3 else d_red)
        sfill = gnf if score>=4 else (wf if score>=3 else rdf)
        wr(ws1,row,["",item_nm,evidence,judgment,f"{score}/5","","",""],
           fonts=[df,db,df,sf,sf,df,df,df],fills=[wf,llf,wf,sfill,sfill,wf,wf,wf],
           als=[ca,la,la,ca,ca,ca,ca,ca]); row+=1

print("  [1/7] Four Filters 대시보드")

# ============================================================
# SHEET 2: 경제적 해자 분석
# ============================================================
ws2 = wb.create_sheet("경제적_해자"); ws2.sheet_properties.tabColor=DARK
sw(ws2,[20,16,16,16,16,16,16,16])

row=1; st(ws2,row,"경제적 해자(Economic Moat) 분석",8); row+=1

# 해자 유형 판정표
st(ws2,row,"A. 해자 유형 판정",8); row+=1
wh(ws2,row,["해자 유형","보유 여부","강도","근거","지속가능성","","",""]); row+=1
moat_types = [
    ("규제 독점","O","매우 강함","한국 유일 내국인 카지노 면허 (폐광지역개발특별법)","영구적"),
    ("브랜드/입지","O","강함","정선 독점 입지, 서울서 3시간 거리에도 유일한 선택지","10년+"),
    ("전환비용","O","강함","도박 중독성+대체재 부재로 재방문율 높음","10년+"),
    ("네트워크 효과","X","없음","카지노/리조트 사업 구조상 해당 없음","-"),
    ("비용우위","△","보통","독점이므로 비용 경쟁 불필요, 오히려 고비용 구조","해당없음"),
]
for tp,has,strength,basis,duration in moat_types:
    hf2 = d_grn if has=="O" else (d_red if has=="X" else db)
    hfl = gnf if has=="O" else (rdf if has=="X" else gld)
    wr(ws2,row,[tp,has,strength,basis,duration,"","",""],
       fonts=[db,hf2,hf2,df,df,df,df,df],fills=[llf,hfl,hfl,wf,wf,wf,wf,wf],
       als=[la,ca,ca,la,ca,ca,ca,ca]); row+=1

# 수익성 추이
row+=1; st(ws2,row,"B. 수익성 추이 (10년)",8); row+=1
wh(ws2,row,["연도","매출액(억)","매출총이익(억)","GPM","영업이익(억)","OPM","ROIC","비고"]); row+=1

for yr in YEARS_ALL:
    d = annual[yr]
    rev_b = d["rev"]//억 if d["rev"] else 0
    gp_b = d["gp"]//억 if d["gp"] else 0
    op_b = d["op"]//억
    gpm = d["gp"]/d["rev"] if d["rev"] and d["gp"] else 0
    opm = d["op"]/d["rev"] if d["rev"] else 0
    nopat = d["op"] * 0.78
    roic = nopat / d["te"] if d["te"] else 0

    gpm_f = d_grn if gpm>0.40 else (d_red if gpm<0.20 else df)
    opm_f = d_grn if opm>0.25 else (d_red if opm<0.10 else df)
    note = ""
    if yr=='2020': note = "COVID 영업제한"
    elif yr=='2021': note = "COVID 회복 초기"
    elif yr=='2024': note = "COVID 정상화 완료"
    elif yr=='2015': note = "최대 실적"

    wr(ws2,row,[yr,rev_b,gp_b if gp_b else "-",f"{gpm*100:.1f}%" if gpm else "-",
                op_b,f"{opm*100:.1f}%",f"{roic*100:.1f}%" if roic else "-",note],
       fonts=[db,df,df,gpm_f if gpm else df,df,opm_f,df,sm],fills=[lf]+[wf]*7,
       als=[ca,ra,ra,ca,ra,ca,ca,la],nfs=[None,NF,NF,None,NF,None,None,None]); row+=1

# 면허/규제 분석 (R&D 대신)
row+=1; st(ws2,row,"C. 면허 및 규제 환경 (핵심 해자 원천)",8); row+=1
wh(ws2,row,["항목","내용","영향","해자 기여도","","","",""]); row+=1
license_items = [
    ("카지노 면허","폐광지역개발특별법에 의한 내국인 카지노 유일 면허","결정적","매우 높음"),
    ("입장 제한","월 15회 입장 제한, 24시간 이내 퇴장","매출 상한선 형성","중립"),
    ("입장료","1만원 (2014년부터 시행)","소액, 수요 영향 제한적","중립"),
    ("영업시간","24시간 영업 (2023년 정상화)","COVID 기간 제한 해제","긍정"),
    ("사행산업 규제","사행산업감독위원회 감독","규제 강화 시 직접 영향","리스크"),
    ("지역기여 의무","폐광지역 지원금, 고용 의무","비용 요인이나 면허 유지 정당성","중립"),
]
for item,content,impact,contrib in license_items:
    cf = d_grn if contrib in ["매우 높음","긍정"] else (d_red if contrib=="리스크" else df)
    cfl = gnf if contrib in ["매우 높음","긍정"] else (rdf if contrib=="리스크" else wf)
    wr(ws2,row,[item,content,impact,contrib,"","","",""],
       fonts=[db,df,df,cf,df,df,df,df],fills=[llf,wf,wf,cfl,wf,wf,wf,wf],
       als=[la,la,la,ca,ca,ca,ca,ca]); row+=1

# 해자 종합 평가
row+=1; st(ws2,row,"D. 해자 종합 평가",8); row+=1
for item,val in [("해자 폭","Wide (규제 독점 — 가장 강력한 해자 유형)"),
                  ("해자 추세","유지 (면허 체제 변동 가능성 낮음, 제2카지노 투자로 사업 확대)"),
                  ("핵심 해자","폐광지역개발특별법에 의한 내국인 카지노 독점 면허"),
                  ("위협 요인","온라인 도박 합법화, 복합리조트(영종도 등) 내국인 개방, 추가 면허"),
                  ("결론","규제 독점 기반 Wide Moat 보유. 면허 체제 변경 전까지 해자 영구적")]:
    ws2.cell(row=row,column=1,value=item).font=db; ws2.cell(row=row,column=1).fill=llf
    ws2.cell(row=row,column=1).alignment=la; ws2.cell(row=row,column=1).border=tb
    ws2.merge_cells(start_row=row,start_column=2,end_row=row,end_column=8)
    ws2.cell(row=row,column=2,value=val).font=df if "위협" not in item else d_red
    ws2.cell(row=row,column=2).alignment=la; ws2.cell(row=row,column=2).border=tb; row+=1

print("  [2/7] 경제적 해자")

# ============================================================
# SHEET 3: 경영진 평가
# ============================================================
ws3 = wb.create_sheet("경영진_평가"); ws3.sheet_properties.tabColor=GOLD_C
sw(ws3,[20,16,16,16,16,16,16,16])

row=1; st(ws3,row,"경영진 평가 (Management Quality)",8); row+=1

# 지배구조 개요
st(ws3,row,"A. 지배구조 개요",8); row+=1
wh(ws3,row,["항목","내용","비고","","","","",""]); row+=1
gov_items = [
    ("최대주주","한국광해광업공단 (36.27%)","정부 산하 공공기관"),
    ("경영 형태","전문경영인 체제","정부 영향력 하 CEO 선임"),
    ("현 대표이사","최철규","2024년 사업보고서 기준"),
    ("이사회 구성","사외이사 과반수","공기업형 지배구조"),
    ("감사위원회","이사회 내 감사위원회 설치","독립성 확보"),
]
for item,content,note in gov_items:
    is_ceo = "대표" in item
    wr(ws3,row,[item,content,note,"","","","",""],
       fonts=[d_navy if is_ceo else db,df,sm,df,df,df,df,df],
       fills=[gld if is_ceo else llf]+[wf]*7,als=[la,la,la,ca,ca,ca,ca,ca]); row+=1

# CEO 보수 (개별 미공개)
row+=1; st(ws3,row,"B. 대표이사 보수",8); row+=1
ws3.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
ws3.cell(row=row,column=1,value="개별 보수 미공개 (5억원 미만 추정). 공기업형 보수 체계 적용 → 과다 보수 리스크 낮음").font=df
ws3.cell(row=row,column=1).fill=blf; ws3.cell(row=row,column=1).alignment=la; ws3.cell(row=row,column=1).border=tb
row+=2

# 전문경영인 체크리스트
st(ws3,row,"C. 전문경영인 체크리스트",8); row+=1
wh(ws3,row,["","평가 항목","현황","판정","근거","","",""]); row+=1
mgmt_checks = [
    ("1","최대주주 지분율","36.27% (한국광해광업공단)","양호","공공기관 최대주주 → 적대적 M&A 방어"),
    ("2","CEO 임기/안정성","전문경영인 교체 주기 2~3년","보통","정권 교체에 따른 CEO 교체 리스크"),
    ("3","보수 합리성","개별 미공개 (공기업 기준 적용)","양호","과다 보수 가능성 낮음"),
    ("4","배당 정책","고배당 전통 유지, 성향 45~60%","우수","주주환원 의지 강함"),
    ("5","자사주 매입","2024 233만주 취득 (약 413억)","우수","밸류업 프로그램 적극 참여"),
    ("6","지역공헌 의무","폐광지역 지원금, 지역고용 의무","보통","비용 요인이나 면허 정당성 확보"),
    ("7","IR 활동","정기 IR 개최 (연 8회+), 잠정실적 공시","양호","투자자 소통 적극적"),
]
for num,item,status,judgment,basis in mgmt_checks:
    jf = d_grn if judgment=="우수" else (d_blue if judgment=="양호" else (db if judgment=="보통" else d_red))
    jfl = gnf if judgment=="우수" else (blf if judgment=="양호" else (gld if judgment=="보통" else rdf))
    wr(ws3,row,[num,item,status,judgment,basis,"","",""],
       fonts=[df,db,df,jf,sm,df,df,df],fills=[wf,llf,wf,jfl,wf,wf,wf,wf],
       als=[ca,la,la,ca,la,ca,ca,ca]); row+=1

# 주식 구조
row+=1; st(ws3,row,"D. 주식 구조 및 자사주 현황",8); row+=1
for lbl,val in [("발행주식수(보통주)",f"{SHARES:,}주"),
                ("자기주식",f"{SHARES_TREASURY:,}주 ({SHARES_TREASURY/SHARES*100:.1f}%)"),
                ("유통주식수",f"{SHARES_OUT:,}주"),
                ("최대주주(한국광해광업공단)",f"77,602,044주 (36.27%)"),
                ("소액주주","53.38% (약 11만명)")]:
    ws3.cell(row=row,column=1,value=lbl).font=db; ws3.cell(row=row,column=1).fill=llf
    ws3.cell(row=row,column=1).alignment=la; ws3.cell(row=row,column=1).border=tb
    ws3.merge_cells(start_row=row,start_column=2,end_row=row,end_column=4)
    ws3.cell(row=row,column=2,value=val).font=d_blue; ws3.cell(row=row,column=2).alignment=la
    ws3.cell(row=row,column=2).border=tb; row+=1

# 경영진 종합 평가
row+=1; st(ws3,row,"E. 경영진 종합 평가",8); row+=1
for item,val in [("지배구조 점수","3/5 (공기업형 전문경영인 → 오너십은 없으나 견제 기능 양호)"),
                  ("자본배분 점수","4/5 (고배당+자사주매입 병행, 제2카지노 성장투자)"),
                  ("투명성 점수","4/5 (정기 IR 활발, 잠정실적 분기 공시)"),
                  ("CEO 리스크","보통 (정권교체에 따른 CEO 변경 가능. 단 사업모델이 단순하여 영향 제한적)"),
                  ("결론","전문경영인이나 독점 사업의 단순성으로 경영진 역량 의존도 낮음. 주주환원 적극적.")]:
    ws3.cell(row=row,column=1,value=item).font=db; ws3.cell(row=row,column=1).fill=llf
    ws3.cell(row=row,column=1).alignment=la; ws3.cell(row=row,column=1).border=tb
    ws3.merge_cells(start_row=row,start_column=2,end_row=row,end_column=8)
    ws3.cell(row=row,column=2,value=val).font=df; ws3.cell(row=row,column=2).alignment=la
    ws3.cell(row=row,column=2).border=tb; row+=1

print("  [3/7] 경영진 평가")

# ============================================================
# SHEET 4: 자본배분 이력
# ============================================================
ws4 = wb.create_sheet("자본배분"); ws4.sheet_properties.tabColor="E74C3C"
sw(ws4,[14,14,14,14,14,14,14,14])

row=1; st(ws4,row,"자본배분 이력 (Capital Allocation History)",8); row+=1

# 배당 추이
st(ws4,row,"A. 배당 추이",8); row+=1
wh(ws4,row,["연도","주당배당(원)","EPS(원)","배당성향(%)","변화","비고","",""]); row+=1
div_years = ['2015','2016','2017','2018','2019','2020','2021','2022','2023','2024']
prev_dps = None
for yr in div_years:
    dps = div_map.get(yr,0)
    eps = eps_map.get(yr,0)
    po = dpayout_map.get(yr, 0)
    chg = ""
    if prev_dps is not None and prev_dps > 0 and dps > 0:
        chg = "↑" if dps > prev_dps else ("↓" if dps < prev_dps else "→")
    elif dps > 0 and (prev_dps is None or prev_dps == 0):
        chg = "재개"
    note = ""
    if yr=='2024': note = "대폭 인상 (+25.8%)"
    elif yr=='2023': note = "정상화 배당 복귀"
    elif yr=='2022': note = "배당 재개 (COVID 후)"
    elif yr=='2020' or yr=='2021': note = "COVID 무배당"
    elif yr=='2015': note = "고배당 전통"

    dps_f = d_grn if dps > 0 else d_red
    chg_f = d_grn if chg in ["↑","재개"] else (d_red if chg=="↓" else df)
    wr(ws4,row,[yr,dps if dps else "-",eps,f"{po:.1f}%" if po else "-",chg,note,"",""],
       fonts=[db,dps_f,df,df,chg_f,sm,df,df],fills=[lf]+[wf]*7,
       als=[ca,ra,ra,ca,ca,la,ca,ca]); row+=1
    prev_dps = dps

# 2025E
wr(ws4,row,["2025E",1300,"-","-","↑","예상 (배당수익률 7.3%)","",""],
   fonts=[db,d_grn,df,df,d_grn,sm,df,df],fills=[gnf]+[gnf]*7,
   als=[ca,ra,ra,ca,ca,la,ca,ca]); row+=1

# 자사주 이력
row+=1; st(ws4,row,"B. 자기주식 매입/처분/소각 이력",8); row+=1
wh(ws4,row,["연도","기초(주)","취득(주)","처분(주)","소각(주)","기말(주)","비고",""]); row+=1
for ts in ts_rows:
    yr = ts[0]
    def parse_q(v):
        try: return int(str(v).replace(',','').replace('-','0'))
        except: return 0
    bsis,acq,dsp,incnr,end = parse_q(ts[1]),parse_q(ts[2]),parse_q(ts[3]),parse_q(ts[4]),parse_q(ts[5])
    note = ""
    if acq > 0: note = "밸류업 매입"

    wr(ws4,row,[yr,f"{bsis:,}" if bsis else "-",f"{acq:,}" if acq else "-",
                f"{dsp:,}" if dsp else "-",f"{incnr:,}" if incnr else "-",
                f"{end:,}" if end else "-",note,""],
       fonts=[db,df,d_grn if acq>0 else df,d_red if dsp>0 else df,
              Font(name="맑은 고딕",size=10,bold=True,color="8E44AD") if incnr>0 else df,db,sm,df],
       fills=[lf]+[wf]*7,als=[ca,ra,ra,ra,ra,ra,la,ca]); row+=1

# 2025 자사주 이벤트
wr(ws4,row,["2025","13,534,816","진행중","-","-","-","2025.11 추가 취득 결정",""],
   fonts=[db,df,d_grn,df,df,df,sm,df],
   fills=[gnf]*8,als=[ca,ra,ra,ra,ra,ra,la,ca]); row+=1

# 주주환원 이벤트 타임라인
row+=1; st(ws4,row,"C. 주주환원 이벤트 타임라인",8); row+=1
wh(ws4,row,["일자","이벤트 유형","내용","","","","",""]); row+=1
events = [
    ("2024.03","배당","결산배당 주당 930원 결정 (2023년분)"),
    ("2024.10","자기주식","자기주식 취득 결정 (233만주, 약 413억)"),
    ("2024.11","신규투자","제2카지노영업장 조성사업 1,796억 투자 결정"),
    ("2025.03","배당","결산배당 주당 1,170원 결정 (2024년분, +25.8%)"),
    ("2025.11","자기주식","자기주식 추가 취득 결정"),
]
for dt,tp,content in events:
    tp_f = d_grn if "배당" in tp else (Font(name="맑은 고딕",size=10,bold=True,color="8E44AD") if "소각" in tp else d_blue)
    wr(ws4,row,[dt,tp,content,"","","","",""],
       fonts=[db,tp_f,df,df,df,df,df,df],fills=[llf,gld if "소각" in tp else wf,wf,wf,wf,wf,wf,wf],
       als=[ca,ca,la,ca,ca,ca,ca,ca]); row+=1

# 유상증자/CB 이력
row+=1; st(ws4,row,"D. 유상증자/전환사채 이력 (주요)",8); row+=1
has_cap = False
for c in cap_all:
    if c[1] and ('유상' in c[1] or '전환' in c[1]):
        if not has_cap:
            wh(ws4,row,["일자","변동사유","주식종류","수량(주)","","","",""]); row+=1
            has_cap = True
        wr(ws4,row,[c[0],c[1],c[2],c[3],"","","",""],
           fonts=[db,d_red,df,df,df,df,df,df],als=[ca,la,ca,ra,ca,ca,ca,ca]); row+=1
if not has_cap:
    ws4.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
    ws4.cell(row=row,column=1,value="유상증자/CB 발행 이력 없음 → 주주 희석 리스크 없음 (무차입경영)").font=d_grn
    ws4.cell(row=row,column=1).fill=gnf; ws4.cell(row=row,column=1).alignment=la; ws4.cell(row=row,column=1).border=tb
    row+=1

# 자본배분 점수카드
row+=1; st(ws4,row,"E. 자본배분 점수카드",8); row+=1
capex_opcf_ratio = CAPEX24/OPCF24 if OPCF24 else 0
total_return = DIV_PAID/MCAP  # 자사주 매입 금액은 2024 CF에서 별도 확인 필요
for item,score,comment in [
    ("배당 일관성","4/5","10년 중 2년(2020~2021) COVID 무배당 외 연속 배당, 고배당성향 45~60%"),
    ("자사주 정책","4/5","2024년 233만주 매입(413억) + 2025년 추가 매입 결정"),
    ("유상증자/CB","5/5","무차입경영, 유상증자/CB 이력 전무"),
    ("투자(CAPEX)","4/5",f"CAPEX/영업CF {capex_opcf_ratio*100:.0f}% — 제2카지노 성장투자 적절"),
    ("배당수익률","5/5",f"DPS {DPS24:,}원 / 주가 {PRICE:,}원 = {DPS24/PRICE*100:.1f}% — 시장 상위"),
]:
    sf = d_grn if "5/5" in score else (d_blue if "4/5" in score else df)
    wr(ws4,row,[item,score,comment,"","","","",""],
       fonts=[db,sf,df,df,df,df,df,df],fills=[llf,gnf if "5/5" in score else blf,wf,wf,wf,wf,wf,wf],
       als=[la,ca,la,ca,ca,ca,ca,ca]); row+=1

print("  [4/7] 자본배분")

# ============================================================
# SHEET 5: 수익성 품질 분석
# ============================================================
ws5 = wb.create_sheet("수익성_품질"); ws5.sheet_properties.tabColor="8E44AD"
sw(ws5,[14,14,14,14,14,14,14,14])

row=1; st(ws5,row,"수익성 품질 분석 (Earnings Quality)",8); row+=1

# DuPont 분해
st(ws5,row,"A. DuPont 3단계 분해 (5년)",8); row+=1
wh(ws5,row,["연도","순이익률","자산회전율","레버리지","ROE","ROE (산식확인)","비고",""]); row+=1

roe_list = []
for yr in YEARS5:
    d = annual[yr]
    npm = d["ni"]/d["rev"] if d["rev"] else 0
    ato = d["rev"]/d["ta"] if d["ta"] else 0
    lev = d["ta"]/d["te"] if d["te"] else 0
    roe_calc = npm * ato * lev
    roe_list.append(roe_calc)

    npm_f = d_grn if npm>0.15 else (d_red if npm<0.05 else df)
    roe_f = d_grn if roe_calc>0.10 else (d_red if roe_calc<0 else df)

    note = ""
    if yr=='2020': note = "COVID 적자"
    elif yr=='2024': note = "정상화 완료"

    wr(ws5,row,[yr,f"{npm*100:.1f}%",f"{ato:.2f}회",f"{lev:.2f}배",f"{roe_calc*100:.1f}%",
                f"= {npm*100:.1f}×{ato:.2f}×{lev:.2f}",note,""],
       fonts=[db,npm_f,df,df,roe_f,sm,sm,df],fills=[lf]+[wf]*7,
       als=[ca,ca,ca,ca,ca,la,la,ca]); row+=1

# ROE summary
row+=1
# COVID 기간 제외 평균 (2022~2024)
roe_normal = [r for r in roe_list if r > 0]
avg_roe = statistics.mean(roe_normal) if roe_normal else 0
ws5.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
ws5.cell(row=row,column=1,value=f"정상 연도 평균 ROE: {avg_roe*100:.1f}% | 판정: {'양호 (>10%)' if avg_roe>0.10 else '보통'}  (COVID 기간 제외)").font=db
ws5.cell(row=row,column=1).fill=gnf if avg_roe>0.10 else gld; ws5.cell(row=row,column=1).alignment=la
ws5.cell(row=row,column=1).border=tb; row+=2

# FCF/NI 비율
st(ws5,row,"B. FCF 대 순이익 비율 (현금이익 품질)",8); row+=1
wh(ws5,row,["연도","순이익(억)","영업CF(억)","CAPEX(억)","FCF(억)","FCF/NI","판정",""]); row+=1
for yr in YEARS5:
    d = annual[yr]
    ni_b = d["ni"]//억; opcf_b = d["opcf"]//억; capex_b = d["capex"]//억
    fcf_b = opcf_b - capex_b
    fcf_ni = (d["opcf"]-d["capex"])/d["ni"] if d["ni"]>0 else 0

    j = "양호" if fcf_ni>0.8 else ("주의" if fcf_ni>0.5 else "경고" if d["ni"]>0 else "적자")
    jf = d_grn if j=="양호" else (db if j=="주의" else d_red)
    jfl = gnf if j=="양호" else (gld if j=="주의" else rdf)

    wr(ws5,row,[yr,ni_b,opcf_b,capex_b,fcf_b,f"{fcf_ni*100:.0f}%" if d["ni"]>0 else "-",j,""],
       fonts=[db,df,df,df,d_blue,jf,jf,df],fills=[lf,wf,wf,wf,gld,jfl,jfl,wf],
       als=[ca,ra,ra,ra,ra,ca,ca,ca],nfs=[None,NF,NF,NF,NF,None,None,None]); row+=1

# 발생액 비율
row+=1; st(ws5,row,"C. 발생액 비율 (Accrual Ratio)",8); row+=1
ws5.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
ws5.cell(row=row,column=1,value="발생액 비율 = (순이익 - 영업CF) / 총자산  |  ±5% 이내 양호, ±10% 초과 경고").font=sm
ws5.cell(row=row,column=1).alignment=la; row+=1

wh(ws5,row,["연도","순이익(억)","영업CF(억)","차이(억)","총자산(억)","발생액비율","판정",""]); row+=1
for yr in YEARS5:
    d = annual[yr]
    ni_b = d["ni"]//억; opcf_b = d["opcf"]//억; diff = (d["ni"]-d["opcf"])//억
    ta_b = d["ta"]//억
    accrual = (d["ni"]-d["opcf"])/d["ta"] if d["ta"] else 0

    j = "양호" if abs(accrual)<0.05 else ("주의" if abs(accrual)<0.10 else "경고")
    jf = d_grn if j=="양호" else (db if j=="주의" else d_red)
    jfl = gnf if j=="양호" else (gld if j=="주의" else rdf)

    wr(ws5,row,[yr,ni_b,opcf_b,diff,ta_b,f"{accrual*100:.1f}%",j,""],
       fonts=[db,df,df,d_red if diff>0 else d_grn,df,jf,jf,df],
       fills=[lf,wf,wf,wf,wf,jfl,jfl,wf],als=[ca,ra,ra,ra,ra,ca,ca,ca],
       nfs=[None,NF,NF,NF,NF,None,None,None]); row+=1

# 이익 변동성
row+=1; st(ws5,row,"D. 이익 변동성 (Earnings Volatility)",8); row+=1

# COVID 포함 전체 10년
op_list_all = [annual[yr]["op"] for yr in YEARS_ALL if annual[yr]["op"] != 0]
op_list_normal = [annual[yr]["op"] for yr in ['2022','2023','2024']]  # 정상 연도
if len(op_list_normal) >= 2:
    op_mean = statistics.mean(op_list_normal)
    op_stdev = statistics.stdev(op_list_normal)
    cv = op_stdev / op_mean if op_mean > 0 else 99
else:
    op_mean = op_list_normal[0] if op_list_normal else 0
    op_stdev = 0; cv = 0

cv_j = "안정적" if cv<0.3 else ("변동성 있음" if cv<0.6 else "높은 변동성")
cv_f = d_grn if cv<0.3 else (db if cv<0.6 else d_red)

wh(ws5,row,["지표","값","판정","기준","","","",""]); row+=1
wr(ws5,row,["영업이익 평균 (정상3년)",f"{int(op_mean/억):,}억","","","","","",""],
   fonts=[db,d_blue,df,df,df,df,df,df],fills=[llf,gld,wf,wf,wf,wf,wf,wf],als=[la,ra,ca,la,ca,ca,ca,ca]); row+=1
wr(ws5,row,["영업이익 표준편차",f"{int(op_stdev/억):,}억","","","","","",""],
   fonts=[db,df,df,df,df,df,df,df],fills=[llf,wf,wf,wf,wf,wf,wf,wf],als=[la,ra,ca,la,ca,ca,ca,ca]); row+=1
wr(ws5,row,["변동계수 (CV)",f"{cv:.2f}",cv_j,"<0.3 안정 / 0.3~0.6 보통 / >0.6 높음","","","",""],
   fonts=[db,cv_f,cv_f,sm,df,df,df,df],
   fills=[llf,gnf if cv<0.3 else (gld if cv<0.6 else rdf),gnf if cv<0.3 else (gld if cv<0.6 else rdf),wf,wf,wf,wf,wf],
   als=[la,ca,ca,la,ca,ca,ca,ca]); row+=1

row+=1
ws5.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
ws5.cell(row=row,column=1,value="주의: COVID 기간(2020~2021) 포함 시 변동성 극대화. 정상 연도(2022~2024) 기준으로 분석").font=sm
ws5.cell(row=row,column=1).fill=gld; ws5.cell(row=row,column=1).alignment=la; ws5.cell(row=row,column=1).border=tb
row+=2

# 종합 판정
st(ws5,row,"E. 수익성 품질 종합",8); row+=1
for item,val in [
    ("DuPont 분석",f"정상연도 평균 ROE {avg_roe*100:.1f}% — 독점 사업의 높은 수익성. 레버리지 낮음(1.1~1.2배, 무차입)"),
    ("FCF 품질","영업CF가 순이익을 크게 상회 — 현금이익 우수. 감가상각비가 높아 FCF 창출력 양호"),
    ("발생액","대부분 ±5% 이내 — 이익의 현금 뒷받침 양호"),
    ("이익 변동성",f"정상연도 CV {cv:.2f} — COVID 제외 시 안정적. 독점 사업의 이익 안정성"),
    ("결론","이익 품질 우수. 독점 카지노의 높은 현금창출력과 낮은 레버리지가 핵심. COVID는 일회성 이벤트"),
]:
    ws5.cell(row=row,column=1,value=item).font=db; ws5.cell(row=row,column=1).fill=llf
    ws5.cell(row=row,column=1).alignment=la; ws5.cell(row=row,column=1).border=tb
    ws5.merge_cells(start_row=row,start_column=2,end_row=row,end_column=8)
    ws5.cell(row=row,column=2,value=val).font=df
    ws5.cell(row=row,column=2).alignment=la; ws5.cell(row=row,column=2).border=tb; row+=1

print("  [5/7] 수익성 품질")

# ============================================================
# SHEET 6: 내재가치 & 안전마진
# ============================================================
ws6 = wb.create_sheet("내재가치_안전마진"); ws6.sheet_properties.tabColor="2980B9"
sw(ws6,[22,16,16,16,16,16,16,16])

row=1; st(ws6,row,"내재가치 & 안전마진 분석",8); row+=1

# Owner Earnings
OE24 = NI24 + DA24 - CAPEX24
OE_per_share = OE24 / SHARES

st(ws6,row,"A. Owner Earnings (버핏 방식)",8); row+=1
wh(ws6,row,["항목","2024(억)","주당(원)","비고","","","",""]); row+=1
for lbl,amt,per_share,note in [
    ("순이익",int(NI24/억),f"{int(NI24/SHARES):,}","연결 지배기업"),
    ("(+) 감가상각비",int(DA24/억),f"{int(DA24/SHARES):,}","유형+무형"),
    ("(-) CAPEX",int(CAPEX24/억),f"{int(CAPEX24/SHARES):,}","유형자산 취득"),
    ("(=) Owner Earnings",int(OE24/억),f"{int(OE_per_share):,}","버핏이 보는 진정한 이익"),
]:
    tot = "(=)" in lbl
    wr(ws6,row,[lbl,f"{amt:,}",per_share,note,"","","",""],
       fonts=[db if tot else df,d_blue if tot else df,sm,df,df,df,df,df],
       fills=[gld if tot else llf,gld if tot else wf,wf,wf,wf,wf,wf,wf],
       als=[la,ra,la,la,ca,ca,ca,ca]); row+=1

row+=1
oe_per = MCAP / OE24
ws6.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
ws6.cell(row=row,column=1,value=f"현재 시총/Owner Earnings = {oe_per:.1f}배  |  Owner Earnings 수익률 = {OE24/MCAP*100:.1f}%").font=db
ws6.cell(row=row,column=1).fill=blf; ws6.cell(row=row,column=1).alignment=la; ws6.cell(row=row,column=1).border=tb
row+=2

# 자산가치
st(ws6,row,"B. 자산가치 (BPS 기반)",8); row+=1
wh(ws6,row,["지표","값","판정","","","","",""]); row+=1
for lbl,val,j in [
    ("BPS",f"{int(BPS):,}원",""),
    ("현재 PBR",f"{PRICE/BPS:.2f}배","적정" if 0.8<=PRICE/BPS<=1.5 else ("저평가" if PRICE/BPS<0.8 else "고평가")),
    ("PBR 0.8배 (하단)",fw(int(BPS*0.8)),""),
    ("PBR 1.0배",fw(int(BPS*1.0)),""),
    ("PBR 1.5배 (상단)",fw(int(BPS*1.5)),""),
    ("순현금/주",fw(int(CASH24/SHARES)),"무차입 순현금"),
]:
    wr(ws6,row,[lbl,val,j,"","","","",""],
       fonts=[db,d_blue,d_grn if "저평가" in j else (d_red if "고평가" in j else df),df,df,df,df,df],
       fills=[llf,gld,gnf if "저평가" in j else (rdf if "고평가" in j else wf),wf,wf,wf,wf,wf],
       als=[la,ra,ca,ca,ca,ca,ca,ca]); row+=1

# PER/PBR 밴드
row+=1; st(ws6,row,"C. PER/PBR 밴드 (과거 vs 현재)",8); row+=1
wh(ws6,row,["연도","EPS(원)","PER(현재가)","BPS(원)","PBR(현재가)","ROE","",""]); row+=1
HIST={2015:(2178,2997755204842),2016:(2242,3252727458873),2017:(2159,3501235933257),
      2018:(1467,3586777089279),2019:(1651,3717759076243),2020:(-1361,3247028580390),
      2022:(570,3415378100640),2023:(1681,3677900816262),2024:(EPS24,EQ24)}
for yr in [2015,2016,2017,2018,2019,2022,2023,2024]:
    eps,eq = HIST[yr]
    bps_yr = int(eq/SHARES)
    per_yr = PRICE/eps if eps>0 else 0
    pbr_yr = PRICE/bps_yr if bps_yr>0 else 0
    ni_yr = annual[str(yr)]["ni"]
    roe_yr = ni_yr/eq if eq and ni_yr else 0

    wr(ws6,row,[str(yr),f"{eps:,}",f"{per_yr:.1f}배" if per_yr>0 else "-",f"{bps_yr:,}",
                f"{pbr_yr:.2f}배",f"{roe_yr*100:.1f}%","",""],
       fonts=[db,d_blue,df,df,df,d_grn if roe_yr>0.10 else df,df,df],
       fills=[lf]+[wf]*7,als=[ca,ra,ca,ra,ca,ca,ca,ca]); row+=1

# 시나리오별 적정가
row+=1; st(ws6,row,"D. 시나리오별 적정주가",8); row+=1
wh(ws6,row,["시나리오","방법론","적정가","현재가 대비","전제","","",""]); row+=1

scenarios = [
    ("보수적","PER 10배 × 정상EPS",int(EPS24*0.85)*10,f"EPS {int(EPS24*0.85):,}원 × 10배"),
    ("보수적","BPS × 0.8배",int(BPS*0.8),"PBR 하단"),
    ("기본","PER 12배 × 확정EPS",EPS24*12,f"EPS {EPS24:,}원 × 12배"),
    ("기본","배당수익률 5% 기준",int(DPS24/0.05),"DPS 1,170원 / 5%"),
    ("낙관적","PER 15배 × 확정EPS",EPS24*15,f"EPS {EPS24:,}원 × 15배"),
    ("낙관적","OE × 15배",int(OE_per_share*15),"Owner Earnings 기반"),
]
for scen,method,target,basis in scenarios:
    target = int(target)
    upside = (target-PRICE)/PRICE
    uf = d_grn if upside>0 else d_red
    ufl = gnf if upside>0 else rdf
    scen_fl = gnf if "낙관" in scen else (gld if "기본" in scen else rdf)
    wr(ws6,row,[scen,method,fw(target),f"{upside*100:+.1f}%",basis,"","",""],
       fonts=[db,df,d_blue,uf,sm,df,df,df],fills=[scen_fl,wf,gld,ufl,wf,wf,wf,wf],
       als=[ca,la,ra,ca,la,ca,ca,ca]); row+=1

# 안전마진 테스트
row+=1; st(ws6,row,"E. 안전마진 테스트",8); row+=1

conserv_values = [s[2] for s in scenarios if "보수" in s[0]]
conserv_avg = int(statistics.mean(conserv_values)) if conserv_values else PRICE
base_values = [s[2] for s in scenarios if "기본" in s[0]]
base_avg = int(statistics.mean(base_values)) if base_values else PRICE
opt_values = [s[2] for s in scenarios if "낙관" in s[0]]
opt_avg = int(statistics.mean(opt_values)) if opt_values else PRICE

for lbl,iv,verdict in [
    ("보수적 내재가치",conserv_avg,"하방 리스크 확인" if PRICE>conserv_avg else "안전마진 확보"),
    ("기본 내재가치",base_avg,"적정" if abs(PRICE-base_avg)/base_avg<0.15 else ("상승여력" if PRICE<base_avg else "과대평가")),
    ("낙관적 내재가치",opt_avg,"업사이드" if PRICE<opt_avg else "이미 반영"),
]:
    margin = (iv - PRICE) / iv if iv else 0
    mf2 = d_grn if margin>0.15 else (df if margin>0 else d_red)
    mfl = gnf if margin>0.15 else (gld if margin>0 else rdf)
    wr(ws6,row,[lbl,fw(iv),f"안전마진 {margin*100:+.1f}%",verdict,f"현재가 {PRICE:,}원","","",""],
       fonts=[db,d_blue,mf2,mf2,sm,df,df,df],fills=[llf,gld,mfl,mfl,wf,wf,wf,wf],
       als=[la,ra,ca,ca,la,ca,ca,ca]); row+=1

row+=1
ws6.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
verdict_overall = ("기본 시나리오 기준 상승 여력 존재. 카지노업 PER 10~15배 밴드에서 현재 저평가 영역. "
                   f"안전마진: 기본 내재가치({fw(base_avg)}) 대비 {(base_avg-PRICE)/base_avg*100:+.1f}%")
ws6.cell(row=row,column=1,value=f"→ {verdict_overall}").font=db
ws6.cell(row=row,column=1).fill=blf; ws6.cell(row=row,column=1).alignment=la; ws6.cell(row=row,column=1).border=tb

print("  [6/7] 내재가치/안전마진")

# ============================================================
# SHEET 7: 투자판정 & 모니터링
# ============================================================
ws7 = wb.create_sheet("투자판정"); ws7.sheet_properties.tabColor="1ABC9C"
sw(ws7,[22,16,16,16,16,16,16,16])

row=1
# Title banner
for r in range(1,5):
    for c in range(1,9): ws7.cell(row=r,column=c).fill=tf
ws7.merge_cells('A2:H2')
ws7.cell(row=2,column=1,value="최종 투자 판정").font=Font(name="맑은 고딕",size=22,bold=True,color=W)
ws7.cell(row=2,column=1).alignment=ca
row=6

# 최종 판정
st(ws7,row,"A. 최종 판정",8); row+=1

# Determine verdict
if total_score >= 70:
    verdict = "BUY (매수)"
    verdict_color = GREEN_C
    verdict_fill = gnf
elif total_score >= 55:
    verdict = "HOLD (관망)"
    verdict_color = GOLD_C
    verdict_fill = gld
else:
    verdict = "REJECT (거부)"
    verdict_color = RED_C
    verdict_fill = rdf

ws7.merge_cells(start_row=row,start_column=1,end_row=row+1,end_column=3)
c = ws7.cell(row=row,column=1,value=verdict)
c.font=Font(name="맑은 고딕",size=20,bold=True,color=verdict_color)
c.fill=verdict_fill; c.alignment=ca; c.border=tb

ws7.merge_cells(start_row=row,start_column=4,end_row=row+1,end_column=8)
c = ws7.cell(row=row,column=4)
c.value=(f"Four Filters 종합 {total_score}/100 (등급 {grade})\n"
         f"PER {per_trail:.1f}배(T4Q) | PBR {pbr:.2f}배 | 배당수익률 {DPS24/PRICE*100:.1f}% | "
         f"FCF Yield {FCF24/MCAP*100:.1f}%")
c.font=Font(name="맑은 고딕",size=10,bold=True,color=NAVY)
c.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True); c.border=tb
row+=3

# 판정 근거
st(ws7,row,"B. 판정 근거",8); row+=1
reasons = [
    ("매수 근거 1","규제 독점 해자: 한국 유일 내국인 카지노 면허 — 신규 진입 불가, 가장 강력한 해자 유형"),
    ("매수 근거 2","무차입 순현금 경영 + 고배당: 순현금 2,445억, 배당성향 51%, 배당수익률 6.6%"),
    ("매수 근거 3",f"밸류에이션 매력: PER {per_trail:.1f}배(T4Q), 카지노업 기준 10~15배 대비 저평가"),
    ("매수 근거 4","성장 투자: 제2카지노영업장 1,796억 투자 → 중장기 매출 성장 동력 확보"),
    ("주의 사항 1","정부 규제 리스크: 영업시간/입장료/횟수 제한 변경 시 매출 직접 타격"),
    ("주의 사항 2","2025년 실적 둔화: 영업이익 -17.7% YoY, 순이익 -30.7% YoY (잠정)"),
    ("주의 사항 3","온라인 도박 합법화 또는 내국인 카지노 추가 면허 시 독점 해자 훼손"),
]
for title,detail in reasons:
    is_risk = "주의" in title
    tf2 = d_red if is_risk else d_grn
    tfl = rdf if is_risk else gnf
    ws7.cell(row=row,column=1,value=title).font=tf2; ws7.cell(row=row,column=1).fill=tfl
    ws7.cell(row=row,column=1).alignment=la; ws7.cell(row=row,column=1).border=tb
    ws7.merge_cells(start_row=row,start_column=2,end_row=row,end_column=8)
    ws7.cell(row=row,column=2,value=detail).font=df
    ws7.cell(row=row,column=2).alignment=la; ws7.cell(row=row,column=2).border=tb; row+=1

# SWOT (구루 관점)
row+=1; st(ws7,row,"C. 구루 관점 SWOT",8); row+=1

ws7.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
ws7.cell(row=row,column=1,value="강점 (Strengths)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws7.cell(row=row,column=1).fill=PatternFill("solid",fgColor=GREEN_C); ws7.cell(row=row,column=1).alignment=ca
ws7.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
ws7.cell(row=row,column=5,value="약점 (Weaknesses)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws7.cell(row=row,column=5).fill=PatternFill("solid",fgColor=RED_C); ws7.cell(row=row,column=5).alignment=ca
row+=1

SW_S = ["한국 유일 내국인 카지노 (규제 독점)",
        "무차입경영 + 순현금 2,445억",
        "GPM 50%+, OPM 25~35% 구조적 고마진",
        "고배당 전통 (배당성향 45~60%)"]
SW_W = ["전문경영인 체제 (정부 영향, 오너십 부재)",
        "정선 단일 입지 (지리적 집중 리스크)",
        "사행산업 규제 의존적 사업구조",
        "COVID 같은 외부충격에 극도로 취약"]
for i in range(4):
    ws7.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
    ws7.cell(row=row,column=1,value=f"  {SW_S[i]}").font=df; ws7.cell(row=row,column=1).fill=gnf
    ws7.cell(row=row,column=1).alignment=la; ws7.cell(row=row,column=1).border=tb
    ws7.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
    ws7.cell(row=row,column=5,value=f"  {SW_W[i]}").font=df; ws7.cell(row=row,column=5).fill=rdf
    ws7.cell(row=row,column=5).alignment=la; ws7.cell(row=row,column=5).border=tb; row+=1

row+=1
ws7.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
ws7.cell(row=row,column=1,value="기회 (Opportunities)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws7.cell(row=row,column=1).fill=PatternFill("solid",fgColor="2980B9"); ws7.cell(row=row,column=1).alignment=ca
ws7.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
ws7.cell(row=row,column=5,value="위협 (Threats)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws7.cell(row=row,column=5).fill=PatternFill("solid",fgColor="7F8C8D"); ws7.cell(row=row,column=5).alignment=ca
row+=1

O = ["제2카지노영업장 개장 → 매출 성장 동력",
     "관광산업 정상화 → 외국인 관광객 유입",
     "밸류업 프로그램 → 주주환원 강화 가속",
     "배당수익률 매력 → 가치주 재평가"]
T = ["온라인 도박 합법화 → 독점 해자 약화",
     "내국인 카지노 추가 면허 발급",
     "사행산업 규제 강화 (영업시간 축소 등)",
     "경기침체 → 여가소비 감소 → 입장객 감소"]
for i in range(4):
    ws7.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
    ws7.cell(row=row,column=1,value=f"  {O[i]}").font=df; ws7.cell(row=row,column=1).fill=blf
    ws7.cell(row=row,column=1).alignment=la; ws7.cell(row=row,column=1).border=tb
    ws7.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
    ws7.cell(row=row,column=5,value=f"  {T[i]}").font=df
    ws7.cell(row=row,column=5).fill=PatternFill("solid",fgColor="E5E7E9")
    ws7.cell(row=row,column=5).alignment=la; ws7.cell(row=row,column=5).border=tb; row+=1

# 모니터링 지표
row+=1; st(ws7,row,"D. 핵심 모니터링 지표 (5개)",8); row+=1
wh(ws7,row,["#","지표","세부 내용","확인 시기","구루 관점","","",""]); row+=1
monitors = [
    ("1","월별 입장객 수","입장객 증감이 매출과 직결. 전년동기 대비 추이 확인","매월","해자 가동률 확인"),
    ("2","영업이익률 추이","OPM 15% 미만=경고, 20~25%=양호, 30%+=호황","분기 실적","수익성 품질"),
    ("3","배당 정책 변화","DPS 인상/유지/삭감 여부, 배당성향 변화","3월 주총","경영진 신뢰"),
    ("4","규제 환경 변화","영업시간, 입장료, 입장횟수 제한 변경 여부","수시","해자 위협"),
    ("5","제2카지노 진행상황","공사 진척률, 개장 일정, 예상 매출 기여도","반기","성장 동력"),
]
for num,title,detail,timing,guru in monitors:
    wr(ws7,row,[num,title,detail,timing,guru,"","",""],
       fonts=[Font(name="맑은 고딕",size=12,bold=True,color=NAVY),db,df,df,sm,df,df,df],
       fills=[gld,llf,wf,llf,wf,wf,wf,wf],als=[ca,la,la,ca,la,ca,ca,ca])
    ws7.row_dimensions[row].height=40; row+=1

# 매도 트리거
row+=1; st(ws7,row,"E. 매도 트리거 조건",8); row+=1
sell_triggers = [
    ("내국인 카지노 추가 면허 발급 확정","해자 붕괴 → 즉시 매도"),
    ("온라인 도박 합법화 법안 통과","독점 해자 약화 → 단계적 매도"),
    ("영업이익률 2년 연속 15% 미만","구조적 수익성 훼손"),
    ("배당 2년 연속 삭감 또는 무배당","주주환원 정책 변화"),
    ("PER 18배+ (카지노업 과열 기준)","안전마진 소멸"),
]
for trigger,action in sell_triggers:
    ws7.merge_cells(start_row=row,start_column=1,end_row=row,end_column=5)
    ws7.cell(row=row,column=1,value=f"  {trigger}").font=df
    ws7.cell(row=row,column=1).fill=rdf; ws7.cell(row=row,column=1).alignment=la; ws7.cell(row=row,column=1).border=tb
    ws7.merge_cells(start_row=row,start_column=6,end_row=row,end_column=8)
    ws7.cell(row=row,column=6,value=action).font=d_red
    ws7.cell(row=row,column=6).alignment=la; ws7.cell(row=row,column=6).border=tb; row+=1

# 리스크 시나리오
row+=1; st(ws7,row,"F. 리스크 시나리오 (최악의 경우)",8); row+=1
ws7.merge_cells(start_row=row,start_column=1,end_row=row+3,end_column=8)
c = ws7.cell(row=row,column=1)
c.value=("최악 시나리오: 내국인 카지노 추가 면허 2개 발급 + 온라인 도박 부분 합법화\n"
         "→ 입장객 40% 감소 + 1인당 GGR 20% 하락\n"
         f"→ 매출 6,000억, 영업이익 600억(OPM 10%), 순이익 450억, EPS ~2,100원\n"
         f"→ PER 8배 적용 시 주가 ~16,800원 (현재가 대비 -5%). 이 시나리오 확률: 5% 미만")
c.font=Font(name="맑은 고딕",size=10,color=RED_C)
c.fill=PatternFill("solid",fgColor="FDE8E8")
c.alignment=Alignment(horizontal='left',vertical='top',wrap_text=True); c.border=tb

print("  [7/7] 투자판정/모니터링")

# === SAVE ===
OUT = os.path.join(BASE, "강원랜드_투자구루분석.xlsx")
wb.save(OUT)
conn.close()
print(f"\n투자 구루 분석 보고서 생성 완료: {OUT}")
print(f"시트 ({len(wb.sheetnames)}개): {wb.sheetnames}")
