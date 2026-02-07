# -*- coding: utf-8 -*-
"""엠씨넥스 투자 구루 분석 보고서 (Buffett/Munger 한국형 Four Filters)"""
import sqlite3, sys, os, statistics
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ai.db")
BASE = os.path.dirname(os.path.abspath(__file__))
conn = sqlite3.connect(DB)

# === KEY CONSTANTS (create_combined.py와 동일) ===
PRICE = 28100
SHARES = 17977732
SHARES_TREASURY = 1110000
SHARES_WA = 17373105
억 = 100_000_000
MCAP = PRICE * SHARES

# 2024 Annual
REV24 = 1057058423929; OP24 = 44384671816; NI24 = 63604930146; EPS24 = 3661
EQ24 = 372224146153; EQ23 = 323612702499; ASSETS24 = 574103551032; LIAB24 = 201879404879
CASH24 = 35881602429; ST_DEBT = 44402388000; LT_DEBT = 1308000000
DA24 = 44255618814 + 462932726 + 1740746569
OPCF24 = 82470298068; CAPEX24 = 36541908724
DIV_PAID = 10570639200; TREAS_BUY = 15656989850; DPS24 = 800

# Quarters
Q24 = [{"rev":287835057669,"op":15204634567,"ni":15929655236},
       {"rev":225759412794,"op":7650453061,"ni":18422813873},
       {"rev":245070688152,"op":3077406264,"ni":2504912286},
       {"rev":REV24-287835057669-225759412794-245070688152,
        "op":OP24-15204634567-7650453061-3077406264,
        "ni":NI24-15929655236-18422813873-2504912286}]
Q25 = [{"rev":347917e6,"op":20918e6,"ni":20068e6},
       {"rev":326424e6,"op":16171e6,"ni":13050e6},
       {"rev":303537e6,"op":6708e6,"ni":9537e6}]
CUM25 = {"rev":977877e6,"op":43797e6,"ni":42656e6}

# Derived
BPS = EQ24 / SHARES
NET_DEBT = ST_DEBT + LT_DEBT - CASH24
EBITDA24 = OP24 + DA24; FCF24 = OPCF24 - CAPEX24; EV = MCAP + NET_DEBT
TRAIL_NI = Q24[3]["ni"]+Q25[0]["ni"]+Q25[1]["ni"]+Q25[2]["ni"]
TRAIL_EPS = TRAIL_NI / SHARES_WA
AVG_EQ = (EQ24+EQ23)/2; ROE24 = NI24/AVG_EQ

# === STYLE (create_combined.py와 동일) ===
NAVY="1B2A4A"; DARK="2C3E6B"; MID="3A5BA0"; LB="D6E4F0"; LLB="EBF1F8"; W="FFFFFF"
GOLD_C="D4A843"; RED_C="C0392B"; GREEN_C="27AE60"; GRAY_C="F2F2F2"

title_font=Font(name="맑은 고딕",size=22,bold=True,color=W)
sub_font=Font(name="맑은 고딕",size=11,color="B0C4DE")
sec_font=Font(name="맑은 고딕",size=14,bold=True,color=NAVY)
hdr_font=Font(name="맑은 고딕",size=10,bold=True,color=W)
df=Font(name="맑은 고딕",size=10)
db=Font(name="맑은 고딕",size=10,bold=True)
d_blue=Font(name="맑은 고딕",size=10,bold=True,color="0000FF")
d_grn=Font(name="맑은 고딕",size=10,bold=True,color=GREEN_C)
d_red=Font(name="맑은 고딕",size=10,bold=True,color=RED_C)
d_navy=Font(name="맑은 고딕",size=12,bold=True,color=NAVY)
sm=Font(name="맑은 고딕",size=9,color="666666")

tf=PatternFill("solid",fgColor=NAVY)
hf=PatternFill("solid",fgColor=DARK)
mf=PatternFill("solid",fgColor=MID)
lf=PatternFill("solid",fgColor=LB)
llf=PatternFill("solid",fgColor=LLB)
gf=PatternFill("solid",fgColor=GRAY_C)
wf=PatternFill("solid",fgColor=W)
gld=PatternFill("solid",fgColor="FFF3CD")
rdf=PatternFill("solid",fgColor="F8D7DA")
gnf=PatternFill("solid",fgColor="D4EDDA")
blf=PatternFill("solid",fgColor="D6EAF8")

ca=Alignment(horizontal='center',vertical='center',wrap_text=True)
la=Alignment(horizontal='left',vertical='center',wrap_text=True)
ra=Alignment(horizontal='right',vertical='center')
tb=Border(left=Side('thin',color="D9D9D9"),right=Side('thin',color="D9D9D9"),
          top=Side('thin',color="D9D9D9"),bottom=Side('thin',color="D9D9D9"))
bb=Border(bottom=Side('medium',color=NAVY))
NF='#,##0'; PF='0.0%'

def sw(ws,w):
    for i,v in enumerate(w,1): ws.column_dimensions[get_column_letter(i)].width=v
def setup_print(ws):
    ws.page_setup.paperSize = Worksheet.PAPERSIZE_LETTER
    ws.page_setup.orientation = Worksheet.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.75, bottom=0.75,
        header=0.3, footer=0.3
    )
    ws.print_options.horizontalCentered = True
def wh(ws,r,h,fills=None):
    for i,v in enumerate(h,1):
        c=ws.cell(row=r,column=i,value=v); c.font=hdr_font; c.fill=fills[i-1] if fills else hf; c.alignment=ca; c.border=tb
def wr(ws,r,d,fonts=None,fills=None,als=None,nfs=None):
    for i,v in enumerate(d,1):
        c=ws.cell(row=r,column=i,value=v)
        c.font=fonts[i-1] if fonts else df; c.fill=fills[i-1] if fills else wf
        c.alignment=als[i-1] if als else ca; c.border=tb
        if nfs and i<=len(nfs) and nfs[i-1]: c.number_format=nfs[i-1]
def st(ws,r,t,ce=8):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ce)
    c=ws.cell(row=r,column=1,value=t); c.font=sec_font; c.border=bb; return r+1
def fmt(v): return f"{v/억:,.0f}억"
def fw(v): return f"{v:,.0f}원"
def pct(v): return f"{v*100:.1f}%"

# === LOAD DATA ===
perf = conn.execute("SELECT * FROM v_annual_performance").fetchall()

# 매출총이익
gp_rows = conn.execute("""SELECT bsns_year, thstrm_amount FROM financial_statements
    WHERE account_nm='매출총이익' AND sj_div='CIS' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    ORDER BY bsns_year""").fetchall()
gp_map = {r[0]: int(r[1]) for r in gp_rows if r[1]}

# 영업CF
opcf_rows = conn.execute("""SELECT bsns_year, thstrm_amount FROM financial_statements
    WHERE account_nm='영업활동현금흐름' AND sj_div='CF' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    ORDER BY bsns_year""").fetchall()
opcf_map = {r[0]: int(r[1]) for r in opcf_rows if r[1]}

# CAPEX
capex_rows = conn.execute("""SELECT bsns_year, thstrm_amount FROM financial_statements
    WHERE account_nm LIKE '%유형자산%취득%' AND sj_div='CF' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    ORDER BY bsns_year""").fetchall()
capex_map = {r[0]: abs(int(r[1])) for r in capex_rows if r[1]}

# 감가상각비 (유형+투자부동산+무형)
da_rows = conn.execute("""SELECT bsns_year, account_nm, thstrm_amount FROM financial_statements
    WHERE account_nm IN ('감가상각비','투자부동산감가상각비','무형자산상각비')
    AND sj_div='CF' AND reprt_code='11011' AND reprt_nm LIKE '%연결%'
    ORDER BY bsns_year""").fetchall()
da_map = {}
for r in da_rows:
    if r[2]:
        da_map[r[0]] = da_map.get(r[0], 0) + int(r[2])

# 배당 (주당 현금배당금)
div_rows = conn.execute("""SELECT bsns_year, thstrm FROM dividends
    WHERE se='주당 현금배당금(원)' AND stock_knd='보통주' ORDER BY bsns_year""").fetchall()
div_map = {}
for r in div_rows:
    try: div_map[r[0]] = int(r[1])
    except: div_map[r[0]] = 0

# 배당성향
dpayout_rows = conn.execute("""SELECT bsns_year, thstrm FROM dividends
    WHERE se='(연결)현금배당성향(%)' ORDER BY bsns_year""").fetchall()
dpayout_map = {}
for r in dpayout_rows:
    try: dpayout_map[r[0]] = float(r[1])
    except: pass

# 자사주 총계
ts_rows = conn.execute("""SELECT bsns_year, bsis_qy, change_qy_acqs, change_qy_dsps, change_qy_incnr, trmend_qy
    FROM treasury_stock WHERE acqs_mth1='총계' AND stock_knd='보통주' ORDER BY bsns_year""").fetchall()

# 유상증자/CB
cap_rows = conn.execute("""SELECT isu_dcrs_de, isu_dcrs_stle, isu_dcrs_stock_knd, isu_dcrs_qy, isu_dcrs_mstvdv_amount
    FROM capital_changes WHERE bsns_year='2024' AND isu_dcrs_stle IS NOT NULL
    ORDER BY isu_dcrs_de""").fetchall()
# All capital changes for significant events
cap_all = conn.execute("""SELECT DISTINCT isu_dcrs_de, isu_dcrs_stle, isu_dcrs_stock_knd, isu_dcrs_qy
    FROM capital_changes WHERE isu_dcrs_stle LIKE '%유상증자%' OR isu_dcrs_stle LIKE '%전환%' OR isu_dcrs_stle LIKE '%무상%'
    ORDER BY isu_dcrs_de""").fetchall()

# 임원 (2024)
exec_rows = conn.execute("""SELECT nm, ofcps, rgist_exctv_at, fte_at FROM executives
    WHERE bsns_year='2024' LIMIT 20""").fetchall()

# CEO 보수
pay_rows = conn.execute("""SELECT bsns_year, nm, ofcps, mendng_totamt FROM individual_pay
    WHERE nm='민동욱' ORDER BY bsns_year""").fetchall()

# 특허
patent_rows = conn.execute("SELECT rcept_dt, patent_name, patent_detail, patent_plan FROM patents ORDER BY rcept_dt").fetchall()

# 주요 이벤트
event_rows = conn.execute("""SELECT rcept_dt, event_type, SUBSTR(event_summary,1,80) FROM key_events
    WHERE rcept_dt>='20220101' ORDER BY rcept_dt DESC LIMIT 20""").fetchall()

# Build annual data dict for analysis
annual = {}
for p in perf:
    yr = p[0]
    rev = int(p[1]) if p[1] else 0
    op = int(p[2]) if p[2] else 0
    ni = int(p[3]) if p[3] else 0
    eps = int(p[4]) if p[4] else 0
    ta = int(p[5]) if p[5] else 0
    tl = int(p[6]) if p[6] else 0
    te = int(p[7]) if p[7] else 0
    annual[yr] = {"rev":rev,"op":op,"ni":ni,"eps":eps,"ta":ta,"tl":tl,"te":te,
                  "gp":gp_map.get(yr,0),"opcf":opcf_map.get(yr,0),
                  "capex":capex_map.get(yr,0),"da":da_map.get(yr,0)}

# Last 5 years for trend analysis
YEARS5 = [y for y in sorted(annual.keys()) if y >= '2020']

wb = Workbook()

# ============================================================
# SHEET 1: Four Filters 종합 대시보드
# ============================================================
ws1 = wb.active; ws1.title="Four Filters"; ws1.sheet_properties.tabColor=NAVY
sw(ws1,[4,22,14,14,14,14,14,14,4])
setup_print(ws1)

# Title banner
for r in range(1,6):
    for c in range(1,10): ws1.cell(row=r,column=c).fill=tf
ws1.merge_cells('B2:H2')
ws1.cell(row=2,column=2,value="엠씨넥스(MCNEX) 투자 구루 분석").font=Font(name="맑은 고딕",size=24,bold=True,color=W)
ws1.cell(row=2,column=2).alignment=ca
ws1.merge_cells('B4:H4')
ws1.cell(row=4,column=2,value=f"Four Filters 종합 대시보드  |  현재가 {PRICE:,}원  |  2026.02.06 기준").font=sub_font
ws1.cell(row=4,column=2).alignment=ca

row = 7

# === Four Filters Score ===
st(ws1,row,"A. Four Filters 스코어카드 (100점 만점)",8); row+=1

# Filter 1: 사업 이해도 - 소항목
f1_items = [
    ("사업모델 단순성","CCM+전장 카메라, 2개 사업부","높음",4),
    ("매출구조 명확성","모바일70%+전장25%+기타5%","높음",4),
    ("10년 일관성","2004년 설립 이래 카메라모듈 일관","높음",5),
    ("기술변화 예측성","카메라 고화소·다안화 트렌드 명확","보통",4),
    ("해외법인 이해도","베트남VINA 3공장, 연결의 핵심","높음",4),
]
f1_score = sum(x[3] for x in f1_items)

# Filter 2: 경제적 해자
f2_items = [
    ("전환비용","삼성/현대 인증 2~3년, 공급사 교체 어려움","강함",5),
    ("비용우위","베트남 3공장, 원가경쟁력","보통",3),
    ("무형자산(기술)","특허24건, R&D인력57.7%, 액츄에이터 내재화","보통",4),
    ("규모의 경제","국내 CCM 1~2위, 전장 카메라 성장","보통",3),
    ("수익성 지속성","GPM 10.8%, OPM 4.2% (부품업 평균 수준)","보통",3),
]
f2_score = sum(x[3] for x in f2_items)

# Filter 3: 경영진
f3_items = [
    ("오너 지분율","민동욱 25.93% - 적정 수준","양호",4),
    ("보수 합리성","연 524M원 - 과도하지 않음","양호",4),
    ("자사주 정책","2024 75만주 취득, 2025 58만주 소각","우수",5),
    ("배당 인상","500→600→800원, 꾸준한 인상 추세","양호",4),
    ("소통/IR","정기공시 충실, 밸류업 방향 일치","보통",3),
]
f3_score = sum(x[3] for x in f3_items)

# Filter 4: 안전마진
per24 = PRICE/EPS24; per_trail = PRICE/TRAIL_EPS; pbr = PRICE/BPS
f4_items = [
    ("PER 수준",f"확정 {per24:.1f}배 / T4Q {per_trail:.1f}배 (기준 8~12배)","저평가" if per_trail<8 else "적정",4 if per_trail<8 else 3),
    ("PBR 수준",f"{pbr:.2f}배 (기준 0.8~1.5배)","적정",3),
    ("FCF 수익률",f"{FCF24/MCAP*100:.1f}% (기준 >5%)","양호" if FCF24/MCAP>0.05 else "보통",4 if FCF24/MCAP>0.05 else 3),
    ("EV/EBITDA",f"{EV/EBITDA24:.1f}배 (기준 6~10배)","저평가" if EV/EBITDA24<6 else "적정",4 if EV/EBITDA24<6 else 3),
    ("Owner Earnings 대비",f"OE {fmt(NI24+DA24-CAPEX24)}, OE배수 {MCAP/(NI24+DA24-CAPEX24):.1f}배","양호",4),
]
f4_score = sum(x[3] for x in f4_items)

total_score = f1_score + f2_score + f3_score + f4_score
grade = "A" if total_score>=80 else ("B" if total_score>=60 else ("C" if total_score>=40 else "D"))
grade_color = GREEN_C if grade in ["A","B"] else (GOLD_C if grade=="C" else RED_C)

# Score summary
wh(ws1,row,["","필터","소항목1","소항목2","소항목3","소항목4","소항목5","소계"]); row+=1
for fname,items,score in [("F1: 사업 이해도",f1_items,f1_score),
                           ("F2: 경제적 해자",f2_items,f2_score),
                           ("F3: 경영진 품질",f3_items,f3_score),
                           ("F4: 안전마진",f4_items,f4_score)]:
    scores = [x[3] for x in items]
    sc_fill = gnf if score>=20 else (gld if score>=15 else rdf)
    wr(ws1,row,["",fname]+scores+[f"{score}/25"],
       fonts=[df,db]+[df]*5+[Font(name="맑은 고딕",size=11,bold=True,color=GREEN_C if score>=20 else NAVY)],
       fills=[wf,llf]+[wf]*5+[sc_fill],als=[ca,la]+[ca]*6)
    row+=1

# Total
row+=1
ws1.merge_cells(start_row=row,start_column=2,end_row=row,end_column=5)
ws1.cell(row=row,column=2,value="종합점수").font=d_navy; ws1.cell(row=row,column=2).alignment=la; ws1.cell(row=row,column=2).border=tb
ws1.merge_cells(start_row=row,start_column=6,end_row=row,end_column=7)
ws1.cell(row=row,column=6,value=f"{total_score}/100 (등급: {grade})").font=Font(name="맑은 고딕",size=14,bold=True,color=grade_color)
ws1.cell(row=row,column=6).fill=gnf if grade in ["A","B"] else gld
ws1.cell(row=row,column=6).alignment=ca; ws1.cell(row=row,column=6).border=tb
row+=2

# === 투자 테제 ===
st(ws1,row,"B. 투자 테제",8); row+=1
ws1.merge_cells(start_row=row,start_column=2,end_row=row+1,end_column=8)
c=ws1.cell(row=row,column=2)
c.value=("엠씨넥스는 삼성전자 CCM 핵심 공급사+현대모비스 전장 카메라 성장이라는 이중 성장축을 가진 "
         "기술 부품기업으로, 오너경영진의 적극적 주주환원(배당인상+자사주소각)이 확인되며, "
         f"현재 PER {per_trail:.1f}배(T4Q)는 한국 부품업 평균 대비 저평가 영역. "
         f"FCF 수익률 {FCF24/MCAP*100:.1f}%로 현금창출력 대비 매력적.")
c.font=Font(name="맑은 고딕",size=10,bold=True,color=NAVY); c.alignment=Alignment(horizontal='left',vertical='top',wrap_text=True)
row+=3

# === 핵심 강점/리스크 ===
st(ws1,row,"C. 핵심 강점 vs 리스크",8); row+=1

# Strengths
ws1.merge_cells(start_row=row,start_column=2,end_row=row,end_column=4)
ws1.cell(row=row,column=2,value="핵심 강점 (3)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws1.cell(row=row,column=2).fill=PatternFill("solid",fgColor=GREEN_C); ws1.cell(row=row,column=2).alignment=ca
ws1.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
ws1.cell(row=row,column=5,value="핵심 리스크 (3)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws1.cell(row=row,column=5).fill=PatternFill("solid",fgColor=RED_C); ws1.cell(row=row,column=5).alignment=ca
row+=1

strengths = ["삼성·현대 핵심 공급사 지위 (전환비용 해자)",
             "적극적 주주환원: 배당인상+자사주매입/소각",
             "전장(자동차) 카메라 비중 확대 → 밸류에이션 리레이팅 잠재력"]
risks = ["삼성전자 매출 편중 ~70% (단일 고객 리스크)",
         "조립 중심 저마진(OPM 3~5%) 사업 구조",
         "중국 업체 추격 + 삼성 모듈 내재화 가능성"]

for i in range(3):
    ws1.merge_cells(start_row=row,start_column=2,end_row=row,end_column=4)
    ws1.cell(row=row,column=2,value=f"  {i+1}. {strengths[i]}").font=df
    ws1.cell(row=row,column=2).fill=gnf; ws1.cell(row=row,column=2).alignment=la; ws1.cell(row=row,column=2).border=tb
    ws1.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
    ws1.cell(row=row,column=5,value=f"  {i+1}. {risks[i]}").font=df
    ws1.cell(row=row,column=5).fill=rdf; ws1.cell(row=row,column=5).alignment=la; ws1.cell(row=row,column=5).border=tb
    row+=1

# === 소항목 상세 ===
row+=1; st(ws1,row,"D. 필터별 소항목 상세",8); row+=1
wh(ws1,row,["","항목","평가 근거","판정","점수","","",""]); row+=1
for fname,items in [("Filter 1: 사업 이해도",f1_items),("Filter 2: 경제적 해자",f2_items),
                     ("Filter 3: 경영진 품질",f3_items),("Filter 4: 안전마진",f4_items)]:
    ws1.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
    ws1.cell(row=row,column=1,value=fname).font=Font(name="맑은 고딕",size=11,bold=True,color=MID)
    ws1.cell(row=row,column=1).fill=llf; ws1.cell(row=row,column=1).border=tb; row+=1
    for item_nm,evidence,judgment,score in items:
        sf = d_grn if score>=4 else (df if score>=3 else d_red)
        sfill = gnf if score>=4 else (wf if score>=3 else rdf)
        wr(ws1,row,["",item_nm,evidence,judgment,f"{score}/5","","",""],
           fonts=[df,db,df,sf,sf,df,df,df],fills=[wf,llf,wf,sfill,sfill,wf,wf,wf],
           als=[ca,la,la,ca,ca,ca,ca,ca]); row+=1

print("  [1/7] Four Filters 대시보드")

# ============================================================
# SHEET 2: 경제적 해자 분석
# ============================================================
ws2 = wb.create_sheet("경제적_해자"); ws2.sheet_properties.tabColor=DARK
sw(ws2,[20,16,16,16,16,16,16,16])
setup_print(ws2)

row=1; st(ws2,row,"경제적 해자(Economic Moat) 분석",8); row+=1

# 해자 유형 판정표
st(ws2,row,"A. 해자 유형 판정",8); row+=1
wh(ws2,row,["해자 유형","보유 여부","강도","근거","지속가능성","","",""]); row+=1
moat_types = [
    ("비용우위","O","보통","베트남 VINA 3공장 원가경쟁력","5년+"),
    ("전환비용","O","강함","삼성/현대 인증 2~3년, 공급사 교체 비용 높음","10년+"),
    ("네트워크 효과","X","없음","B2B 사업 구조상 해당 없음","-"),
    ("무형자산(기술)","O","보통","특허 24건, R&D 인력 57.7%, 액츄에이터 내재화","5년+"),
    ("규모의 경제","△","보통","국내 CCM 1~2위이나 글로벌 경쟁 치열","3~5년"),
]
for tp,has,strength,basis,duration in moat_types:
    hf2 = d_grn if has=="O" else (d_red if has=="X" else db)
    hfl = gnf if has=="O" else (rdf if has=="X" else gld)
    wr(ws2,row,[tp,has,strength,basis,duration,"","",""],
       fonts=[db,hf2,hf2,df,df,df,df,df],fills=[llf,hfl,hfl,wf,wf,wf,wf,wf],
       als=[la,ca,ca,la,ca,ca,ca,ca]); row+=1

# 수익성 추이
row+=1; st(ws2,row,"B. 수익성 추이 (5년)",8); row+=1
wh(ws2,row,["연도","매출액(억)","매출총이익(억)","GPM","영업이익(억)","OPM","ROIC","비고"]); row+=1

prev_eq = None
for yr in YEARS5:
    d = annual[yr]
    rev_b = d["rev"]//억; gp_b = d["gp"]//억; op_b = d["op"]//억
    gpm = d["gp"]/d["rev"] if d["rev"] else 0
    opm = d["op"]/d["rev"] if d["rev"] else 0
    # ROIC = NOPAT / Invested Capital (simplified: op*(1-0.22) / (equity + net_debt))
    nopat = d["op"] * 0.78
    ic = d["te"] + max(0, d["tl"] - d["ta"]*0.1)  # simplified
    roic = nopat / d["te"] if d["te"] else 0

    gpm_f = d_grn if gpm>0.12 else (d_red if gpm<0.08 else df)
    opm_f = d_grn if opm>0.05 else (d_red if opm<0.03 else df)
    note = ""
    if yr=='2024': note = "역대 최고 매출"
    elif yr=='2023': note = "매출 저점"

    wr(ws2,row,[yr,rev_b,gp_b,f"{gpm*100:.1f}%",op_b,f"{opm*100:.1f}%",f"{roic*100:.1f}%",note],
       fonts=[db,df,df,gpm_f,df,opm_f,df,sm],fills=[lf]+[wf]*7,
       als=[ca,ra,ra,ca,ra,ca,ca,la],nfs=[None,NF,NF,None,NF,None,None,None]); row+=1
    prev_eq = d["te"]

# R&D
row+=1; st(ws2,row,"C. R&D 투자 및 특허",8); row+=1
wh(ws2,row,["연도","R&D비용(억)","매출대비","비고","","","",""]); row+=1
for yr,rd,pct_v,note in [("2022",397,"3.58%",""),("2023",356,"3.82%",""),("2024",327,"3.10%","연구인력 284명")]:
    wr(ws2,row,[yr,rd,pct_v,note,"","","",""],fonts=[db,df,df,sm,df,df,df,df],
       fills=[lf,wf,wf,wf,wf,wf,wf,wf],als=[ca,ra,ca,la,ca,ca,ca,ca]); row+=1

# Patent summary
row+=1
ws2.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
ws2.cell(row=row,column=1,value=f"특허 포트폴리오: 총 {len(patent_rows)}건 — 액츄에이터, 전장카메라(AVM), 광학모듈 등").font=db
ws2.cell(row=row,column=1).fill=blf; ws2.cell(row=row,column=1).alignment=la; ws2.cell(row=row,column=1).border=tb
row+=1

wh(ws2,row,["취득일","특허명/내용","분류","활용","","","",""]); row+=1
for p in patent_rows[:12]:  # Top 12
    dt=p[0]; nm=(p[1]or"")[:50]; det=(p[2]or"")[:70]; plan=(p[3]or"")[:30]
    cat="액츄에이터" if "액" in det or "ctuator" in det else "전장" if "차량" in det or "주차" in det or "AVM" in det else "광학/제조" if "광축" in det or "프리즘" in det else "보안" if "홍채" in det else "기타"
    disp = nm if nm and "특허" not in nm[:3] else det[:50]
    wr(ws2,row,[dt,disp,cat,plan[:25] if plan else "","","","",""],
       fonts=[df,df,db,sm,df,df,df,df],als=[ca,la,ca,la,ca,ca,ca,ca]); row+=1

if len(patent_rows) > 12:
    ws2.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
    ws2.cell(row=row,column=1,value=f"... 외 {len(patent_rows)-12}건").font=sm; row+=1

# 해자 종합 평가
row+=1; st(ws2,row,"D. 해자 종합 평가",8); row+=1
for item,val in [("해자 폭","Narrow ~ Medium (전환비용 + 기술력 기반)"),
                  ("해자 추세","확대 중 (전장 비중 증가 → 고객 다변화)"),
                  ("핵심 해자","삼성·현대 공급사 인증 (전환비용) + 액츄에이터 내재화 (기술)"),
                  ("위협 요인","중국 업체 추격, 삼성 모듈 내재화 가능성"),
                  ("결론","전환비용 기반 Narrow Moat 보유. 전장 확대 시 해자 강화 가능")]:
    ws2.cell(row=row,column=1,value=item).font=db; ws2.cell(row=row,column=1).fill=llf
    ws2.cell(row=row,column=1).alignment=la; ws2.cell(row=row,column=1).border=tb
    ws2.merge_cells(start_row=row,start_column=2,end_row=row,end_column=8)
    ws2.cell(row=row,column=2,value=val).font=df if "위협" not in item else d_red
    ws2.cell(row=row,column=2).alignment=la; ws2.cell(row=row,column=2).border=tb; row+=1

print("  [2/7] 경제적 해자")

# ============================================================
# SHEET 3: 경영진 평가
# ============================================================
ws3 = wb.create_sheet("경영진_평가"); ws3.sheet_properties.tabColor=GOLD_C
sw(ws3,[20,16,16,16,16,16,16,16])
setup_print(ws3)

row=1; st(ws3,row,"경영진 평가 (Management Quality)",8); row+=1

# 임원 명단
st(ws3,row,"A. 등기임원 명단 (2024년 사업보고서 기준)",8); row+=1
wh(ws3,row,["성명","직위","등기임원","상근여부","","","",""]); row+=1
for e in exec_rows:
    nm,ofcps,reg,fte = e[0],e[1],e[2],e[3]
    if not nm or nm=='-': continue
    is_ceo = "대표" in (ofcps or "")
    wr(ws3,row,[nm,ofcps,reg,fte,"","","",""],
       fonts=[d_navy if is_ceo else db,df,df,df,df,df,df,df],
       fills=[gld if is_ceo else llf]+[wf]*7,als=[la,la,ca,ca,ca,ca,ca,ca]); row+=1

# CEO 보수 추이
row+=1; st(ws3,row,"B. 대표이사 보수 추이",8); row+=1
wh(ws3,row,["연도","성명","직위","총보수","비고","","",""]); row+=1
for pr in pay_rows:
    yr,nm,ofcps,amt = pr[0],pr[1],pr[2],pr[3]
    if nm=='-': continue
    # Format amount
    try:
        amt_v = int(str(amt).replace(',',''))
        amt_str = f"{amt_v:,}원" if amt_v > 10000 else str(amt)
        amt_m = f"(월 ~{amt_v//12:,}원)" if amt_v > 10000 else ""
    except:
        amt_str = str(amt); amt_m = ""
    wr(ws3,row,[yr,nm,ofcps,amt_str,amt_m,"","",""],
       fonts=[db,db,df,d_blue,sm,df,df,df],fills=[lf,wf,wf,gld,wf,wf,wf,wf],
       als=[ca,la,la,ra,la,ca,ca,ca]); row+=1

# 오너경영 체크리스트
row+=1; st(ws3,row,"C. 오너경영 체크리스트",8); row+=1
wh(ws3,row,["","평가 항목","현황","판정","근거","","",""]); row+=1
owner_checks = [
    ("1","대표이사 지분율","25.93% (4,661,000주)","양호","20~30% 적정 범위 내"),
    ("2","경영 안정성","2004년 설립 이래 민동욱 대표 체제","양호","창업자 오너경영 20년+"),
    ("3","보수 합리성","연 524M원 (2024)","양호","순이익 636억 대비 0.8%"),
    ("4","자사주 매입/소각","2024 75만주 매입, 2025 58만주 소각","우수","적극적 주주환원 의지"),
    ("5","배당 정책","500→600→800원 연속 인상","양호","배당성향 21% (한국 기준 적정)"),
    ("6","일감 몰아주기","구내식당 자회사(엠씨넥스F&B) 외 특이사항 없음","양호","사익편취 리스크 낮음"),
    ("7","소통","정기공시 충실, 잠정실적 공시","보통","별도 IR 활동은 제한적"),
]
for num,item,status,judgment,basis in owner_checks:
    jf = d_grn if judgment=="우수" else (d_blue if judgment=="양호" else (db if judgment=="보통" else d_red))
    jfl = gnf if judgment=="우수" else (blf if judgment=="양호" else (gld if judgment=="보통" else rdf))
    wr(ws3,row,[num,item,status,judgment,basis,"","",""],
       fonts=[df,db,df,jf,sm,df,df,df],fills=[wf,llf,wf,jfl,wf,wf,wf,wf],
       als=[ca,la,la,ca,la,ca,ca,ca]); row+=1

# 주식 구조
row+=1; st(ws3,row,"D. 주식 구조 및 자사주 현황",8); row+=1
for lbl,val in [("발행주식수(보통주)",f"{SHARES:,}주"),
                ("자기주식",f"{SHARES_TREASURY:,}주 ({SHARES_TREASURY/SHARES*100:.1f}%)"),
                ("유통주식수",f"{SHARES-SHARES_TREASURY:,}주"),
                ("최대주주(민동욱 외)",f"4,661,000주 (25.93%)"),
                ("소액주주","66.33% (58,267명)")]:
    ws3.cell(row=row,column=1,value=lbl).font=db; ws3.cell(row=row,column=1).fill=llf
    ws3.cell(row=row,column=1).alignment=la; ws3.cell(row=row,column=1).border=tb
    ws3.merge_cells(start_row=row,start_column=2,end_row=row,end_column=4)
    ws3.cell(row=row,column=2,value=val).font=d_blue; ws3.cell(row=row,column=2).alignment=la
    ws3.cell(row=row,column=2).border=tb; row+=1

# 경영진 종합 평가
row+=1; st(ws3,row,"E. 경영진 종합 평가",8); row+=1
for item,val in [("오너십 점수","4/5 (주주와 이해관계 일치하는 오너경영)"),
                  ("자본배분 점수","4/5 (배당인상+자사주매입소각 병행)"),
                  ("투명성 점수","3/5 (공시 충실하나 별도 주주서한 등 없음)"),
                  ("승계 리스크","보통 (민동욱 대표 장기 체제, 후계 계획 미공개)"),
                  ("결론","주주 친화적 오너경영. 자본배분 개선 추세 뚜렷.")]:
    ws3.cell(row=row,column=1,value=item).font=db; ws3.cell(row=row,column=1).fill=llf
    ws3.cell(row=row,column=1).alignment=la; ws3.cell(row=row,column=1).border=tb
    ws3.merge_cells(start_row=row,start_column=2,end_row=row,end_column=8)
    ws3.cell(row=row,column=2,value=val).font=df; ws3.cell(row=row,column=2).alignment=la
    ws3.cell(row=row,column=2).border=tb; row+=1

print("  [3/7] 경영진 평가")

# ============================================================
# SHEET 4: 자본배분 이력
# ============================================================
ws4 = wb.create_sheet("자본배분"); ws4.sheet_properties.tabColor="E74C3C"
sw(ws4,[14,14,14,14,14,14,14,14])
setup_print(ws4)

row=1; st(ws4,row,"자본배분 이력 (Capital Allocation History)",8); row+=1

# 배당 추이
st(ws4,row,"A. 배당 추이",8); row+=1
wh(ws4,row,["연도","주당배당(원)","EPS(원)","배당성향(%)","변화","비고","",""]); row+=1
div_years = ['2015','2016','2017','2018','2019','2020','2021','2022','2023','2024']
prev_dps = None
for yr in div_years:
    dps = div_map.get(yr,0)
    eps = annual[yr]["eps"] if yr in annual else 0
    po = dpayout_map.get(yr, 0)
    chg = ""
    if prev_dps is not None and prev_dps > 0 and dps > 0:
        chg = "↑" if dps > prev_dps else ("↓" if dps < prev_dps else "→")
    elif dps > 0 and (prev_dps is None or prev_dps == 0):
        chg = "재개"
    note = ""
    if yr=='2024': note = "대폭 인상 (+33%)"
    elif yr=='2023': note = "인상 (+20%)"
    elif yr=='2019': note = "최대실적 배당"
    elif yr=='2016': note = "적자 무배당"

    dps_f = d_grn if dps > 0 else d_red
    chg_f = d_grn if chg in ["↑","재개"] else (d_red if chg=="↓" else df)
    wr(ws4,row,[yr,dps if dps else "-",eps,f"{po:.1f}%" if po else "-",chg,note,"",""],
       fonts=[db,dps_f,df,df,chg_f,sm,df,df],fills=[lf]+[wf]*7,
       als=[ca,ra,ra,ca,ca,la,ca,ca]); row+=1
    prev_dps = dps

# 2025E
wr(ws4,row,["2025E",1000,"-","-","↑","예상 (배당수익률 3.6%)","",""],
   fonts=[db,d_grn,df,df,d_grn,sm,df,df],fills=[gnf]+[gnf]*7,
   als=[ca,ra,ra,ca,ca,la,ca,ca]); row+=1

# 자사주 이력
row+=1; st(ws4,row,"B. 자기주식 매입/처분/소각 이력",8); row+=1
wh(ws4,row,["연도","기초(주)","취득(주)","처분(주)","소각(주)","기말(주)","비고",""]); row+=1
for ts in ts_rows:
    yr = ts[0]
    def parse_q(v):
        try: return int(str(v).replace(',','').replace('-','0'))
        except: return 0
    bsis,acq,dsp,incnr,end = parse_q(ts[1]),parse_q(ts[2]),parse_q(ts[3]),parse_q(ts[4]),parse_q(ts[5])
    note = ""
    if acq > 0 and yr in ['2024','2023','2022']: note = "적극 매입"
    if incnr > 0: note = "소각 실행"

    wr(ws4,row,[yr,f"{bsis:,}" if bsis else "-",f"{acq:,}" if acq else "-",
                f"{dsp:,}" if dsp else "-",f"{incnr:,}" if incnr else "-",
                f"{end:,}" if end else "-",note,""],
       fonts=[db,df,d_grn if acq>0 else df,d_red if dsp>0 else df,
              Font(name="맑은 고딕",size=10,bold=True,color="8E44AD") if incnr>0 else df,db,sm,df],
       fills=[lf]+[wf]*7,als=[ca,ra,ra,ra,ra,ra,la,ca]); row+=1

# 2025 소각 이벤트
wr(ws4,row,["2025","-","-","-","583,482","-","2025.11 소각 결정",""],
   fonts=[db,df,df,df,Font(name="맑은 고딕",size=10,bold=True,color="8E44AD"),df,sm,df],
   fills=[gnf]*8,als=[ca,ra,ra,ra,ra,ra,la,ca]); row+=1

# 주주환원 이벤트 타임라인
row+=1; st(ws4,row,"C. 주주환원 이벤트 타임라인",8); row+=1
wh(ws4,row,["일자","이벤트 유형","내용","","","","",""]); row+=1
events = [
    ("2024.03","자기주식","자기주식 취득 결정"),
    ("2024.07","자기주식","자기주식 취득 결정 (추가)"),
    ("2024.11","자기주식","자기주식 추가 취득"),
    ("2024.12","배당","결산배당 주당 800원 결정"),
    ("2025.01","자기주식","자기주식 취득"),
    ("2025.09","자기주식","자기주식 취득"),
    ("2025.11","주식소각","자기주식 583,482주 소각 (~170억)"),
    ("2025.12","배당","결산배당 주당 1,000원 결정"),
]
for dt,tp,content in events:
    tp_f = d_grn if "배당" in tp else (Font(name="맑은 고딕",size=10,bold=True,color="8E44AD") if "소각" in tp else d_blue)
    wr(ws4,row,[dt,tp,content,"","","","",""],
       fonts=[db,tp_f,df,df,df,df,df,df],fills=[llf,gld if "소각" in tp else wf,wf,wf,wf,wf,wf,wf],
       als=[ca,ca,la,ca,ca,ca,ca,ca]); row+=1

# 유상증자/CB 이력 (최근 주요)
row+=1; st(ws4,row,"D. 유상증자/전환사채 이력 (주요)",8); row+=1
has_cap = False
for c in cap_all:
    if c[1] and ('유상' in c[1] or '전환' in c[1]):
        if not has_cap:
            wh(ws4,row,["일자","변동사유","주식종류","수량(주)","","","",""]); row+=1
            has_cap = True
        wr(ws4,row,[c[0],c[1],c[2],c[3],"","","",""],
           fonts=[db,d_red,df,df,df,df,df,df],als=[ca,la,ca,ra,ca,ca,ca,ca]); row+=1
if not has_cap:
    ws4.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
    ws4.cell(row=row,column=1,value="최근 유상증자/CB 발행 없음 → 주주 희석 리스크 낮음").font=d_grn
    ws4.cell(row=row,column=1).fill=gnf; ws4.cell(row=row,column=1).alignment=la; ws4.cell(row=row,column=1).border=tb
    row+=1

# 자본배분 점수카드
row+=1; st(ws4,row,"E. 자본배분 점수카드",8); row+=1
for item,score,comment in [
    ("배당 일관성","4/5","10년 중 1년(2016) 무배당 외 연속 배당, 최근 3년 연속 인상"),
    ("자사주 정책","5/5","매입+소각 병행 실행 — 한국 기업 중 드문 적극성"),
    ("유상증자/CB","5/5","최근 희석 이벤트 없음"),
    ("투자(CAPEX)","4/5","CAPEX/영업CF 44% — 성장투자와 주주환원 균형"),
    ("총주주환원율","4/5",f"{pct((DIV_PAID+TREAS_BUY)/MCAP)} (배당+자사주) — 업종 상위"),
]:
    sf = d_grn if "5/5" in score else (d_blue if "4/5" in score else df)
    wr(ws4,row,[item,score,comment,"","","","",""],
       fonts=[db,sf,df,df,df,df,df,df],fills=[llf,gnf if "5/5" in score else blf,wf,wf,wf,wf,wf,wf],
       als=[la,ca,la,ca,ca,ca,ca,ca]); row+=1

print("  [4/7] 자본배분")

# ============================================================
# SHEET 5: 수익성 품질 분석
# ============================================================
ws5 = wb.create_sheet("수익성_품질"); ws5.sheet_properties.tabColor="8E44AD"
sw(ws5,[14,14,14,14,14,14,14,14])
setup_print(ws5)

row=1; st(ws5,row,"수익성 품질 분석 (Earnings Quality)",8); row+=1

# DuPont 분해
st(ws5,row,"A. DuPont 3단계 분해 (5년)",8); row+=1
wh(ws5,row,["연도","순이익률","자산회전율","레버리지","ROE","ROE (산식확인)","비고",""]); row+=1

roe_list = []
for yr in YEARS5:
    d = annual[yr]
    npm = d["ni"]/d["rev"] if d["rev"] else 0
    ato = d["rev"]/d["ta"] if d["ta"] else 0
    lev = d["ta"]/d["te"] if d["te"] else 0
    roe_calc = npm * ato * lev
    roe_list.append(roe_calc)

    npm_f = d_grn if npm>0.05 else (d_red if npm<0.02 else df)
    roe_f = d_grn if roe_calc>0.12 else (d_red if roe_calc<0.05 else df)

    wr(ws5,row,[yr,f"{npm*100:.1f}%",f"{ato:.2f}회",f"{lev:.2f}배",f"{roe_calc*100:.1f}%",
                f"= {npm*100:.1f}×{ato:.2f}×{lev:.2f}","",""],
       fonts=[db,npm_f,df,df,roe_f,sm,df,df],fills=[lf]+[wf]*7,
       als=[ca,ca,ca,ca,ca,la,ca,ca]); row+=1

# ROE summary
row+=1
avg_roe = statistics.mean(roe_list)
ws5.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
ws5.cell(row=row,column=1,value=f"5년 평균 ROE: {avg_roe*100:.1f}% | 판정: {'양호 (>10%)' if avg_roe>0.10 else '보통 (한국 평균 수준)'}").font=db
ws5.cell(row=row,column=1).fill=gnf if avg_roe>0.10 else gld; ws5.cell(row=row,column=1).alignment=la
ws5.cell(row=row,column=1).border=tb; row+=2

# FCF/NI 비율
st(ws5,row,"B. FCF 대 순이익 비율 (현금이익 품질)",8); row+=1
wh(ws5,row,["연도","순이익(억)","영업CF(억)","CAPEX(억)","FCF(억)","FCF/NI","판정",""]); row+=1
for yr in YEARS5:
    d = annual[yr]
    ni_b = d["ni"]//억; opcf_b = d["opcf"]//억; capex_b = d["capex"]//억
    fcf_b = opcf_b - capex_b
    fcf_ni = (d["opcf"]-d["capex"])/d["ni"] if d["ni"]>0 else 0

    j = "양호" if fcf_ni>0.8 else ("주의" if fcf_ni>0.5 else "경고")
    jf = d_grn if j=="양호" else (db if j=="주의" else d_red)
    jfl = gnf if j=="양호" else (gld if j=="주의" else rdf)

    wr(ws5,row,[yr,ni_b,opcf_b,capex_b,fcf_b,f"{fcf_ni*100:.0f}%",j,""],
       fonts=[db,df,df,df,d_blue,jf,jf,df],fills=[lf,wf,wf,wf,gld,jfl,jfl,wf],
       als=[ca,ra,ra,ra,ra,ca,ca,ca],nfs=[None,NF,NF,NF,NF,None,None,None]); row+=1

# 발생액 비율
row+=1; st(ws5,row,"C. 발생액 비율 (Accrual Ratio)",8); row+=1
ws5.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
ws5.cell(row=row,column=1,value="발생액 비율 = (순이익 - 영업CF) / 총자산  |  ±5% 이내 양호, ±10% 초과 경고").font=sm
ws5.cell(row=row,column=1).alignment=la; row+=1

wh(ws5,row,["연도","순이익(억)","영업CF(억)","차이(억)","총자산(억)","발생액비율","판정",""]); row+=1
for yr in YEARS5:
    d = annual[yr]
    ni_b = d["ni"]//억; opcf_b = d["opcf"]//억; diff = (d["ni"]-d["opcf"])//억
    ta_b = d["ta"]//억
    accrual = (d["ni"]-d["opcf"])/d["ta"] if d["ta"] else 0

    j = "양호" if abs(accrual)<0.05 else ("주의" if abs(accrual)<0.10 else "경고")
    jf = d_grn if j=="양호" else (db if j=="주의" else d_red)
    jfl = gnf if j=="양호" else (gld if j=="주의" else rdf)

    wr(ws5,row,[yr,ni_b,opcf_b,diff,ta_b,f"{accrual*100:.1f}%",j,""],
       fonts=[db,df,df,d_red if diff>0 else d_grn,df,jf,jf,df],
       fills=[lf,wf,wf,wf,wf,jfl,jfl,wf],als=[ca,ra,ra,ra,ra,ca,ca,ca],
       nfs=[None,NF,NF,NF,NF,None,None,None]); row+=1

# 이익 변동성
row+=1; st(ws5,row,"D. 이익 변동성 (Earnings Volatility)",8); row+=1

op_list = [annual[yr]["op"] for yr in YEARS5]
if len(op_list) >= 2:
    op_mean = statistics.mean(op_list)
    op_stdev = statistics.stdev(op_list)
    cv = op_stdev / op_mean if op_mean > 0 else 99
else:
    cv = 0

cv_j = "안정적" if cv<0.3 else ("변동성 있음" if cv<0.6 else "높은 변동성")
cv_f = d_grn if cv<0.3 else (db if cv<0.6 else d_red)

wh(ws5,row,["지표","값","판정","기준","","","",""]); row+=1
wr(ws5,row,["영업이익 평균 (5년)",f"{int(op_mean/억):,}억","","","","","",""],
   fonts=[db,d_blue,df,df,df,df,df,df],fills=[llf,gld,wf,wf,wf,wf,wf,wf],als=[la,ra,ca,la,ca,ca,ca,ca]); row+=1
wr(ws5,row,["영업이익 표준편차",f"{int(op_stdev/억):,}억","","","","","",""],
   fonts=[db,df,df,df,df,df,df,df],fills=[llf,wf,wf,wf,wf,wf,wf,wf],als=[la,ra,ca,la,ca,ca,ca,ca]); row+=1
wr(ws5,row,["변동계수 (CV)",f"{cv:.2f}",cv_j,"<0.3 안정 / 0.3~0.6 보통 / >0.6 높음","","","",""],
   fonts=[db,cv_f,cv_f,sm,df,df,df,df],
   fills=[llf,gnf if cv<0.3 else (gld if cv<0.6 else rdf),gnf if cv<0.3 else (gld if cv<0.6 else rdf),wf,wf,wf,wf,wf],
   als=[la,ca,ca,la,ca,ca,ca,ca]); row+=1

# 종합 판정
row+=1; st(ws5,row,"E. 수익성 품질 종합",8); row+=1
for item,val in [
    ("DuPont 분석",f"평균 ROE {avg_roe*100:.1f}% — 순이익률 개선이 ROE 드라이버. 레버리지 적정(1.5~1.7배)"),
    ("FCF 품질","영업CF > 순이익 (현금이익 우수). 2020년 CAPEX 과다 외 양호"),
    ("발생액","대부분 ±5% 이내 — 이익의 현금 뒷받침 양호"),
    ("이익 변동성",f"CV {cv:.2f} — {'경기민감주 특성. 삼성 갤럭시 사이클에 연동' if cv>=0.3 else '비교적 안정'}"),
    ("결론","이익 품질 양호. 현금창출력이 회계이익을 뒷받침. 단, 삼성 사이클 의존으로 변동성 존재"),
]:
    ws5.cell(row=row,column=1,value=item).font=db; ws5.cell(row=row,column=1).fill=llf
    ws5.cell(row=row,column=1).alignment=la; ws5.cell(row=row,column=1).border=tb
    ws5.merge_cells(start_row=row,start_column=2,end_row=row,end_column=8)
    ws5.cell(row=row,column=2,value=val).font=df
    ws5.cell(row=row,column=2).alignment=la; ws5.cell(row=row,column=2).border=tb; row+=1

print("  [5/7] 수익성 품질")

# ============================================================
# SHEET 6: 내재가치 & 안전마진
# ============================================================
ws6 = wb.create_sheet("내재가치_안전마진"); ws6.sheet_properties.tabColor="2980B9"
sw(ws6,[22,16,16,16,16,16,16,16])
setup_print(ws6)

row=1; st(ws6,row,"내재가치 & 안전마진 분석",8); row+=1

# Owner Earnings
OE24 = NI24 + DA24 - CAPEX24
OE_per_share = OE24 / SHARES

st(ws6,row,"A. Owner Earnings (버핏 방식)",8); row+=1
wh(ws6,row,["항목","2024(억)","주당(원)","비고","","","",""]); row+=1
for lbl,amt,per_share,note in [
    ("순이익",int(NI24/억),f"{int(NI24/SHARES):,}",""),
    ("(+) 감가상각비",int(DA24/억),f"{int(DA24/SHARES):,}","유형+투자부동산+무형"),
    ("(-) CAPEX",int(CAPEX24/억),f"{int(CAPEX24/SHARES):,}","유형자산 취득"),
    ("(=) Owner Earnings",int(OE24/억),f"{int(OE_per_share):,}","버핏이 보는 진정한 이익"),
]:
    tot = "(=)" in lbl
    wr(ws6,row,[lbl,f"{amt:,}",per_share,note,"","","",""],
       fonts=[db if tot else df,d_blue if tot else df,sm,df,df,df,df,df],
       fills=[gld if tot else llf,gld if tot else wf,wf,wf,wf,wf,wf,wf],
       als=[la,ra,la,la,ca,ca,ca,ca]); row+=1

row+=1
oe_per = MCAP / OE24
ws6.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
ws6.cell(row=row,column=1,value=f"현재 시총/Owner Earnings = {oe_per:.1f}배  |  Owner Earnings 수익률 = {OE24/MCAP*100:.1f}%").font=db
ws6.cell(row=row,column=1).fill=blf; ws6.cell(row=row,column=1).alignment=la; ws6.cell(row=row,column=1).border=tb
row+=2

# 자산가치
st(ws6,row,"B. 자산가치 (BPS 기반)",8); row+=1
wh(ws6,row,["지표","값","판정","","","","",""]); row+=1
for lbl,val,j in [
    ("BPS",f"{int(BPS):,}원",""),
    ("현재 PBR",f"{PRICE/BPS:.2f}배","적정" if 0.8<=PRICE/BPS<=1.5 else ("저평가" if PRICE/BPS<0.8 else "고평가")),
    ("PBR 0.8배 (하단)",fw(int(BPS*0.8)),""),
    ("PBR 1.0배",fw(int(BPS*1.0)),""),
    ("PBR 1.5배 (상단)",fw(int(BPS*1.5)),""),
]:
    wr(ws6,row,[lbl,val,j,"","","","",""],
       fonts=[db,d_blue,d_grn if "저평가" in j else (d_red if "고평가" in j else df),df,df,df,df,df],
       fills=[llf,gld,gnf if "저평가" in j else (rdf if "고평가" in j else wf),wf,wf,wf,wf,wf],
       als=[la,ra,ca,ca,ca,ca,ca,ca]); row+=1

# PER/PBR 밴드
row+=1; st(ws6,row,"C. PER/PBR 밴드 (과거 vs 현재)",8); row+=1
wh(ws6,row,["연도","EPS(원)","PER(현재가)","BPS(원)","PBR(현재가)","ROE","",""]); row+=1
HIST={2019:(5151,224567081980),2020:(2180,252934743419),2021:(2229,298966196208),
      2022:(1293,309066577971),2023:(1585,323612702499),2024:(EPS24,EQ24)}
for yr in [2019,2020,2021,2022,2023,2024]:
    eps,eq = HIST[yr]
    bps_yr = int(eq/SHARES)
    per_yr = PRICE/eps if eps>0 else 0
    pbr_yr = PRICE/bps_yr if bps_yr>0 else 0
    roe_yr = annual[str(yr)]["ni"]/eq if eq else 0

    wr(ws6,row,[str(yr),f"{eps:,}",f"{per_yr:.1f}배" if per_yr>0 else "-",f"{bps_yr:,}",
                f"{pbr_yr:.2f}배",f"{roe_yr*100:.1f}%","",""],
       fonts=[db,d_blue,df,df,df,d_grn if roe_yr>0.10 else df,df,df],
       fills=[lf]+[wf]*7,als=[ca,ra,ca,ra,ca,ca,ca,ca]); row+=1

# 시나리오별 적정가
row+=1; st(ws6,row,"D. 시나리오별 적정주가",8); row+=1
wh(ws6,row,["시나리오","방법론","적정가","현재가 대비","전제","","",""]); row+=1

scenarios = [
    ("보수적","PER 7배 × 보수EPS",int(TRAIL_EPS*0.9)*7,f"EPS {int(TRAIL_EPS*0.9):,}원 × 7배"),
    ("보수적","BPS × 0.8배",int(BPS*0.8),"PBR 하단"),
    ("기본","PER 10배 × T4Q EPS",int(TRAIL_EPS)*10,f"EPS {int(TRAIL_EPS):,}원 × 10배"),
    ("기본","RIM (ROE 12%)",int(BPS*(1+0.02/0.08)),"ke=10%, g=2%"),
    ("낙관적","PER 12배 × T4Q EPS",int(TRAIL_EPS)*12,f"EPS {int(TRAIL_EPS):,}원 × 12배"),
    ("낙관적","OE × 12배",int(OE_per_share*12),"Owner Earnings 기반"),
]
for scen,method,target,basis in scenarios:
    upside = (target-PRICE)/PRICE
    uf = d_grn if upside>0 else d_red
    ufl = gnf if upside>0 else rdf
    scen_fl = gnf if "낙관" in scen else (gld if "기본" in scen else rdf)
    wr(ws6,row,[scen,method,fw(target),f"{upside*100:+.1f}%",basis,"","",""],
       fonts=[db,df,d_blue,uf,sm,df,df,df],fills=[scen_fl,wf,gld,ufl,wf,wf,wf,wf],
       als=[ca,la,ra,ca,la,ca,ca,ca]); row+=1

# 안전마진 테스트
row+=1; st(ws6,row,"E. 안전마진 테스트",8); row+=1

# Conservative intrinsic value = average of conservative scenarios
conserv_values = [s[2] for s in scenarios if "보수" in s[0]]
conserv_avg = int(statistics.mean(conserv_values)) if conserv_values else PRICE
base_values = [s[2] for s in scenarios if "기본" in s[0]]
base_avg = int(statistics.mean(base_values)) if base_values else PRICE
opt_values = [s[2] for s in scenarios if "낙관" in s[0]]
opt_avg = int(statistics.mean(opt_values)) if opt_values else PRICE

for lbl,iv,verdict in [
    ("보수적 내재가치",conserv_avg,"하방 리스크 확인" if PRICE>conserv_avg else "안전마진 확보"),
    ("기본 내재가치",base_avg,"적정" if abs(PRICE-base_avg)/base_avg<0.15 else ("상승여력" if PRICE<base_avg else "과대평가")),
    ("낙관적 내재가치",opt_avg,"업사이드" if PRICE<opt_avg else "이미 반영"),
]:
    margin = (iv - PRICE) / iv if iv else 0
    mf2 = d_grn if margin>0.15 else (df if margin>0 else d_red)
    mfl = gnf if margin>0.15 else (gld if margin>0 else rdf)
    wr(ws6,row,[lbl,fw(iv),f"안전마진 {margin*100:+.1f}%",verdict,f"현재가 {PRICE:,}원","","",""],
       fonts=[db,d_blue,mf2,mf2,sm,df,df,df],fills=[llf,gld,mfl,mfl,wf,wf,wf,wf],
       als=[la,ra,ca,ca,la,ca,ca,ca]); row+=1

row+=1
ws6.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
verdict_overall = ("기본 시나리오 기준 상승 여력 존재. 보수적 시나리오에서도 큰 하방 리스크 제한적. "
                   f"안전마진: 기본 내재가치({fw(base_avg)}) 대비 {(base_avg-PRICE)/base_avg*100:+.1f}%")
ws6.cell(row=row,column=1,value=f"→ {verdict_overall}").font=db
ws6.cell(row=row,column=1).fill=blf; ws6.cell(row=row,column=1).alignment=la; ws6.cell(row=row,column=1).border=tb

print("  [6/7] 내재가치/안전마진")

# ============================================================
# SHEET 7: 투자판정 & 모니터링
# ============================================================
ws7 = wb.create_sheet("투자판정"); ws7.sheet_properties.tabColor="1ABC9C"
sw(ws7,[22,16,16,16,16,16,16,16])
setup_print(ws7)

row=1
# Title banner
for r in range(1,5):
    for c in range(1,9): ws7.cell(row=r,column=c).fill=tf
ws7.merge_cells('A2:H2')
ws7.cell(row=2,column=1,value="최종 투자 판정").font=Font(name="맑은 고딕",size=22,bold=True,color=W)
ws7.cell(row=2,column=1).alignment=ca
row=6

# 최종 판정
st(ws7,row,"A. 최종 판정",8); row+=1

# Determine verdict
if total_score >= 70:
    verdict = "BUY (매수)"
    verdict_color = GREEN_C
    verdict_fill = gnf
elif total_score >= 55:
    verdict = "HOLD (관망)"
    verdict_color = GOLD_C
    verdict_fill = gld
else:
    verdict = "REJECT (거부)"
    verdict_color = RED_C
    verdict_fill = rdf

ws7.merge_cells(start_row=row,start_column=1,end_row=row+1,end_column=3)
c = ws7.cell(row=row,column=1,value=verdict)
c.font=Font(name="맑은 고딕",size=20,bold=True,color=verdict_color)
c.fill=verdict_fill; c.alignment=ca; c.border=tb

ws7.merge_cells(start_row=row,start_column=4,end_row=row+1,end_column=8)
c = ws7.cell(row=row,column=4)
c.value=(f"Four Filters 종합 {total_score}/100 (등급 {grade})\n"
         f"PER {per_trail:.1f}배(T4Q) | PBR {pbr:.2f}배 | FCF Yield {FCF24/MCAP*100:.1f}% | "
         f"총주주환원 {pct((DIV_PAID+TREAS_BUY)/MCAP)}")
c.font=Font(name="맑은 고딕",size=10,bold=True,color=NAVY)
c.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True); c.border=tb
row+=3

# 판정 근거
st(ws7,row,"B. 판정 근거",8); row+=1
reasons = [
    ("매수 근거 1","전환비용 기반 해자: 삼성·현대 핵심 공급사 지위는 2~3년 인증 과정으로 보호"),
    ("매수 근거 2","적극적 주주환원: 배당 연속인상(500→800원) + 자사주 매입·소각 병행"),
    ("매수 근거 3",f"밸류에이션 매력: PER {per_trail:.1f}배(T4Q), FCF Yield {FCF24/MCAP*100:.1f}% — 한국 부품업 평균 대비 저평가"),
    ("매수 근거 4","전장(자동차) 비중 확대: ADAS 의무화 → 차량당 카메라 수 증가 → 성장 구조적"),
    ("주의 사항 1","삼성전자 매출 편중 ~70% — 갤럭시 부진 시 직접적 타격"),
    ("주의 사항 2","구조적 저마진(OPM 3~5%) — 원가 상승기에 수익성 급락 리스크"),
    ("주의 사항 3","중국 CCM 업체 가격 경쟁 심화 — 중장기 해자 약화 가능성"),
]
for title,detail in reasons:
    is_risk = "주의" in title
    tf2 = d_red if is_risk else d_grn
    tfl = rdf if is_risk else gnf
    ws7.cell(row=row,column=1,value=title).font=tf2; ws7.cell(row=row,column=1).fill=tfl
    ws7.cell(row=row,column=1).alignment=la; ws7.cell(row=row,column=1).border=tb
    ws7.merge_cells(start_row=row,start_column=2,end_row=row,end_column=8)
    ws7.cell(row=row,column=2,value=detail).font=df
    ws7.cell(row=row,column=2).alignment=la; ws7.cell(row=row,column=2).border=tb; row+=1

# SWOT (구루 관점)
row+=1; st(ws7,row,"C. 구루 관점 SWOT",8); row+=1

ws7.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
ws7.cell(row=row,column=1,value="강점 (Strengths)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws7.cell(row=row,column=1).fill=PatternFill("solid",fgColor=GREEN_C); ws7.cell(row=row,column=1).alignment=ca
ws7.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
ws7.cell(row=row,column=5,value="약점 (Weaknesses)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws7.cell(row=row,column=5).fill=PatternFill("solid",fgColor=RED_C); ws7.cell(row=row,column=5).alignment=ca
row+=1

SW_S = ["삼성/현대 핵심 공급사 (전환비용 해자)",
        "오너경영 + 적극적 주주환원",
        "베트남 3공장 원가경쟁력",
        "R&D인력 57.7% + 특허 24건"]
SW_W = ["삼성 매출 편중 70%+ (단일 고객)",
        "조립 중심 저마진 3~5%",
        "스마트폰 성숙기 시장",
        "자체 브랜드 없는 OEM 구조"]
for i in range(4):
    ws7.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
    ws7.cell(row=row,column=1,value=f"  {SW_S[i]}").font=df; ws7.cell(row=row,column=1).fill=gnf
    ws7.cell(row=row,column=1).alignment=la; ws7.cell(row=row,column=1).border=tb
    ws7.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
    ws7.cell(row=row,column=5,value=f"  {SW_W[i]}").font=df; ws7.cell(row=row,column=5).fill=rdf
    ws7.cell(row=row,column=5).alignment=la; ws7.cell(row=row,column=5).border=tb; row+=1

row+=1
ws7.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
ws7.cell(row=row,column=1,value="기회 (Opportunities)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws7.cell(row=row,column=1).fill=PatternFill("solid",fgColor="2980B9"); ws7.cell(row=row,column=1).alignment=ca
ws7.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
ws7.cell(row=row,column=5,value="위협 (Threats)").font=Font(name="맑은 고딕",size=11,bold=True,color=W)
ws7.cell(row=row,column=5).fill=PatternFill("solid",fgColor="7F8C8D"); ws7.cell(row=row,column=5).alignment=ca
row+=1

O = ["ADAS 의무화 → 차량당 카메라 증가",
     "전장 비중 30%+ 시 밸류에이션 리레이팅",
     "프리즘줌·XR 고부가 신사업",
     "밸류업 프로그램 → 주주환원 강화"]
T = ["중국 CCM 업체 가격 경쟁",
     "삼성 모듈 내재화 가능성",
     "환율/인건비 상승 압박",
     "글로벌 경기침체 → 스마트폰 수요 감소"]
for i in range(4):
    ws7.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
    ws7.cell(row=row,column=1,value=f"  {O[i]}").font=df; ws7.cell(row=row,column=1).fill=blf
    ws7.cell(row=row,column=1).alignment=la; ws7.cell(row=row,column=1).border=tb
    ws7.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
    ws7.cell(row=row,column=5,value=f"  {T[i]}").font=df
    ws7.cell(row=row,column=5).fill=PatternFill("solid",fgColor="E5E7E9")
    ws7.cell(row=row,column=5).alignment=la; ws7.cell(row=row,column=5).border=tb; row+=1

# 모니터링 지표
row+=1; st(ws7,row,"D. 핵심 모니터링 지표 (5개)",8); row+=1
wh(ws7,row,["#","지표","세부 내용","확인 시기","구루 관점","","",""]); row+=1
monitors = [
    ("1","삼성 갤럭시 카메라 사양","화소·개수·OIS/줌이 ASP와 매출 직접 결정","1~2월, 7~8월","해자 유지 여부 확인"),
    ("2","전장 매출 비중","30%+ 넘으면 밸류에이션 리레이팅. ADAS 채택률","매 분기","해자 확대 여부"),
    ("3","영업이익률 추이","3% 미만=경고, 4~5%=양호, 6%+=호황","분기 실적","수익성 품질"),
    ("4","자사주/배당 정책","매입·소각·배당인상 지속 여부","3월 주총, 수시","경영진 신뢰"),
    ("5","중국 업체 동향","LG이노텍/삼성전기 대비 점유율 변화","반기","해자 위협"),
]
for num,title,detail,timing,guru in monitors:
    wr(ws7,row,[num,title,detail,timing,guru,"","",""],
       fonts=[Font(name="맑은 고딕",size=12,bold=True,color=NAVY),db,df,df,sm,df,df,df],
       fills=[gld,llf,wf,llf,wf,wf,wf,wf],als=[ca,la,la,ca,la,ca,ca,ca])
    ws7.row_dimensions[row].height=40; row+=1

# 매도 트리거
row+=1; st(ws7,row,"E. 매도 트리거 조건",8); row+=1
sell_triggers = [
    ("삼성전자 CCM 공급사 변경 or 내재화 본격화","해자 붕괴 → 즉시 매도"),
    ("영업이익률 2년 연속 2% 미만","구조적 수익성 훼손"),
    ("ROE 3년 연속 5% 미만","자본비용 미달 → 가치 파괴"),
    ("유상증자/CB 발행 (비합리적 사유)","주주 희석 → 경영진 신뢰 하락"),
    ("PER 15배+ (한국 부품업 기준 과열)","안전마진 소멸"),
]
for trigger,action in sell_triggers:
    ws7.merge_cells(start_row=row,start_column=1,end_row=row,end_column=5)
    ws7.cell(row=row,column=1,value=f"  {trigger}").font=df
    ws7.cell(row=row,column=1).fill=rdf; ws7.cell(row=row,column=1).alignment=la; ws7.cell(row=row,column=1).border=tb
    ws7.merge_cells(start_row=row,start_column=6,end_row=row,end_column=8)
    ws7.cell(row=row,column=6,value=action).font=d_red
    ws7.cell(row=row,column=6).alignment=la; ws7.cell(row=row,column=6).border=tb; row+=1

# 리스크 시나리오
row+=1; st(ws7,row,"F. 리스크 시나리오 (최악의 경우)",8); row+=1
ws7.merge_cells(start_row=row,start_column=1,end_row=row+3,end_column=8)
c = ws7.cell(row=row,column=1)
c.value=("최악 시나리오: 삼성전자가 CCM 모듈 내재화를 본격 추진 + 중국 업체가 전장 카메라 시장 진입\n"
         "→ 모바일 매출 50% 감소 + 전장 성장 정체\n"
         f"→ 매출 5,000억, 영업이익 100억(OPM 2%), 순이익 80억, EPS ~450원\n"
         f"→ PER 7배 적용 시 주가 ~3,150원 (현재가 대비 -89%). 이 시나리오 확률: 5% 미만")
c.font=Font(name="맑은 고딕",size=10,color=RED_C)
c.fill=PatternFill("solid",fgColor="FDE8E8")
c.alignment=Alignment(horizontal='left',vertical='top',wrap_text=True); c.border=tb

print("  [7/7] 투자판정/모니터링")

# === SAVE ===
OUT = os.path.join(BASE, "엠씨넥스_투자구루분석.xlsx")
wb.save(OUT)
conn.close()
print(f"\n투자 구루 분석 보고서 생성 완료: {OUT}")
print(f"시트 ({len(wb.sheetnames)}개): {wb.sheetnames}")
