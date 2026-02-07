# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "엠씨넥스_밸류에이션.xlsx")
wb = Workbook()

# === CONSTANTS ===
PRICE = 28100
SHARES_OUTSTANDING = 17977732  # 발행-소각
SHARES_TREASURY = 1110000
SHARES_WA = 17373105  # 가중평균 (2024 EPS 역산)
MARKET_CAP = PRICE * SHARES_OUTSTANDING  # 505,174M

# 2024 Annual
REV_2024 = 1057058423929
OP_2024 = 44384671816
NI_2024 = 63604930146
EPS_2024 = 3661
EQUITY_2024 = 372224146153
EQUITY_2023 = 323612702499
ASSETS_2024 = 574103551032
LIAB_2024 = 201879404879
CASH_2024 = 35881602429
ST_DEBT_2024 = 44402388000
LT_DEBT_2024 = 1308000000
DA_2024 = 44255618814 + 462932726 + 1740746569  # 감가상각+투자부동산+무형자산
OPCF_2024 = 82470298068
CAPEX_2024 = 36541908724
DIV_PAID_2024 = 10570639200
TREASURY_BUY_2024 = 15656989850
DPS_2024 = 800

# 2024 Quarterly
Q1_24 = {"rev": 287835057669, "op": 15204634567, "ni": 15929655236}
Q2_24 = {"rev": 225759412794, "op": 7650453061, "ni": 18422813873}
Q3_24 = {"rev": 245070688152, "op": 3077406264, "ni": 2504912286}
Q4_24_rev = REV_2024 - Q1_24["rev"] - Q2_24["rev"] - Q3_24["rev"]
Q4_24_op = OP_2024 - Q1_24["op"] - Q2_24["op"] - Q3_24["op"]
Q4_24_ni = NI_2024 - Q1_24["ni"] - Q2_24["ni"] - Q3_24["ni"]
Q4_24 = {"rev": Q4_24_rev, "op": Q4_24_op, "ni": Q4_24_ni}

# 2025 Quarterly (잠정)
Q1_25 = {"rev": 347917e6, "op": 20918e6, "ni": 20068e6}
Q2_25 = {"rev": 326424e6, "op": 16171e6, "ni": 13050e6}
Q3_25 = {"rev": 303537e6, "op": 6708e6, "ni": 9537e6}
CUM_25_9M = {"rev": 977877e6, "op": 43797e6, "ni": 42656e6}

# Historical
HIST = {
    2019: {"rev": 1267712135419, "op": 113067104268, "ni": 84507307910, "eps": 5151, "equity": 224567081980},
    2020: {"rev": 1311342738149, "op": 59167161589, "ni": 38443583445, "eps": 2180, "equity": 252934743419},
    2021: {"rev": 1009273639991, "op": 23878951946, "ni": 39505208461, "eps": 2229, "equity": 298966196208},
    2022: {"rev": 1108637950325, "op": 10702196295, "ni": 22973446783, "eps": 1293, "equity": 309066577971},
    2023: {"rev": 932490110758, "op": 18224255538, "ni": 27923690827, "eps": 1585, "equity": 323612702499},
    2024: {"rev": REV_2024, "op": OP_2024, "ni": NI_2024, "eps": EPS_2024, "equity": EQUITY_2024},
}

억 = 100_000_000

# === STYLES ===
NAVY = "1B2A4A"
DARK = "2C3E6B"
MID = "3A5BA0"
WHITE = "FFFFFF"
GOLD_BG = "FFF3CD"
GREEN_BG = "D4EDDA"
RED_BG = "F8D7DA"
BLUE_BG = "D6E4F0"
LBLUE = "EBF1F8"

title_font = Font(name="맑은 고딕", size=18, bold=True, color=WHITE)
section_font = Font(name="맑은 고딕", size=13, bold=True, color=NAVY)
h_font = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
d_font = Font(name="맑은 고딕", size=10)
d_bold = Font(name="맑은 고딕", size=10, bold=True)
d_blue = Font(name="맑은 고딕", size=10, bold=True, color="0000FF")
d_green = Font(name="맑은 고딕", size=10, bold=True, color="27AE60")
d_red = Font(name="맑은 고딕", size=10, bold=True, color="C0392B")
note_font = Font(name="맑은 고딕", size=9, color="666666")
big_font = Font(name="맑은 고딕", size=14, bold=True, color=NAVY)
huge_font = Font(name="맑은 고딕", size=20, bold=True, color=NAVY)

title_fill = PatternFill("solid", fgColor=NAVY)
header_fill = PatternFill("solid", fgColor=DARK)
mid_fill = PatternFill("solid", fgColor=MID)
blue_fill = PatternFill("solid", fgColor=BLUE_BG)
lblue_fill = PatternFill("solid", fgColor=LBLUE)
gold_fill = PatternFill("solid", fgColor=GOLD_BG)
green_fill = PatternFill("solid", fgColor=GREEN_BG)
red_fill = PatternFill("solid", fgColor=RED_BG)
white_fill = PatternFill("solid", fgColor=WHITE)
gray_fill = PatternFill("solid", fgColor="F2F2F2")

center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")
thin_border = Border(
    left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"))
bottom_border = Border(bottom=Side(style="medium", color=NAVY))


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def setup_print(ws):
    """Letter 용지 가로(Landscape) 인쇄 최적화 설정"""
    ws.page_setup.paperSize = Worksheet.PAPERSIZE_LETTER
    ws.page_setup.orientation = Worksheet.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.5, bottom=0.5,
        header=0.3, footer=0.3
    )
    ws.print_options.horizontalCentered = True

def write_header(ws, row, vals, fills=None):
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = h_font
        c.fill = fills[i-1] if fills else header_fill
        c.alignment = center
        c.border = thin_border

def write_row(ws, row, vals, fonts=None, fills=None, aligns=None):
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = fonts[i-1] if fonts else d_font
        c.fill = fills[i-1] if fills else white_fill
        c.alignment = aligns[i-1] if aligns else center
        c.border = thin_border

def section_title(ws, row, title, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=title)
    c.font = section_font
    c.border = bottom_border
    return row + 1

def fmt(v):
    """숫자를 억원 문자열로"""
    return f"{v/억:,.0f}억"

def fmt_won(v):
    return f"{v:,.0f}원"

def pct(v):
    return f"{v*100:.1f}%"


# === DERIVED CALCULATIONS ===
BPS = EQUITY_2024 / SHARES_OUTSTANDING  # 20,706원
NET_DEBT = ST_DEBT_2024 + LT_DEBT_2024 - CASH_2024
EBITDA_2024 = OP_2024 + DA_2024
FCF_2024 = OPCF_2024 - CAPEX_2024
EV = MARKET_CAP + NET_DEBT

# Trailing 4Q
TRAIL_NI = Q4_24["ni"] + Q1_25["ni"] + Q2_25["ni"] + Q3_25["ni"]
TRAIL_OP = Q4_24["op"] + Q1_25["op"] + Q2_25["op"] + Q3_25["op"]
TRAIL_REV = Q4_24["rev"] + Q1_25["rev"] + Q2_25["rev"] + Q3_25["rev"]
TRAIL_EPS = TRAIL_NI / SHARES_WA

# 2025E variants
E25_ANNUAL_NI = CUM_25_9M["ni"] * 4 / 3
E25_ANNUAL_EPS = E25_ANNUAL_NI / SHARES_WA
E25_Q4_LIKE_24Q4 = CUM_25_9M["ni"] + Q4_24["ni"]
E25_EPS_LIKE_24Q4 = E25_Q4_LIKE_24Q4 / SHARES_WA
E25_CONSERVATIVE_Q4 = (Q2_25["ni"] + Q3_25["ni"]) / 2
E25_CONSERVATIVE = CUM_25_9M["ni"] + E25_CONSERVATIVE_Q4
E25_CONSERVATIVE_EPS = E25_CONSERVATIVE / SHARES_WA

# ROE
AVG_EQ = (EQUITY_2024 + EQUITY_2023) / 2
ROE_2024 = NI_2024 / AVG_EQ


# ============================================================
# SHEET 1: 밸류에이션 종합 대시보드
# ============================================================
ws1 = wb.active
ws1.title = "종합"
ws1.sheet_properties.tabColor = "1B2A4A"
set_widths(ws1, [22, 18, 18, 18, 18, 22])
setup_print(ws1)

# Title bar
row = 1
ws1.merge_cells('A1:F2')
c = ws1.cell(row=1, column=1, value="엠씨넥스(097520) 밸류에이션 분석")
c.font = title_font
c.fill = title_fill
c.alignment = Alignment(horizontal="center", vertical="center")

row = 3
ws1.merge_cells('A3:F3')
ws1.cell(row=3, column=1, value=f"기준일: 2026.02.06 | 현재가: {PRICE:,}원 | 시가총액: {fmt(MARKET_CAP)}").font = Font(name="맑은 고딕", size=11, bold=True, color=DARK)
ws1.cell(row=3, column=1).alignment = center

# Key metrics box
row = 5
row = section_title(ws1, row, "기본 정보", 6)
info_data = [
    ["현재 주가", fmt_won(PRICE), "발행주식수", f"{SHARES_OUTSTANDING:,}주", "시가총액", fmt(MARKET_CAP)],
    ["자기주식", f"{SHARES_TREASURY:,}주", "유통주식수", f"{SHARES_OUTSTANDING - SHARES_TREASURY:,}주", "유통시총", fmt(PRICE * (SHARES_OUTSTANDING - SHARES_TREASURY))],
    ["2024 EPS", fmt_won(EPS_2024), "2024 BPS", fmt_won(int(BPS)), "2024 DPS", fmt_won(DPS_2024)],
    ["2024 매출", fmt(REV_2024), "2024 영업이익", fmt(OP_2024), "2024 순이익", fmt(NI_2024)],
]
for data in info_data:
    for i in range(0, 6, 2):
        ws1.cell(row=row, column=i+1, value=data[i]).font = d_bold
        ws1.cell(row=row, column=i+1).fill = blue_fill
        ws1.cell(row=row, column=i+1).alignment = left
        ws1.cell(row=row, column=i+1).border = thin_border
        ws1.cell(row=row, column=i+2, value=data[i+1]).font = d_blue
        ws1.cell(row=row, column=i+2).fill = white_fill
        ws1.cell(row=row, column=i+2).alignment = center
        ws1.cell(row=row, column=i+2).border = thin_border
    row += 1

# Valuation Summary
row += 1
row = section_title(ws1, row, "밸류에이션 멀티플 종합", 6)
write_header(ws1, row, ["지표", "산출 방식", "값", "판정", "업종 평균(참고)", "비고"])
row += 1

valuation_rows = [
    ["PER (2024)", f"주가 / EPS({fmt_won(EPS_2024)})", f"{PRICE/EPS_2024:.1f}배", "저평가", "10~15배", "확정 실적 기준"],
    ["PER (Trailing 4Q)", f"주가 / T4Q EPS({fmt_won(int(TRAIL_EPS))})", f"{PRICE/TRAIL_EPS:.1f}배", "저평가", "10~15배", "24Q4+25Q1~Q3"],
    ["PER (2025E 단순연환산)", f"주가 / 연환산EPS({fmt_won(int(E25_ANNUAL_EPS))})", f"{PRICE/E25_ANNUAL_EPS:.1f}배", "적정~저평가", "10~15배", "9M×4/3"],
    ["PER (2025E 보수적)", f"주가 / EPS({fmt_won(int(E25_CONSERVATIVE_EPS))})", f"{PRICE/E25_CONSERVATIVE_EPS:.1f}배", "적정", "10~15배", "Q4=Q2Q3평균"],
    ["PBR", f"주가 / BPS({fmt_won(int(BPS))})", f"{PRICE/BPS:.2f}배", "적정", "1.0~2.0배", "2024말 자본 기준"],
    ["EV/EBITDA (2024)", f"EV({fmt(EV)}) / EBITDA({fmt(EBITDA_2024)})", f"{EV/EBITDA_2024:.1f}배", "저평가", "7~12배", ""],
    ["PSR (2024)", f"시총 / 매출({fmt(REV_2024)})", f"{MARKET_CAP/REV_2024:.2f}배", "매우 저평가", "0.5~1.5배", "매출 1조 기업"],
    ["PSR (Trailing)", f"시총 / T4Q매출({fmt(TRAIL_REV)})", f"{MARKET_CAP/TRAIL_REV:.2f}배", "매우 저평가", "0.5~1.5배", "직전4분기"],
    ["PCR", f"주가 / OpCF/주({fmt_won(int(OPCF_2024/SHARES_OUTSTANDING))})", f"{PRICE/(OPCF_2024/SHARES_OUTSTANDING):.1f}배", "저평가", "8~15배", "영업CF 기준"],
    ["배당수익률", f"DPS({fmt_won(DPS_2024)}) / 주가", pct(DPS_2024/PRICE), "양호", "1~3%", "2024년 배당"],
    ["FCF 수익률", f"FCF/주({fmt_won(int(FCF_2024/SHARES_OUTSTANDING))}) / 주가", pct(FCF_2024/SHARES_OUTSTANDING/PRICE), "매우 양호", "3~6%", ""],
    ["총주주환원율", f"배당+자사주({fmt(DIV_PAID_2024+TREASURY_BUY_2024)}) / 시총", pct((DIV_PAID_2024+TREASURY_BUY_2024)/MARKET_CAP), "양호", "2~5%", ""],
    ["ROE (2024)", f"순이익 / 평균자본", pct(ROE_2024), "양호", "8~15%", "2023-2024 평균자본"],
]

for vals in valuation_rows:
    judge = vals[3]
    if "매우 저" in judge or "매우 양" in judge:
        judge_font = d_green
        judge_fill = green_fill
    elif "저평가" in judge or "양호" in judge:
        judge_font = Font(name="맑은 고딕", size=10, bold=True, color="2E86C1")
        judge_fill = PatternFill("solid", fgColor="D6EAF8")
    elif "적정" in judge:
        judge_font = d_bold
        judge_fill = gold_fill
    else:
        judge_font = d_red
        judge_fill = red_fill

    write_row(ws1, row, vals,
              fonts=[d_bold, d_font, d_blue, judge_font, d_font, note_font],
              fills=[lblue_fill, white_fill, gold_fill, judge_fill, gray_fill, white_fill],
              aligns=[left, left, center, center, center, left])
    row += 1

row += 1
ws1.cell(row=row, column=1, value="* 판정 기준: 업종(전자부품/카메라모듈) 평균 대비 상대 평가").font = note_font
row += 1
ws1.cell(row=row, column=1, value="* EPS 가중평균주식수: 17,373,105주 (2024 사업보고서 기준)").font = note_font

print("  [1/5] 종합 대시보드 완료")


# ============================================================
# SHEET 2: PER 다각도 분석
# ============================================================
ws2 = wb.create_sheet("PER분석")
ws2.sheet_properties.tabColor = "2980B9"
set_widths(ws2, [24, 16, 16, 16, 16, 20])
setup_print(ws2)

row = 1
row = section_title(ws2, row, "PER 다각도 분석 (현재가 28,100원 기준)", 6)

# A. EPS 산출 방식별 PER
row = section_title(ws2, row, "A. EPS 산출 방식별 PER", 6)
write_header(ws2, row, ["산출 방식", "순이익(억)", "EPS(원)", "PER(배)", "의미", "비고"])
row += 1

per_methods = [
    ["2024 확정 (사업보고서)", int(NI_2024/억), EPS_2024, PRICE/EPS_2024,
     "가장 신뢰 높은 확정치", "reprt_code 11011"],
    ["Trailing 4Q (24Q4+25Q1~Q3)", int(TRAIL_NI/억), int(TRAIL_EPS), PRICE/TRAIL_EPS,
     "직전 4분기 실적 합산", "가장 최신 실적 반영"],
    ["2025E 연환산 (9M×4/3)", int(E25_ANNUAL_NI/억), int(E25_ANNUAL_EPS), PRICE/E25_ANNUAL_EPS,
     "단순 연율화", "Q4 계절성 미반영"],
    ["2025E (9M+24Q4 대입)", int(E25_Q4_LIKE_24Q4/억), int(E25_EPS_LIKE_24Q4), PRICE/E25_EPS_LIKE_24Q4,
     "전년Q4 실적 대입", "Q4 통상 강세 반영"],
    ["2025E 보수적 (Q4=Q2Q3평균)", int(E25_CONSERVATIVE/억), int(E25_CONSERVATIVE_EPS), PRICE/E25_CONSERVATIVE_EPS,
     "하반기 둔화 반영", "보수적 추정"],
]

for vals in per_methods:
    per_val = vals[3]
    if per_val < 7.5:
        per_font = d_green
        per_fill = green_fill
    elif per_val < 9:
        per_font = d_blue
        per_fill = PatternFill("solid", fgColor="D6EAF8")
    else:
        per_font = d_bold
        per_fill = gold_fill

    write_row(ws2, row,
              [vals[0], vals[1], f"{vals[2]:,}", f"{per_val:.2f}배", vals[4], vals[5]],
              fonts=[d_bold, d_font, d_blue, per_font, d_font, note_font],
              fills=[lblue_fill, white_fill, gold_fill, per_fill, white_fill, gray_fill],
              aligns=[left, right, right, center, left, left])
    row += 1

# B. 분기별 실적 추이
row += 1
row = section_title(ws2, row, "B. 분기별 실적 추이 (백만원)", 6)
write_header(ws2, row, ["분기", "매출액", "영업이익", "OPM", "순이익", "YoY 순이익"])
row += 1

quarters = [
    ("24Q1", Q1_24), ("24Q2", Q2_24), ("24Q3", Q3_24), ("24Q4", Q4_24),
    ("25Q1", Q1_25), ("25Q2", Q2_25), ("25Q3", Q3_25),
]
prev_year_q = {"25Q1": Q1_24, "25Q2": Q2_24, "25Q3": Q3_24}

for name, q in quarters:
    rev_m = int(q["rev"] / 1e6)
    op_m = int(q["op"] / 1e6)
    ni_m = int(q["ni"] / 1e6)
    opm = q["op"] / q["rev"] if q["rev"] else 0
    yoy = ""
    if name in prev_year_q:
        prev = prev_year_q[name]
        if prev["ni"] > 0:
            yoy_val = (q["ni"] - prev["ni"]) / prev["ni"]
            yoy = f"{yoy_val*100:+.1f}%"

    is_2025 = name.startswith("25")
    row_fill = green_fill if is_2025 else white_fill
    opm_font = d_green if opm > 0.05 else (d_red if opm < 0.03 else d_font)

    write_row(ws2, row, [name, f"{rev_m:,}", f"{op_m:,}", f"{opm*100:.1f}%", f"{ni_m:,}", yoy],
              fonts=[d_bold, d_font, d_font, opm_font, d_font, d_bold],
              fills=[blue_fill if is_2025 else lblue_fill, row_fill, row_fill, row_fill, row_fill, row_fill],
              aligns=[center, right, right, center, right, center])
    row += 1

# Trailing 4Q summary row
row_fill = gold_fill
write_row(ws2, row, ["Trailing 4Q 합계",
                      f"{int(TRAIL_REV/1e6):,}", f"{int(TRAIL_OP/1e6):,}",
                      f"{TRAIL_OP/TRAIL_REV*100:.1f}%",
                      f"{int(TRAIL_NI/1e6):,}", ""],
          fonts=[d_bold, d_bold, d_bold, d_bold, d_bold, d_font],
          fills=[gold_fill]*6, aligns=[center, right, right, center, right, center])

# C. Historical PER
row += 2
row = section_title(ws2, row, "C. 역사적 EPS & 적정주가 (목표 PER 적용)", 6)
write_header(ws2, row, ["기준", "EPS(원)", "PER 7배", "PER 8배", "PER 10배", "PER 12배"])
row += 1

eps_scenarios = [
    ("2024 확정", EPS_2024),
    ("Trailing 4Q", int(TRAIL_EPS)),
    ("2025E (24Q4대입)", int(E25_EPS_LIKE_24Q4)),
    ("2025E 보수적", int(E25_CONSERVATIVE_EPS)),
]

for label, eps in eps_scenarios:
    write_row(ws2, row,
              [label, f"{eps:,}", fmt_won(eps*7), fmt_won(eps*8), fmt_won(eps*10), fmt_won(eps*12)],
              fonts=[d_bold, d_blue, d_font, d_font, d_green, d_green],
              fills=[lblue_fill, gold_fill, white_fill, green_fill if eps*8 > PRICE else white_fill,
                     green_fill, green_fill],
              aligns=[left, right, right, right, right, right])
    row += 1

row += 1
ws2.cell(row=row, column=1, value="→ 현재 28,100원은 Trailing EPS 기준 PER 7배 수준. 목표 PER 10배 적용 시 적정가 약 36,000~40,000원").font = d_bold

print("  [2/5] PER 분석 완료")


# ============================================================
# SHEET 3: PBR/ROE/RIM 분석
# ============================================================
ws3 = wb.create_sheet("PBR_ROE")
ws3.sheet_properties.tabColor = "8E44AD"
set_widths(ws3, [20, 16, 16, 16, 16, 22])
setup_print(ws3)

row = 1
row = section_title(ws3, row, "PBR / ROE / 잔여이익모델(RIM) 분석", 6)

# A. Historical BPS & PBR
row = section_title(ws3, row, "A. 연도별 자본/BPS 추이", 6)
write_header(ws3, row, ["연도", "자본(억)", "BPS(원)", "ROE", "EPS(원)", "순이익(억)"])
row += 1

prev_eq = None
for yr in [2019, 2020, 2021, 2022, 2023, 2024]:
    h = HIST[yr]
    eq = h["equity"]
    bps = int(eq / SHARES_OUTSTANDING)
    ni = h["ni"]
    eps = h["eps"]
    if prev_eq:
        roe = ni / ((eq + prev_eq) / 2)
    else:
        roe = ni / eq
    prev_eq = eq

    roe_font = d_green if roe > 0.12 else (d_red if roe < 0.05 else d_font)
    write_row(ws3, row,
              [str(yr), int(eq/억), f"{bps:,}", f"{roe*100:.1f}%", f"{eps:,}", int(ni/억)],
              fonts=[d_bold, d_font, d_blue, roe_font, d_font, d_font],
              fills=[lblue_fill, white_fill, gold_fill, white_fill, white_fill, white_fill],
              aligns=[center, right, right, center, right, right])
    row += 1

# Current PBR
row += 1
row = section_title(ws3, row, "B. 현재 PBR 분석", 6)
write_header(ws3, row, ["항목", "값", "", "", "", ""])
row += 1

pbr_info = [
    ("2024말 자본총계 (지배)", fmt(EQUITY_2024)),
    ("BPS (지배지분/발행주식)", fmt_won(int(BPS))),
    ("현재 PBR", f"{PRICE/BPS:.2f}배"),
    ("PBR 1.0배 주가", fmt_won(int(BPS))),
    ("PBR 1.5배 주가", fmt_won(int(BPS * 1.5))),
    ("PBR 2.0배 주가", fmt_won(int(BPS * 2.0))),
]
for label, val in pbr_info:
    ws3.cell(row=row, column=1, value=label).font = d_bold
    ws3.cell(row=row, column=1).fill = lblue_fill
    ws3.cell(row=row, column=1).alignment = left
    ws3.cell(row=row, column=1).border = thin_border
    ws3.cell(row=row, column=2, value=val).font = d_blue
    ws3.cell(row=row, column=2).fill = gold_fill
    ws3.cell(row=row, column=2).alignment = center
    ws3.cell(row=row, column=2).border = thin_border
    row += 1

# C. RIM (Residual Income Model)
row += 1
row = section_title(ws3, row, "C. 잔여이익모델(RIM) 적정주가", 6)
ws3.cell(row=row, column=1, value="산식: 적정가 = BPS + BPS × (ROE - ke) / (ke - g)").font = note_font
row += 1
ws3.cell(row=row, column=1, value="ke(자기자본비용) = 무위험이자율 3.5% + β(1.0) × ERP(6.5%) = 10.0%").font = note_font
row += 1

write_header(ws3, row, ["시나리오", "지속ROE", "ke", "성장률(g)", "적정주가", "현재가 대비"])
row += 1

rim_scenarios = [
    ("보수적 (ROE=ke)", 0.10, 0.10, 0.02, "ROE가 자본비용과 동일"),
    ("기본 (과거평균)", 0.12, 0.10, 0.02, "2019~2024 평균 ROE ~12%"),
    ("적극적 (최근수준)", 0.15, 0.10, 0.02, "2024년 ROE 18% 할인 적용"),
    ("낙관적 (ROE유지)", 0.18, 0.10, 0.03, "2024년 ROE 유지 가정"),
]

for label, roe, ke, g, note in rim_scenarios:
    fair = BPS * (1 + (roe - ke) / (ke - g))
    upside = (fair - PRICE) / PRICE
    upside_str = f"{upside*100:+.1f}%"
    up_font = d_green if upside > 0 else d_red
    up_fill = green_fill if upside > 0 else red_fill

    write_row(ws3, row,
              [label, f"{roe*100:.0f}%", f"{ke*100:.0f}%", f"{g*100:.0f}%", fmt_won(int(fair)), upside_str],
              fonts=[d_bold, d_font, d_font, d_font, d_blue, up_font],
              fills=[lblue_fill, white_fill, white_fill, white_fill, gold_fill, up_fill],
              aligns=[left, center, center, center, right, center])
    row += 1

row += 1
ws3.cell(row=row, column=1, value="→ 기본 시나리오(ROE 12%) 적정가 약 25,900원. 적극적(15%) 시 33,600원. 현재가 28,100원은 ROE 12~15% 사이 반영.").font = d_bold

print("  [3/5] PBR/ROE 완료")


# ============================================================
# SHEET 4: EV/EBITDA & FCF
# ============================================================
ws4 = wb.create_sheet("EV_EBITDA_FCF")
ws4.sheet_properties.tabColor = "E67E22"
set_widths(ws4, [24, 18, 18, 18, 24])
setup_print(ws4)

row = 1
row = section_title(ws4, row, "EV/EBITDA & FCF 밸류에이션", 5)

# A. EV 산출
row = section_title(ws4, row, "A. Enterprise Value 산출", 5)
write_header(ws4, row, ["항목", "금액(억)", "비고", "", ""])
row += 1

ev_items = [
    ("시가총액", int(MARKET_CAP/억), f"주가 {PRICE:,}원 × {SHARES_OUTSTANDING:,}주"),
    ("(+) 총차입금", int((ST_DEBT_2024+LT_DEBT_2024)/억), f"단기 {int(ST_DEBT_2024/억)}억 + 장기 {int(LT_DEBT_2024/억)}억"),
    ("(-) 현금성자산", int(CASH_2024/억), "현금및현금성자산"),
    ("(=) 순차입금", int(NET_DEBT/억), "차입금 - 현금"),
    ("(=) EV", int(EV/억), "시가총액 + 순차입금"),
]
for label, amt, note in ev_items:
    is_total = label.startswith("(=)")
    write_row(ws4, row, [label, f"{amt:,}", note, "", ""],
              fonts=[d_bold if is_total else d_font, d_blue if is_total else d_font, note_font, d_font, d_font],
              fills=[gold_fill if is_total else lblue_fill, gold_fill if is_total else white_fill, white_fill, white_fill, white_fill],
              aligns=[left, right, left, left, left])
    row += 1

# B. EBITDA
row += 1
row = section_title(ws4, row, "B. EBITDA 산출", 5)
write_header(ws4, row, ["항목", "2024(억)", "비고", "", ""])
row += 1

ebitda_items = [
    ("영업이익", int(OP_2024/억), "CIS 영업이익(손실)"),
    ("(+) 감가상각비", int(44255618814/억), "유형자산"),
    ("(+) 투자부동산감가상각", int(462932726/억), ""),
    ("(+) 무형자산상각비", int(1740746569/억), ""),
    ("(=) EBITDA", int(EBITDA_2024/억), ""),
    ("EBITDA 마진", "", f"{EBITDA_2024/REV_2024*100:.1f}%"),
]
for label, amt, note in ebitda_items:
    is_total = label.startswith("(=)") or label == "EBITDA 마진"
    val = f"{amt:,}" if isinstance(amt, int) else note
    write_row(ws4, row, [label, val if isinstance(amt, int) else "", note if isinstance(amt, int) else val, "", ""],
              fonts=[d_bold if is_total else d_font, d_blue if is_total else d_font, note_font, d_font, d_font],
              fills=[gold_fill if is_total else lblue_fill, gold_fill if is_total else white_fill, white_fill, white_fill, white_fill],
              aligns=[left, right, left, left, left])
    row += 1

# C. EV/EBITDA
row += 1
row = section_title(ws4, row, "C. EV/EBITDA 밸류에이션", 5)
write_header(ws4, row, ["기준", "EBITDA(억)", "EV/EBITDA", "적정EV(8배)", "적정주가(8배)"])
row += 1

trail_ebitda_est = TRAIL_OP + DA_2024  # trailing OP + 2024 D&A 유지 가정
e25_ebitda_ann = (CUM_25_9M["op"] * 4/3) + DA_2024

ev_scenarios = [
    ("2024 확정", int(EBITDA_2024/억)),
    ("Trailing 4Q (추정)", int(trail_ebitda_est/억)),
    ("2025E 연환산 (추정)", int(e25_ebitda_ann/억)),
]

for label, ebitda_b in ev_scenarios:
    ev_ebitda = (EV/억) / ebitda_b
    fair_ev_8x = ebitda_b * 8
    fair_eq = (fair_ev_8x - int(NET_DEBT/억)) * 억
    fair_price = int(fair_eq / SHARES_OUTSTANDING)

    write_row(ws4, row,
              [label, f"{ebitda_b:,}", f"{ev_ebitda:.1f}배", f"{fair_ev_8x:,}억", fmt_won(fair_price)],
              fonts=[d_bold, d_font, d_blue, d_font, d_green],
              fills=[lblue_fill, white_fill, gold_fill, white_fill, green_fill],
              aligns=[left, right, center, right, right])
    row += 1

# D. FCF 분석
row += 1
row = section_title(ws4, row, "D. FCF(잉여현금흐름) 분석", 5)
write_header(ws4, row, ["항목", "2024(억)", "비고", "", ""])
row += 1

fcf_items = [
    ("영업활동현금흐름", int(OPCF_2024/억), ""),
    ("(-) 설비투자(CAPEX)", int(CAPEX_2024/억), "유형자산 취득"),
    ("(=) FCF", int(FCF_2024/억), ""),
    ("FCF/주", "", fmt_won(int(FCF_2024/SHARES_OUTSTANDING))),
    ("FCF 수익률", "", f"{FCF_2024/SHARES_OUTSTANDING/PRICE*100:.1f}%"),
    ("", "", ""),
    ("배당금 지급", int(DIV_PAID_2024/억), fmt_won(DPS_2024) + "/주"),
    ("자기주식 취득", int(TREASURY_BUY_2024/억), "2024년 취득분"),
    ("총 주주환원", int((DIV_PAID_2024+TREASURY_BUY_2024)/억), ""),
    ("총주주환원율", "", f"{(DIV_PAID_2024+TREASURY_BUY_2024)/MARKET_CAP*100:.1f}%"),
]

for label, amt, note in fcf_items:
    if not label:
        row += 1
        continue
    is_total = "(=)" in label or "수익률" in label or "환원율" in label or "총" in label
    val_str = f"{amt:,}" if isinstance(amt, int) and amt > 0 else note
    write_row(ws4, row, [label, val_str if isinstance(amt, int) and amt > 0 else "",
                          note if isinstance(amt, int) and amt > 0 else val_str, "", ""],
              fonts=[d_bold if is_total else d_font, d_blue if is_total else d_font, note_font, d_font, d_font],
              fills=[gold_fill if is_total else lblue_fill, gold_fill if is_total else white_fill, white_fill, white_fill, white_fill],
              aligns=[left, right, left, left, left])
    row += 1

row += 1
ws4.cell(row=row, column=1, value="→ FCF 수익률 9.1%로 매우 양호. 시총 5,052억 대비 연간 460억 잉여현금 창출.").font = d_bold

print("  [4/5] EV/EBITDA/FCF 완료")


# ============================================================
# SHEET 5: 시나리오별 목표주가 종합
# ============================================================
ws5 = wb.create_sheet("목표주가")
ws5.sheet_properties.tabColor = "2ECC71"
set_widths(ws5, [22, 18, 18, 18, 22])
setup_print(ws5)

row = 1
row = section_title(ws5, row, "시나리오별 목표주가 종합 (현재가 28,100원)", 5)

# A. 방법론별 적정가 범위
row = section_title(ws5, row, "A. 밸류에이션 방법론별 적정가 레인지", 5)
write_header(ws5, row, ["방법론", "보수적", "기본", "적극적", "산출 근거"])
row += 1

methods = [
    ("PER 방식", fmt_won(int(E25_CONSERVATIVE_EPS*7)),
     fmt_won(int(TRAIL_EPS*10)), fmt_won(int(TRAIL_EPS*12)),
     "EPS×목표PER (7/10/12배)"),
    ("PBR 방식", fmt_won(int(BPS*1.0)),
     fmt_won(int(BPS*1.5)), fmt_won(int(BPS*2.0)),
     "BPS×목표PBR (1.0/1.5/2.0배)"),
    ("EV/EBITDA 방식", fmt_won(int((int(EBITDA_2024/억)*6-int(NET_DEBT/억))*억/SHARES_OUTSTANDING)),
     fmt_won(int((int(EBITDA_2024/억)*8-int(NET_DEBT/억))*억/SHARES_OUTSTANDING)),
     fmt_won(int((int(EBITDA_2024/억)*10-int(NET_DEBT/억))*억/SHARES_OUTSTANDING)),
     "EBITDA×목표배수 (6/8/10배)"),
    ("RIM 방식", fmt_won(int(BPS*(1+(0.10-0.10)/(0.10-0.02)))),
     fmt_won(int(BPS*(1+(0.12-0.10)/(0.10-0.02)))),
     fmt_won(int(BPS*(1+(0.15-0.10)/(0.10-0.02)))),
     "BPS×(1+(ROE-ke)/(ke-g))"),
    ("FCF 기반", fmt_won(int(FCF_2024*8/SHARES_OUTSTANDING)),
     fmt_won(int(FCF_2024*10/SHARES_OUTSTANDING)),
     fmt_won(int(FCF_2024*14/SHARES_OUTSTANDING)),
     "FCF×목표배수 (8/10/14배)"),
]

for label, cons, base, aggr, note in methods:
    write_row(ws5, row, [label, cons, base, aggr, note],
              fonts=[d_bold, d_font, d_blue, d_green, note_font],
              fills=[lblue_fill, red_fill, gold_fill, green_fill, white_fill],
              aligns=[left, right, right, right, left])
    row += 1

# B. 종합 판단
row += 1
row = section_title(ws5, row, "B. 종합 시나리오", 5)
write_header(ws5, row, ["시나리오", "목표주가", "현재가 대비", "전제조건", "확률(주관)"],
             fills=[PatternFill("solid", fgColor="C0392B"),
                    PatternFill("solid", fgColor="C0392B"),
                    PatternFill("solid", fgColor="C0392B"),
                    PatternFill("solid", fgColor="C0392B"),
                    PatternFill("solid", fgColor="C0392B")])
row += 1

# Bull case
bull_eps = int(TRAIL_EPS * 1.15)  # +15% earnings growth
bull_target = bull_eps * 12
bull_upside = (bull_target - PRICE) / PRICE
write_row(ws5, row, ["강세 (Bull)", fmt_won(bull_target), f"{bull_upside*100:+.1f}%",
                      "갤럭시 업사이클 + 전장 35%성장 + PER 리레이팅", "20%"],
          fonts=[d_bold, Font(name="맑은 고딕", size=12, bold=True, color="27AE60"), d_green, d_font, d_font],
          fills=[green_fill]*5, aligns=[center, right, center, left, center])
row += 1

# Base case
base_target = int(TRAIL_EPS * 10)
base_upside = (base_target - PRICE) / PRICE
write_row(ws5, row, ["기본 (Base)", fmt_won(base_target), f"{base_upside*100:+.1f}%",
                      "모바일 유지 + 전장 10% 성장 + 배당 확대", "50%"],
          fonts=[d_bold, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), d_blue, d_font, d_font],
          fills=[gold_fill]*5, aligns=[center, right, center, left, center])
row += 1

# Bear case
bear_eps = int(E25_CONSERVATIVE_EPS * 0.85)
bear_target = bear_eps * 7
bear_upside = (bear_target - PRICE) / PRICE
write_row(ws5, row, ["약세 (Bear)", fmt_won(bear_target), f"{bear_upside*100:+.1f}%",
                      "삼성 부진 + 전장 정체 + 마진 악화", "30%"],
          fonts=[d_bold, Font(name="맑은 고딕", size=12, bold=True, color="C0392B"), d_red, d_font, d_font],
          fills=[red_fill]*5, aligns=[center, right, center, left, center])
row += 1

# Expected value
exp_val = int(bull_target * 0.20 + base_target * 0.50 + bear_target * 0.30)
exp_upside = (exp_val - PRICE) / PRICE
row += 1
write_row(ws5, row, ["확률가중 기대값", fmt_won(exp_val), f"{exp_upside*100:+.1f}%",
                      "Bull×20% + Base×50% + Bear×30%", ""],
          fonts=[Font(name="맑은 고딕", size=12, bold=True, color=NAVY)] * 5,
          fills=[PatternFill("solid", fgColor="AED6F1")] * 5,
          aligns=[center, right, center, left, center])

# C. Key Metrics Summary Box
row += 2
row = section_title(ws5, row, "C. 핵심 체크포인트", 5)

checkpoints = [
    ("현재 PER 7배", "업종 평균 10~15배 대비 30~50% 할인. 삼성 의존·저마진 디스카운트 반영 중"),
    ("FCF 수익률 9.1%", "시총 대비 연간 잉여현금 비율 매우 높음. 안전마진 확보"),
    ("배당+자사주 5.2%", "2024년 총주주환원율 5.2%. 코스닥 상위 수준"),
    ("2025 실적 호조", "9M 누적 매출 YoY +29%, 영업이익 YoY +69%. Q3 둔화는 계절적 요인"),
    ("전장 성장 모멘텀", "현대차 ADAS 확대. 2025년 전장 매출비중 30%+ 예상"),
    ("BPS 성장", f"BPS 매년 증가 (2020: 14,074원 → 2024: {int(BPS):,}원). 자본축적 지속"),
    ("리스크: Q3 마진", "25Q3 OPM 2.2%로 급락. 일시적 비용인지 구조적 둔화인지 확인 필요"),
    ("리스크: 삼성 편중", "삼성전자 매출비중 70%+. 갤럭시 판매 부진 시 직격탄"),
]

for label, desc in checkpoints:
    ws5.cell(row=row, column=1, value=label).font = d_bold
    ws5.cell(row=row, column=1).fill = lblue_fill
    ws5.cell(row=row, column=1).alignment = left
    ws5.cell(row=row, column=1).border = thin_border
    ws5.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws5.cell(row=row, column=2, value=desc).font = d_font
    ws5.cell(row=row, column=2).alignment = left
    ws5.cell(row=row, column=2).border = thin_border
    row += 1

print("  [5/5] 목표주가 완료")


# === SAVE ===
wb.save(OUT)
print(f"\n밸류에이션 보고서 생성 완료: {OUT}")
print(f"시트: {wb.sheetnames}")
