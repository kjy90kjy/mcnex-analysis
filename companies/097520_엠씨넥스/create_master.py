"""엠씨넥스 마스터 보고서 생성 - 모바일용 제외 전체 통합

종합보고서(12시트) + 투자구루분석(7시트) = 19시트
기업분석보고서/밸류에이션은 종합보고서에 이미 포함되므로 제외.
"""
import copy
import os
from openpyxl import load_workbook

DIR = os.path.dirname(os.path.abspath(__file__))

SOURCES = [
    "엠씨넥스_종합보고서.xlsx",
    "엠씨넥스_투자구루분석.xlsx",
]

OUTPUT = os.path.join(DIR, "엠씨넥스_마스터보고서.xlsx")


def copy_sheet(src_ws, dst_wb, title):
    dst_ws = dst_wb.create_sheet(title=title)
    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column)
            dst_cell.value = cell.value
            if cell.has_style:
                dst_cell.font = copy.copy(cell.font)
                dst_cell.fill = copy.copy(cell.fill)
                dst_cell.border = copy.copy(cell.border)
                dst_cell.alignment = copy.copy(cell.alignment)
                dst_cell.number_format = cell.number_format
                dst_cell.protection = copy.copy(cell.protection)
    for merged in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged))
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
        if dim.hidden:
            dst_ws.column_dimensions[col_letter].hidden = True
    for row_num, dim in src_ws.row_dimensions.items():
        if dim.height:
            dst_ws.row_dimensions[row_num].height = dim.height
        if dim.hidden:
            dst_ws.row_dimensions[row_num].hidden = True
    dst_ws.sheet_properties = copy.copy(src_ws.sheet_properties)
    if src_ws.print_options:
        dst_ws.print_options = copy.copy(src_ws.print_options)
    return dst_ws


def main():
    master_wb = load_workbook(os.path.join(DIR, SOURCES[0]))

    for fname in SOURCES[1:]:
        fpath = os.path.join(DIR, fname)
        if not os.path.exists(fpath):
            print(f"  [SKIP] {fname} 파일 없음")
            continue
        src_wb = load_workbook(fpath)
        for sname in src_wb.sheetnames:
            dst_title = sname
            if dst_title in master_wb.sheetnames:
                suffix = fname.replace("엠씨넥스_", "").replace(".xlsx", "")
                dst_title = f"{sname}_{suffix}"
            print(f"  + {dst_title}")
            copy_sheet(src_wb[sname], master_wb, dst_title)
        src_wb.close()

    master_wb.save(OUTPUT)
    master_wb.close()
    total = len(load_workbook(OUTPUT, read_only=True).sheetnames)
    print(f"\n>>> {os.path.basename(OUTPUT)} 저장 완료 ({total}시트)")


if __name__ == "__main__":
    print("=" * 60)
    print("  엠씨넥스 마스터 보고서 생성")
    print("  종합보고서 + 투자구루분석")
    print("=" * 60)
    main()
